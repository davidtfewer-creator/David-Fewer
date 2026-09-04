"""
build_9stock_performance.py — companion performance workbook (user, 4 Sep 2026;
r2 same day: MIN plan includes the pooling credit, capital-additions column,
EOY projection block).

Creates 9stock_performance.xlsx: a SEPARATE workbook that reads the live
trading workbook (a mirror that must never be modified) by external links and
compares realised P&L against a planned-return band.

Performance sheet
  - assumptions (editable): starting capital (match the trading book's
    Performance!C4); MIN planned return 74%/yr (the settled plan — anchors x
    execution x pooling at its weaker half); MAX planned return 88.5%/yr (the
    pooled full-sample ceiling reference).
  - CAPITAL ADDED column (editable): money paid into IBKR during a week is
    typed against that week ending. Additions compound into BOTH plan lines
    from that week on (plan equity path: (prev + addition) x (1 + weekly
    rate)), and into invested capital / fund size. Actual P&L needs no
    adjustment — the blotter records dollars either way.
  - Table 1: weekly + cumulative P&L as recorded (blotter + closed
    discretionary, Active Trading cols P and R) vs cumulative MIN/MAX plan.
  - Table 2: the same EXCLUDING the discretionary log.
  - summary block: invested capital to date, current equity, projected
    end-of-year profit and year-end fund size (current equity compounded at
    the MIN and MAX plan rates over the remaining weeks; future additions
    join the projection when their week arrives).
  - two charts: cumulative actual vs the plan band, all trading / model-only.

Feed sheet
  - direct external references to '[1]Active Trading' (O/P/R 42:93, disc log
    AD/AE 42:241). Direct cell references are the ONE link type Excel
    refreshes from a CLOSED source — SUMIFS/INDIRECT against closed workbooks
    fail, so raw columns are mirrored here and everything computes locally.

The external-link plumbing (xl/externalLinks/*, workbook externalReferences,
content types) is injected into the saved zip because openpyxl does not
manage external workbooks; the link cache is seeded with the source's own
Excel-cached values so the file shows real numbers before the first refresh.
The link target is the source's bare filename — the pair works in any folder
as long as the two files sit together (Data -> Edit Links -> Change Source if
the live file is renamed).

Usage:
    python build_9stock_performance.py <source_trading_book.xlsx> <out.xlsx>
The source is opened READ-ONLY; it is never written.
"""
import datetime as dt
import os
import re
import shutil
import sys
import zipfile

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY, GOLD, TEAL, GREY = 'FF182644', 'FFC6A04A', 'FF176E78', 'FF5F5F5F'
EDIT_FILL = PatternFill(fill_type='solid', start_color='FFF6ECD8',
                        end_color='FFF6ECD8')
EPOCH = dt.date(1899, 12, 30)
WK_ROWS = range(42, 94)          # weekly log rows in Active Trading
DISC_ROWS = range(42, 242)       # discretionary log rows
NROWS = len(WK_ROWS)             # 52
R0, R1 = 11, 11 + NROWS - 1      # table rows in this workbook


def serial(v):
    if isinstance(v, dt.datetime):
        v = v.date()
    return (v - EPOCH).days


def read_cache(src_path):
    wb = openpyxl.load_workbook(src_path, data_only=True)
    at = wb['Active Trading']
    cache = {}
    for col in ('O', 'P', 'R'):
        for r in WK_ROWS:
            cache[f'{col}{r}'] = at[f'{col}{r}'].value
    for col in ('AD', 'AE'):
        for r in DISC_ROWS:
            cache[f'{col}{r}'] = at[f'{col}{r}'].value
    return cache


def build(src_name, out_path, start_cap):
    wb = openpyxl.Workbook()
    pf = wb.active
    pf.title = 'Performance'
    fd = wb.create_sheet('Feed')
    hdr_fill = PatternFill(fill_type='solid', start_color=NAVY, end_color=NAVY)
    hdr_font = Font(bold=True, color='FFFFFFFF')
    note = Font(italic=True, size=9, color=GREY)

    # ---------------- Feed: direct external references only
    fd['A1'], fd['B1'], fd['C1'] = 'Week ending', 'Weekly P&L', 'Cumulative'
    fd['E1'], fd['F1'] = 'Disc net P&L', 'Disc week ending'
    for c in ('A1', 'B1', 'C1', 'E1', 'F1'):
        fd[c].font = Font(bold=True, color=NAVY)
    for i, r_src in enumerate(WK_ROWS):
        r = 2 + i
        fd.cell(row=r, column=1, value=f"='[1]Active Trading'!O{r_src}").number_format = 'dd/mm/yyyy'
        fd.cell(row=r, column=2, value=f"='[1]Active Trading'!P{r_src}")
        fd.cell(row=r, column=3, value=f"='[1]Active Trading'!R{r_src}")
    for i, r_src in enumerate(DISC_ROWS):
        r = 2 + i
        fd.cell(row=r, column=5, value=f"='[1]Active Trading'!AD{r_src}")
        f = fd.cell(row=r, column=6, value=f"='[1]Active Trading'!AE{r_src}")
        f.number_format = 'dd/mm/yyyy'
    fd['H2'] = ('Direct external references to the trading workbook (same folder). '
                'Do not edit; everything downstream computes on the Performance sheet.')
    fd['H2'].font = note

    # ---------------- Performance sheet
    t = pf.cell(row=1, column=1, value='BAYESIAN CAPITAL — 9-STOCK BOOK vs PLAN BAND')
    t.font = Font(bold=True, size=13, color=NAVY)
    pf.cell(row=2, column=1, value=(
        f'Feeds from {src_name} in the same folder (never modified). Actual = weekly '
        'realised P&L as recorded (blotter + closed discretionary); the second table '
        'strips the discretionary log. Type IBKR capital additions in the shaded '
        'column — they compound into both plan lines from that week. Enable link '
        'updates when Excel asks.')).font = Font(italic=True, color=GREY)

    pf.cell(row=4, column=2, value='Starting capital').font = Font(bold=True)
    c = pf.cell(row=4, column=3, value=start_cap)
    c.number_format = '$#,##0'
    c.fill = EDIT_FILL
    pf.cell(row=5, column=2, value='MIN planned return (p.a.)').font = Font(bold=True)
    c = pf.cell(row=5, column=3, value=0.74)
    c.number_format = '0.0%'
    c.fill = EDIT_FILL
    pf.cell(row=6, column=2, value='MAX planned return (p.a.)').font = Font(bold=True)
    c = pf.cell(row=6, column=3, value=0.885)
    c.number_format = '0.0%'
    c.fill = EDIT_FILL
    pf.cell(row=7, column=2, value='Weekly rate MIN / MAX').font = Font(bold=True)
    pf.cell(row=7, column=3, value='=(1+$C$5)^(1/52)-1').number_format = '0.000%'
    pf.cell(row=7, column=4, value='=(1+$C$6)^(1/52)-1').number_format = '0.000%'
    pf.cell(row=8, column=2, value=(
        'MIN = the plan (anchors × execution × pooling, 74%/yr); MAX = pooled '
        'full-sample ceiling (88.5%/yr). Match starting capital to the trading '
        'workbook, Performance!C4. Shaded cells are yours to edit.')).font = note

    # summary block
    pf.cell(row=4, column=6, value='Invested capital to date').font = Font(bold=True, color=NAVY)
    pf.cell(row=4, column=8,
            value=f'=$C$4+SUMIFS($C${R0}:$C${R1},$B${R0}:$B${R1},"<="&TODAY()+6)')
    pf.cell(row=5, column=6, value='Equity now (capital + P&L)').font = Font(bold=True, color=NAVY)
    pf.cell(row=5, column=8, value=f'=$H$4+IFERROR(LOOKUP(9.99E+307,$E${R0}:$E${R1}),0)')
    pf.cell(row=6, column=6, value='Projected EOY profit  (MIN | MAX)').font = Font(bold=True, color=NAVY)
    proj = (f'=IFERROR(LOOKUP(9.99E+307,$E${R0}:$E${R1}),0)'
            f'+$H$5*((1+{{rate}})^(52-COUNT($D${R0}:$D${R1}))-1)')
    pf.cell(row=6, column=8, value=proj.format(rate='$C$7'))
    pf.cell(row=6, column=9, value=proj.format(rate='$D$7'))
    pf.cell(row=7, column=6, value='Projected year-end fund size').font = Font(bold=True, color=NAVY)
    pf.cell(row=7, column=8, value='=$H$4+$H$6')
    pf.cell(row=7, column=9, value='=$H$4+$I$6')
    # extrapolation at the book's own run-rate: realized weekly growth to date,
    # (equity / invested)^(1/weeks elapsed) - 1, compounded over the rest of the 52
    live = f'(($H$5/$H$4)^(1/COUNT($D${R0}:$D${R1}))-1)'
    pf.cell(row=8, column=6, value='At current run-rate  (profit | fund)').font = \
        Font(bold=True, color=NAVY)
    pf.cell(row=8, column=8, value='=' + proj.format(rate=live)[1:])
    pf.cell(row=8, column=9, value='=$H$4+$H$8')
    j8 = pf.cell(row=8, column=10,
                 value=f'=IFERROR("pace "&TEXT((1+{live})^52-1,"0%")&"/yr","")')
    j8.font = note
    for addr in ('H4', 'H5', 'H6', 'I6', 'H7', 'I7', 'H8', 'I8'):
        pf[addr].number_format = '$#,##0'
        pf[addr].font = Font(bold=True, color=NAVY)
    pf.cell(row=9, column=6, value=(
        'Projections: current equity compounded over the remaining weeks of the 52 — '
        'at the plan rates, and (last row) at the book\'s own realized weekly growth '
        'to date. Early in the year the run-rate is noisy; it firms up as weeks '
        'accumulate. Future additions typed below lift the plan lines now and join '
        'the projections once their week arrives.')).font = note

    heads = ['Week #', 'Week ending', 'Capital added', 'Weekly P&L',
             'Cumulative P&L', 'Cum plan MIN', 'Cum plan MAX', 'vs MIN', 'vs MAX',
             '', 'Weekly P&L excl disc', 'Cumulative excl disc',
             'excl disc vs MIN', 'excl disc vs MAX', '',
             'plan equity MIN (helper)', 'plan equity MAX (helper)']
    for j, h in enumerate(heads, start=1):
        if not h:
            continue
        cell = pf.cell(row=10, column=j, value=h)
        if j >= 16:
            cell.font = note
        else:
            cell.fill = hdr_fill
            cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    widths = [7, 12, 13, 13, 14, 14, 14, 12, 12, 2, 15, 15, 13, 13, 2, 13, 13]
    for j, w in enumerate(widths, start=1):
        pf.column_dimensions[get_column_letter(j)].width = w

    for i in range(NROWS):
        r, rf = R0 + i, 2 + i
        pf.cell(row=r, column=1, value=f'=IF($B{r}="","",ROW()-10)')
        b = pf.cell(row=r, column=2,
                    value=f'=IF(OR(Feed!A{rf}="",Feed!A{rf}=0),"",Feed!A{rf})')
        b.number_format = 'dd/mm/yyyy'
        cadd = pf.cell(row=r, column=3)                    # editable additions
        cadd.fill = EDIT_FILL
        pf.cell(row=r, column=4,
                value=f'=IF(OR($B{r}="",$B{r}>TODAY()+6),"",N(Feed!B{rf}))')
        pf.cell(row=r, column=5, value=f'=IF($D{r}="","",N(Feed!C{rf}))')
        base_min = '$C$4' if r == R0 else f'P{r - 1}'
        base_max = '$C$4' if r == R0 else f'Q{r - 1}'
        pf.cell(row=r, column=16,
                value=f'=IF($B{r}="","",({base_min}+N($C{r}))*(1+$C$7))')
        pf.cell(row=r, column=17,
                value=f'=IF($B{r}="","",({base_max}+N($C{r}))*(1+$D$7))')
        pf.cell(row=r, column=6,
                value=f'=IF($B{r}="","",P{r}-$C$4-SUM($C${R0}:$C{r}))')
        pf.cell(row=r, column=7,
                value=f'=IF($B{r}="","",Q{r}-$C$4-SUM($C${R0}:$C{r}))')
        pf.cell(row=r, column=8, value=f'=IF(OR($E{r}="",$F{r}=""),"",$E{r}-$F{r})')
        pf.cell(row=r, column=9, value=f'=IF(OR($E{r}="",$G{r}=""),"",$E{r}-$G{r})')
        pf.cell(row=r, column=11,
                value=(f'=IF($D{r}="","",$D{r}'
                       f'-SUMIFS(Feed!$E$2:$E$201,Feed!$F$2:$F$201,$B{r}))'))
        pf.cell(row=r, column=12, value=f'=IF($K{r}="","",SUM($K${R0}:$K{r}))')
        pf.cell(row=r, column=13, value=f'=IF(OR($L{r}="",$F{r}=""),"",$L{r}-$F{r})')
        pf.cell(row=r, column=14, value=f'=IF(OR($L{r}="",$G{r}=""),"",$L{r}-$G{r})')
        for j in (3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14, 16, 17):
            pf.cell(row=r, column=j).number_format = '$#,##0'

    def mk_chart(title, val_col):
        ch = LineChart()
        ch.title = title
        ch.style = 2
        ch.height, ch.width = 9, 16
        cats = Reference(pf, min_col=2, min_row=R0, max_row=R1)
        for col, rgb, wd, dash in ((val_col, TEAL[2:], 28000, None),
                                   (6, GREY[2:], 16000, 'dash'),
                                   (7, GOLD[2:], 16000, 'dash')):
            ref = Reference(pf, min_col=col, min_row=10, max_row=R1)
            ch.add_data(ref, titles_from_data=True)
            s = ch.series[-1]
            s.graphicalProperties.line.solidFill = rgb
            s.graphicalProperties.line.width = wd
            if dash:
                s.graphicalProperties.line.dashStyle = dash
        ch.set_categories(cats)
        ch.y_axis.numFmt = '$#,##0'
        return ch

    pf.add_chart(mk_chart('Cumulative P&L vs plan band — all trading', 5),
                 f'A{R1 + 3}')
    pf.add_chart(mk_chart('Cumulative P&L vs plan band — excluding discretionary', 12),
                 f'K{R1 + 3}')

    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)


# ---------------- external-link injection (openpyxl does not manage these)
CT_EL = ('application/vnd.openxmlformats-officedocument.spreadsheetml.'
         'externalLink+xml')
NS_MAIN = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
REL_EL = f'{NS_R}/externalLink'
REL_ELPATH = f'{NS_R}/externalLinkPath'


def cached_xml(cache):
    by_row = {}
    for ref, v in cache.items():
        if v is None:
            continue
        m = re.match(r'([A-Z]+)(\d+)', ref)
        by_row.setdefault(int(m.group(2)), []).append((ref, v))
    rows = []
    for rn in sorted(by_row):
        cells = []
        for ref, v in sorted(by_row[rn]):
            if isinstance(v, (dt.datetime, dt.date)):
                cells.append(f'<cell r="{ref}"><v>{serial(v)}</v></cell>')
            elif isinstance(v, (int, float)):
                cells.append(f'<cell r="{ref}"><v>{v!r}</v></cell>')
            else:
                s = str(v).replace('&', '&amp;').replace('<', '&lt;')
                cells.append(f'<cell r="{ref}" t="str"><v>{s}</v></cell>')
        rows.append(f'<row r="{rn}">{"".join(cells)}</row>')
    return ''.join(rows)


def inject_links(path, src_name, cache):
    tmp = path + '.tmp'
    zin = zipfile.ZipFile(path)
    zout = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename == '[Content_Types].xml':
            data = data.replace(
                b'</Types>',
                f'<Override PartName="/xl/externalLinks/externalLink1.xml" '
                f'ContentType="{CT_EL}"/></Types>'.encode())
        elif item.filename == 'xl/workbook.xml':
            data = data.replace(
                b'</sheets>',
                b'</sheets><externalReferences><externalReference '
                b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/'
                b'2006/relationships" r:id="rIdEL1"/></externalReferences>')
        elif item.filename == 'xl/_rels/workbook.xml.rels':
            data = data.replace(
                b'</Relationships>',
                f'<Relationship Id="rIdEL1" Type="{REL_EL}" '
                f'Target="externalLinks/externalLink1.xml"/></Relationships>'.encode())
        zout.writestr(item, data)
    zout.writestr(
        'xl/externalLinks/externalLink1.xml',
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<externalLink xmlns="{NS_MAIN}" xmlns:r="{NS_R}">'
        f'<externalBook r:id="rId1">'
        f'<sheetNames><sheetName val="Active Trading"/></sheetNames>'
        f'<sheetDataSet><sheetData sheetId="0">{cached_xml(cache)}</sheetData>'
        f'</sheetDataSet></externalBook></externalLink>')
    zout.writestr(
        'xl/externalLinks/_rels/externalLink1.xml.rels',
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
        f'relationships"><Relationship Id="rId1" Type="{REL_ELPATH}" '
        f'Target="{src_name}" TargetMode="External"/></Relationships>')
    zout.close()
    zin.close()
    shutil.move(tmp, path)


def main():
    src, out = sys.argv[1], sys.argv[2]
    src_name = re.sub(r'^[0-9a-f]{8}-', '', os.path.basename(src))
    cache = read_cache(src)
    wbp = openpyxl.load_workbook(src, data_only=True)
    start_cap = wbp['Performance']['C4'].value or 5_000_000
    build(src_name, out, start_cap)
    inject_links(out, src_name, cache)

    import xml.dom.minidom
    z = zipfile.ZipFile(out)
    assert z.testzip() is None
    for part in ('xl/externalLinks/externalLink1.xml',
                 'xl/externalLinks/_rels/externalLink1.xml.rels',
                 'xl/workbook.xml', '[Content_Types].xml'):
        xml.dom.minidom.parseString(z.read(part))
    z.close()
    wb2 = openpyxl.load_workbook(out)
    assert wb2.sheetnames == ['Performance', 'Feed']
    assert wb2['Feed']['A2'].value == "='[1]Active Trading'!O42"
    print(f'built {out} (links -> {src_name}, start capital {start_cap:,.0f}); '
          f'validation passed')


if __name__ == '__main__':
    main()
