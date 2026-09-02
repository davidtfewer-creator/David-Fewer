"""
backfill_query.py — fill missing daily OHLC in the workbook's Query sheet
from IBKR (the interface Scripts 1 and 2 already use).

Built for the RKLB -> AVGO swap (2 Sep 2026): the AVGO history wired into
Query runs to 3 Aug 2026, leaving 4-31 Aug empty. This fills any EMPTY
OHLC cells for a ticker whose date row already exists — it never overwrites
a filled cell and never adds rows (new days remain Script 1's job).

Usage (trading machine, TWS/Gateway logged in, workbook CLOSED in Excel):

    python backfill_query.py "C:\\path\\to\\TradingExcel.xlsx" AVGO
    python backfill_query.py book.xlsx AVGO --dry-run     # print, write nothing

The ticker's columns are located by the Query header row (e.g. AVGO_O..AVGO_C),
so this works for any name without edits. Requires: pip install openpyxl
ib_insync. Saves in place after a .bak copy beside the workbook.
"""
import argparse
import datetime as dt
import shutil
import sys

import openpyxl

HOST, PORT, CLIENT_ID = '127.0.0.1', 7496, 31   # 7497 for paper TWS
EPOCH = dt.date(1899, 12, 30)


def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, (int, float)):
        return EPOCH + dt.timedelta(days=int(v))
    return None


def fetch_daily(ticker, days_back):
    from ib_insync import IB, Stock
    ib = IB()
    ib.connect(HOST, PORT, clientId=CLIENT_ID, timeout=20)
    c = Stock(ticker, 'SMART', 'USD')
    ib.qualifyContracts(c)
    bars = ib.reqHistoricalData(c, endDateTime='', durationStr=f'{days_back} D',
                                barSizeSetting='1 day', whatToShow='TRADES',
                                useRTH=True, formatDate=1)
    ib.disconnect()
    return {b.date: (b.open, b.high, b.low, b.close) for b in bars}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('workbook')
    ap.add_argument('ticker')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    t = args.ticker.upper()

    wb = openpyxl.load_workbook(args.workbook)
    q = wb['Query']
    heads = {q.cell(row=1, column=c).value: c for c in range(1, q.max_column + 1)}
    try:
        cols = [heads[f'{t}_{k}'] for k in 'OHLC']
    except KeyError:
        sys.exit(f'{t}_O..{t}_C headers not found in Query — is {t} in the book?')

    empty = []
    for r in range(2, q.max_row + 1):
        d = as_date(q.cell(row=r, column=1).value)
        if d and all(q.cell(row=r, column=c).value is None for c in cols):
            empty.append((r, d))
    if not empty:
        print(f'no empty {t} rows in Query — nothing to do.')
        return
    span = (dt.date.today() - empty[0][1]).days + 7
    print(f'{len(empty)} empty {t} rows ({empty[0][1]} .. {empty[-1][1]}); '
          f'fetching ~{span} days from IBKR...')
    bars = fetch_daily(t, span)

    filled, missing = 0, []
    for r, d in empty:
        if d in bars:
            for c, v in zip(cols, bars[d]):
                q.cell(row=r, column=c).value = round(float(v), 4)
            filled += 1
        else:
            missing.append(d)
    print(f'filled {filled}/{len(empty)} rows' +
          (f'; IBKR returned nothing for: {", ".join(map(str, missing))}' if missing else ''))

    if args.dry_run:
        print('[dry-run] not saved.')
        return
    shutil.copy2(args.workbook, args.workbook + '.bak')
    wb.save(args.workbook)
    print(f'saved {args.workbook} (backup: {args.workbook}.bak). Open in Excel '
          f'once so formulas recalculate before the next trading morning.')


if __name__ == '__main__':
    main()
