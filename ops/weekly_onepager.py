"""
weekly_onepager.py — the Sunday one-pager (Book Operations Manual, section 4).

Reads the live trading workbook and writes a single-page branded HTML report
for the week just ended:

  1. Scorecard        — weekly and cumulative realised P&L against the plan
                        line (compounding target from the Performance tab
                        assumptions, C4 starting capital / C5 target rate),
                        book equity vs plan equity, ahead/behind.
  2. Activity         — fills and exits this week by name; exits classified
                        (target/gap vs 50-day stop by hold length; losses
                        flagged); discretionary trades listed separately.
  3. Gates            — 200dma breadth count for EACH day of the week
                        (reconstructed from the Query sheet's close history,
                        the same trailing-200 average the sheet uses) with
                        gate-armed days marked against the K threshold in
                        Active Trading M8; ATH proximity per name (cap
                        territory); MU pause status from the next-earnings
                        date in Active Trading G11; earnings inside 21 days.
  4. Open positions   — every OPEN blotter row with its age against the
                        50-day stop; discretionary opens listed below.
  5. Hold tracker     — live average hold per name vs the back-test
                        reference (mirrors the Performance tab table).
  6. Checks           — Query freshness, OHLC gaps in the report week (the
                        VRT-style dropout), blotter integrity, weekly-log
                        capacity, and the Performance-tab recording reminder.

Everything is recomputed from TYPED cells (blotter entries, Query history,
assumption cells) — no dependence on Excel-cached formula values, so it works
on any copy of the workbook, saved by Excel or not. The two deliberate
approximations are labelled on the page: ATH is the maximum high in the Query
window (the model's own ATH lives in the Model sheets), and pre-market vetoes
are not reconstructable because the sheet keeps no 9:00 history.

Usage (trading machine, Sunday, after the Friday close is in the workbook):

    python weekly_onepager.py "C:\\path\\to\\TradingExcel.xlsx"
    python weekly_onepager.py book.xlsx --week 2026-08-28   # a past week
    python weekly_onepager.py book.xlsx --out report.html

Schedule (Windows Task Scheduler): Program pythonw, weekly, Sunday 08:00.
Requires: pip install openpyxl. Output lands beside the workbook as
BC_onepager_<week-ending>.html — print or archive it as-is.
"""
import argparse
import datetime as dt
import html
import os
import statistics
import sys

import openpyxl

# slot order matches the Query sheet's 4-column blocks (AVGO replaced RKLB
# in slot 4 on 2 Sep 2026; hold references re-run with AVGO in the book)
NAMES = ['TSM', 'VRT', 'VST', 'AVGO', 'MU', 'GM', 'VLO', 'CF', 'MRVL']
HIST_HOLD = {'TSM': 4.2, 'VRT': 4.4, 'VST': 4.4, 'AVGO': 6.5, 'MU': 6.5,
             'GM': 6.2, 'VLO': 14.1, 'CF': 16.7, 'MRVL': 7.3, 'BOOK': 6.5}
STOP_DAYS = 50
DMA_WIN = 200            # trailing closes, inclusive — matches AT col N
ATH_NEAR = 0.02          # "cap territory" when close is within 2% of ATH
EXCEL_EPOCH = dt.date(1899, 12, 30)

NAVY, GOLD, TEAL, GREY, RED = '#182644', '#C6A04A', '#176E78', '#5F5F5F', '#A03C32'
PAPER = '#F7F7F4'


def as_date(v):
    """Workbook date cell -> date. Handles datetime, date and Excel serials."""
    if v is None or v == '':
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)):
        return EXCEL_EPOCH + dt.timedelta(days=int(v))
    return None


def friday_of(d):
    """The sheet's week-ending convention: H - WEEKDAY(H,2) + 5."""
    return d - dt.timedelta(days=d.isoweekday()) + dt.timedelta(days=5)


def money(x, signed=False):
    if x is None:
        return '—'
    s = '+' if (signed and x >= 0) else ('−' if x < 0 else '')
    return f'{s}${abs(x):,.0f}'


def read_book(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    at = wb['Active Trading']
    buy_fee = float(at['B3'].value or 0.005)
    sell_fee = float(at['E3'].value or 0.0095)

    trades = []                       # model blotter
    for r in range(42, 1042):
        stock = at.cell(row=r, column=2).value
        if not stock:
            continue
        buy_d = as_date(at.cell(row=r, column=4).value)
        sell_d = as_date(at.cell(row=r, column=8).value)
        px, sh = at.cell(row=r, column=5).value, at.cell(row=r, column=6).value
        spx = at.cell(row=r, column=9).value
        t = dict(row=r, stock=str(stock).strip(), tranche=str(at.cell(row=r, column=3).value or ''),
                 buy=buy_d, px=px, sh=sh, sell=sell_d, spx=spx, problems=[])
        if buy_d is None or px is None or sh is None:
            t['problems'].append('missing buy date/price/shares')
        else:
            t['cost'] = sh * px + sh * buy_fee
        if sell_d is not None:
            if spx is None:
                t['problems'].append('CLOSED but no sell price')
            elif 'cost' in t:
                t['pnl'] = sh * spx - sh * sell_fee - t['cost']
                t['week'] = friday_of(sell_d)
                t['hold'] = (sell_d - buy_d).days
        trades.append(t)

    disc = []                         # discretionary log
    for r in range(42, 242):
        stock = at.cell(row=r, column=21).value          # U
        if not stock:
            continue
        buy_d = as_date(at.cell(row=r, column=23).value)  # W
        sell_d = as_date(at.cell(row=r, column=27).value)  # AA
        px, sh = at.cell(row=r, column=24).value, at.cell(row=r, column=25).value  # X, Y
        spx = at.cell(row=r, column=28).value             # AB
        d = dict(stock=str(stock).strip(), buy=buy_d, px=px, sh=sh, sell=sell_d, spx=spx)
        if None not in (buy_d, px, sh):
            d['cost'] = sh * px + sh * buy_fee
            if sell_d is not None and spx is not None:
                d['pnl'] = sh * spx - sh * sell_fee - d['cost']
                d['week'] = friday_of(sell_d)
        disc.append(d)

    pf = wb['Performance']
    start_cap = float(pf['C4'].value or 6_150_000)
    target = float(pf['C5'].value or 0.80)

    gates = dict(pm_pct=at['M6'].value, ath_eps=at['M7'].value, breadth_k=at['M8'].value)
    earn = {}                          # AT G7:G15 next-earnings, typed by hand
    for i, nm in enumerate(NAMES):
        d = as_date(at.cell(row=7 + i, column=7).value)
        if d:
            earn[nm] = d

    q = wb['Query']
    days, series = [], {nm: dict(o=[], h=[], l=[], c=[]) for nm in NAMES}
    for r in range(2, q.max_row + 1):
        d = as_date(q.cell(row=r, column=1).value)
        if d is None:
            continue
        days.append(d)
        for i, nm in enumerate(NAMES):
            base = 2 + i * 4
            for j, k in enumerate('ohlc'):
                v = q.cell(row=r, column=base + j).value
                series[nm][k].append(float(v) if isinstance(v, (int, float)) else None)

    cash = wb['Allocation']['B5'].value
    return dict(trades=trades, disc=disc, start_cap=start_cap, target=target,
                gates=gates, earn=earn, qdays=days, q=series, cash=cash,
                fees=(buy_fee, sell_fee))


def weekly_series(book, report_week):
    closed = [t for t in book['trades'] + book['disc'] if 'pnl' in t]
    if not closed:
        return [], {}, 0.0
    first = min(t['week'] for t in closed)
    weeks = []
    w = first
    while w <= report_week:
        weeks.append(w)
        w += dt.timedelta(days=7)
    actual = {w: sum(t['pnl'] for t in closed if t['week'] == w) for w in weeks}
    wk_rate = (1 + book['target']) ** (1 / 52) - 1
    return weeks, actual, wk_rate


def breadth_by_day(book, report_week):
    """Per trading day of the report week: names below their trailing-200dma."""
    days, q = book['qdays'], book['q']
    mon = report_week - dt.timedelta(days=4)
    out, gaps = [], set()
    for idx, d in enumerate(days):
        if not (mon <= d <= report_week):
            continue
        below = []
        for nm in NAMES:
            closes = q[nm]['c']
            if closes[idx] is None:
                gaps.add((d, nm))
                continue
            hist = [c for c in closes[:idx + 1] if c is not None][-DMA_WIN:]
            if len(hist) >= DMA_WIN and closes[idx] < statistics.fmean(hist):
                below.append(nm)
        out.append((d, below))
    return out, sorted(gaps)


def ath_proximity(book, report_week):
    days, q = book['qdays'], book['q']
    upto = [i for i, d in enumerate(days) if d <= report_week]
    if not upto:
        return []
    last = upto[-1]
    rows = []
    for nm in NAMES:
        highs = [h for h in q[nm]['h'][:last + 1] if h is not None]
        close = next((q[nm]['c'][i] for i in reversed(upto) if q[nm]['c'][i] is not None), None)
        if not highs or close is None:
            continue
        ath = max(highs)
        rows.append((nm, close, ath, (ath - close) / ath))
    return sorted(rows, key=lambda r: r[3])


def mu_pause(book, report_week):
    d = book['earn'].get('MU')
    if d is None:
        return None, ('No MU report date entered (Active Trading G11) — enter it '
                      'as soon as MU confirms; the pause window derives from it.')
    monday = d - dt.timedelta(days=d.isoweekday() - 1)
    lo, hi = monday - dt.timedelta(days=7), monday + dt.timedelta(days=13)
    if lo <= report_week <= hi:
        state = f'ACTIVE — no new MU bids until {hi:%d %b %Y}'
    elif report_week < lo:
        state = f'clear — window opens {lo:%d %b %Y} (report {d:%d %b %Y})'
    else:
        state = f'clear — last window closed {hi:%d %b %Y}'
    return (lo, hi), state


CSS = f"""
@page {{ size: A4; margin: 10mm; }}
* {{ box-sizing: border-box; }}
body {{ font-family: 'Source Sans 3', 'Segoe UI', Arial, sans-serif; color: {NAVY};
       background: {PAPER}; margin: 0; padding: 18px 26px; font-size: 12.5px; }}
h1 {{ font-family: Spectral, Georgia, serif; font-size: 21px; margin: 0; letter-spacing: .2px; }}
h2 {{ font-size: 11px; text-transform: uppercase; letter-spacing: 1.4px; color: {TEAL};
     border-bottom: 1.5px solid {TEAL}; padding-bottom: 3px; margin: 16px 0 6px; }}
.sub {{ color: {GREY}; margin-top: 2px; font-size: 12px; }}
.rule {{ height: 3px; background: linear-gradient(90deg, {NAVY}, {TEAL} 60%, {GOLD}); margin: 10px 0 2px; }}
.cards {{ display: flex; gap: 10px; margin-top: 8px; }}
.card {{ flex: 1; background: #fff; border: 1px solid #e4e2da; border-top: 3px solid {GOLD};
        padding: 8px 10px; }}
.card .k {{ font-size: 10px; text-transform: uppercase; letter-spacing: 1px; color: {GREY}; }}
.card .v {{ font-size: 17px; font-weight: 700; margin-top: 2px; font-variant-numeric: tabular-nums; }}
.card .d {{ font-size: 11px; color: {GREY}; margin-top: 2px; }}
table {{ border-collapse: collapse; width: 100%; background: #fff; }}
th {{ background: {NAVY}; color: #fff; font-size: 10.5px; text-transform: uppercase;
     letter-spacing: .6px; padding: 4px 7px; text-align: left; }}
td {{ padding: 3.5px 7px; border-bottom: 1px solid #ecebe4; font-variant-numeric: tabular-nums; }}
tr:last-child td {{ border-bottom: none; }}
.num {{ text-align: right; }}
.pos {{ color: {TEAL}; font-weight: 600; }}
.neg {{ color: {RED}; font-weight: 600; }}
.warn {{ color: {RED}; font-weight: 600; }}
.ok {{ color: {TEAL}; }}
.tag {{ display: inline-block; padding: 1px 7px; border-radius: 9px; font-size: 10.5px;
       font-weight: 700; letter-spacing: .4px; }}
.tag.on {{ background: {RED}; color: #fff; }}
.tag.off {{ background: #e7ede9; color: {TEAL}; }}
.cols {{ display: flex; gap: 16px; align-items: flex-start; }}
.cols > div {{ flex: 1; min-width: 0; }}
.note {{ font-size: 11px; color: {GREY}; font-style: italic; margin-top: 4px; }}
ul.checks {{ margin: 4px 0 0; padding-left: 18px; }}
ul.checks li {{ margin: 2px 0; }}
.foot {{ margin-top: 14px; border-top: 1px solid #d9d7cd; padding-top: 6px;
        font-size: 10.5px; color: {GREY}; display: flex; justify-content: space-between; }}
"""


def build_html(book, report_week, wb_name):
    weeks, actual, wk_rate = weekly_series(book, report_week)
    cap, target = book['start_cap'], book['target']
    n = len(weeks)
    cum_act = sum(actual.values())
    cum_plan = cap * ((1 + wk_rate) ** n - 1)
    wk_act = actual.get(report_week, 0.0)
    wk_plan = cap * (1 + wk_rate) ** (n - 1) * wk_rate if n else 0.0
    equity = cap + cum_act
    gap = cum_act - cum_plan

    mon = report_week - dt.timedelta(days=4)
    buys = [t for t in book['trades'] if t.get('buy') and mon <= t['buy'] <= report_week]
    exits = [t for t in book['trades'] if t.get('week') == report_week]
    stops = [t for t in exits if t.get('hold', 0) >= STOP_DAYS]
    losses = [t for t in exits if t.get('pnl', 0) < 0]
    opens = [t for t in book['trades'] if t.get('buy') and t['sell'] is None]
    disc_wk = [d for d in book['disc'] if d.get('week') == report_week]
    disc_open = [d for d in book['disc'] if d.get('buy') and d['sell'] is None]

    breadth, qgaps = breadth_by_day(book, report_week)
    K = book['gates']['breadth_k']
    ath = ath_proximity(book, report_week)
    _, mu_state = mu_pause(book, report_week)
    today = dt.date.today()

    def cls(x):
        return 'pos' if x >= 0 else 'neg'

    # -- activity by name (former book names, e.g. RKLB before the swap,
    #    still show for weeks in which their blotter rows closed)
    extras = sorted({t['stock'] for t in buys + exits + opens} - set(NAMES))
    act_rows = []
    for nm in NAMES + extras:
        b = [t for t in buys if t['stock'] == nm]
        e = [t for t in exits if t['stock'] == nm]
        s = [t for t in e if t in stops]
        pnl = sum(t['pnl'] for t in e)
        o = [t for t in opens if t['stock'] == nm]
        if not (b or e or o):
            continue
        act_rows.append(
            f"<tr><td><b>{nm}</b></td><td class='num'>{len(b) or '—'}</td>"
            f"<td class='num'>{money(sum(t.get('cost', 0) for t in b)) if b else '—'}</td>"
            f"<td class='num'>{len(e) or '—'}</td>"
            f"<td class='num'>{len(s) or '—'}</td>"
            f"<td class='num {cls(pnl)}'>{money(pnl, signed=True) if e else '—'}</td>"
            f"<td class='num'>{len(o) or '—'}</td></tr>")

    # -- breadth table
    br_rows = []
    for d, below in breadth:
        armed = isinstance(K, (int, float)) and len(below) >= K
        tag = f"<span class='tag on'>GATE ON</span>" if armed else "<span class='tag off'>off</span>"
        br_rows.append(f"<tr><td>{d:%a %d %b}</td><td class='num'>{len(below)}</td>"
                       f"<td>{', '.join(below) or '—'}</td><td>{tag}</td></tr>")
    armed_days = sum(1 for d, below in breadth
                     if isinstance(K, (int, float)) and len(below) >= K)

    # -- ATH proximity (show the four tightest)
    ath_rows = []
    for nm, close, a, pct in ath[:4]:
        flag = " <span class='warn'>cap territory</span>" if pct <= ATH_NEAR else ''
        ath_rows.append(f"<tr><td><b>{nm}</b></td><td class='num'>{close:,.2f}</td>"
                        f"<td class='num'>{a:,.2f}</td>"
                        f"<td class='num'>{pct * 100:.1f}%{flag}</td></tr>")

    # -- open positions
    op_rows = []
    for t in sorted(opens, key=lambda t: t['buy']):
        age = (today - t['buy']).days
        left = STOP_DAYS - age
        state = (f"<span class='warn'>STOP DUE</span>" if left <= 0 else
                 f"<span class='warn'>{left}d left</span>" if left <= 7 else f'{left}d left')
        op_rows.append(f"<tr><td><b>{t['stock']}</b></td><td>{t['tranche']}</td>"
                       f"<td>{t['buy']:%d %b}</td><td class='num'>{t['px']:,.2f}</td>"
                       f"<td class='num'>{t['sh']:,.0f}</td>"
                       f"<td class='num'>{money(t.get('cost'))}</td>"
                       f"<td class='num'>{age}</td><td>{state}</td></tr>")
    for d in disc_open:
        age = (today - d['buy']).days
        op_rows.append(f"<tr><td><b>{d['stock']}</b></td><td>DISC</td>"
                       f"<td>{d['buy']:%d %b}</td><td class='num'>{d['px']:,.2f}</td>"
                       f"<td class='num'>{d['sh']:,.0f}</td>"
                       f"<td class='num'>{money(d.get('cost'))}</td>"
                       f"<td class='num'>{age}</td><td>discretionary — no stop</td></tr>")

    # -- hold tracker
    hold_rows = []
    all_holds = []
    for nm in NAMES:
        hs = [t['hold'] for t in book['trades'] if t.get('hold') is not None and t['stock'] == nm]
        all_holds += hs
        live = f'{statistics.fmean(hs):.1f}' if hs else '—'
        drift = ''
        if hs and statistics.fmean(hs) > 1.6 * HIST_HOLD[nm] and len(hs) >= 3:
            drift = " <span class='warn'>▲</span>"
        hold_rows.append(f"<tr><td><b>{nm}</b></td><td class='num'>{HIST_HOLD[nm]}</td>"
                         f"<td class='num'>{live}{drift}</td><td class='num'>{len(hs)}</td></tr>")
    live_book = f'{statistics.fmean(all_holds):.1f}' if all_holds else '—'
    hold_rows.append(f"<tr><td><b>BOOK</b></td><td class='num'>{HIST_HOLD['BOOK']}</td>"
                     f"<td class='num'><b>{live_book}</b></td>"
                     f"<td class='num'>{len(all_holds)}</td></tr>")

    # -- checks
    checks = []
    qlast = max((d for d in book['qdays']), default=None)
    if qlast is None or qlast < report_week:
        checks.append(f"<li class='warn'>Query history ends {qlast or 'never'} — the Friday "
                      f"close is missing. Run Script 1, save, and regenerate this page.</li>")
    else:
        checks.append(f"<li class='ok'>Query current through {qlast:%d %b %Y}.</li>")
    if qgaps:
        gap_txt = '; '.join(f'{nm} on {d:%a %d %b}' for d, nm in qgaps)
        checks.append(f"<li class='warn'>Missing closes in the report week: {gap_txt} "
                      f"(the VRT-style dropout) — breadth counts on those days exclude them.</li>")
    bad = [t for t in book['trades'] if t['problems']]
    if bad:
        rows_txt = ', '.join(f"row {t['row']} ({'; '.join(t['problems'])})" for t in bad[:5])
        checks.append(f"<li class='warn'>Blotter integrity: {rows_txt}.</li>")
    else:
        checks.append("<li class='ok'>Blotter integrity: every row complete.</li>")
    if n > 46:
        checks.append(f"<li class='warn'>Weekly log at {n}/52 rows — extend the log block "
                      f"before it fills.</li>")
    checks.append(f"<li>Performance tab: confirm week-{n} row shows "
                  f"<b>{money(wk_act, signed=True)}</b> actual and the cumulative-actual "
                  f"line has extended. Hold-table live column should match the tracker here.</li>")
    if losses:
        ltxt = ', '.join(f"{t['stock']} {money(t['pnl'], signed=True)}" for t in losses)
        checks.append(f"<li class='warn'>Losing exits this week: {ltxt} — confirm each was a "
                      f"stop or logged discretionary decision, not an override.</li>")
    upcoming = sorted((d, nm) for nm, d in book['earn'].items()
                      if report_week < d <= report_week + dt.timedelta(days=21))
    if upcoming:
        checks.append('<li>Earnings inside 21 days: ' +
                      ', '.join(f'<b>{nm}</b> {d:%d %b}' for d, nm in upcoming) + '.</li>')

    gate_cfg = (f"PM rule {book['gates']['pm_pct'] * 100:.0f}% at 9:00 · ATH ε = "
                f"{book['gates']['ath_eps']} · breadth K = {K}")

    return f"""<!doctype html><html><head><meta charset="utf-8">
<title>BC weekly one-pager — {report_week:%d %b %Y}</title><style>{CSS}</style></head><body>
<h1>Bayesian Capital — Weekly Book One-Pager</h1>
<div class="sub">Week ending <b>{report_week:%A %d %B %Y}</b> · generated {today:%d %b %Y}
 · workbook: {html.escape(wb_name)} · live config: {gate_cfg}</div>
<div class="rule"></div>

<div class="cards">
 <div class="card"><div class="k">Week P&amp;L (realised)</div>
  <div class="v {cls(wk_act)}">{money(wk_act, signed=True)}</div>
  <div class="d">plan {money(wk_plan)} · {len(exits)} exits, {len(buys)} fills, {len(stops)} stops</div></div>
 <div class="card"><div class="k">Cumulative vs plan</div>
  <div class="v {cls(gap)}">{money(gap, signed=True)}</div>
  <div class="d">actual {money(cum_act, signed=True)} vs plan {money(cum_plan)} over {n} wk</div></div>
 <div class="card"><div class="k">Book equity (realised)</div>
  <div class="v">{money(equity)}</div>
  <div class="d">plan line {money(cap + cum_plan)} · target {target * 100:.0f}%/yr on {money(cap)}</div></div>
 <div class="card"><div class="k">Open positions</div>
  <div class="v">{len(opens)}{f' +{len(disc_open)} disc' if disc_open else ''}</div>
  <div class="d">oldest {max(((today - t['buy']).days for t in opens), default=0)}d of {STOP_DAYS}d stop</div></div>
</div>

<div class="cols">
<div>
<h2>This week's activity by name</h2>
<table><tr><th>Stock</th><th class="num">Fills</th><th class="num">$ deployed</th>
<th class="num">Exits</th><th class="num">Stops</th><th class="num">Week P&amp;L</th>
<th class="num">Open</th></tr>{''.join(act_rows) or '<tr><td colspan=7>no activity</td></tr>'}</table>
{f"<div class='note'>Discretionary closed this week: " + ', '.join(f"{d['stock']} {money(d['pnl'], signed=True)}" for d in disc_wk) + '.</div>' if disc_wk else ''}
<div class="note">Exits classified as stops when hold ≥ {STOP_DAYS} calendar days (the model's
only forced sale); everything else is a target or gap exit.</div>

<h2>Open positions vs the {STOP_DAYS}-day stop</h2>
<table><tr><th>Stock</th><th>Tranche</th><th>Bought</th><th class="num">Price</th>
<th class="num">Shares</th><th class="num">Cost</th><th class="num">Age (d)</th><th>Stop</th></tr>
{''.join(op_rows) or '<tr><td colspan=8>book fully in cash</td></tr>'}</table>
</div>

<div>
<h2>Gates this week</h2>
<table><tr><th>Day</th><th class="num">Below 200dma</th><th>Names</th><th>Breadth gate</th></tr>
{''.join(br_rows) or '<tr><td colspan=4>no Query data for the week</td></tr>'}</table>
<div class="note">Gate armed {armed_days} of {len(breadth)} days (fires at K = {K} of nine names,
trailing-{DMA_WIN} close average, the sheet's own convention).
<b>MU pause:</b> {mu_state}
Pre-market vetoes are not reconstructable — the sheet keeps no 9:00 history; check them live
on the morning run.</div>

<h2>ATH proximity (cap check)</h2>
<table><tr><th>Stock</th><th class="num">Friday close</th><th class="num">ATH*</th>
<th class="num">Below ATH</th></tr>{''.join(ath_rows)}</table>
<div class="note">*ATH proxied as the highest high in the workbook's Query window; the buy cap
binds as a name approaches it (ε = {book['gates']['ath_eps']}). Four tightest shown.</div>

<h2>Hold tracker (calendar days)</h2>
<table><tr><th>Stock</th><th class="num">Historical</th><th class="num">Live</th>
<th class="num">Closed trades</th></tr>{''.join(hold_rows)}</table>
<div class="note">Historical = pooled back-test on verified fills. ▲ marks live &gt; 1.6×
reference with ≥ 3 closes — a drifting hold flags a harvest change before returns can.
Only CLOSED trades count, so early weeks read short: the long holds are still open.</div>
</div>
</div>

<h2>Checks before filing</h2>
<ul class="checks">{''.join(checks)}</ul>

<div class="foot"><span>Bayesian Capital · internal — book operations</span>
<span>Equity is realised P&amp;L only; open positions carried at cost.
Plan from Performance-tab assumptions.</span></div>
</body></html>"""


def main():
    ap = argparse.ArgumentParser(description='Bayesian Capital weekly one-pager')
    ap.add_argument('workbook')
    ap.add_argument('--week', help='week-ending Friday, YYYY-MM-DD (default: last Friday)')
    ap.add_argument('--out', help='output HTML path (default: beside the workbook)')
    args = ap.parse_args()

    if args.week:
        report_week = dt.date.fromisoformat(args.week)
        if report_week.isoweekday() != 5:
            print(f'note: {report_week} is not a Friday; using its week-ending '
                  f'{friday_of(report_week)}', file=sys.stderr)
            report_week = friday_of(report_week)
    else:
        today = dt.date.today()
        report_week = friday_of(today)
        if report_week > today:                      # Mon-Thu: report the week just ended
            report_week -= dt.timedelta(days=7)

    book = read_book(args.workbook)
    page = build_html(book, report_week, os.path.basename(args.workbook))
    out = args.out or os.path.join(os.path.dirname(os.path.abspath(args.workbook)),
                                   f'BC_onepager_{report_week.isoformat()}.html')
    with open(out, 'w', encoding='utf-8') as f:
        f.write(page)
    print(f'wrote {out}  (week ending {report_week:%A %d %B %Y})')


if __name__ == '__main__':
    main()
