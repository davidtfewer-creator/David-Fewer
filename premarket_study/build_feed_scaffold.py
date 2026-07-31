"""
Build a scaffold Excel workbook that hosts the Power Query replica of the Apps Script feed.
Creates the Config and Tickers tables the query reads, the Data header (query load target),
and a ReadMe with the M code + setup steps. No API key is written anywhere.
"""
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment

OUT = '/home/user/David-Fewer/Hybrid10_Feed_PowerQuery.xlsx'
TICKERS = ['NVDA','TSM','TSLA','VRT','VST','AVGO','PLTR','RKLB','SOFI','SPOT']
HDR = Font(name='Arial', size=10, bold=True)
BLUE = PatternFill('solid', fgColor='FFD9E1F2')
GOLD = PatternFill('solid', fgColor='FFF7F1E4')

wb = openpyxl.Workbook()

# ---- Config sheet: table "Config" (Setting | Value) ----
cfg = wb.active; cfg.title = 'Config'
cfg['A1'] = 'Setting'; cfg['B1'] = 'Value'
cfg['A2'] = 'ApiKey'                               # B2 left genuinely empty (user pastes key)
cfg['A3'] = 'StartDate'; cfg['B3'] = '2024-04-01'
for c in ('A1','B1'): cfg[c].font = HDR
cfg['B2'].fill = GOLD; cfg['B3'].fill = BLUE
cfg['B2'].number_format = '@'; cfg['B3'].number_format = '@'
cfg['D2'] = '<-- paste your Massive API key here (kept out of the query code)'
cfg['D3'] = '<-- feed start date (yyyy-mm-dd), matches the script'
cfg.column_dimensions['A'].width = 12; cfg.column_dimensions['B'].width = 24
cfg.column_dimensions['D'].width = 60
tbl_cfg = Table(displayName='Config', ref='A1:B3')
tbl_cfg.tableStyleInfo = TableStyleInfo(name='TableStyleLight9', showRowStripes=True)
cfg.add_table(tbl_cfg)

# ---- Tickers sheet: table "Tickers" (Ticker) in column order ----
tk = wb.create_sheet('Tickers')
tk['A1'] = 'Ticker'; tk['A1'].font = HDR
for i, t in enumerate(TICKERS, start=2):
    tk.cell(i, 1, t)
tk.column_dimensions['A'].width = 12
tbl_tk = Table(displayName='Tickers', ref=f'A1:A{1+len(TICKERS)}')
tbl_tk.tableStyleInfo = TableStyleInfo(name='TableStyleLight9', showRowStripes=True)
tk.add_table(tbl_tk)

# ---- Data sheet: header only (query load target) ----
data = wb.create_sheet('Data')
hdr = ['Date']
for t in TICKERS:
    hdr += [f'{t}_O', f'{t}_H', f'{t}_L', f'{t}_C']
for j, h in enumerate(hdr, start=1):
    c = data.cell(1, j, h); c.font = HDR
data['A3'] = ('(Load the Hybrid10Feed query here: Close & Load To > Table > existing worksheet '
              '= Data!$A$1. It overwrites this header with the live feed.)')
data.column_dimensions['A'].width = 12

# ---- ReadMe sheet: M code + steps ----
rm = wb.create_sheet('ReadMe')
rm.column_dimensions['A'].width = 120
lines = [
    ('Hybrid10 Feed — Power Query replica of the Google Apps Script', True),
    ('', False),
    ('WHAT THIS DOES', True),
    ('Pulls daily OHLC for the ten tickers from the Massive aggregates API (same endpoint the', False),
    ('Apps Script used) and builds the Date | <TK>_O/_H/_L/_C table, keeping only dates common', False),
    ('to all ten names — identical output to the script, refreshable from Excel.', False),
    ('', False),
    ('SECURITY', True),
    ('The key is NOT stored in the query. Paste it into Config!B2. The key you shared in chat', False),
    ('should be rotated in your Massive account, since it was exposed.', False),
    ('', False),
    ('ONE-TIME SETUP', True),
    ('1. Config!B2  : paste your Massive API key.   Config!B3 : start date (default 2024-04-01).', False),
    ('2. Data tab   : Get Data > Launch Power Query Editor > New Query > Blank Query.', False),
    ('3. Advanced Editor : delete everything, paste the M code below, Done. Rename it Hybrid10Feed.', False),
    ('4. Close & Load To... > Table > Existing worksheet > =Data!$A$1.', False),
    ('5. First refresh: if prompted, set the api.massive.com credential to Anonymous', False),
    ('   (the key travels in the URL) and privacy level to Public/Organizational.', False),
    ('', False),
    ('TO REFRESH LATER : Data > Refresh All (or right-click the table > Refresh).', False),
    ('TO fully match the model feed, Date is emitted as text dd/MM/yyyy (as the script did).', False),
    ('', False),
    ('====================  M CODE (paste into Advanced Editor)  ====================', True),
]
r = 1
for text, bold in lines:
    if text:                                       # skip empty spacer rows (no empty-string cells)
        c = rm.cell(r, 1, text)
        if bold: c.font = Font(name='Arial', size=11, bold=True, color='0B1F3A')
    r += 1
with open('/home/user/David-Fewer/premarket_study/HybridFeed.m') as f:
    for code_line in f.read().splitlines():
        if code_line == '':                        # blank M lines -> leave row empty, no cell
            r += 1; continue
        cc = rm.cell(r, 1, code_line)
        cc.font = Font(name='Consolas', size=9)
        r += 1

wb.save(OUT)
print('saved', OUT)
print('sheets:', wb.sheetnames)
