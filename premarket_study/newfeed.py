"""
Analyse the Hybrid10_Feed_new candidate set (OXY DVN FSLR COIN MSTR HOOD RBLX MRNA ARM SHOP).
Feed carries OHLC only (no volume, no fundamentals).

Fast pass: volatility profile (avg daily range %, annualised close-to-close vol) as the
high-beta proxy, and each name's return correlation to the existing ten-name book (the
diversification payoff). Engine model-fit metrics are produced separately (needs optimisation).
"""
import openpyxl, datetime, math, statistics

NEW = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/1c7fbf38-Hybrid10_Feed_new.xlsx'
BOOK = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/1b35dd3e-Hybrid9_Bayesian__OU_tranche_LIVE_MU.xlsx'.replace('Bayesian__OU_tranche_LIVE_MU','')  # placeholder
BOOK = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/1b35dd3e-Hybrid9_Bayesian_OULIVE_PLTR_nopq.xlsx'
NEW_TK = ['OXY','DVN','FSLR','COIN','MSTR','HOOD','RBLX','MRNA','ARM','SHOP']
BOOK_TK = ['NVDA','TSM','TSLA','VRT','VST','AVGO','PLTR','RKLB','SOFI','SPOT']


def parse_date(d):
    if isinstance(d, datetime.datetime): return d.date()
    if isinstance(d, datetime.date): return d
    if isinstance(d, (int,float)): return datetime.date(1899,12,30)+datetime.timedelta(days=int(d))
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%m/%d/%Y'):
        try: return datetime.datetime.strptime(str(d)[:10], fmt).date()
        except Exception: pass
    return None


def load(path, tickers):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Data'] if 'Data' in wb.sheetnames else wb['Query']
    hdr = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    col = {}
    for i,h in enumerate(hdr, start=1):
        if h and h.endswith('_O'): col[h[:-2]] = i
    out = {}
    for t in tickers:
        c0 = col[t]; dts=[]; O=[];H=[];L=[];C=[]
        for r in range(2, ws.max_row+1):
            o = ws.cell(r, c0).value
            if not isinstance(o,(int,float)) or o<=0: continue
            d = parse_date(ws.cell(r,1).value)
            dts.append(d); O.append(o)
            H.append(ws.cell(r,c0+1).value); L.append(ws.cell(r,c0+2).value); C.append(ws.cell(r,c0+3).value)
        out[t] = (dts,O,H,L,C)
    return out


def rets(C):
    return [C[i]/C[i-1]-1 for i in range(1,len(C))]

def align_rets(dtsA, CA, dtsB, CB):
    ib = {d:i for i,d in enumerate(dtsB)}
    xa=[]; xb=[]
    for i in range(1,len(dtsA)):
        d=dtsA[i]
        if d in ib and ib[d]>0:
            j=ib[d]
            xa.append(CA[i]/CA[i-1]-1); xb.append(CB[j]/CB[j-1]-1)
    return xa,xb

def corr(a,b):
    n=min(len(a),len(b)); a,b=a[:n],b[:n]
    if n<3: return 0.0
    ma,mb=statistics.mean(a),statistics.mean(b)
    cov=sum((a[i]-ma)*(b[i]-mb) for i in range(n))
    va=sum((x-ma)**2 for x in a); vb=sum((x-mb)**2 for x in b)
    return cov/math.sqrt(va*vb) if va>0 and vb>0 else 0.0


if __name__ == '__main__':
    new = load(NEW, NEW_TK)
    book = load(BOOK, BOOK_TK)

    print('=== VOLATILITY PROFILE (high-vol proxy; no index so not literal beta) ===')
    print(f'{"tick":5s}{"lastpx":>9s}{"avgRange%":>10s}{"annVol%":>9s}{"maxDD%(BH)":>11s}')
    for t in NEW_TK:
        dts,O,H,L,C = new[t]
        rng = statistics.mean([(H[i]-L[i])/C[i] for i in range(len(C))])*100
        rr = rets(C); av = statistics.pstdev(rr)*math.sqrt(252)*100
        peak=-1e9; mdd=0
        for c in C:
            peak=max(peak,c); mdd=max(mdd,(peak-c)/peak) if peak>0 else mdd
        print(f'{t:5s}{C[-1]:>9.2f}{rng:>10.2f}{av:>9.0f}{mdd*100:>11.0f}')

    print('\n=== CORRELATION TO EXISTING BOOK (daily close-to-close returns) ===')
    # book average return series (equal weight) aligned to each new name
    print(f'{"tick":5s}{"vs BOOK avg":>13s}{"vs AI-core":>12s}   (AI-core = NVDA,TSM,AVGO,PLTR,VRT,VST)')
    ai = ['NVDA','TSM','AVGO','PLTR','VRT','VST']
    for t in NEW_TK:
        dts,O,H,L,C = new[t]
        # build book-average return on common dates
        # align each book name to this ticker's dates, average
        cbook=[]; cai=[]
        # collect per-date book returns
        # simple: correlate to equal-weight average of all book names / AI subset
        # compute aligned return vectors for each book name, then average elementwise on intersection
        # easier: intersect dates across this ticker and ALL book names
        idx = {d:i for i,d in enumerate(dts)}
        common = [d for d in dts if all(d in {x:1 for x in book[b][0]} for b in BOOK_TK)]
        # build map date->ret for each book name
        def rmap(b):
            bd,_,_,_,bc = book[b]; m={}
            bi={d:i for i,d in enumerate(bd)}
            for k in range(1,len(bd)): m[bd[k]]=bc[k]/bc[k-1]-1
            return m
        rm = {b:rmap(b) for b in BOOK_TK}
        tret = {dts[k]:C[k]/C[k-1]-1 for k in range(1,len(dts))}
        va=[]; vb=[]; vc=[]
        for d in tret:
            if all(d in rm[b] for b in BOOK_TK):
                va.append(tret[d])
                vb.append(statistics.mean(rm[b][d] for b in BOOK_TK))
                vc.append(statistics.mean(rm[b][d] for b in ai))
        print(f'{t:5s}{corr(va,vb):>13.2f}{corr(va,vc):>12.2f}')
