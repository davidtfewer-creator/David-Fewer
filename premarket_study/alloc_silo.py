"""
Silo each stock's profit to that stock, and never let a held sleeve concentrate its stock.

The rule
--------
Every stock has an ENTITLEMENT -- its weighted share of the book's base capital plus the profit
it has earned itself:

    book value   = cash on hand (B5) + capital deployed in open positions
    base capital = book value - total profit to date
    entitlement  = weight x base capital + own profit

Entitlements sum to book value by construction, so no subtraction of profit from cash is
involved and nothing can go negative however much capital is deployed. The earlier version
computed base as B5 - profit, which went negative once enough capital was tied up.

Today's cash is then handed out PRO RATA to entitlement, with one restriction that keeps
position sizes honest:

    a sleeve that is HOLDING cannot take new money, and its entitlement is shared EVENLY
    across the OTHER stocks -- never across its own stock's second sleeve.

That last clause is the point. Giving a held stock's share to its own free sleeve is what would
have put over $1m into VST: the OU sleeve already holds $131k of stock and the Bayes sleeve
would have been handed the whole stock allocation on top. Now VST's held half goes to the other
four names and VST's free sleeve claims only its own half.

Within a stock, whatever it ends up claiming is split across its FREE sleeves in proportion to
their own entitlement shares.

Conservation: allocations always sum to the cash available, so the existing B38 reconciliation
still reads OK and the blotter is neither starved nor over-committed.

Where it is written
-------------------
Allocation columns S:AB rows 11-15 carry the working, rows 42-52 the summary and guards, and
C25:C34 is rewired to read it. Active Trading D19:D28 and H19:H28 already INDEX into C25:C34,
so the blotter picks it up untouched.

Safety notes. This workbook has no Power Query, no charts and no pivots -- checked -- which is
the only reason openpyxl is usable on it; it is NOT safe on the per-ticker LIVE workbooks.
openpyxl also re-serialises floats on save, moving thousands of literals by one ULP including
manually entered price overrides, so restore_untouched puts every literal outside the intended
edit set back to the source bytes. And openpyxl blanks cached results, so the delivered file
shows blanks until Excel opens and recalculates; this script therefore mirrors the arithmetic
in Python rather than reading its own output back.

Run:  python3 alloc_silo.py
"""
import shutil

import openpyxl

SRC = ('/root/.claude/uploads/822d405e-f99b-5b59-9c6b-87e725054402/'
       'c655e5aa-TradingExcel_5stock_live_silo.xlsx')
DST = '/home/user/David-Fewer/TradingExcel_5stock_live_siloed.xlsx'

STOCK_ROWS = (11, 12, 13, 14, 15)
SLEEVE_TOP = 25
LOG = "'Active Trading'!$G$42:$G$1041"
LOGST = "'Active Trading'!$M$42:$M$1041"

COLS = [('S', 'Profit to date ($)'), ('T', 'Entitlement ($)'), ('U', 'Free sleeves'),
        ('V', 'Released entitlement'), ('W', 'Retained entitlement'),
        ('X', 'Can receive?'), ('Y', 'Other recipients'), ('Z', 'Released per recipient'),
        ('AA', 'Inbound from others'), ('AB', 'Claim ($)')]


def sleeve_rows(i):
    b = SLEEVE_TOP + 2 * (i - STOCK_ROWS[0])
    return b, b + 1


def write(ws):
    for col, head in COLS:
        c = ws[f'{col}10']
        c.value = head
        c.font = openpyxl.styles.Font(bold=True, size=9)

    for i in STOCK_ROWS:
        b, o = sleeve_rows(i)
        ws[f'S{i}'] = ("=IFERROR(INDEX('Active Trading'!$C$7:$C$11,"
                       f"MATCH($A{i},'Active Trading'!$A$7:$A$11,0)),0)")
        ws[f'T{i}'] = f'=$K{i}*$B$45+$S{i}'
        ws[f'U{i}'] = f'=$B{i}*((1-$D${b})+(1-$D${o}))'
        ws[f'V{i}'] = (f'=IF($B{i}=0,$T{i},$T{i}*($R{i}*$D${b}+(1-$R{i})*$D${o}))')
        ws[f'W{i}'] = (f'=IF($B{i}=0,0,$T{i}*($R{i}*(1-$D${b})+(1-$R{i})*(1-$D${o})))')
        ws[f'X{i}'] = f'=IF(AND($B{i}=1,$U{i}=2),1,0)'
        ws[f'Y{i}'] = f'=SUM($X$11:$X$15)-$X{i}'
        ws[f'Z{i}'] = f'=IF($Y{i}=0,0,$V{i}/$Y{i})'
        ws[f'AA{i}'] = f'=$X{i}*(SUM($Z$11:$Z$15)-$Z{i})'
        ws[f'AB{i}'] = f'=$W{i}+$AA{i}'
        for col in ('S', 'T', 'V', 'W', 'Z', 'AA', 'AB'):
            ws[f'{col}{i}'].number_format = '#,##0.00'

    block = [
        (42, 'SILO ACCOUNTING  —  entitlement = weight x base capital + own profit', None),
        (43, 'Capital deployed in open positions ($)',
             f'=SUMIFS({LOG},{LOGST},"OPEN")'),
        (44, 'Book value  =  cash + deployed', '=$B$5+$B$43'),
        (45, 'Base capital  =  book value less total profit', '=$B$44-$B$46'),
        (46, 'Total profit to date ($)', '=SUM($S$11:$S$15)'),
        (47, 'Total claim ($)', '=SUM($AB$11:$AB$15)'),
        (48, 'Scale factor  =  MIN(1, cash / total claim)',
             '=IF($B$47=0,0,MIN(1,$B$5/$B$47))'),
        (49, 'Cash left idle today ($)', '=$B$5-SUM($C$25:$C$34)'),
        (50, 'Check: allocated + idle = cash available',
             '=IF(ABS(SUM($C$25:$C$34)+$B$49-$B$5)<0.01,"OK","MISMATCH")'),
        (51, 'Guard: any negative allocation?',
             '=IF(MIN($C$25:$C$34)<-0.01,"NEGATIVE ALLOCATION - check inputs","")'),
        (52, 'Guard: base capital sane?',
             '=IF($B$45<=0,"BASE CAPITAL NOT POSITIVE - check B5 and the trade log","")'),
    ]
    for row, label, formula in block:
        ws[f'A{row}'] = label
        if formula is None:
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=10)
        else:
            ws[f'B{row}'] = formula
            if row in (43, 44, 45, 46, 47, 49):
                ws[f'B{row}'].number_format = '#,##0.00'

    for i in STOCK_ROWS:
        b, o = sleeve_rows(i)
        den = f'($R{i}*(1-$D{b})+(1-$R{i})*(1-$D{o}))'
        for row, share in ((b, f'$R{i}*(1-$D{b})'), (o, f'(1-$R{i})*(1-$D{o})')):
            ws[f'C{row}'] = (f'=IF(OR({den}=0,$B$47=0),0,'
                             f'$B$48*$AB{i}*{share}/{den})')

    ws['E24'] = 'Base wt (ref)'
    ws['F24'] = 'Free wt (ref)'
    ws['A54'] = ('C25:C34 = cash x stock claim / total claim, split over that stock\'s FREE '
                 'sleeves. A HOLDING sleeve\'s entitlement goes evenly to the OTHER stocks, '
                 'never to its own stock. E and F are reference only.')


def restore_untouched(src, dst, intended):
    """Put back any numeric literal openpyxl re-serialised outside the intended edit set.

    Saving through openpyxl rewrites floats with Python's repr, moving the last bit on ~2,300
    values workbook-wide -- including manually entered trade-price overrides, which are
    deliberate. Matches literal value cells ONLY: the lazy <c ...>.*?</c> form runs past a
    self-closing <c .../> and swallows whole rows, which is what produced an unreadable file
    the first time this was attempted.
    """
    import re
    import zipfile

    src_z, dst_z = zipfile.ZipFile(src), zipfile.ZipFile(dst)
    parts, fixed = {}, 0
    cell_re = re.compile(rb'<c r="([A-Z]+\d+)"([^>/]*)><v>([^<]*)</v></c>')
    for name in dst_z.namelist():
        data = dst_z.read(name)
        if not (name.startswith('xl/worksheets/sheet') and name.endswith('.xml')):
            parts[name] = data
            continue
        try:
            sdata = src_z.read(name)
        except KeyError:
            parts[name] = data
            continue
        srcvals = {m.group(1).decode(): m.group(3) for m in cell_re.finditer(sdata)}

        def repl(m):
            nonlocal fixed
            coord = m.group(1).decode()
            want = srcvals.get(coord)
            if coord in intended or want is None or m.group(3) == want:
                return m.group(0)
            fixed += 1
            return (b'<c r="' + m.group(1) + b'"' + m.group(2)
                    + b'><v>' + want + b'</v></c>')

        parts[name] = cell_re.sub(repl, data)
    src_z.close(); dst_z.close()
    with zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED) as z:
        for n, d in parts.items():
            z.writestr(n, d)
    return fixed


def read_inputs():
    wv = openpyxl.load_workbook(SRC, data_only=True)
    a, t = wv['Allocation'], wv['Active Trading']
    names = [a[f'A{i}'].value for i in STOCK_ROWS]
    prof, K, R, inc = {}, {}, {}, {}
    for i, n in zip(STOCK_ROWS, names):
        K[i] = a[f'K{i}'].value or 0.0
        R[i] = a[f'R{i}'].value or 0.5
        inc[i] = a[f'B{i}'].value or 0
        prof[i] = next((t[f'C{r}'].value or 0.0) for r in range(7, 12)
                       if t[f'A{r}'].value == n)
    deployed = sum((t[f'G{r}'].value or 0.0) for r in range(42, 1042)
                   if t[f'M{r}'].value == 'OPEN')
    held = {r: (1 if t[f'C{r-6}'].value == 'HOLDING' else 0) for r in range(25, 35)}
    old = {r: (a[f'C{r}'].value or 0.0) for r in range(25, 35)}
    return dict(names=names, prof=prof, K=K, R=R, inc=inc, B5=a['B5'].value,
                deployed=deployed, held=held, old=old)


def allocate(inp, held=None):
    """Exactly the arithmetic written into S:AB and C25:C34."""
    held = inp['held'] if held is None else held
    B5, K, R, inc, prof = inp['B5'], inp['K'], inp['R'], inp['inc'], inp['prof']
    book = B5 + inp['deployed']
    total_profit = sum(prof.values())
    base = book - total_profit

    T, V, W, X = {}, {}, {}, {}
    for i in STOCK_ROWS:
        b, o = sleeve_rows(i)
        T[i] = K[i] * base + prof[i]
        if inc[i] == 0:
            V[i], W[i] = T[i], 0.0
        else:
            V[i] = T[i] * (R[i] * held[b] + (1 - R[i]) * held[o])
            W[i] = T[i] * (R[i] * (1 - held[b]) + (1 - R[i]) * (1 - held[o]))
        free = inc[i] * ((1 - held[b]) + (1 - held[o]))
        # only a stock with BOTH sleeves free may receive. Letting a part-held stock receive
        # makes redistribution circular when every name holds one sleeve: each gives away half
        # and takes half back, and its free sleeve ends up carrying the whole stock entitlement
        # on top of a position it already has -- the concentration this rule exists to prevent.
        X[i] = 1 if (inc[i] == 1 and free == 2) else 0
    Y = {i: sum(X.values()) - X[i] for i in STOCK_ROWS}
    Z = {i: (V[i] / Y[i] if Y[i] else 0.0) for i in STOCK_ROWS}
    AA = {i: X[i] * (sum(Z.values()) - Z[i]) for i in STOCK_ROWS}
    AB = {i: W[i] + AA[i] for i in STOCK_ROWS}
    tot = sum(AB.values())
    # Fund at entitlement, never above it. Scaling claims up to absorb all cash is what
    # re-concentrates a book that is mostly holding; the surplus stays in cash earning
    # interest, which is the correct answer when every name already has a position.
    factor = min(1.0, B5 / tot) if tot > 0 else 0.0

    sleeve = {}
    for i in STOCK_ROWS:
        b, o = sleeve_rows(i)
        den = R[i] * (1 - held[b]) + (1 - R[i]) * (1 - held[o])
        if den == 0 or tot == 0:
            sleeve[b] = sleeve[o] = 0.0
        else:
            sleeve[b] = factor * AB[i] * R[i] * (1 - held[b]) / den
            sleeve[o] = factor * AB[i] * (1 - R[i]) * (1 - held[o]) / den
    return dict(book=book, base=base, total_profit=total_profit, T=T, V=V, W=W,
                X=X, AB=AB, tot=tot, factor=factor, sleeve=sleeve,
                idle=B5 - sum(sleeve.values()))


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    write(wb['Allocation'])
    wb.save(DST)

    intended = {f'{c}{i}' for c, _ in COLS for i in list(STOCK_ROWS) + [10]}
    intended |= {f'C{r}' for r in range(SLEEVE_TOP, SLEEVE_TOP + 10)}
    intended |= {f'A{r}' for r in range(42, 55)} | {f'B{r}' for r in range(42, 55)}
    intended |= {'E24', 'F24'}
    n = restore_untouched(SRC, DST, intended)
    print(f'written: {DST}')
    print(f'  restored {n} numeric literals openpyxl had re-serialised\n')

    inp = read_inputs()
    r = allocate(inp)
    nm = dict(zip(STOCK_ROWS, inp['names']))
    print(f"  cash available B5        {inp['B5']:>14,.2f}")
    print(f"  deployed in open pos.    {inp['deployed']:>14,.2f}")
    print(f"  book value               {r['book']:>14,.2f}")
    print(f"  total profit to date     {r['total_profit']:>14,.2f}")
    print(f"  base capital             {r['base']:>14,.2f}\n")
    print(f"  {'stock':6s} {'profit':>11s} {'entitlement':>13s} {'released':>12s} "
          f"{'inbound':>11s} {'claim':>13s}")
    for i in STOCK_ROWS:
        inb = r['AB'][i] - r['W'][i]
        print(f"  {nm[i]:6s} {inp['prof'][i]:>11,.0f} {r['T'][i]:>13,.0f} "
              f"{r['V'][i]:>12,.0f} {inb:>11,.0f} {r['AB'][i]:>13,.0f}")

    print(f"\n  {'sleeve':16s} {'status':9s} {'new $':>13s} {'previous $':>13s} {'change':>12s}")
    tot = 0.0
    for i in STOCK_ROWS:
        b, o = sleeve_rows(i)
        for lbl, row in (('Bayes', b), ('OU', o)):
            v = r['sleeve'][row]
            tot += v
            st = 'HOLDING' if inp['held'][row] else 'CASH'
            print(f"  {nm[i]+' '+lbl:16s} {st:9s} {v:>13,.2f} {inp['old'][row]:>13,.2f} "
                  f"{v-inp['old'][row]:>+12,.2f}")
    print(f"\n  allocated {tot:,.2f} vs cash {inp['B5']:,.2f}  "
          f"diff {tot-inp['B5']:+.4f}  "
          f"{'OK' if abs(tot-inp['B5'])<0.01 else 'MISMATCH'}")
    vst = r['sleeve'][29] + r['sleeve'][30]
    print(f"\n  VST exposure if its Bayes bid fills: {vst:,.0f} new + "
          f"{inp['deployed']:,.0f} already held = {vst+inp['deployed']:,.0f}")
    print('\n  NOTE: openpyxl blanks cached results. Values appear only once Excel opens the')
    print('  file and recalculates (fullCalcOnLoad). Compare against this table.')


if __name__ == '__main__':
    main()
