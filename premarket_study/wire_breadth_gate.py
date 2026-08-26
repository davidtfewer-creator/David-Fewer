"""
Wire the breadth-conditional 200dma gate into the live nine-stock workbook
(adoption of HANDOVER 3.23, 26 Aug 2026): the gate fires only when the bear is
BROAD — a stock below its own 200-day average is gated only when at least K of
the nine names are below theirs. K=4 (two names of margin inside the 2022
protection boundary K<=5).

Layout-preserving by construction — nothing is inserted or moved, the order
block A18:J36 and every script-read cell keep their addresses. Changes:

  1. K variable: Active Trading M8 = 4 (label L8, explanation N8) — the Gate
     Variables block grows one row in previously empty cells.
  2. Breadth count: Allocation G43 (previously empty) = sum of the nine Bayes-row
     breach flags G25,G27..G41 (both sleeve rows carry identical flags; summing
     one row per name counts NAMES below their 200dmas). Label in A43.
  3. The gate application: Allocation F25:F42 become
       =E*(1-D)*(1-G*($G$43>='Active Trading'!$M$8))*(1-H)
     G stays the RAW per-name breach flag (unchanged formulas, still fails open
     to 0 on lookup errors), so col G remains readable as "which names are
     below". A blank M8 makes the condition count>=0 = TRUE, i.e. the rule
     fails CLOSED to the old per-name gate — the failure direction is more
     protection, never less.
  4. Notes: an UPDATE — 26 AUGUST 2026 section.

Traffic lights on Active Trading col N are untouched: red still means "this
stock is below its own 200dma"; whether that red is enforced now depends on the
breadth count G43 vs M8.

Verification: full-workbook cell diff (the changed set must be exactly the
intended cells) + a Python mirror computing every 200dma, breach flag, the
breadth count and the resulting gate state from the Query closes.
"""
import copy
import sys

import openpyxl

STOCKS = ['TSM', 'VRT', 'VST', 'RKLB', 'MU', 'GM', 'VLO', 'CF', 'MRVL']
QCOL = {s: 5 + 4 * i for i, s in enumerate(STOCKS)}
BAYES_ROWS = [25 + 2 * i for i in range(9)]                 # Allocation, one per name
COUNT_FORMULA = '=' + '+'.join(f'G{r}' for r in BAYES_ROWS)


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    at = wb['Active Trading']
    al = wb['Allocation']
    notes = wb['Notes']

    # ---- preconditions: the layout this edit assumes
    assert at['L5'].value == 'Gate Variables'
    assert at['L6'].value == 'PM gate %' and at['L7'].value == 'ATH gate ε'
    assert all(at.cell(row=8, column=c).value is None for c in (12, 13, 14))
    assert al['G24'].value == 'DMA gate' and al['H24'].value == 'PM gate'
    for r in range(25, 43):
        f = al.cell(row=r, column=6).value
        assert f == f'=E{r}*(1-D{r})*(1-G{r})*(1-H{r})', (r, f)
        g = al.cell(row=r, column=7).value
        assert g.startswith('=IFERROR(IF(INDEX(Dashboard!$B$4:$B$12'), (r, g)
    assert all(al.cell(row=43, column=c).value is None for c in range(1, 10))

    # ---- 1. K in the Gate Variables block
    for col, src, val in ((12, 'L7', 'DMA breadth K'), (13, 'M6', 4),
                          (14, 'N7',
                           'The 200dma gate fires only when at least this many of the nine '
                           'names sit below their own 200-day averages (breadth = a bear, an '
                           'isolated breach = a recovery that should trade). Allocation G43 '
                           'holds the live count. Blank = 1 name is enough (the old per-name '
                           'gate — fails closed to MORE protection). Tested: K=4 keeps the '
                           'full 2022 save and removes the April-2025 concentration penalty; '
                           'the 2022 protection boundary is K<=5.')):
        ref = at[src]
        c = at.cell(row=8, column=col, value=val)
        c.font = copy.copy(ref.font)
        c.fill = copy.copy(ref.fill)
        c.border = copy.copy(ref.border)
        c.alignment = copy.copy(ref.alignment)
    at['M8'].number_format = '0'

    # ---- 2. breadth count in Allocation G43 (+ label A43)
    la = al.cell(row=43, column=1, value='Names below 200dma (gate arms at K, AT M8)')
    ref = al['A44']
    la.font = copy.copy(ref.font)
    la.alignment = copy.copy(ref.alignment)
    cc = al.cell(row=43, column=7, value=COUNT_FORMULA)
    refc = al['G24']
    cc.font = copy.copy(refc.font)
    cc.alignment = copy.copy(refc.alignment)
    cc.number_format = '0'

    # ---- 3. gate application: F25:F42
    for r in range(25, 43):
        al.cell(row=r, column=6,
                value=f"=E{r}*(1-D{r})*(1-G{r}*($G$43>='Active Trading'!$M$8))*(1-H{r})")

    # ---- 4. Notes
    style_hdr = notes['A32']            # 'UPDATE — 19 AUGUST 2026'
    style_a = notes['A33']
    style_b = notes['B33']
    row = notes.max_row + 2
    hdr = notes.cell(row=row, column=1, value='UPDATE — 26 AUGUST 2026')
    hdr.font = copy.copy(style_hdr.font)
    hdr.alignment = copy.copy(style_hdr.alignment)
    row += 1
    entries = [
        ('Breadth-conditional DMA gate (K = M8)',
         'The 200dma gate now fires only when the bear is BROAD: a stock below its own 200-day '
         'average is gated only when at least K of the nine names are below theirs (K in Active '
         'Trading M8, default 4). An isolated breach — one name recovering from its own drawdown, '
         'e.g. VST in Aug 2026 — trades normally. Allocation col G is unchanged and still shows '
         'the raw per-name breach flags; G43 counts the names currently below; F25:F42 now apply '
         'the gate as breach x (G43 >= K). A blank M8 falls back to the old per-name gate (fails '
         'closed to more protection).'),
        (None,
         'Basis (verified fills, both regimes): bull-sample breaches are mostly isolated (193 '
         'days with exactly 1 name below vs 60 days with 5+), and the trades the per-name gate '
         'forgave on isolated days averaged +2.35% (384 entries, 21 losers) — recoveries taxed. '
         'K=4 keeps the 2022 bear save identical (-2.6% vs -21.9% unprotected, max DD 17.4%) and '
         'removes the April-2025 concentration penalty (book max drawdown 32.2% -> 25.8%). K=5 '
         'tested best in-sample but sits one step from the K=6 protection cliff (2022: -12.6%); '
         'K=4 keeps two names of margin.'),
        ('Reading the lights',
         'Col N red still means that stock is below its own 200dma, but it is only ENFORCED when '
         'the breadth count (Allocation G43) reaches K (M8). Compare G43 with M8 to see whether '
         'the gate is armed book-wide.'),
    ]
    for label, text in entries:
        if label:
            ca = notes.cell(row=row, column=1, value=label)
            ca.font = copy.copy(style_a.font)
            ca.alignment = copy.copy(style_a.alignment)
        cb = notes.cell(row=row, column=2, value=text)
        cb.font = copy.copy(style_b.font)
        cb.alignment = copy.copy(style_b.alignment)
        row += 1

    wb.save(out_path)
    return out_path


def diff(in_path, out_path):
    a = openpyxl.load_workbook(in_path)
    b = openpyxl.load_workbook(out_path)
    changed = []
    for ws_name in b.sheetnames:
        wa, wb_ = a[ws_name], b[ws_name]
        for r in range(1, max(wa.max_row, wb_.max_row) + 1):
            for c in range(1, max(wa.max_column, wb_.max_column) + 1):
                va = wa.cell(row=r, column=c).value
                vb = wb_.cell(row=r, column=c).value
                if va != vb:
                    changed.append((ws_name, wb_.cell(row=r, column=c).coordinate, va, vb))
    return changed


def mirror(in_path):
    """Compute today's breach flags, breadth count and gate state off Query closes."""
    wb = openpyxl.load_workbook(in_path, data_only=True)
    q = wb['Query']
    print(f'{"name":6s}{"last close":>11s}{"200dma":>10s}{"breach":>8s}')
    count = 0
    breach = {}
    for s in STOCKS:
        closes = [q.cell(row=r, column=QCOL[s]).value for r in range(2, q.max_row + 1)]
        closes = [v for v in closes if isinstance(v, (int, float))]
        dma = sum(closes[-200:]) / len(closes[-200:])
        br = closes[-1] < dma
        breach[s] = br
        count += br
        print(f'{s:6s}{closes[-1]:>11.2f}{dma:>10.2f}{"BELOW" if br else "-":>8s}')
    K = 4
    print(f'breadth count = {count};  K = {K};  gate armed: {count >= K}')
    for s in STOCKS:
        if breach[s]:
            print(f'  {s}: below its dma -> {"GATED" if count >= K else "TRADES (isolated)"}')
    return count


if __name__ == '__main__':
    src, dst = sys.argv[1], sys.argv[2]
    wire(src, dst)
    print('=== cell diff (must be exactly the intended set) ===')
    for ws, coord, va, vb in diff(src, dst):
        va_s = (str(va)[:60] + '..') if va and len(str(va)) > 60 else va
        vb_s = (str(vb)[:60] + '..') if vb and len(str(vb)) > 60 else vb
        print(f'  {ws}!{coord}: {va_s!r} -> {vb_s!r}')
    print('\n=== mirror: live gate state from Query closes ===')
    mirror(src)
