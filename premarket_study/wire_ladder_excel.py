"""Wire Option C (Bayes-only ladder) into the Dashboard of the trading workbook.
Non-destructive: adds rung-2/-3 price columns + a config block; leaves Model sheets untouched.
Rung1 = existing Bayes BUY (col C); shared take-profit = existing Bayes target (col D)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/2323ed8f-TradingExcel_s1_laddering.xlsx'
OUT = '/home/user/David-Fewer/TradingExcel_s1_laddering_OptionC.xlsx'
ROWS = range(4, 14)
BLUE = 'FFD9E1F2'   # existing input-cell fill
HDR = Font(name='Arial', size=10, bold=True)
REG = Font(name='Arial', size=10)
CUR = '"$"#,##0.00'

wb = openpyxl.load_workbook(SRC, data_only=False)
ws = wb['Dashboard']

# --- config block (editable inputs, blue) : cols R-U, rows 15-20 ---
ws['R15'] = 'LADDER — Option C  (Bayes sleeve only; OU unchanged)'; ws['R15'].font = HDR
ws['R16'] = 'rung2 depth (×k)'; ws['S16'] = 1.3
ws['R17'] = 'rung3 depth (×k)'; ws['S17'] = 1.7
ws['R18'] = 'weight rung1'; ws['S18'] = 0.80
ws['R19'] = 'weight rung2'; ws['S19'] = 0.15
ws['R20'] = 'weight rung3'; ws['S20'] = 0.05
ws['R21'] = 'Place 3 Bayes limit BUYs (rung1/2/3 at the weights above); GTC SELL the whole'
ws['R22'] = 'Bayes position at the Bayes target (col D = rung1 target). OU: bid col E, sell col F.'
for r in (16, 17):
    ws[f'S{r}'].font = REG; ws[f'S{r}'].fill = PatternFill('solid', fgColor=BLUE)
    ws[f'S{r}'].number_format = '0.00'
for r in (18, 19, 20):
    ws[f'S{r}'].font = REG; ws[f'S{r}'].fill = PatternFill('solid', fgColor=BLUE)
    ws[f'S{r}'].number_format = '0%'
for r in (16, 17, 18, 19, 20):
    ws[f'R{r}'].font = REG
for a in ('R21', 'R22'):
    ws[a].font = Font(name='Arial', size=9, italic=True)

# --- rung price columns R,S,T (rows 4-13) ---
ws['R3'] = 'Bayes Rung1'; ws['S3'] = 'Bayes Rung2'; ws['T3'] = 'Bayes Rung3'
for a in ('R3', 'S3', 'T3'):
    ws[a].font = HDR; ws[a].alignment = Alignment(horizontal='center')
for r in ROWS:
    cform = ws.cell(row=r, column=3).value            # existing Bayes BUY (rung1)
    # rung1 mirrors col C
    ws.cell(row=r, column=18, value=f'=C{r}')
    # rung2/rung3: insert depth-multiplier ref right after the "J{r}-" term
    ws.cell(row=r, column=19, value=cform.replace(f'J{r}-', f'J{r}-$S$16*', 1))
    ws.cell(row=r, column=20, value=cform.replace(f'J{r}-', f'J{r}-$S$17*', 1))
    for c in (18, 19, 20):
        cell = ws.cell(row=r, column=c); cell.font = REG; cell.number_format = CUR
for col in ('R', 'S', 'T'):
    ws.column_dimensions[col].width = 12

# --- refreshed instruction line ---
ws['A2'] = ('Each morning: for BAYES place 3 limit BUYs at the rung prices (cols R/S/T, weights '
            'in the config block) and one GTC SELL of the whole Bayes position at the Bayes target '
            '(col D). For OU place a limit BUY at col E and a GTC SELL at col F.')

wb.calculation.calcMode = 'auto'; wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print('saved', OUT)
