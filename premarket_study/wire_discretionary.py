"""
Wire the discretionary carve-out into the Allocation sheet (15 Aug 2026).

The user wants the option of carving a discretionary fund out of the main
capital for opportunistic stock picks: a 1/0 flag and an amount, netted off the
cash the machine block divides across the eighteen sleeves. Allocation-sheet
only; nothing moves, so the scripts and the order block see the same layout.

Wiring (all in previously empty cells of the INPUTS block):
  D4  'DISCRETIONARY'          label
  D5  'Flag (1/0)'             E5 = 0   (blue input cell)
  D6  'Amount ($)'             E6 = 0   (blue input cell)
  D7  'Deployed to book ($)'   E7 = MAX(0, B5 - E5*E6)   (computed)
and one in-place edit repeated over the machine block, C25:C42:
  $B$5  ->  $E$7
so the sleeves divide the NET capital. With the flag at 0 (or the amount at 0)
E7 equals B5 and every allocation is exactly what it was before. The MAX(0,..)
guard means an over-sized carve-out parks the book in cash rather than going
negative.
"""
import copy
import sys

import openpyxl


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    al = wb['Allocation']
    notes = wb['Notes']

    assert all(al.cell(row=r, column=c).value is None for r in range(4, 10) for c in range(3, 7))
    for r in range(25, 43):
        assert al.cell(row=r, column=3).value == \
            f'=IF(SUM($F$25:$F$42)=0,0,$B$5*F{r}/SUM($F$25:$F$42))', al.cell(row=r, column=3).value

    lab = al['A5']            # label style
    inp = al['B5']            # blue input style

    def put(coord, value, style_from, numfmt=None):
        c = al[coord]
        c.value = value
        c.font = copy.copy(style_from.font)
        c.fill = copy.copy(style_from.fill)
        c.border = copy.copy(style_from.border)
        c.alignment = copy.copy(style_from.alignment)
        if numfmt:
            c.number_format = numfmt
        return c

    put('D4', 'DISCRETIONARY', al['A4'])
    put('D5', 'Flag (1/0)', lab)
    put('E5', 0, inp, '0')
    put('D6', 'Amount ($)', lab)
    put('E6', 0, inp, '"$"#,##0')
    put('D7', 'Deployed to book ($)', lab)
    e7 = put('E7', '=MAX(0,$B$5-$E$5*$E$6)', lab, '"$"#,##0')
    e7.fill = copy.copy(al['A7'].fill)      # computed, not a blue input

    for r in range(25, 43):
        al.cell(row=r, column=3,
                value=f'=IF(SUM($F$25:$F$42)=0,0,$E$7*F{r}/SUM($F$25:$F$42))')

    style_a, style_b = notes['A8'], notes['B8']
    row = notes.max_row + 2
    ca = notes.cell(row=row, column=1, value='Discretionary carve-out')
    ca.font = copy.copy(style_a.font)
    ca.alignment = copy.copy(style_a.alignment)
    cb = notes.cell(
        row=row, column=2,
        value=('Allocation E5 (flag 1/0) and E6 (amount $) carve a discretionary fund out of the '
               'cash before it is divided across the sleeves: E7 = MAX(0, B5 - E5*E6) and the '
               'machine block C25:C42 now divides E7 instead of B5. Flag 0 or amount 0 reproduces '
               'the old behaviour exactly; the MAX guard means an over-sized carve-out parks the '
               'book in cash rather than going negative. Discretionary trades themselves live '
               'outside the workbook.'))
    cb.font = copy.copy(style_b.font)
    cb.alignment = copy.copy(style_b.alignment)

    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)


def diff(in_path, out_path):
    a = openpyxl.load_workbook(in_path)
    b = openpyxl.load_workbook(out_path)
    out = []
    for sn in a.sheetnames:
        wa, wc = a[sn], b[sn]
        for r in range(1, max(wa.max_row, wc.max_row) + 1):
            for c in range(1, max(wa.max_column, wc.max_column) + 1):
                va = wa.cell(row=r, column=c).value
                vb = wc.cell(row=r, column=c).value
                if va != vb:
                    out.append((sn, wa.cell(row=r, column=c).coordinate, str(va)[:45], str(vb)[:60]))
    return out


if __name__ == '__main__':
    in_path, out_path = sys.argv[1], sys.argv[2]
    wire(in_path, out_path)
    for row in diff(in_path, out_path):
        print(row)
