"""
Fill MU's missing daily OHLC in the Query sheet from the 5-minute bars.

The workbook's MU column has a 23-session gap between 25 June and 28 July 2026, where the other
four names carry data. The 5-minute file covers those sessions, so the daily bar is rebuilt from
them: open = first regular-hours bar's open, high/low = extremes across regular hours, close =
last regular-hours bar's close.

Regular hours only. The 5-minute file also carries pre- and post-market bars, and including them
reproduces the existing daily rows to a median error of 1.27% against 0.015% for regular hours
alone -- so the daily series this workbook uses is a regular-session series, and the aggregation
is validated against the 560 sessions where both sources already agree before any cell is written.
"""
import openpyxl, datetime, collections, statistics, sys

SRC = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/79803cea-TradingExcel_5stock.xlsx'
FIVE = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/94f1080f-MU_5min_Apr2024Aug2026.xlsx'
OUT = '/home/user/David-Fewer/TradingExcel_5stock_MUfilled.xlsx'
MU_COL = 18                                            # R = MU_O

# Sessions present in the workbook but rebuilt anyway. 24 June 2026 was the last row of MU's
# original source and reads as a session captured before the close: its low of 1031.00 against
# 991.10 in the 5-minute bars is a 3.87% error, the largest in the 564-session comparison, where
# the next worst is 3.20% and 99.6% fall inside 0.5%. That low feeds the all-time-high and OU
# calculations, so it is replaced rather than left.
REPAIR = {datetime.date(2026, 6, 24)}
RTH0, RTH1 = datetime.time(9, 30), datetime.time(16, 0)


def to_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    if isinstance(v, (int, float)):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
    return None


def daily_from_five():
    wb = openpyxl.load_workbook(FIVE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first: first = False; continue
        dt, o, h, l, c = row[0], row[1], row[2], row[3], row[4]
        if not isinstance(dt, datetime.datetime) or o is None: continue
        if RTH0 <= dt.time() < RTH1:
            by[dt.date()].append((dt, o, h, l, c))
    wb.close()
    out = {}
    for d, v in by.items():
        v.sort(key=lambda x: x[0])
        out[d] = (v[0][1], max(x[2] for x in v), min(x[3] for x in v), v[-1][4])
    return out


if __name__ == '__main__':
    five = daily_from_five()
    wb = openpyxl.load_workbook(SRC)
    q = wb['Query']

    # validate against every session where the workbook already has MU
    errs = []
    for r in range(2, q.max_row + 1):
        d = to_date(q.cell(r, 1).value)
        o = q.cell(r, MU_COL).value
        if d in five and isinstance(o, (int, float)):
            cur = [q.cell(r, MU_COL + j).value for j in range(4)]
            s = five[d]
            errs.append(max(abs(s[j]/cur[j] - 1) for j in range(4) if cur[j]))
    errs.sort()
    print(f'validation on {len(errs)} overlapping sessions: median '
          f'{statistics.median(errs)*100:.4f}%, 95th {errs[int(len(errs)*0.95)]*100:.3f}%')
    if statistics.median(errs) > 0.005:
        sys.exit('aggregation does not reproduce the existing rows; refusing to write')

    filled, repaired, absent = [], [], []
    for r in range(2, q.max_row + 1):
        d = to_date(q.cell(r, 1).value)
        has = isinstance(q.cell(r, MU_COL).value, (int, float))
        if has and d not in REPAIR:
            continue                                   # already has data, leave it alone
        if d not in five:
            if not has: absent.append(d)
            continue
        before = [q.cell(r, MU_COL + j).value for j in range(4)] if has else None
        for j, v in enumerate(five[d]):
            q.cell(r, MU_COL + j).value = round(v, 4)
        (repaired if has else filled).append((d, before, [round(v, 4) for v in five[d]])
                                             if has else d)

    print(f'filled {len(filled)} sessions'
          + (f': {filled[0]} -> {filled[-1]}' if filled else ''))
    for d, before, after in repaired:
        print(f'repaired {d}: {before} -> {after}')
    if absent:
        print(f'still missing (no 5-minute coverage): {len(absent)} {absent[:5]}')
    n = sum(1 for r in range(2, q.max_row + 1)
            if isinstance(q.cell(r, MU_COL).value, (int, float)))
    print(f'MU rows with data: {n} of {q.max_row - 1}')

    nt = wb['Notes']
    nt['B6'] = ('MU daily OHLC for 25 June to 28 July 2026 was rebuilt from the 5-minute bars '
                '(regular hours, 09:30-16:00 ET), and 24 June 2026 was replaced on the same '
                'basis: its stored low of 1031.00 was a part-session value against 991.10 in '
                'the bars. The aggregation reproduces every other stored session to a '
                'median 0.015%.')
    wb.save(OUT)
    print('written', OUT)
