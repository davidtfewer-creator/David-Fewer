"""
Map every stop-loss exit (date, stock, tranche, days held, loss) from the attached workbook,
then test whether the AI-linked names' stops coincide in time.
"""
import openpyxl, datetime, collections
from engine import Params, run_model

F = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/1b35dd3e-Hybrid9_Bayesian_OULIVE_PLTR_nopq.xlsx'
STOCKS = ['NVDA', 'TSM', 'TSLA', 'VRT', 'VST', 'AVGO', 'PLTR', 'RKLB', 'SOFI', 'SPOT']
# AI-linked = chips/AI-software/AI-infrastructure(power+cooling); non-AI = space, fintech, music
AI = {'NVDA', 'TSM', 'TSLA', 'VRT', 'VST', 'AVGO', 'PLTR'}
NONAI = {'RKLB', 'SOFI', 'SPOT'}


def load_book():
    wbv = openpyxl.load_workbook(F, data_only=True)
    q = wbv['Query']
    data = {}
    for i, s in enumerate(STOCKS):
        co = 2 + 4 * i
        dts, O, H, L, C = [], [], [], [], []
        for r in range(2, q.max_row + 1):
            d = q.cell(r, 1).value
            o = q.cell(r, co).value
            if not isinstance(o, (int, float)) or o <= 0:
                continue
            if isinstance(d, datetime.datetime):
                d = d.date()
            elif isinstance(d, (int, float)):
                d = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(d))
            dts.append(d); O.append(o)
            H.append(q.cell(r, co + 1).value); L.append(q.cell(r, co + 2).value)
            C.append(q.cell(r, co + 3).value)
        data[s] = (dts, O, H, L, C)
    params, cached = {}, {}
    for s in STOCKS:
        ws = wbv[f'Model {s}']; g = lambda a: ws[a].value
        params[s] = Params(lam=g('B2'), phi_L=g('D2'), psi=g('F2'), k=g('H2'),
            premium=g('J2'), peak_cap=g('L2'), comm=g('N2'), capital=g('P2'),
            interest=g('R2'), stop_days=int(g('T2')), bayes_pct=g('V2'),
            ou_W=int(g('B3')), ou_buf_k=g('D3'), ou_prem=g('H3'), ou_cap=g('J3'), years=2.2)
        cached[s] = int(g('AC4'))
    return data, params, cached


def stop_events(s, dts, O, H, L, C, p):
    r = run_model(dts, O, H, L, C, p, collect=True)
    out = []
    for tkey, tname in (('t1', 'Bayes'), ('t2', 'OU')):
        t = r.frames[tkey]
        AD, AC, AB, AV = t['AD'], t['AC'], t['AB'], t['AV']
        for i in range(len(dts)):
            if AD[i] == 1 and AC[i] is not None and AC[i] < AB[i]:
                entry = AV[i - 1] if i > 0 and AV[i - 1] is not None else None
                days = (dts[i] - entry).days if entry else None
                loss = AC[i] / AB[i] - 1.0
                out.append(dict(stock=s, tranche=tname, date=dts[i], entry=entry,
                                days=days, sale=AC[i], target=AB[i], loss=loss))
    return out, r.stop_loss_exits


if __name__ == '__main__':
    data, params, cached = load_book()
    all_ev, okall = [], True
    for s in STOCKS:
        ev, n = stop_events(s, *data[s], params[s])
        ok = (n == cached[s])
        okall &= ok
        all_ev.extend(ev)
    print('validation stop-counts vs cached AC4:', 'ALL OK' if okall else 'MISMATCH')
    print(f'current T2 (per stock):', {s: params[s].stop_days for s in STOCKS})
    all_ev.sort(key=lambda e: e['date'])

    print(f'\n=== ALL STOP-LOSS EXITS (chronological), n={len(all_ev)} ===')
    print(f'{"date":12s}{"stock":6s}{"grp":5s}{"tranche":8s}{"held(d)":>8s}{"loss%":>8s}')
    for e in all_ev:
        grp = 'AI' if e['stock'] in AI else 'non'
        print(f'{e["date"].isoformat():12s}{e["stock"]:6s}{grp:5s}{e["tranche"]:8s}'
              f'{e["days"]:>8}{e["loss"]*100:>7.1f}')

    # coincidence: group into ISO-week buckets
    print('\n=== CLUSTERING (by ISO week) ===')
    wk = collections.defaultdict(list)
    for e in all_ev:
        y, w, _ = e['date'].isocalendar()
        wk[(y, w)].append(e)
    for (y, w), evs in sorted(wk.items()):
        names = [f'{e["stock"]}' for e in evs]
        ai_n = sum(1 for e in evs if e['stock'] in AI)
        d0 = min(e['date'] for e in evs)
        flag = '  <== CLUSTER' if len(evs) >= 2 else ''
        print(f'{y}-W{w:02d} (wk of {d0.isoformat()}): {len(evs)} stops '
              f'[{ai_n} AI] {", ".join(names)}{flag}')

    # month buckets too (coarser coincidence) - by EXIT date
    print('\n=== CLUSTERING by EXIT month ===')
    mo = collections.defaultdict(list)
    for e in all_ev:
        mo[(e['date'].year, e['date'].month)].append(e)
    for (y, m), evs in sorted(mo.items()):
        names = [e['stock'] for e in evs]
        ai_n = sum(1 for e in evs if e['stock'] in AI)
        flag = '  <== CLUSTER' if len(evs) >= 3 else ''
        print(f'{y}-{m:02d}: {len(evs)} stops [{ai_n} AI] {", ".join(names)}{flag}')

    # by ENTRY month - when the losing positions were actually opened (stress date)
    print('\n=== CLUSTERING by ENTRY month (the ~50d-earlier stress) ===')
    en = collections.defaultdict(list)
    for e in all_ev:
        if e['entry']:
            en[(e['entry'].year, e['entry'].month)].append(e)
    for (y, m), evs in sorted(en.items()):
        names = [e['stock'] for e in evs]
        ai_n = sum(1 for e in evs if e['stock'] in AI)
        flag = '  <== CLUSTER' if len(evs) >= 3 else ''
        print(f'{y}-{m:02d}: {len(evs)} entries-that-stopped [{ai_n} AI] {", ".join(names)}{flag}')
