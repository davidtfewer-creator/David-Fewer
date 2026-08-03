"""
Pooled book with pre-market-directed capital allocation. NO LEVERAGE: one shared cash account,
total capital fixed, allocated each morning across the sleeves that have a live bid.

Each session, before the bell:
  * every flat sleeve computes its uncapped bid from data through yesterday
  * the pre-market VWAP gives a fill probability p (calibrated on TRAIN data only)
  * available cash is divided among the candidates by policy
      equal : evenly across all candidates          (mimics the siloed book)
      prob  : in proportion to p                    (steer toward likely fillers)
      topk  : concentrated in the K highest-p names
  * orders totalling no more than available cash are placed; cash committed to a name that
    does not fill simply stays in cash

Exits are unchanged: target sell, or the 50-day calendar stop. Idle cash earns the model rate.
"""
import openpyxl, datetime, statistics, copy
from stop_sweep import load_book
from engine import run_model

PM = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/39840daf-PreMarket_04000900ET_20240401_to_20260728.xlsx'
data, params, cached = load_book()
STOCKS = list(data)
BAYES_W = 0.75


def load_pm():
    wb = openpyxl.load_workbook(PM, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it)); col = {h: i for i, h in enumerate(hdr) if h}
    out = {}
    for row in it:
        d = row[0]
        if not isinstance(d, datetime.datetime): continue
        rec = {}
        for s in STOCKS:
            c = col.get(f'{s}_PM_VWAP'); cl = col.get(f'{s}_PM_Low')
            if c is None: continue
            v, lo = row[c], row[cl]
            if isinstance(v, (int, float)): rec[s] = (v, lo)
        out[d.date()] = rec
    wb.close(); return out


PMD = load_pm()


def build_signals():
    """Per stock: list over dates of (date, bayes_bid, ou_bid, sigma, prevclose, O,H,L,C)."""
    sig = {}
    for s in STOCKS:
        dts, O, H, L, C = data[s]
        p = params[s]
        fr = run_model(dts, O, H, L, C, p, collect=True).frames
        Lvl, Slp, W, G = fr['Lvl'], fr['Slp'], fr['W'], fr['G']
        OUf, OUsig = fr['OUf'], fr['OUsig']
        rows = []
        for i in range(1, len(C)):
            bb = min(Lvl[i-1] + Slp[i-1] - p.k*W[i-1], G[i-1]*(1-p.peak_cap))
            ob = None
            if OUf[i] is not None and OUsig[i] is not None:
                ob = min(OUf[i] - p.ou_buf_k*OUsig[i], G[i-1]*(1-p.ou_cap))
            rows.append(dict(d=dts[i], bb=bb, ob=ob, sig=W[i-1], pc=C[i-1],
                             O=O[i], H=H[i], L=L[i], C=C[i]))
        sig[s] = rows
    return sig


SIG = build_signals()
DATES = sorted({r['d'] for s in STOCKS for r in SIG[s]})
IDX = {s: {r['d']: r for r in SIG[s]} for s in STOCKS}


def calibrate(train_dates):
    """z-bucket -> empirical fill rate, from TRAIN dates only."""
    b = {}
    for s in STOCKS:
        for r in SIG[s]:
            if r['d'] not in train_dates: continue
            rec = PMD.get(r['d'], {}).get(s)
            if rec is None or r['sig'] <= 0: continue
            z = (rec[0] - r['bb'])/r['sig']
            k = min(int(z//1), 5) if z >= 0 else -1
            f, t = b.get(k, (0, 0))
            b[k] = (f + (r['L'] <= r['bb']), t + 1)
    return {k: (f/t if t else 0.3) for k, (f, t) in b.items()}


def prob(cal, s, r):
    rec = PMD.get(r['d'], {}).get(s)
    if rec is None or r['sig'] <= 0: return 0.30
    z = (rec[0] - r['bb'])/r['sig']
    k = min(int(z//1), 5) if z >= 0 else -1
    return cal.get(k, 0.30)


def run_pool(policy, cal, lo_i, hi_i, capital=1_000_000, topk=4, maxfrac=0.34):
    """Simulate the pooled book over DATES[lo_i:hi_i+1]. Returns (final equity, deployment)."""
    cash = capital
    pos = {}                                  # (stock, sleeve) -> dict(shares,target,bd)
    p0 = params[STOCKS[0]]
    deploy = []
    prev = None
    for di in range(lo_i, hi_i + 1):
        d = DATES[di]
        # ---- interest on idle cash ----
        if prev is not None:
            cash += cash * p0.interest * (d - prev).days / 365.0
        prev = d
        # ---- exits ----
        for key in list(pos):
            s, sl = key
            r = IDX[s].get(d)
            if r is None: continue
            P = pos[key]
            stop = (d - P['bd']).days >= params[s].stop_days
            if r['H'] >= P['tgt']:
                cash += P['sh'] * (P['tgt'] - params[s].comm); del pos[key]
            elif stop:
                cash += P['sh'] * (r['O'] - params[s].comm); del pos[key]
        # ---- candidates ----
        cand = []
        for s in STOCKS:
            r = IDX[s].get(d)
            if r is None: continue
            for sl, bid, w in (('B', r['bb'], BAYES_W), ('O', r['ob'], 1-BAYES_W)):
                if bid is None or (s, sl) in pos: continue
                cand.append((s, sl, bid, w, prob(cal, s, r), r))
        if not cand or cash <= 0:
            deploy.append(0.0); continue
        # ---- allocate ----
        if policy == 'equal':
            wts = {(c[0], c[1]): 1.0 for c in cand}
        elif policy == 'prob':
            wts = {(c[0], c[1]): max(c[4], 1e-6) for c in cand}
        elif policy == 'topk':
            top = sorted(cand, key=lambda c: -c[4])[:topk]
            keys = {(c[0], c[1]) for c in top}
            wts = {(c[0], c[1]): (1.0 if (c[0], c[1]) in keys else 0.0) for c in cand}
        tw = sum(wts.values()) or 1.0
        alloc = {k: cash*min(v/tw, maxfrac) for k, v in wts.items()}
        # ---- fills ----
        used = 0.0
        for (s, sl, bid, w, p_, r) in cand:
            a = alloc.get((s, sl), 0.0)
            if a <= 0: continue
            if r['L'] <= bid:                                   # fills
                sh = a/(bid + params[s].comm)
                prem = params[s].premium if sl == 'B' else params[s].ou_prem
                pos[(s, sl)] = dict(sh=sh, tgt=bid + r['pc']*prem, bd=d)
                cash -= a; used += a
        # ---- mark ----
        mtm = sum(P['sh']*IDX[s][d]['C'] for (s, sl), P in pos.items() if d in IDX[s])
        eq = cash + mtm
        deploy.append(mtm/eq if eq > 0 else 0)
    mtm = sum(P['sh']*IDX[s][DATES[hi_i]]['C'] for (s, sl), P in pos.items() if DATES[hi_i] in IDX[s])
    return cash + mtm, statistics.mean(deploy)


if __name__ == '__main__':
    N = len(DATES); mid = N//2
    train = set(DATES[:mid]); cal = calibrate(train)
    print('calibration (train half): z-bucket -> fill rate')
    print('  ', {k: round(v, 2) for k, v in sorted(cal.items())})
    yrs = (DATES[-1] - DATES[mid]).days/365.25
    print(f'\n=== POOLED BOOK, TEST HALF ({DATES[mid]} -> {DATES[-1]}, {yrs:.2f}y) ===')
    print(f'{"policy":16s}{"final":>14s}{"annualised":>12s}{"mean deployed":>15s}')
    print('-'*57)
    for pol, kw in (('equal', {}), ('prob', {}), ('topk (K=4)', {'topk': 4}),
                    ('topk (K=6)', {'topk': 6})):
        p = pol.split()[0]
        fin, dep = run_pool(p, cal, mid, N-1, **kw)
        ann = (fin/1_000_000)**(1/yrs) - 1
        print(f'{pol:16s}{fin:>14,.0f}{ann*100:>11.0f}%{dep*100:>14.0f}%')
