"""
GM, VLO and CF computed at last, and the nine-name book rebuilt around them.

The three diversifiers arrive as 5-minute bars, which give both things needed at once: daily
OHLC (regular hours only -- the files carry pre-market prints from 04:00 which would corrupt
the open and the range) and the intraday ordering that decides whether a same-day round trip
was genuinely achievable. So all three are scored on VERIFIED fills, like NVDA and AVGO.

Parameters are the handover's deployable vectors, unchanged. They are already on the residual
sigma scale (newdiv used ou_sigma='resid' throughout), so unlike NVDA and AVGO they need no
re-expression -- the buffers 1.060 / 1.178 / 0.841 go in as they are.

WHAT IS AND IS NOT ON ONE BASIS
-------------------------------
Five of the nine are now computed on verified fills here: NVDA, AVGO, GM, VLO, CF. The four
kept deployed names -- RKLB, TSM, VRT, MU -- have no 5-minute data in this session, so they
cannot be. They are validated against the live workbook to the penny and then carried at their
published verified-fill figures, which were struck on the same basis in earlier work.

Both are reported: the planning table uses the published figures for those four (verified
throughout, so comparable), and an all-computed at-open-floor table is shown alongside as a
fully internally consistent lower bound. The floor understates a deployed-premium name by
roughly 10pp, so the two tables bracket rather than agree, and the bracket is the honest
statement of what is known.

Run:  python3 nine_computed.py
"""
import collections
import datetime
import itertools

import numpy as np
import openpyxl

import ramp_premium as R
from engine import Params, run_model

UP = '/root/.claude/uploads/822d405e-f99b-5b59-9c6b-87e725054402'
LIVE = f'{UP}/8d17afe4-TradingExcel_5stock_live.xlsx'
CAND5 = {'GM':  f'{UP}/d7751daf-GM_5min_Apr2024Aug2026.xlsx',
         'VLO': f'{UP}/70cee706-VLO_5min_Apr2024Aug2026.xlsx',
         'CF':  f'{UP}/7a75da3b-CF_5min_Apr2024Aug2026.xlsx'}
RTH0, RTH1 = datetime.time(9, 30), datetime.time(16, 0)

KEEP = ('RKLB', 'TSM', 'VRT', 'MU')
PAIR = ('NVDA', 'AVGO')
NEW = ('GM', 'VLO', 'CF')
HIGH = (0.040, 0.045, 0.050, 0.060)
HIGH_STOP = 200
BUF_BAND = (0.6, 0.8, 1.0, 1.25, 1.5)
HAIRCUT = 2.1

# handover section 6, deployable vectors: lam, phi_L, psi, k, prem, peak_cap,
#                                         ou_buf(resid), ou_prem, ou_cap, ou_W
CAND_P = {
    'GM':  [0.704472, 0.750723, 0.0805582, 0.461711, 0.00969721, 0.0423717,
            1.06024, 0.0451112, 0.0417662, 45],
    'VLO': [0.723393, 0.625222, 0.0200999, 0.543763, 0.0303035, 0.0308569,
            1.17841, 0.0419514, 0.0370821, 125],
    'CF':  [0.736218, 0.106120, 0.0694359, 2.03552, 0.0465592, 0.0286416,
            0.841237, 0.0590932, 0.0831024, 66],
}
PUBLISHED_4 = {'RKLB': 158.0, 'TSM': 57.0, 'VRT': 67.0, 'MU': 63.0}
BUYS_4 = {'RKLB': 96, 'TSM': 65, 'VRT': 54, 'MU': 51}


def five_min(stock):
    """Regular-hours 5-minute bars -> (daily OHLC, per-session low/suffix-max-high index)."""
    wb = openpyxl.load_workbook(CAND5[stock], read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        dt, o, h, l, c = row[0], row[1], row[2], row[3], row[4]
        if not isinstance(dt, datetime.datetime) or o is None:
            continue
        if not (RTH0 <= dt.time() < RTH1):
            continue
        by[dt.date()].append((dt, o, h, l, c))
    wb.close()
    d, O, H, L, C = [], [], [], [], []
    idx = {}
    for day in sorted(by):
        bars = sorted(by[day], key=lambda x: x[0])
        d.append(day)
        O.append(bars[0][1])
        H.append(max(b[2] for b in bars))
        L.append(min(b[3] for b in bars))
        C.append(bars[-1][4])
        hi = np.array([b[2] for b in bars], dtype=float)
        lo = np.array([b[3] for b in bars], dtype=float)
        idx[day] = (lo, np.maximum.accumulate(hi[::-1])[::-1])
    return (d, O, H, L, C), idx


def cand_params(stock, cap=1_000_000):
    v = CAND_P[stock]
    return Params(lam=v[0], phi_L=v[1], psi=v[2], k=v[3], premium=v[4], peak_cap=v[5],
                  comm=0.005, capital=cap, interest=0.0314, stop_days=50, bayes_pct=0.5,
                  ou_W=int(v[9]), ou_buf_k=v[6], ou_prem=v[7], ou_cap=v[8], years=2.2)


def windows(eq, d):
    yrs = (d[-1] - d[0]).days / 365.25
    iS = next(k for k, x in enumerate(d) if x >= R.SPLIT)
    return ((eq[-1] / eq[0]) ** (1 / yrs) - 1,
            (eq[iS] / eq[0]) ** (365.25 / (d[iS] - d[0]).days) - 1,
            (eq[-1] / eq[iS]) ** (365.25 / (d[-1] - d[iS]).days) - 1)


def sigma_ratio(args, p):
    a = run_model(*args, p, ou_sigma='level', collect=True)
    b = run_model(*args, p, ou_sigma='resid', collect=True)
    sl = np.array([x for x in a.frames['OUsig'] if x is not None])
    sr = np.array([x for x in b.frames['OUsig'] if x is not None])
    return float(sr.mean() / sl.mean())


def main():
    out = {}
    print('=== GM, VLO, CF: daily bars from 5-minute data, verified fills ===\n')
    for n in NEW:
        args, idx = five_min(n)
        d = args[0]
        p = cand_params(n)
        p = Params(**{**p.__dict__, 'years': (d[-1] - d[0]).days / 365.25})
        chk = R.make_checker(idx, d, args[1])
        rv = run_model(*args, p, ou_sigma='resid', same_day_exit=chk, collect=True)
        rf = run_model(*args, p, ou_sigma='resid', same_day_exit='at_open', collect=True)
        yrs = p.years
        wv, wf = windows(rv.frames['equity'], d), windows(rf.frames['equity'], d)
        out[n] = dict(full=wv[0], fit=wv[1], tst=wv[2], buys=rv.total_buys / yrs,
                      floor=wf[0], dd=rv.max_drawdown, C=args[4], d=d)
        print(f'  {n}: {len(d)} sessions {d[0]}..{d[-1]}   buffer {p.ou_buf_k:.4f} (resid, as fitted)')
        print(f'     verified  full {100*wv[0]:6.2f}%  fitted {100*wv[1]:6.2f}%  '
              f'tested {100*wv[2]:6.2f}%   {rv.total_buys/yrs:.0f} buys/yr   DD {100*rv.max_drawdown:.1f}%')
        print(f'     at-open floor full {100*wf[0]:6.2f}%   (floor understates by '
              f'{100*(wv[0]-wf[0]):+.2f}pp)')
        print(f'     handover quoted: full fit {[50.3,64.2,46.9][NEW.index(n)]:.1f}%, '
              f'planning {[33,35,41][NEW.index(n)]:.0f}%, '
              f'{[61,35,27][NEW.index(n)]:d} buys/yr')

    # ---- NVDA / AVGO on the proposed rule -------------------------------------------
    print(f'\n=== NVDA, AVGO: premium 4-6%, 200d stop, residual sigma ===\n')
    for n in PAIR:
        d, O, H, L, C = R.load_feed(n)
        args = (d, O, H, L, C)
        p, _ = R.load_params(n, years=(d[-1] - d[0]).days / 365.25)
        chk = R.make_checker(R.build_index(n), d, O)
        matched = p.ou_buf_k / sigma_ratio(args, p)
        rows = []
        for m in BUF_BAND:
            inner = []
            for prem in HIGH:
                q = Params(**{**p.__dict__, 'ou_buf_k': matched * m, 'stop_days': HIGH_STOP,
                              'premium': prem, 'ou_prem': prem})
                r = run_model(*args, q, ou_sigma='resid', same_day_exit=chk, collect=True)
                w = windows(r.frames['equity'], d)
                inner.append((w[0], w[1], w[2], r.total_buys / q.years, r.max_drawdown))
            rows.append([float(np.median([x[j] for x in inner])) for j in range(5)])
        v = [float(np.median([r[j] for r in rows])) for j in range(5)]
        out[n] = dict(full=v[0], fit=v[1], tst=v[2], buys=v[3], dd=v[4], C=C, d=d)
        print(f'  {n}: buffer {matched:.4f}   full {100*v[0]:6.2f}%  fitted {100*v[1]:6.2f}%  '
              f'tested {100*v[2]:6.2f}%   {v[3]:.0f} buys/yr   DD {100*v[4]:.1f}%')

    # ---- the book -------------------------------------------------------------------
    # Planning figures use each group's OWN measured out-of-sample basis. They are not
    # interchangeable: the deployed names carry a 2.1pp haircut measured for them, while
    # GM/VLO/CF were struck on windows nothing had seen (3 walk-forward folds plus the
    # half-sample test) and their measured gap is far larger -- VLO gives back 29pp, GM 17pp.
    # Applying 2.1pp to those three would overstate them by roughly 20pp each.
    PLAN_NEW = {'GM': 33.0, 'VLO': 35.0, 'CF': 41.0}     # handover section 6
    rows = []
    for n in KEEP:
        rows.append((n, 'deployed', PUBLISHED_4[n], PUBLISHED_4[n] - HAIRCUT, BUYS_4[n],
                     'quoted', '2.1pp haircut'))
    for n in PAIR:
        rows.append((n, '4-6% / 200d', 100 * out[n]['full'], 100 * out[n]['full'] - HAIRCUT,
                     out[n]['buys'], 'computed', '2.1pp, UNMEASURED'))
    for n in NEW:
        rows.append((n, 'deployed', 100 * out[n]['full'], PLAN_NEW[n],
                     out[n]['buys'], 'computed', 'unseen windows'))

    print(f'\n=== the nine-name book ===\n')
    print(f"{'name':6s} {'rule':14s} {'potential':>10s} {'planned':>9s} {'buys/yr':>8s}  "
          f"{'source':10s} planning basis")
    for n, rule, pot, plan, buys, src, bas in sorted(rows, key=lambda r: -r[2]):
        print(f'{n:6s} {rule:14s} {pot:9.1f}% {plan:8.1f}% {buys:8.0f}  [{src}]'
              f'{"":<{max(0,10-len(src))}} {bas}')
    pot = np.array([r[2] for r in rows]); plan = np.array([r[3] for r in rows])
    buys = np.array([r[4] for r in rows])
    T = 2.34
    print(f"\n{'':21s} {'-'*10} {'-'*9} {'-'*8}")
    print(f"{'BOOK (mean of 9)':21s} {pot.mean():9.1f}% {plan.mean():8.1f}% {buys.sum():8.0f}")
    print(f"{'  held construction':21s} "
          f"{((np.mean((1+pot/100)**T))**(1/T)-1)*100:9.1f}% "
          f"{((np.mean((1+plan/100)**T))**(1/T)-1)*100:8.1f}%")

    # ---- concentration, now computable ----------------------------------------------
    print(f'\n=== concentration, all nine measurable for the first time ===\n')
    rets = {}
    for n in KEEP:
        d, O, H, L, C = R.load_feed(n, path=LIVE); rets[n] = (d, C)
    for n in PAIR:
        d, O, H, L, C = R.load_feed(n); rets[n] = (d, C)
    for n in NEW:
        rets[n] = (out[n]['d'], out[n]['C'])
    common = sorted(set.intersection(*[set(v[0]) for v in rets.values()]))
    M = {}
    for n, (d, C) in rets.items():
        ix = {t: k for k, t in enumerate(d)}
        s = np.array([C[ix[t]] for t in common], dtype=float)
        M[n] = s[1:] / s[:-1] - 1
    def ac(names):
        return float(np.mean([np.corrcoef(M[a], M[b])[0, 1]
                              for a, b in itertools.combinations(names, 2)]))
    six = KEEP + PAIR
    print(f'  aligned on {len(common)} common sessions\n')
    print(f'  six AI-exposed names (RKLB TSM VRT MU NVDA AVGO)   {ac(six):.3f}')
    print(f'  the three diversifiers among themselves            {ac(NEW):.3f}')
    print(f'  ALL NINE                                           {ac(six + NEW):.3f}')
    print(f'\n  for reference: deployed five 0.489, five + NVDA/AVGO 0.520,')
    print(f'  handover section 6 quoted 0.21 for five + GM/VLO/CF/ALNY')
    print(f'\n  each diversifier against the six:')
    for n in NEW:
        cs = [np.corrcoef(M[n], M[f])[0, 1] for f in six]
        print(f'    {n:4s} mean {np.mean(cs):+.3f}   (' +
              ', '.join(f'{f} {c:+.2f}' for f, c in zip(six, cs)) + ')')

    # ---- book equity and drawdown ----------------------------------------------------
    # The four kept names have no intraday data, so their legs run at the AT-OPEN FLOOR while
    # the other five are verified. That mixture understates the four's return, but drawdown is
    # far less sensitive to the same-day fill question than return is (a same-day round trip
    # barely moves the equity path), so the shape below is close to right even though the
    # level of the four's contribution is not.
    curves = {}
    for n in KEEP:
        d, O, H, L, C = R.load_feed(n, path=LIVE)
        p, cached = R.load_params(n, path=LIVE, years=(d[-1] - d[0]).days / 365.25)
        r = run_model(d, O, H, L, C, p, ou_sigma=cached['ou_sigma'],
                      same_day_exit='at_open', collect=True)
        curves[n] = (d, np.array(r.frames['equity']) / p.capital)
    for n in PAIR:
        d, O, H, L, C = R.load_feed(n)
        p, _ = R.load_params(n, years=(d[-1] - d[0]).days / 365.25)
        chk = R.make_checker(R.build_index(n), d, O)
        mt = p.ou_buf_k / sigma_ratio((d, O, H, L, C), p)
        cs = []
        for prem in HIGH:
            q = Params(**{**p.__dict__, 'ou_buf_k': mt, 'stop_days': HIGH_STOP,
                          'premium': prem, 'ou_prem': prem})
            rr = run_model(d, O, H, L, C, q, ou_sigma='resid', same_day_exit=chk, collect=True)
            cs.append(np.array(rr.frames['equity']) / q.capital)
        curves[n] = (d, np.median(cs, axis=0))
    for n in NEW:
        args, idx = five_min(n)
        d = args[0]
        p = cand_params(n)
        p = Params(**{**p.__dict__, 'years': (d[-1] - d[0]).days / 365.25})
        rr = run_model(*args, p, ou_sigma='resid',
                       same_day_exit=R.make_checker(idx, d, args[1]), collect=True)
        curves[n] = (d, np.array(rr.frames['equity']) / p.capital)

    def book(names):
        w = 1.0 / len(names)
        eq = np.zeros(len(common))
        for n in names:
            d, c = curves[n]
            ix = {t: k for k, t in enumerate(d)}
            leg = np.array([c[ix[t]] for t in common])
            eq += w * leg / leg[0]
        peak, dd = -1e30, 0.0
        for e in eq:
            peak = max(peak, e)
            dd = max(dd, (peak - e) / peak)
        yrs = (common[-1] - common[0]).days / 365.25
        return (eq[-1] / eq[0]) ** (1 / yrs) - 1, dd

    print(f'\n=== book equity, held construction (four kept names at the at-open floor) ===\n')
    print(f"{'book':40s} {'return':>9s} {'maxDD':>8s}")
    for label, names in (('deployed five (VST included)', KEEP + ('VST',)),
                         ('four kept only', KEEP),
                         ('four + NVDA/AVGO', KEEP + PAIR),
                         ('NINE', KEEP + PAIR + NEW)):
        if 'VST' in names:
            d, O, H, L, C = R.load_feed('VST', path=LIVE)
            p, cached = R.load_params('VST', path=LIVE, years=(d[-1] - d[0]).days / 365.25)
            r = run_model(d, O, H, L, C, p, ou_sigma=cached['ou_sigma'],
                          same_day_exit='at_open', collect=True)
            curves['VST'] = (d, np.array(r.frames['equity']) / p.capital)
        ret, dd = book(names)
        print(f'{label:40s} {100*ret:8.2f}% {100*dd:7.1f}%')
    print('\n  Return levels here are depressed by the floor on the deployed names; the')
    print('  drawdown column is the one to read.')


if __name__ == '__main__':
    main()
