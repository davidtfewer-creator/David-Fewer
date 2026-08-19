"""
The midday re-pool: a second allocation session at 13:00.

Theme (user, 19 Aug 2026): the survivors of the overlay program were capital-
routing rules, not model changes. The pool currently allocates ONCE, at the
open; after that, two kinds of capital sit dead for the rest of the day --
allocations behind orders that will not fill, and proceeds from round trips
completed in the morning. A 13:00 session re-routes both.

THE SESSION, mechanical and price-free (veto, never amend):
  cancel   unfilled orders whose model bid sits more than y below the 13:00
           tape (last 5-minute close before 13:00) -- their cash frees up;
  re-admit sleeves benched by the 09:00 pre-market rule whose bid is now
           within y of the tape (the veto is per-session, not per-day);
  re-pool  freed cash + proceeds banked before 13:00 (same-day round trips
           completed in the morning, held-position target exits provably in
           the morning, stop exits at the open) as an equal top-up across all
           live afternoon orders -- at their UNCHANGED model prices.
  (option) re-arm sleeves whose same-day round trip completed before 13:00:
           the same bid goes back in the book for the afternoon.

Everything runs on top of the adopted pre-market rule (09:00 cutoff, 4%). The
only fitted quantity is y: train-half pick, frozen, tested-half verdict.
Fills are verified with intraday ordering inside each segment: a fill exits
same-day only if a high AFTER the bid was touched reaches the target; where
the morning and afternoon are ambiguous the exit is booked to the afternoon
(conservative for the re-pool, which only recycles provably-morning cash).

VALIDATION: with the midday session off, the simulator must closely reproduce
book_sim's pooled result under the same pre-market rule (118.4% tested half).

Days without 5-minute coverage use the conservative daily fallback (same-day
exit only when the fill is provably at the open) and skip the midday session,
as do early-close days with no afternoon bars.
"""
import collections
import datetime
import json
import os
import pickle

import numpy as np

from engine import Params
from fresh_opt import SPLIT, annualise
from fresh_opt_cands import aw_params
from book_sim import load_all, NAMES as N8
from minute_index import DIR, RTH_START, RTH_END

NAMES = ['TSM', 'VRT', 'VST', 'RKLB', 'MU', 'GM', 'VLO', 'CF', 'MRVL']
NOON = datetime.time(13, 0)
COMM = 0.005
INTEREST = 0.0314
STOP_DAYS = 50
SEG_CACHE = os.path.join(DIR, 'midday_index.pkl')


# ------------------------------------------------------------ segment index
def build_segments():
    """name -> date -> (am_lows, am_suf, pm_lows, pm_suf, px13).
    am_suf / pm_suf are PURE within-segment suffix maxima of the highs."""
    if os.path.exists(SEG_CACHE):
        with open(SEG_CACHE, 'rb') as f:
            return pickle.load(f)
    import openpyxl
    out = {}
    for s in NAMES:
        wb = openpyxl.load_workbook(os.path.join(DIR, f'{s}_5min.xlsx'),
                                    read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        by = collections.defaultdict(list)
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            dt, h, l, c = row[0], row[2], row[3], row[4]
            if isinstance(dt, str):
                try:
                    dt = datetime.datetime.fromisoformat(dt)
                except ValueError:
                    continue
            if not isinstance(dt, datetime.datetime) or h is None:
                continue
            t = dt.time()
            if t < RTH_START or t >= RTH_END:
                continue
            by[dt.date()].append((dt, float(h), float(l), float(c)))
        wb.close()
        idx = {}
        for d, bars in by.items():
            bars.sort()
            am = [(h, l, c) for (dt, h, l, c) in bars if dt.time() < NOON]
            pm = [(h, l, c) for (dt, h, l, c) in bars if dt.time() >= NOON]
            if not am:
                continue
            am_lows = np.array([b[1] for b in am])
            am_suf = np.maximum.accumulate(np.array([b[0] for b in am])[::-1])[::-1]
            pm_lows = np.array([b[1] for b in pm]) if pm else np.array([])
            pm_suf = (np.maximum.accumulate(np.array([b[0] for b in pm])[::-1])[::-1]
                      if pm else np.array([]))
            idx[d] = (am_lows, am_suf, pm_lows, pm_suf, am[-1][2])
        out[s] = idx
        print(f'  segments {s}: {len(idx)} days', flush=True)
    with open(SEG_CACHE, 'wb') as f:
        pickle.dump(out, f)
    return out


def load_pm0900():
    with open('data_pm/pm_last_cuts.pkl', 'rb') as f:
        return pickle.load(f)['09:00']


def load_book_data():
    t0 = Params(capital=1_000_000, comm=0.005, interest=0.0314, stop_days=50,
                bayes_pct=0.5, years=2.2, ou_W=80)
    mrvl = aw_params(json.load(open('fresh_opt_cands.json'))['MRVL']['reference']['vec'], t0)
    return load_all(names=N8 + ['MRVL'], params_override={'MRVL': mrvl})


# ------------------------------------------------------------ diagnostics
def diagnostics(data, sleeves, seg):
    am_touch, pm_first, tot = 0, 0, 0
    curve = [(-0.01, 0.0, [0, 0]), (-0.02, -0.01, [0, 0]), (-0.03, -0.02, [0, 0]),
             (-0.05, -0.03, [0, 0]), (-1.0, -0.05, [0, 0])]
    for s in sleeves:
        nd = data[s['name']]
        sg = seg[s['name']]
        for i in range(1, len(nd['dts'])):
            bid = nd[s['bids']][i]
            d = nd['dts'][i]
            if bid is None or d not in sg:
                continue
            am_lows, am_suf, pm_lows, pm_suf, px13 = sg[d]
            hit_am = bool((am_lows <= bid + 1e-12).any())
            hit_pm = len(pm_lows) > 0 and bool((pm_lows <= bid + 1e-12).any())
            tot += 1
            if hit_am:
                am_touch += 1
            elif hit_pm:
                pm_first += 1
            if not hit_am and len(pm_lows):
                dep = bid / px13 - 1
                for lo, hi, cnt in curve:
                    if lo <= dep < hi:
                        cnt[0] += 1
                        cnt[1] += 1 if hit_pm else 0
                        break
    print(f'DIAGNOSTIC: {tot} live order-days; touched before 13:00: {am_touch} '
          f'({am_touch/tot*100:.1f}%); first touch after 13:00: {pm_first} '
          f'({pm_first/tot*100:.1f}%)', flush=True)
    print('afternoon fill curve (orders unfilled at 13:00, depth vs 13:00 tape):', flush=True)
    for lo, hi, (n, f) in curve:
        if n:
            print(f'  {lo*100:4.0f}% to {hi*100:3.0f}%: {n:5d} order-days, '
                  f'PM fill rate {f/n*100:5.1f}%', flush=True)


# ------------------------------------------------------------ the simulator
def simulate(data, sleeves, cal, seg, pm0900, capital=9_000_000,
             pm_thr=0.04, midday=None, rearm=False):
    for s in sleeves:
        s.update(holding=False, shares=0.0, target=None, entry=None)
    cash = capital
    eq_curve = []
    fills = stops = 0
    prev_d = cal[0]
    for d in cal:
        gap = (d - prev_d).days
        if gap:
            cash *= 1 + INTEREST * gap / 365.0
        prev_d = d
        evening = 0.0        # usable next morning
        noon_pool = 0.0      # usable at 13:00
        live_pm = []         # orders resting into the afternoon

        # ---- morning state under the pre-market rule
        active, benched = [], []
        for s in sleeves:
            nd = data[s['name']]
            i = nd['idx'][d]
            bid = nd[s['bids']][i]
            s['_i'] = i
            s['_bid'] = bid if (bid is not None and i > 0) else None
            s['_alloc'] = 0.0
            if s['holding'] or s['_bid'] is None:
                continue
            pml = pm0900[s['name']].get(d)
            if pml is not None and s['_bid'] < pml * (1 - pm_thr):
                benched.append(s)
            else:
                active.append(s)
        if active:
            a = cash / len(active)
            for s in active:
                s['_alloc'] = a
            cash = 0.0

        # ---- exits on positions held from before today
        for s in sleeves:
            if not s['holding'] or s['entry'] == d:
                continue
            nd = data[s['name']]
            i = s['_i']
            sg = seg[s['name']].get(d)
            stop = (d - s['entry']).days >= STOP_DAYS
            px = None
            am_exit = False
            if nd['H'][i] >= s['target'] - 1e-12:
                px = s['target']
                if sg is not None and len(sg[1]):
                    am_exit = bool(sg[1][0] >= s['target'] - 1e-12)
            elif stop:
                px = nd['O'][i]
                am_exit = True                     # stop sells at the open
            if px is not None:
                proceeds = s['shares'] * (px - COMM)
                if px < s['target'] - 1e-12:
                    stops += 1
                if am_exit and midday is not None:
                    noon_pool += proceeds
                else:
                    evening += proceeds
                s.update(holding=False, shares=0.0, target=None, entry=None)

        # ---- morning fills
        for s in active:
            nd = data[s['name']]
            i = s['_i']
            bid = s['_bid']
            budget = s['_alloc']
            sg = seg[s['name']].get(d)
            if sg is None:
                # conservative daily fallback, no midday participation
                if nd['L'][i] <= bid + 1e-12:
                    fills += 1
                    shares = budget / (bid + COMM)
                    tgt = bid + nd['C'][i - 1] * s['prem']
                    if bid >= nd['O'][i] - 1e-9 and nd['H'][i] >= tgt - 1e-12:
                        evening += shares * (tgt - COMM)
                    else:
                        s.update(holding=True, shares=shares, target=tgt, entry=d)
                else:
                    evening += budget
                s['_alloc'] = 0.0
                continue
            am_lows, am_suf, pm_lows, pm_suf, px13 = sg
            pm_max = pm_suf[0] if len(pm_suf) else -1e30
            hit = am_lows <= bid + 1e-12
            if hit.any():
                j = int(np.argmax(hit))
                fills += 1
                shares = budget / (bid + COMM)
                tgt = bid + nd['C'][i - 1] * s['prem']
                s['_alloc'] = 0.0
                if am_suf[j] >= tgt - 1e-12:
                    # provably exited before 13:00
                    proceeds = shares * (tgt - COMM)
                    if midday is not None:
                        noon_pool += proceeds
                        if rearm:
                            live_pm.append(s)     # same bid re-armed for the PM
                    else:
                        evening += proceeds
                elif pm_max >= tgt - 1e-12:
                    # exits same day, but in the afternoon
                    evening += shares * (tgt - COMM)
                else:
                    s.update(holding=True, shares=shares, target=tgt, entry=d)
            else:
                # unfilled at 13:00
                if midday is None:
                    # order rests all day (single-session behaviour)
                    if len(pm_lows) and (pm_lows <= bid + 1e-12).any():
                        j = int(np.argmax(pm_lows <= bid + 1e-12))
                        fills += 1
                        shares = budget / (bid + COMM)
                        tgt = bid + nd['C'][i - 1] * s['prem']
                        if pm_suf[j] >= tgt - 1e-12:
                            evening += shares * (tgt - COMM)
                        else:
                            s.update(holding=True, shares=shares, target=tgt, entry=d)
                    else:
                        evening += budget
                    s['_alloc'] = 0.0
                elif len(pm_lows) == 0:
                    evening += budget             # early close: no PM session
                    s['_alloc'] = 0.0
                elif bid >= px13 * (1 - midday):
                    live_pm.append(s)             # survivor keeps its alloc
                else:
                    noon_pool += budget           # cancelled
                    s['_alloc'] = 0.0

        # ---- 13:00 session
        if midday is not None:
            for s in benched:
                if s['holding']:
                    continue
                sg = seg[s['name']].get(d)
                if sg is None or len(sg[2]) == 0:
                    continue
                if s['_bid'] >= sg[4] * (1 - midday):
                    live_pm.append(s)             # re-admitted, funded by top-up
            if live_pm and noon_pool > 0:
                top = noon_pool / len(live_pm)
                for s in live_pm:
                    s['_alloc'] += top
                noon_pool = 0.0
            for s in live_pm:
                nd = data[s['name']]
                i = s['_i']
                bid = s['_bid']
                budget = s['_alloc']
                s['_alloc'] = 0.0
                if budget <= 0:
                    continue
                am_lows, am_suf, pm_lows, pm_suf, px13 = seg[s['name']][d]
                if len(pm_lows) == 0:
                    evening += budget
                    continue
                hit = pm_lows <= bid + 1e-12
                if hit.any():
                    j = int(np.argmax(hit))
                    fills += 1
                    shares = budget / (bid + COMM)
                    tgt = bid + nd['C'][i - 1] * s['prem']
                    if pm_suf[j] >= tgt - 1e-12:
                        evening += shares * (tgt - COMM)
                    else:
                        s.update(holding=True, shares=shares, target=tgt, entry=d)
                else:
                    evening += budget
            evening += noon_pool                  # nothing live to fund

        cash += evening
        mv = sum(s['shares'] * data[s['name']]['C'][s['_i']]
                 for s in sleeves if s['holding'])
        eq_curve.append(cash + mv)

    N = len(cal)
    cut = next((i for i, dd in enumerate(cal) if dd >= SPLIT), N - 1)
    peak, dd_ = -1e30, 0.0
    for e in eq_curve:
        peak = max(peak, e)
        if peak > 0:
            dd_ = max(dd_, (peak - e) / peak)
    return dict(full=annualise(eq_curve[-1] / eq_curve[0] - 1, cal, 0, N - 1),
                train=annualise(eq_curve[cut] / eq_curve[0] - 1, cal, 0, cut),
                test=annualise(eq_curve[-1] / eq_curve[cut] - 1, cal, cut, N - 1),
                maxdd=dd_, fills=fills, stops=stops)


def main():
    print('building segment index...', flush=True)
    seg = build_segments()
    pm0900 = load_pm0900()
    data, sleeves, cal = load_book_data()
    diagnostics(data, sleeves, seg)

    print('\nVALIDATION (midday off, pre-market rule on):', flush=True)
    r = simulate(data, sleeves, cal, seg, pm0900, midday=None)
    print(f'  single-session: full {r["full"]*100:6.1f}%  train {r["train"]*100:5.1f}%  '
          f'test {r["test"]*100:6.1f}%  DD {r["maxdd"]*100:.1f}%  '
          f'(book_sim reference: 87.3 / 59.5 / 118.4 / 25.8)', flush=True)

    print('\nINTERVENTION grid (on top of the 09:00/4% pre-market rule):', flush=True)
    cells = []
    for y in (0.015, 0.02, 0.03, 0.05):
        r = simulate(data, sleeves, cal, seg, pm0900, midday=y)
        cells.append((y, r))
        print(f'  midday y={y*100:.1f}%: full {r["full"]*100:6.1f}%  train {r["train"]*100:5.1f}%  '
              f'TEST {r["test"]*100:6.1f}%  DD {r["maxdd"]*100:4.1f}%  fills {r["fills"]}',
              flush=True)
    best = max(cells, key=lambda c: c[1]['train'])
    print(f'  train-pick y={best[0]*100:.1f}% -> frozen TEST {best[1]["test"]*100:.1f}%', flush=True)
    r = simulate(data, sleeves, cal, seg, pm0900, midday=best[0], rearm=True)
    print(f'  + re-arm at y={best[0]*100:.1f}%: full {r["full"]*100:6.1f}%  '
          f'train {r["train"]*100:5.1f}%  TEST {r["test"]*100:6.1f}%  DD {r["maxdd"]*100:4.1f}%',
          flush=True)


if __name__ == '__main__':
    main()
