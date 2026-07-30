"""
When does the OU sleeve protect us, and should we tilt capital toward Bayes?

Three measurements per stock (frozen params, this workbook):
  1. STANDALONE   - Bayes-only (bayes_pct=1) vs OU-only (bayes_pct=0): ann/Sharpe/maxDD/trades/stops.
  2. CO-MOVEMENT  - on days the Bayes sleeve falls, what does the OU sleeve do? Fraction of
                    down-Bayes days where OU is up (genuine offset), and each sleeve's own maxDD.
  3. SPLIT SWEEP  - book ann/Sharpe/maxDD as bayes_pct goes 0 -> 1. Where is the optimum?

Book-level aggregation uses the Allocation-sheet 'included' names and weights so the split
verdict reflects the actual portfolio, not an equal-weight average.
"""
import openpyxl, datetime, math, statistics, copy
from engine import Params, run_model

F = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/1b35dd3e-Hybrid9_Bayesian_OULIVE_PLTR_nopq.xlsx'
STOCKS = ['NVDA', 'TSM', 'TSLA', 'VRT', 'VST', 'AVGO', 'PLTR', 'RKLB', 'SOFI', 'SPOT']


def load_book():
    wbv = openpyxl.load_workbook(F, data_only=True)
    q = wbv['Query']
    data = {}
    for i, s in enumerate(STOCKS):
        co = 2 + 4 * i
        dts, O, H, L, C = [], [], [], [], []
        for r in range(2, q.max_row + 1):
            d = q.cell(r, 1).value; o = q.cell(r, co).value
            if not isinstance(o, (int, float)) or o <= 0:
                continue
            if isinstance(d, datetime.datetime): d = d.date()
            elif isinstance(d, (int, float)):
                d = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(d))
            dts.append(d); O.append(o)
            H.append(q.cell(r, co + 1).value); L.append(q.cell(r, co + 2).value); C.append(q.cell(r, co + 3).value)
        data[s] = (dts, O, H, L, C)
    params = {}
    for s in STOCKS:
        ws = wbv[f'Model {s}']; g = lambda a: ws[a].value
        params[s] = Params(lam=g('B2'), phi_L=g('D2'), psi=g('F2'), k=g('H2'),
            premium=g('J2'), peak_cap=g('L2'), comm=g('N2'), capital=g('P2'),
            interest=g('R2'), stop_days=int(g('T2')), bayes_pct=g('V2'),
            ou_W=int(g('B3')), ou_buf_k=g('D3'), ou_prem=g('H3'), ou_cap=g('J3'), years=2.2)
    # allocation weights (included names only)
    al = wbv['Allocation']; wts = {}
    for r in range(11, 21):
        nm = al.cell(r, 1).value
        if nm in STOCKS and al.cell(r, 2).value == 1:
            wts[nm] = al.cell(r, 7).value
    return data, params, wts


def tranche_equity(frames, tkey, C):
    t = frames[tkey]
    return [t['AA'][i] * C[i] if t['AE'][i] == 1 else t['Y'][i] for i in range(len(C))]


def stats(eq, years=2.2):
    rets = [eq[i] / eq[i - 1] - 1 for i in range(1, len(eq)) if eq[i - 1] > 0]
    if len(rets) < 3: return 0, 0, 0
    ann = (eq[-1] / eq[0]) ** (1 / years) - 1 if eq[0] > 0 else 0
    sd = statistics.pstdev(rets)
    sh = statistics.mean(rets) / sd * math.sqrt(252) if sd > 0 else 0
    peak = -1e30; mdd = 0
    for e in eq:
        peak = max(peak, e); mdd = max(mdd, (peak - e) / peak) if peak > 0 else mdd
    return ann, sh, mdd


if __name__ == '__main__':
    data, params, wts = load_book()

    print('=== 1. STANDALONE SLEEVES (each run with the full name capital) ===')
    print(f'{"stock":6s} | {"Bayes ann/Sh/DD":>22s} | {"OU ann/Sh/DD":>22s} | {"corr":>6s}')
    print('-' * 68)
    corr = {}
    for s in STOCKS:
        dts, O, H, L, C = data[s]; p = params[s]
        pb = copy.copy(p); pb.bayes_pct = 1.0
        po = copy.copy(p); po.bayes_pct = 0.0
        rb = run_model(dts, O, H, L, C, pb, collect=True)
        ro = run_model(dts, O, H, L, C, po, collect=True)
        eb = tranche_equity(rb.frames, 't1', C)
        eo = tranche_equity(ro.frames, 't2', C)
        ba, bs, bd = stats(eb); oa, os_, od = stats(eo)
        # daily-return correlation between sleeves
        rbd = [eb[i]/eb[i-1]-1 for i in range(1, len(eb)) if eb[i-1] > 0 and eo[i-1] > 0]
        rod = [eo[i]/eo[i-1]-1 for i in range(1, len(eo)) if eb[i-1] > 0 and eo[i-1] > 0]
        n = min(len(rbd), len(rod))
        if n > 2:
            mb, mo = statistics.mean(rbd[:n]), statistics.mean(rod[:n])
            cov = sum((rbd[i]-mb)*(rod[i]-mo) for i in range(n))
            vb = sum((x-mb)**2 for x in rbd[:n]); vo = sum((x-mo)**2 for x in rod[:n])
            corr[s] = cov/math.sqrt(vb*vo) if vb > 0 and vo > 0 else 0
        else:
            corr[s] = 0
        print(f'{s:6s} | {ba*100:6.0f}% {bs:5.2f} {bd*100:5.1f}% | '
              f'{oa*100:6.0f}% {os_:5.2f} {od*100:5.1f}% | {corr[s]:6.2f}')

    print('\n=== 2. DOES OU OFFSET BAYES WEAKNESS? (50/50 run) ===')
    print('down-Bayes days = days Bayes sleeve fell; OU-up-share = of those, how often OU rose')
    print(f'{"stock":6s}{"downBayes days":>16s}{"OU-up-share":>13s}{"OU avg on those":>17s}')
    print('-' * 52)
    for s in STOCKS:
        dts, O, H, L, C = data[s]
        r = run_model(dts, O, H, L, C, params[s], collect=True)
        eb = tranche_equity(r.frames, 't1', C); eo = tranche_equity(r.frames, 't2', C)
        down = 0; ouup = 0; ouvals = []
        for i in range(1, len(C)):
            if eb[i-1] > 0 and eb[i]/eb[i-1]-1 < -1e-6:
                down += 1
                our = eo[i]/eo[i-1]-1 if eo[i-1] > 0 else 0
                ouvals.append(our)
                if our > 1e-6: ouup += 1
        share = ouup/down if down else 0
        avg = statistics.mean(ouvals)*100 if ouvals else 0
        print(f'{s:6s}{down:>16d}{share*100:>12.0f}%{avg:>16.2f}%')

    print('\n=== 3. CAPITAL-SPLIT SWEEP (book-weighted, included names) ===')
    print(f'included/weights: {", ".join(f"{k} {v:.0%}" for k,v in wts.items())}')
    grid = [0.0, 0.25, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
    print(f'\n{"bayes%":>7s}{"book ann":>10s}{"book Sharpe":>13s}{"book maxDD":>12s}')
    print('-' * 42)
    # build book equity per split by weighting each name's combined equity curve
    tot = sum(wts.values())
    for bp in grid:
        booklen = None; booked = None
        for s, w in wts.items():
            dts, O, H, L, C = data[s]; p = copy.copy(params[s]); p.bayes_pct = bp
            r = run_model(dts, O, H, L, C, p, collect=True)
            eq = r.frames['equity']
            norm = [e / eq[0] * (w / tot) for e in eq]   # weight by allocation, normalise to start=weight
            if booked is None:
                booked = norm; booklen = len(norm)
            else:
                m = min(booklen, len(norm))
                booked = [booked[i] + norm[i] for i in range(m)]; booklen = m
        ann, sh, mdd = stats(booked)
        star = '  <-- current book (Bayes 60%)' if abs(bp-0.6) < 1e-9 else (
               '  (Model sheets use 50%)' if abs(bp-0.5) < 1e-9 else '')
        print(f'{bp*100:>6.0f}%{ann*100:>9.0f}%{sh:>13.2f}{mdd*100:>11.1f}%{star}')
