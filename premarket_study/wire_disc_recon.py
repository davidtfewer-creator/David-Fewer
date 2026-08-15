"""
Discretionary carve-out: the double-count reconciliation (15 Aug 2026).

The problem it solves: E6 earmarks discretionary money while it still sits in
the account. The moment a discretionary buy is actually funded, that money has
left the cash balance -- B5 (entered each morning from the account) already
excludes it. If E6 keeps netting the full earmark, the same money is subtracted
twice and the book under-bids.

Wiring (Allocation, previously empty cells; nothing moves):
  D8 'Already deployed ($)'   E8 = 0 (blue input: what has actually been spent
                                     on discretionary positions and thus already
                                     left B5)
  D9 'Reconciliation'         E9 = status text -- flag off / reserve held /
                                     CHECK: deployed exceeds earmark
  E7 becomes  MAX(0, B5 - E5*MAX(0, E6 - E8))
so only the UNSPENT remainder of the earmark is netted off the book. E6 stays
the total earmark; E8 climbs as discretionary buys are funded; when E8 = E6 the
whole earmark is outside the account and the book bids with all of B5 again.
"""
import copy
import sys

import openpyxl

NOTE = ('Allocation E5 (flag 1/0), E6 (earmark $) and E8 (already deployed $) carve a '
        'discretionary fund out of the cash before it is divided across the sleeves: '
        'E7 = MAX(0, B5 - E5*MAX(0, E6 - E8)) and the machine block C25:C42 divides E7 instead '
        'of B5. E6 is the total earmark; E8 is what has actually been spent on discretionary '
        'positions (money already gone from B5) -- netting only the difference prevents the '
        'double-count. E9 reconciles: it shows the reserve still held inside the book cash and '
        'flags CHECK if deployed exceeds the earmark. Flag 0 (or earmark 0) reproduces the old '
        'behaviour exactly; the MAX guards keep the book from going negative. Discretionary '
        'trades themselves live outside the workbook.')


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    al = wb['Allocation']
    notes = wb['Notes']

    assert al['E7'].value == '=MAX(0,$B$5-$E$5*$E$6)', al['E7'].value
    assert all(al.cell(row=r, column=c).value is None for r in (8, 9) for c in (4, 5))
    assert notes['A25'].value == 'Discretionary carve-out'

    lab, inp = al['D5'], al['E5']

    def put(coord, value, style_from, numfmt=None):
        c = al[coord]
        c.value = value
        c.font = copy.copy(style_from.font)
        c.fill = copy.copy(style_from.fill)
        c.border = copy.copy(style_from.border)
        c.alignment = copy.copy(style_from.alignment)
        if numfmt:
            c.number_format = numfmt
        return c

    put('D8', 'Already deployed ($)', lab)
    put('E8', 0, inp, '"$"#,##0')
    put('D9', 'Reconciliation', lab)
    put('E9', ('=IF($E$5=0,"flag off — full cash to book",'
               'IF($E$8>$E$6,"CHECK: deployed exceeds earmark",'
               '"reserve held: "&TEXT(MAX(0,$E$6-$E$8),"$#,##0")))'), al['D7'])
    al['E7'].value = '=MAX(0,$B$5-$E$5*MAX(0,$E$6-$E$8))'

    notes['B25'].value = NOTE

    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)


def diff(in_path, out_path):
    a = openpyxl.load_workbook(in_path)
    b = openpyxl.load_workbook(out_path)
    out = []
    for sn in a.sheetnames:
        wa, wc = a[sn], b[sn]
        for r in range(1, max(wa.max_row, wc.max_row) + 1):
            for c in range(1, max(wa.max_column, wc.max_column) + 1):
                va, vb = wa.cell(row=r, column=c).value, wc.cell(row=r, column=c).value
                if va != vb:
                    out.append((sn, wa.cell(row=r, column=c).coordinate, str(va)[:45], str(vb)[:70]))
    return out


if __name__ == '__main__':
    in_path, out_path = sys.argv[1], sys.argv[2]
    wire(in_path, out_path)
    for row in diff(in_path, out_path):
        print(row)
