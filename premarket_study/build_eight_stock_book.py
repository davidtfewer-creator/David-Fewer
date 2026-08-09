"""
Add GM, VLO and CF to the live five-name workbook -> the eight-name book.

Follows build_five_stock_book.py, but EXTENDS the uploaded live file in place rather
than rebuilding it: the trade log (real fills), Notes, fees, and every user-entered
cell are preserved. What changes:

  Query           GM/VLO/CF OHLC columns appended (V:AG), data 2024-04-01 to
                  2026-08-03 derived from the 5-minute regular-hours bars. The last
                  Query dates (Aug 4-7) are blank for the new names until the IBKR
                  feed refresh backfills them.
  Feed sheets     Feed GM/VLO/CF as copies of Feed TSM repointed at the new columns.
  Model sheets    Model GM/VLO/CF as copies of Model TSM (residual-sigma AZ and all
                  fixes inherited), Feed references retargeted, diversifier-study
                  parameters written into rows 2-3.
  Allocation      rows 16-18; floor and cap 0.125 (pins 8 x 12.5%); every 11:15
                  range widened to 11:18; machine-readable block rebuilt for 16
                  sleeves (rows 25-40); checks widened.
  Dashboard       rows 9-11.
  Active Trading  funds rows 12-14, order rows 29-34, lookup ranges widened;
                  IBKR_Orders -> $A$18:$J$34. Trade log untouched.

Run: python3 build_eight_stock_book.py
"""
import copy as pycopy
import datetime
import re

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from fresh_opt_cands import daily_from_5min

SRC = ('/root/.claude/uploads/9e026445-9e62-588b-af81-7e0c231b0f24/'
       '34b220b8-TradingExcel_5stock_live.xlsx')
OUT = '/home/user/David-Fewer/TradingExcel_8stock_live.xlsx'

OLD = ['TSM', 'VRT', 'VST', 'RKLB', 'MU']
NEW = ['GM', 'VLO', 'CF']
ALL = OLD + NEW
WEIGHT = 0.125

PARAMS = {
    'GM':  dict(lam=0.704472, phi_L=0.750723, psi=0.0805582, k=0.461711,
                premium=0.00969721, peak_cap=0.0423717, ou_W=45,
                ou_buf_k=1.06024, ou_prem=0.0451112, ou_cap=0.0417662),
    'VLO': dict(lam=0.723393, phi_L=0.625222, psi=0.0200999, k=0.543763,
                premium=0.0303035, peak_cap=0.0308569, ou_W=125,
                ou_buf_k=1.17841, ou_prem=0.0419514, ou_cap=0.0370821),
    'CF':  dict(lam=0.736218, phi_L=0.106120, psi=0.0694359, k=2.03552,
                premium=0.0465592, peak_cap=0.0286416, ou_W=66,
                ou_buf_k=0.841237, ou_prem=0.0590932, ou_cap=0.0831024),
}
ANNRET = {'GM': 0.33, 'VLO': 0.35, 'CF': 0.41}     # June planning figures


def to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
    return None


def selfrow(f, src, dst):
    """Shift only references to the template's OWN row (relative-row refs like $C19
    or K19); absolute-row refs ($C$19) and other rows are left alone."""
    return re.sub(r"(?<![!$A-Za-z0-9_])(\$?)([A-Z]{1,2})" + str(src) + r"(?![0-9])",
                  lambda m: f"{m.group(1)}{m.group(2)}{dst}", f)


def snapshot(ws, row, cols):
    return {c: (ws.cell(row, c).value, ws.cell(row, c)._style) for c in cols}


def write_row(tmpl, src_row, ws, dst_row, subs):
    for c, (v, style) in tmpl.items():
        if isinstance(v, str) and v.startswith('='):
            v = Translator(v, origin=f'{get_column_letter(c)}{src_row}') \
                .translate_formula(f'{get_column_letter(c)}{dst_row}')
            for a, b in subs:
                v = v.replace(a, b)
        cell = ws.cell(dst_row, c)
        cell.value = v
        cell._style = pycopy.copy(style)


def retext(ws, cells_cols_rows, old, new):
    """Replace a substring in formulas over a cell region."""
    for r, c in cells_cols_rows:
        v = ws.cell(r, c).value
        if isinstance(v, str) and old in v:
            ws.cell(r, c).value = v.replace(old, new)


def main():
    wb = openpyxl.load_workbook(SRC)

    # ------------------------------------------------------------------ Query
    q = wb['Query']
    dates = {}
    for r in range(2, q.max_row + 1):
        d = to_date(q.cell(r, 1).value)
        if d:
            dates[d] = r
    filled = {}
    for i, s in enumerate(NEW):
        c0 = 22 + 4 * i                                  # V, Z, AD
        for j, sfx in enumerate(['_O', '_H', '_L', '_C']):
            q.cell(1, c0 + j).value = f'{s}{sfx}'
        dts, O, H, L, C = daily_from_5min(s)
        n = 0
        for k, d in enumerate(dts):
            r = dates.get(d)
            if r is None:
                continue
            q.cell(r, c0).value = O[k]
            q.cell(r, c0 + 1).value = H[k]
            q.cell(r, c0 + 2).value = L[k]
            q.cell(r, c0 + 3).value = C[k]
            n += 1
        filled[s] = (n, dts[-1])
        print(f'Query: {s} cols {get_column_letter(c0)}:{get_column_letter(c0+3)}, '
              f'{n} rows, last {dts[-1]}')

    # ------------------------------------------------------------------ Feeds
    for i, s in enumerate(NEW):
        ws = wb.copy_worksheet(wb['Feed TSM'])
        ws.title = f'Feed {s}'
        c0 = 22 + 4 * i
        letters = [get_column_letter(c0 + j) for j in range(4)]
        ws.cell(1, 1).value = '=Query!A1'
        for j in range(4):
            ws.cell(1, 2 + j).value = f'=Query!{letters[j]}1'
        ws.cell(1, 7).value = (f'Daily OHLC for {s}. Loaded from the Query sheet; '
                               f'the Model sheet reads columns A:E.')
        for r in range(2, ws.max_row + 1):
            ws.cell(r, 1).value = '=INDEX(Query!$A:$A,ROW())'
            for j in range(4):
                Lc = letters[j]
                ws.cell(r, 2 + j).value = f'=INDEX(Query!${Lc}:${Lc},ROW())'
        print(f'Feed {s}: -> Query {letters[0]}:{letters[3]}')

    # ------------------------------------------------------------------ Models
    for s in NEW:
        m = wb.copy_worksheet(wb['Model TSM'])
        m.title = f'Model {s}'
        for row in m.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and 'Feed TSM' in cell.value:
                    cell.value = cell.value.replace('Feed TSM', f'Feed {s}')
        p = PARAMS[s]
        m['A1'] = (f'{s} Daily Bayesian – 2 Tranches, 50-Day Stop-Loss  '
                   f'(live Feed | auto-extends)')
        m['B2'] = p['lam'];      m['D2'] = p['phi_L'];   m['F2'] = p['psi']
        m['H2'] = p['k'];        m['J2'] = p['premium']; m['L2'] = p['peak_cap']
        m['V2'] = 0.5
        m['B3'] = p['ou_W'];     m['D3'] = p['ou_buf_k']
        m['H3'] = p['ou_prem'];  m['J3'] = p['ou_cap']
        print(f'Model {s}: params written, Feed retargeted')

    # -------------------------------------------------------------- Allocation
    al = wb['Allocation']
    al_t = snapshot(al, 11, range(1, 19))                # TSM row as template
    for i, s in enumerate(NEW):
        dst = 16 + i
        write_row(al_t, 11, al, dst, [("'Feed TSM'", f"'Feed {s}'")])
        al.cell(dst, 1).value = s
        al.cell(dst, 2).value = 1
        al.cell(dst, 3).value = ANNRET[s]
        al.cell(dst, 5).value = WEIGHT
        al.cell(dst, 17).value = 0.5                     # Q: per-stock Bayes share
    for r in range(11, 19):                              # widen 11:15 -> 11:18
        for c in range(1, 19):
            v = al.cell(r, c).value
            if isinstance(v, str) and '$15' in v:
                for Lc in 'FHJ':
                    v = v.replace(f'${Lc}$11:${Lc}$15', f'${Lc}$11:${Lc}$18')
                al.cell(r, c).value = v
        al.cell(r, 5).value = WEIGHT                     # cap 0.125 on all eight
    al['B8'] = WEIGHT                                    # floor 0.125
    al['B21'] = '=SUM(B11:B18)'
    for c, Lc in ((11, 'K'), (12, 'L'), (13, 'M'), (14, 'N')):
        al.cell(21, c).value = f'=SUM({Lc}11:{Lc}18)'

    # capture the summary rows' text/styles BEFORE the block overwrites rows 36-40
    summ = [(al.cell(r, 1).value, al.cell(r, 1)._style, al.cell(r, 2)._style)
            for r in (36, 37, 38)]
    held_note = al['A40'].value
    cash_note = al['A41'].value
    al['A41'] = None

    mac_t = {c: al.cell(25, c)._style for c in range(1, 7)}
    for j in range(16):                                  # machine block, 16 sleeves
        r = 25 + j
        i = j // 2
        s = ALL[i]
        tranche = 'Bayes' if j % 2 == 0 else 'OU'
        share = f'$R${11+i}' if tranche == 'Bayes' else f'(1-$R${11+i})'
        al.cell(r, 1).value = s
        al.cell(r, 2).value = tranche
        al.cell(r, 3).value = f'=IF(SUM($F$25:$F$40)=0,0,$B$5*F{r}/SUM($F$25:$F$40))'
        al.cell(r, 4).value = f"=IF('Active Trading'!$C${19+j}=\"HOLDING\",1,0)"
        al.cell(r, 5).value = f'=$K${11+i}*{share}'
        al.cell(r, 6).value = f'=E{r}*(1-D{r})'
        for c in range(1, 7):
            al.cell(r, c)._style = pycopy.copy(mac_t[c])
    # summary moves BELOW the block (block now owns rows 25-40)
    formulas = ['=COUNT(D25:D40)-SUM(D25:D40)', '=IF(B42=0,0,$B$5/B42)',
                '=IF(ABS(SUM(C25:C40)-$B$5)<0.01,"OK","MISMATCH")']
    for k, (label, sa, sb) in enumerate(summ):
        r = 42 + k
        al.cell(r, 1).value = label
        al.cell(r, 1)._style = pycopy.copy(sa)
        al.cell(r, 2).value = formulas[k]
        al.cell(r, 2)._style = pycopy.copy(sb)
    al['A46'] = ('Held? is read from the blotter Status in Active Trading C19:C34; '
                 'a sleeve marked HOLDING gets nothing and its weight is spread '
                 'over the free sleeves.')
    al['A47'] = cash_note
    al['A1'] = ('PORTFOLIO ALLOCATION  —  eight daily names, equal weighted. '
                'GM, VLO and CF added as diversifiers per the August 2026 ranking.')
    al['A2'] = ('Edit only the blue cells. Eight daily names, equal weighted: the '
                'floor (B8) and the cap (E11:E18) are both 0.125, which pins each '
                'name to 12.5%.')
    print('Allocation: rows 16-18 added, ranges widened, machine block 16 sleeves')

    # -------------------------------------------------------------- Dashboard
    db = wb['Dashboard']
    db_t = snapshot(db, 4, range(1, 18))
    for i, s in enumerate(NEW):
        R = 9 + i
        for c, (v, style) in db_t.items():
            if isinstance(v, str) and v.startswith('='):
                v = v.replace("'Feed TSM'", f"'Feed {s}'") \
                     .replace("'Model TSM'", f"'Model {s}'")
                v = selfrow(v, 4, R)
            cell = db.cell(R, c)
            cell.value = v
            cell._style = pycopy.copy(style)
        db.cell(R, 1).value = s
    db['A1'] = 'HYBRID DASHBOARD — daily pre-open levels (8 daily names)'
    print('Dashboard: rows 9-11 added')

    # ---------------------------------------------------------- Active Trading
    at = wb['Active Trading']
    # funds block: add rows 12-14, widen lookup ranges everywhere
    at_t = snapshot(at, 7, range(1, 7))
    for i, s in enumerate(NEW):
        R = 12 + i
        for c, (v, style) in at_t.items():
            if isinstance(v, str) and v.startswith('='):
                v = selfrow(v, 7, R)
            cell = at.cell(R, c)
            cell.value = v
            cell._style = pycopy.copy(style)
        at.cell(R, 1).value = s
    for r in range(7, 15):
        for c in range(1, 7):
            v = at.cell(r, c).value
            if isinstance(v, str):
                v = v.replace('$R$11:$R$15', '$R$11:$R$18') \
                     .replace('$A$11:$A$15', '$A$11:$A$18')
                at.cell(r, c).value = v

    # orders block: add rows 29-34 from the TSM Bayes/OU templates
    ord_t = {t: snapshot(at, 19 + t, range(1, 13)) for t in (0, 1)}
    for i, s in enumerate(NEW):
        for t in (0, 1):
            src = 19 + t
            R = 29 + 2 * i + t
            drow = 9 + i                                 # Dashboard row for s
            for c, (v, style) in ord_t[t].items():
                if isinstance(v, str) and v.startswith('='):
                    for Lc in 'CDEF':                    # per-name Dashboard cells
                        v = v.replace(f'Dashboard!{Lc}4', f'<<{Lc}>>')
                    v = v.replace("'Model TSM'", f"'Model {s}'")
                    v = selfrow(v, src, R)
                    for Lc in 'CDEF':
                        v = v.replace(f'<<{Lc}>>', f'Dashboard!{Lc}{drow}')
                cell = at.cell(R, c)
                cell.value = v
                cell._style = pycopy.copy(style)
            at.cell(R, 1).value = s
            at.cell(R, 2).value = 'Bayes' if t == 0 else 'OU'
    # widen lookup ranges in ALL order rows 19-34
    for r in range(19, 35):
        for c in range(1, 13):
            v = at.cell(r, c).value
            if isinstance(v, str):
                v = v.replace('Allocation!$C$25:$C$34', 'Allocation!$C$25:$C$40') \
                     .replace('Dashboard!$B$4:$B$8', 'Dashboard!$B$4:$B$11') \
                     .replace('Dashboard!$A$4:$A$8', 'Dashboard!$A$4:$A$11')
                at.cell(r, c).value = v
    at['A1'] = ('ACTIVE TRADING BLOTTER  —  eight daily names  '
                '(TSM · VRT · VST · RKLB · MU · GM · VLO · CF)')
    print('Active Trading: funds rows 12-14, order rows 29-34, ranges widened')

    # ------------------------------------------------------------ named range
    dn = wb.defined_names.get('IBKR_Orders')
    dn.value = "'Active Trading'!$A$18:$J$34"
    print('IBKR_Orders ->', dn.value)

    # ------------------------------------------------------------------ Notes
    nt = wb['Notes']
    nt['B3'] = ('Eight daily names: TSM, VRT, VST, RKLB, MU + GM, VLO, CF. The three '
                'additions are the diversifiers from the August 2026 verified-fill '
                'ranking (AI beta 0.0-0.17); NVDA and AVGO assessed and not admitted.')
    nt['B4'] = ('Parameters: incumbents unchanged; GM/VLO/CF carry the '
                'diversifier-study vectors (residual OU sigma scale). Bayes share '
                '0.50 on all eight (split mode B9=2 reads Allocation!Q11:Q18).')
    nt['B5'] = ('Allocation is equal weight: floor and cap both 0.125. Enter the '
                'CASH available this morning in Allocation!B5.')
    nt['B6'] = ('GM/VLO/CF price history in the Query sheet runs 1 Apr 2024 to '
                '1 Aug 2026 (derived from regular-hours 5-minute bars). Run the '
                'IBKR feed refresh once before trading them so the last sessions '
                'backfill; their Dashboard rows read #N/A until then.')
    wb.save(OUT)
    print(f'\nwritten {OUT}')


if __name__ == '__main__':
    main()
