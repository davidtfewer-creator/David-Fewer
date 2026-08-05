"""
Exercise the mirror helpers, including the two failure paths that matter in the morning.

Linux has no mandatory file locking, so the "Excel has it open" state is simulated the only way it
can be: by creating the ~$ lock file Excel itself leaves behind. That is the signal the guard
checks first and the one that fires on a network share, so the test is meaningful even though the
open-for-append test cannot fire here.

What is verified:
  a snapshot is a faithful copy -- same sheets, same formulas in the cells that were changed
  the master is refused while the lock file is present, and accepted once it is gone
  a locked stable mirror does not break publish, and the timestamped snapshot still lands
  save_atomic leaves no temp files behind, on success or on failure
  pruning keeps exactly the newest N
"""
import os
import shutil
import sys
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mirror import (assert_writable, is_open_in_excel, lock_path, publish,  # noqa: E402
                    save_atomic)

SRC = '/home/user/David-Fewer/TradingExcel_5stock_live_freesleeves.xlsx'
TMP = ('/tmp/claude-0/-home-user-David-Fewer/'
       '2d71f10a-e19f-51b2-8457-2cd547c34dff/scratchpad/mirrortest')
fails = []


def check(cond, label):
    print(f'  {"ok  " if cond else "FAIL"}  {label}', flush=True)
    if not cond:
        fails.append(label)


def main():
    shutil.rmtree(TMP, ignore_errors=True)
    os.makedirs(f'{TMP}/model', exist_ok=True)
    master = f'{TMP}/model/TradingExcel.xlsx'
    view = f'{TMP}/view'
    shutil.copy2(SRC, master)

    print('lock guard', flush=True)
    check(not is_open_in_excel(master), 'clean master reports unlocked')
    assert_writable(master)
    open(lock_path(master), 'wb').write(b'x')
    check(is_open_in_excel(master), 'master with ~$ lock file reports locked')
    try:
        assert_writable(master)
        check(False, 'assert_writable refuses a locked master')
    except RuntimeError as e:
        check('open in Excel' in str(e), 'assert_writable refuses a locked master')
    os.remove(lock_path(master))
    check(not is_open_in_excel(master), 'unlocked again once the lock file goes')

    print('\nsnapshot fidelity', flush=True)
    snap, ok = publish(master, view, quiet=True)
    check(ok, 'stable mirror written')
    check(os.path.getsize(snap) == os.path.getsize(master), 'snapshot is byte-for-byte the size')
    a = openpyxl.load_workbook(master)
    b = openpyxl.load_workbook(snap)
    check(a.sheetnames == b.sheetnames, 'same sheets')
    same = all(a['Allocation'].cell(r, c).value == b['Allocation'].cell(r, c).value
               for r in range(25, 35) for c in (3, 4, 5, 6))
    check(same, 'allocation block formulas identical')
    check(b['Active Trading']['H19'].value == a['Active Trading']['H19'].value,
          'blotter share formula identical')

    print('\nlocked stable mirror must not stop the snapshot', flush=True)
    stable = os.path.join(view, 'TradingExcel_view_latest.xlsx')
    check(os.path.exists(stable), 'stable file exists')
    # make the rename fail the way Windows would: replace the stable file with a directory
    os.remove(stable)
    os.makedirs(stable)
    snap2, ok2 = publish(master, view, quiet=True)
    check(not ok2, 'publish reports the stable copy was skipped')
    check(os.path.exists(snap2), 'timestamped snapshot still written')
    check(len([f for f in os.listdir(view) if f.endswith('.xlsx')]) >= 2, 'both snapshots present')
    shutil.rmtree(stable)
    leftovers = [f for f in os.listdir(view) if f.startswith('tmp')]
    check(not leftovers, f'no temp files left behind ({leftovers})')

    print('\npruning', flush=True)
    for _ in range(6):
        publish(master, view, keep=3, quiet=True)
    snaps = [f for f in os.listdir(view) if '_view_2' in f]
    check(len(snaps) == 3, f'keep=3 leaves 3 snapshots (found {len(snaps)})')

    print('\nsave_atomic', flush=True)
    wb = openpyxl.load_workbook(master)
    wb['Allocation']['B5'] = 1234567.89
    save_atomic(wb, master)
    check(openpyxl.load_workbook(master)['Allocation']['B5'].value == 1234567.89,
          'atomic save applied')
    check(not [f for f in os.listdir(f'{TMP}/model') if f.startswith('tmp')],
          'no temp file left in the model folder')

    class Boom:
        def save(self, p):
            raise ValueError('simulated failure mid-save')
    before = sorted(os.listdir(f'{TMP}/model'))
    try:
        save_atomic(Boom(), master)
        check(False, 'a failing save propagates')
    except ValueError:
        check(True, 'a failing save propagates')
    check(sorted(os.listdir(f'{TMP}/model')) == before, 'failed save leaves the folder unchanged')
    check(openpyxl.load_workbook(master)['Allocation']['B5'].value == 1234567.89,
          'master survives a failed save intact')

    print(f'\n{"ALL TESTS PASSED" if not fails else str(len(fails)) + " FAILURES"}', flush=True)
    print('DONE', flush=True)


if __name__ == '__main__':
    main()
