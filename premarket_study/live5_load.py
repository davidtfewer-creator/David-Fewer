"""
Loader for the uploaded TradingExcel_5stock_live.xlsx: per-name OHLC from the Query
sheet (headers NAME_O/H/L/C), deployed Params from each Model sheet, and the sheet's
cached summary (Y4 profit, Y5 ann, AA4 buys, AC4 stops) for the mirror check.
"""
import datetime

import openpyxl
from openpyxl.utils import column_index_from_string as cix

from engine import Params

UPLOAD = ('/root/.claude/uploads/9e026445-9e62-588b-af81-7e0c231b0f24/'
          '5385d808-TradingExcel_5stock_live.xlsx')
STOCKS = ['TSM', 'VRT', 'VST', 'RKLB', 'MU']


def _to_date(d):
    if isinstance(d, datetime.datetime):
        return d.date()
    if isinstance(d, datetime.date):
        return d
    if isinstance(d, (int, float)):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(d))
    return datetime.date.fromisoformat(str(d)[:10])


def load(path=UPLOAD):
    """Return (data, params, cached): data[name] = (dates, O, H, L, C)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    q = wb['Query']
    rows = list(q.iter_rows(values_only=True))
    head = rows[0]
    c0 = {h[:-2]: i for i, h in enumerate(head)
          if isinstance(h, str) and h.endswith('_O')}
    data = {}
    for s in STOCKS:
        dts, O, H, L, C = [], [], [], [], []
        for row in rows[1:]:
            if row[0] is None:
                continue
            o = row[c0[s]]
            if not isinstance(o, (int, float)) or o <= 0:
                continue
            dts.append(_to_date(row[0]))
            O.append(float(o)); H.append(float(row[c0[s] + 1]))
            L.append(float(row[c0[s] + 2])); C.append(float(row[c0[s] + 3]))
        data[s] = (dts, O, H, L, C)

    params, cached = {}, {}
    for s in STOCKS:
        ws = wb[f'Model {s}']
        vals = list(ws.iter_rows(min_row=1, max_row=5, max_col=60, values_only=True))
        g = {f'{openpyxl.utils.get_column_letter(c + 1)}{r + 1}': vals[r][c]
             for r in range(len(vals)) for c in range(len(vals[r]))}
        params[s] = Params(
            lam=g['B2'], phi_L=g['D2'], psi=g['F2'], k=g['H2'], premium=g['J2'],
            peak_cap=g['L2'], comm=g['N2'], capital=g['P2'], interest=g['R2'],
            stop_days=int(g['T2']), bayes_pct=g['V2'], ou_W=int(g['B3']),
            ou_buf_k=g['D3'], ou_prem=g['H3'], ou_cap=g['J3'], years=2.2)
        cached[s] = dict(profit=g['Y4'], ann=g['Y5'], buys=g['AA4'], stops=g['AC4'])
    wb.close()
    return data, params, cached


if __name__ == '__main__':
    from engine import run_model
    data, params, cached = load()
    print(f'{"name":6s}{"N":>5s}{"span":>26s} | {"engine ann":>11s}{"sheet Y5":>10s}'
          f'{"eng buys":>9s}{"AA4":>5s}{"eng stops":>10s}{"AC4":>5s}')
    for s in STOCKS:
        dts, O, H, L, C = data[s]
        r = run_model(dts, O, H, L, C, params[s], ou_sigma='resid', same_day_exit=True)
        c = cached[s]
        print(f'{s:6s}{len(C):>5d}{f"{dts[0]} to {dts[-1]}":>26s} | '
              f'{r.annual_return:>10.1%}{c["ann"]:>10.1%}'
              f'{r.total_buys:>9d}{c["buys"]:>5d}{r.stop_loss_exits:>10d}{c["stops"]:>5d}')
