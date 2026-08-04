"""
The "2.5" model: two entry attempts a week instead of one.

Current rule: a bid is set from Monday's open and rests all week; if it never fills, nothing
happens until the next Monday. That ties every entry to one reference price, the Monday open, and
the Monday open is the single most repriced print of the week -- the weekend-gap finding. So every
entry inherits whatever the weekend did.

The 2.5 rule splits the week:

    Monday open      -> bid_1 = MIN(Monday open, ATH x (1-cap)),  rests Mon, Tue, Wed AM
    Wednesday 12:45  -> if unfilled, cancel and set bid_2 = MIN(Wed 12:45 price, ATH x (1-cap)),
                        rests Wed PM, Thu, Fri

The second bid is referenced to a mid-session price that is not an opening print at all, which is
the point: it exposes the model to a non-opening entry pattern.

Three variants are run per name with the SAME parameters, so the comparison is structural rather
than a refit:

    A  Monday only          the deployed rule
    B  2.5                  Monday, then Wednesday midday
    C  Wednesday only       the control -- if C alone is good, the gain is not about having two
                            attempts, it is about Monday being a poor reference

Parameters are fitted once per name on variant A and then frozen across B and C. Fitting each
variant separately would confound the structure with the fit, which is how several earlier results
in this work turned out to be nothing.

Intraday bars are required to know (a) the 12:45 price and (b) whether a fill happened before or
after the cutoff. Same-day exits are disallowed throughout: it costs little (NVDA has 2 in the
whole sample, AVGO none) and removes any dependence on bar ordering within the fill session.
"""
import datetime, collections, os, pickle, statistics, sys
import openpyxl
from stop_sweep import load_book
import five_min

five_min.FILES.setdefault('MU', '/root/.claude/uploads/'
                          '2d71f10a-e19f-51b2-8457-2cd547c34dff/94f1080f-MU_5min_Apr2024Aug2026.xlsx')
five_min.FILES.setdefault('NVDA', '/root/.claude/uploads/'
    '2d71f10a-e19f-51b2-8457-2cd547c34dff/b0d1e498-Nvidia_Minute_Data__Day_Trading_Model_.xlsx')

DATA, _P, _C = load_book()
BOOK = ['NVDA', 'AVGO', 'RKLB', 'TSM', 'VST', 'VRT', 'MU']
CACHE = '/home/user/David-Fewer/premarket_study/halfweek_sessions.pkl'
CUTOFF = datetime.time(12, 45)
RTH0, RTH1 = datetime.time(9, 30), datetime.time(16, 0)
COMM, INTEREST = 0.005, 0.0314
CUT_DATE = datetime.date(2025, 5, 23)


def session_parts(stock):
    """date -> (am_low, am_high, pm_open, pm_low, pm_high) from intraday bars."""
    cache = pickle.load(open(CACHE, 'rb')) if os.path.exists(CACHE) else {}
    if stock in cache: return cache[stock]
    path = five_min.FILES[stock]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first: first = False; continue
        # two layouts: (datetime, O,H,L,C) and (utc string, local datetime, O,H,L,C)
        if isinstance(row[0], datetime.datetime):
            dt, o, h, l, c = row[0], row[1], row[2], row[3], row[4]
        elif len(row) > 5 and isinstance(row[1], datetime.datetime):
            dt, o, h, l, c = row[1], row[2], row[3], row[4], row[5]
        else:
            continue
        if o is None: continue
        if RTH0 <= dt.time() < RTH1:
            by[dt.date()].append((dt, o, h, l, c))
    wb.close()
    out = {}
    for d, v in by.items():
        v.sort(key=lambda x: x[0])
        am = [x for x in v if x[0].time() < CUTOFF]
        pm = [x for x in v if x[0].time() >= CUTOFF]
        if not am or not pm: continue
        out[d] = (min(x[3] for x in am), max(x[2] for x in am),
                  pm[0][1], min(x[3] for x in pm), max(x[2] for x in pm))
    cache[stock] = out
    pickle.dump(cache, open(CACHE, 'wb'))
    return out


def weeks(dts):
    g = collections.OrderedDict()
    for i, d in enumerate(dts):
        g.setdefault(d - datetime.timedelta(days=d.weekday()), []).append(i)
    return [v for v in g.values() if len(v) >= 2]


def run(stock, variant, cap, prem, parts, w0=1, w1=None):
    """variant in {'A','B','C'}. Same-day exits disallowed."""
    dts, O, H, L, C = DATA[stock]
    WK = weeks(dts)
    if w1 is None: w1 = len(WK)-1
    fund, shares, holding = 1.0, 0.0, False
    tgt = None; trades = 0; fills = 0; second_fills = 0
    ath = max(H[i] for i in WK[w0])
    for wi in range(w0+1, min(w1, len(WK)-1)+1):
        prev, cur = WK[wi-1], WK[wi]
        ath = max(ath, max(H[i] for i in prev))
        pc = C[prev[-1]]
        if not holding:
            fund += fund*INTEREST*(dts[cur[-1]]-dts[prev[-1]]).days/365.0
        if holding:                                   # carry: target unchanged
            for i in cur:
                if H[i] >= tgt:
                    fund = shares*(tgt-COMM); shares = 0.0; holding = False; trades += 1; break
            continue
        wed = next((i for i in cur if dts[i].weekday() == 2 and dts[i] in parts), None)
        # ---- first attempt
        if variant in ('A', 'B'):
            bid1 = min(O[cur[0]], ath*(1-cap)); t1 = bid1 + pc*prem
            if variant == 'A':
                window = cur
            else:
                window = []
                for i in cur:
                    if wed is not None and dts[i] > dts[wed]: break
                    window.append(i)
            hit = None
            for k, i in enumerate(window):
                lo = L[i]
                if variant == 'B' and i == wed:
                    lo = parts[dts[i]][0]             # AM low only
                if lo <= bid1: hit = k; break
            if hit is not None:
                shares = fund/(bid1+COMM); fund = 0.0; holding = True; tgt = t1; fills += 1
                for i in cur[cur.index(window[hit])+1:]:
                    if H[i] >= tgt:
                        fund = shares*(tgt-COMM); shares = 0.0; holding = False
                        trades += 1; break
                continue
        # ---- second attempt (B and C)
        if variant in ('B', 'C') and wed is not None:
            am_lo, am_hi, pm_open, pm_lo, pm_hi = parts[dts[wed]]
            bid2 = min(pm_open, ath*(1-cap)); t2 = bid2 + pc*prem
            rest = [i for i in cur if dts[i] > dts[wed]]
            hit = None
            if pm_lo <= bid2:
                hit = ('wed', wed)
            else:
                for i in rest:
                    if L[i] <= bid2: hit = ('day', i); break
            if hit is not None:
                shares = fund/(bid2+COMM); fund = 0.0; holding = True; tgt = t2
                fills += 1; second_fills += 1
                after = rest if hit[0] == 'wed' else rest[rest.index(hit[1])+1:]
                for i in after:
                    if H[i] >= tgt:
                        fund = shares*(tgt-COMM); shares = 0.0; holding = False
                        trades += 1; break
    last = WK[min(w1, len(WK)-1)][-1]
    val = fund + (shares*C[last] if holding else 0.0)
    yrs = (dts[last]-dts[WK[w0][0]]).days/365.25
    return dict(ann=val**(1/yrs)-1, trades=trades, fills=fills, second=second_fills)


CAPS = [round(0.02+0.005*i, 4) for i in range(37)]
PREMS = [round(0.02+0.005*i, 5) for i in range(37)]


def fit_A(stock, parts, w1=None, floor=8):
    best = None
    for c in CAPS:
        for q in PREMS:
            r = run(stock, 'A', c, q, parts, w1=w1)
            if r['trades'] < floor: continue
            if best is None or r['ann'] > best[0]: best = (r['ann'], c, q)
    return (best[1], best[2]) if best else (0.09, 0.05)


if __name__ == '__main__':
    names = [a for a in sys.argv[1:] if a in BOOK] or BOOK
    parts = {}
    for s in names:
        parts[s] = session_parts(s)
        print(f'{s}: intraday sessions with both halves {len(parts[s])}', flush=True)
    print(flush=True)
    print(f'{"stock":6s}{"cap":>7s}{"prem":>7s} | {"A Monday":>10s}{"B 2.5":>9s}{"C Wed":>9s}'
          f' | {"A tr":>6s}{"B tr":>6s}{"C tr":>6s}{"B 2nd":>7s}', flush=True)
    print('-'*82, flush=True)
    agg = {v: [] for v in 'ABC'}
    for s in names:
        cap, prem = fit_A(s, parts[s])
        res = {v: run(s, v, cap, prem, parts[s]) for v in 'ABC'}
        for v in 'ABC': agg[v].append(res[v]['ann']*100)
        print(f'{s:6s}{cap:>7.3f}{prem:>7.3f} | {res["A"]["ann"]*100:>9.1f}%'
              f'{res["B"]["ann"]*100:>8.1f}%{res["C"]["ann"]*100:>8.1f}% | '
              f'{res["A"]["trades"]:>6d}{res["B"]["trades"]:>6d}{res["C"]["trades"]:>6d}'
              f'{res["B"]["second"]:>7d}', flush=True)
    print('-'*82, flush=True)
    print(f'{"mean":20s} | {statistics.mean(agg["A"]):>9.1f}%{statistics.mean(agg["B"]):>8.1f}%'
          f'{statistics.mean(agg["C"]):>8.1f}%', flush=True)
    wins = sum(1 for i in range(len(names)) if agg['B'][i] > agg['A'][i])
    print(f'2.5 beats Monday-only in {wins}/{len(names)} names; mean '
          f'{statistics.mean(agg["B"][i]-agg["A"][i] for i in range(len(names))):+.1f}pp',
          flush=True)
    print('DONE', flush=True)
