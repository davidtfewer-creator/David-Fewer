"""
Rebuild TradingExcel_s1.xlsx as the five-name daily book.

Keeps TSM, VRT, VST and RKLB, adds MU, and removes NVDA, TSLA, AVGO, PLTR, SOFI and SPOT.
NVDA and AVGO are not dropped from the book -- they move to the weekly single-tranche model,
which is not run from this workbook.

The file carries no Power Query parts, tables or connections (verified by unpacking it), so an
openpyxl round trip is safe here.

What changes:
  Query          columns rebuilt as Date + TSM, VRT, VST, RKLB, MU; MU prices imported from the
                 MU candidate workbook
  Feed sheets    repointed at the new Query columns; Feed MU added
  Model sheets   Model MU added as a copy of Model TSM with its Feed reference retargeted and
                 MU's own fitted parameters written into rows 2 and 3
  Allocation     five rows, equal weighting pinned by setting floor and cap both to 0.20, Bayes
                 fraction 0.75, return assumptions set to the planning figures
  Dashboard      five rows
  Active Trading fund block and order block cut to five names

Row formulas are moved with openpyxl's Translator so relative references follow the row, rather
than by string surgery.
"""
import copy, datetime, re, shutil
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from newcands import load as load_cand

SRC = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/9a719ef3-TradingExcel_s1.xlsx'
OUT = '/home/user/David-Fewer/TradingExcel_5stock.xlsx'

KEEP = ['TSM', 'VRT', 'VST', 'RKLB', 'MU']
DROP = ['NVDA', 'TSLA', 'AVGO', 'PLTR', 'SOFI', 'SPOT']

# planning annualised returns from the seven-name specification
ANNRET = {'TSM': 0.55, 'VRT': 0.53, 'VST': 0.57, 'RKLB': 1.15, 'MU': 0.54}
BAYES_FLAT = 0.75
WEIGHT = 0.20                       # 1/5; floor and cap both set here to pin equal weighting

MU_PARAMS = dict(lam=0.6, phi_L=0.2636, psi=0.01, k=1.0548, premium=0.0243, peak_cap=0.038,
                 ou_W=91, ou_buf_k=0.3, ou_prem=0.025, ou_cap=0.0292)


def to_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    if isinstance(v, (int, float)): return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
    return None


def snapshot(ws, row, cols):
    """Capture a row before anything is written, so a template cannot overwrite itself."""
    return {c: ws.cell(row, c).value for c in cols}


SELF = re.compile(r"(?<![!$A-Za-z0-9_])(\$?)([A-Z]{1,2})(%d)(?![0-9])")


def selfrow(f, src, dst):
    """Shift only references to the template's OWN row; leave every other row reference alone."""
    return re.sub(r"(?<![!A-Za-z0-9_])(\$?)([A-Z]{1,2})" + str(src) + r"(?![0-9])",
                  lambda m: f"{m.group(1)}{m.group(2)}{dst}", f)


def move_row(tmpl, src_row, ws, dst_row, subs):
    """Write a captured row to another row, translating relative refs, then substituting."""
    for c, v in tmpl.items():
        if isinstance(v, str) and v.startswith('='):
            v = Translator(v, origin=f'{get_column_letter(c)}{src_row}') \
                .translate_formula(f'{get_column_letter(c)}{dst_row}')
            for a, b in subs:
                v = v.replace(a, b)
        ws.cell(dst_row, c).value = v


def main():
    wb = openpyxl.load_workbook(SRC)

    # ---------------------------------------------------------------- MU prices
    mdts, mO, mH, mL, mC, _p, _c = load_cand('MU')
    mu = {d: (mO[i], mH[i], mL[i], mC[i]) for i, d in enumerate(mdts)}

    # ---------------------------------------------------------------- Query
    q = wb['Query']
    old_col = {'NVDA': 2, 'TSM': 6, 'TSLA': 10, 'VRT': 14, 'VST': 18, 'AVGO': 22,
               'PLTR': 26, 'RKLB': 30, 'SOFI': 34, 'SPOT': 38}
    nrows = q.max_row
    grab = {}
    for s in KEEP:
        if s == 'MU': continue
        c0 = old_col[s]
        grab[s] = [[q.cell(r, c0 + j).value for j in range(4)] for r in range(2, nrows + 1)]
    dates = [to_date(q.cell(r, 1).value) for r in range(2, nrows + 1)]

    for r in range(2, nrows + 1):                       # clear every price column
        for c in range(2, 42):
            q.cell(r, c).value = None
    for c in range(2, 42):
        q.cell(1, c).value = None

    new_col = {}
    for i, s in enumerate(KEEP):
        c0 = 2 + 4*i
        new_col[s] = c0
        for j, sfx in enumerate(['_O', '_H', '_L', '_C']):
            q.cell(1, c0 + j).value = f'{s}{sfx}'
        for r in range(2, nrows + 1):
            if s == 'MU':
                d = dates[r-2]
                vals = mu.get(d)
                if vals is None: continue
            else:
                vals = grab[s][r-2]
            for j in range(4):
                q.cell(r, c0 + j).value = vals[j]

    # ---------------------------------------------------------------- Feed sheets
    if 'Feed MU' not in wb.sheetnames:
        fmu = wb.copy_worksheet(wb['Feed TSM']); fmu.title = 'Feed MU'
    for s in KEEP:
        ws = wb[f'Feed {s}']
        c0 = new_col[s]
        letters = [get_column_letter(c0 + j) for j in range(4)]
        ws.cell(1, 1).value = '=Query!A1'
        for j in range(4):
            ws.cell(1, 2 + j).value = f'=Query!{letters[j]}1'
        ws.cell(1, 7).value = (f'Daily OHLC for {s}. Loaded from the Query sheet; '
                               f'the Model sheet reads columns A:E.')
        for r in range(2, ws.max_row + 1):
            ws.cell(r, 1).value = '=INDEX(Query!$A:$A,ROW())'
            for j in range(4):
                L = letters[j]
                ws.cell(r, 2 + j).value = f'=INDEX(Query!${L}:${L},ROW())'

    # ---------------------------------------------------------------- Model MU
    if 'Model MU' not in wb.sheetnames:
        mm = wb.copy_worksheet(wb['Model TSM']); mm.title = 'Model MU'
        for row in mm.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and 'Feed TSM' in cell.value:
                    cell.value = cell.value.replace('Feed TSM', 'Feed MU')
        mm['A1'] = ('MU Daily Bayesian – 2 Tranches, 50-Day Stop-Loss  '
                    '(live Feed | auto-extends)')
        mm['B2'] = MU_PARAMS['lam'];      mm['D2'] = MU_PARAMS['phi_L']
        mm['F2'] = MU_PARAMS['psi'];      mm['H2'] = MU_PARAMS['k']
        mm['J2'] = MU_PARAMS['premium'];  mm['L2'] = MU_PARAMS['peak_cap']
        mm['B3'] = MU_PARAMS['ou_W'];     mm['D3'] = MU_PARAMS['ou_buf_k']
        mm['H3'] = MU_PARAMS['ou_prem'];  mm['J3'] = MU_PARAMS['ou_cap']

    for s in KEEP:                                   # uniform Bayes share and title
        m = wb[f'Model {s}']
        m['V2'] = BAYES_FLAT
        m['A1'] = f'{s} Daily Bayesian – 2 Tranches, 50-Day Stop-Loss  (live Feed | auto-extends)'

    # ---------------------------------------------------------------- Allocation
    al = wb['Allocation']
    src_row = 12                                     # TSM row, used as the template
    al_t = snapshot(al, src_row, range(1, 19))
    for i, s in enumerate(KEEP):
        dst = 11 + i
        move_row(al_t, src_row, al, dst, [("'Feed TSM'", f"'Feed {s}'")])
        al.cell(dst, 1).value = s
        al.cell(dst, 2).value = 1
        al.cell(dst, 3).value = ANNRET[s]
        al.cell(dst, 5).value = WEIGHT
        al.cell(dst, 17).value = BAYES_FLAT
    for r in range(16, 21):                          # clear the six vacated rows
        for c in range(1, 19):
            al.cell(r, c).value = None
    for r in range(11, 16):                          # ranges 11:20 -> 11:15
        for c in range(1, 19):
            v = al.cell(r, c).value
            if isinstance(v, str) and '$20' in v:
                for L in 'FHJ':
                    v = v.replace(f'${L}$20', f'${L}$15')
                al.cell(r, c).value = v
    al['B6'] = BAYES_FLAT
    al['B8'] = WEIGHT
    al['B9'] = 1
    al['B21'] = '=SUM(B11:B15)'
    for c, L in ((11, 'K'), (12, 'L'), (13, 'M'), (14, 'N')):
        al.cell(21, c).value = f'=SUM({L}11:{L}15)'
    al['A2'] = ('Edit only the blue cells. Five daily names, equal weighted: the floor (B8) and '
                'the cap (E11:E15) are both 0.20, which pins each name to 20%. Bayes fraction '
                '0.75, flat (split mode 1).')
    al['A1'] = ('PORTFOLIO ALLOCATION  —  five daily names, equal weighted. NVDA and AVGO are '
                'run on the weekly Monday model and are not funded from this sheet.')
    for i, s in enumerate(KEEP):                     # machine-readable block
        r = 25 + 2*i
        al.cell(r, 1).value = s;     al.cell(r, 2).value = 'Bayes'
        al.cell(r, 3).value = f'=M{11+i}'
        al.cell(r+1, 1).value = s;   al.cell(r+1, 2).value = 'OU'
        al.cell(r+1, 3).value = f'=N{11+i}'
    for r in range(35, 45):
        for c in range(1, 4):
            al.cell(r, c).value = None

    # ---------------------------------------------------------------- Dashboard
    db = wb['Dashboard']
    db_t = snapshot(db, 4, range(1, 18))             # NVDA row, before anything is written
    for i, s in enumerate(KEEP):
        R = 4 + i
        for c, v in db_t.items():
            if isinstance(v, str) and v.startswith('='):
                v = v.replace("'Feed NVDA'", f"'Feed {s}'").replace("'Model NVDA'", f"'Model {s}'")
                v = selfrow(v, 4, R)                 # only this row's own refs move
            db.cell(R, c).value = v
        db.cell(R, 1).value = s
    for r in range(9, 14):
        for c in range(1, 18):
            db.cell(r, c).value = None
    db['A1'] = 'HYBRID DASHBOARD — daily pre-open levels (5 daily names)'

    # ---------------------------------------------------------------- Active Trading
    at = wb['Active Trading']
    for i, s in enumerate(KEEP):                     # stock funds block, rows 7-11
        R = 7 + i; A = 11 + i
        at.cell(R, 1).value = s
        at.cell(R, 2).value = f'=Allocation!L{A}'
        at.cell(R, 3).value = f'=Allocation!R{A}'
        at.cell(R, 4).value = f'=B{R}*C{R}'
        at.cell(R, 5).value = f'=B{R}*(1-C{R})'
        at.cell(R, 6).value = f'=D{R}+E{R}'
        at.cell(R, 7).value = f'=F{R}-B{R}'
        at.cell(R, 8).value = f'=IFERROR(F{R}/B{R}-1,"")'
        at.cell(R, 9).value = f"='Model {s}'!J2"
        at.cell(R, 10).value = f"='Model {s}'!H3"
    for r in range(12, 17):
        for c in range(1, 11):
            at.cell(r, c).value = None

    ord_t = {t: snapshot(at, 19 + t, range(1, 11)) for t in (0, 1)}
    LAST = 19 + 2*len(KEEP) - 1                      # last order row after the cut
    for i, s in enumerate(KEEP):
        for t in (0, 1):
            src = 19 + t; R = 19 + 2*i + t
            for c, v in ord_t[t].items():
                if isinstance(v, str) and v.startswith('='):
                    # protect the block ranges and cross-sheet row refs from the row shift
                    v = (v.replace('$C$19:$C$38', '<<CR>>').replace('$D$19:$D$38', '<<DR>>')
                          .replace('Dashboard!$B$4:$B$13', '<<DB>>')
                          .replace('Dashboard!$A$4:$A$13', '<<DA>>')
                          .replace('$D$7', '<<BF>>').replace('$E$7', '<<OF>>'))
                    for L in 'CDEF':
                        v = v.replace(f'Dashboard!{L}4', f'<<D{L}>>')
                    v = selfrow(v, src, R)
                    v = (v.replace('<<CR>>', f'$C$19:$C${LAST}')
                          .replace('<<DR>>', f'$D$19:$D${LAST}')
                          .replace('<<DB>>', f'Dashboard!$B$4:$B${4+len(KEEP)-1}')
                          .replace('<<DA>>', f'Dashboard!$A$4:$A${4+len(KEEP)-1}')
                          .replace('<<BF>>', f'$D${7+i}').replace('<<OF>>', f'$E${7+i}'))
                    for L in 'CDEF':
                        v = v.replace(f'<<D{L}>>', f'Dashboard!{L}{4+i}')
                    v = v.replace("'Model NVDA'", f"'Model {s}'")
                at.cell(R, c).value = v
            at.cell(R, 1).value = s
    for r in range(LAST + 1, 39):
        for c in range(1, 11):
            at.cell(r, c).value = None
    at['A1'] = ('ACTIVE TRADING BLOTTER  —  five daily names  '
                '(TSM · VRT · VST · RKLB · MU)')

    # ---------------------------------------------------------------- drop sheets
    for s in DROP:
        for pre in ('Feed', 'Model'):
            nm = f'{pre} {s}'
            if nm in wb.sheetnames: del wb[nm]

    order = ['Notes', 'Allocation', 'Active Trading', 'Dashboard', 'Query']
    order += [f'Feed {s}' for s in KEEP] + [f'Model {s}' for s in KEEP]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames] + \
                 [ws for ws in wb._sheets if ws.title not in order]

    # the IBKR order range still spanned twenty rows; cut it to the ten now in use
    dn = wb.defined_names.get('IBKR_Orders')
    if dn is not None:
        dn.value = f"'Active Trading'!$A$18:$J${LAST}"

    nt = wb['Notes']
    nt['B3'] = ('Five daily names: TSM, VRT, VST, RKLB and MU. NVDA, TSLA, AVGO, PLTR, SOFI and '
                'SPOT removed — NVDA and AVGO moved to the weekly Monday single-tranche model, '
                'the other four dropped after failing the half-sample test.')
    nt['B4'] = ('Parameters unchanged from the fitted workbooks. MU added with its own fitted '
                'parameters. Bayes share 0.75 flat on all five.')
    nt['B5'] = ('Allocation is equal weight: floor and cap both 0.20. Enter the DAILY portion of '
                'capital in Allocation!B5 — if the two weekly names are funded from the same '
                'pot, that is five sevenths of the total.')
    nt['B6'] = ('MU price history in the Query sheet ends 24 June 2026, where its source '
                'workbook ends; a feed refresh will extend it.')

    wb.save(OUT)
    print(f'written {OUT}')
    print('sheets:', wb.sheetnames)


if __name__ == '__main__':
    main()
