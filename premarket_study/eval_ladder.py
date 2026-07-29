"""
Targeted evaluator for the laddered-column formulas. Reads the written formula strings from
the output sheet, pulls signal/param values from the source workbook's cache, and evaluates
the recursion row-by-row. If terminal fund / buys / annual match the engine, the Excel
formula transcription is verified (LibreOffice can't recalc this workbook in-sandbox).
"""
import re, datetime, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from engine import Params
from ladder_engine import run_ladder
from validate_nvda import load

OUT = '/home/user/David-Fewer/TradingExcel_s1_laddering_OptionC_backtest.xlsx'
SRC = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/2323ed8f-TradingExcel_s1_laddering.xlsx'
SHEET = 'Model NVDA'
LAD_START = 60                                   # first laddered column index
NAMES = ['FRESH','F0','PR1','PR2','PR3','B1','B2','B3','FIL1','FIL2','FIL3','DSH','SHMID',
         'FUNDMID','ANCM','TGT','STOP','EXIT','SALE','SH','FUND','F1','F2','F3','ANC','BD',
         'NB','INT','EQ']
COL = {n: get_column_letter(LAD_START + i) for i, n in enumerate(NAMES)}
EVAL_ORDER = ['FRESH','F0','PR1','PR2','PR3','B1','B2','B3','FIL1','FIL2','FIL3','DSH','SHMID',
              'FUNDMID','ANCM','TGT','BD','STOP','EXIT','SALE','SH','FUND','F1','F2','F3','ANC',
              'NB','INT','EQ']          # dependency-correct order (BD before STOP)
LAD_COLS = set(COL.values())
EPOCH = datetime.date(1899, 12, 30)


def to_serial(v):
    if isinstance(v, datetime.datetime): v = v.date()
    if isinstance(v, datetime.date): return (v - EPOCH).days
    return v


# ---- load caches ----
srcws = openpyxl.load_workbook(SRC, data_only=True)[SHEET]
cache = {}                                        # coordinate -> value (signal/param/config)
for row in srcws.iter_rows():
    for c in row:
        if c.value is not None and c.value != '':      # treat formula-"" cells as blank
            cache[c.coordinate] = to_serial(c.value)
outws_f = openpyxl.load_workbook(OUT, data_only=False)[SHEET]
CFG_VAL_COL = get_column_letter(LAD_START + len(NAMES) + 2)
for r in range(2, 7):                             # config values live in OUT as literals
    cache[f'{CFG_VAL_COL}{r}'] = outws_f[f'{CFG_VAL_COL}{r}'].value

grid = {}                                         # laddered coordinate -> computed value

def cv(ref):
    ref = ref.replace('$', '')
    col = re.match(r'[A-Z]+', ref).group()
    if col in LAD_COLS:
        return grid.get(ref)
    return cache.get(ref)

def IF(a, b, c): return b if a else c
def AND(*a): return all(a)
def OR(*a): return any(a)
def ISNUMBER(x): return isinstance(x, (int, float)) and not isinstance(x, bool)
MIN, MAX = min, max

REF = re.compile(r'\$?[A-Z]{1,3}\$?[0-9]+')

def _split_args(content):
    args, depth, cur = [], 0, ''
    for ch in content:
        if ch == '(':
            depth += 1; cur += ch
        elif ch == ')':
            depth -= 1; cur += ch
        elif ch == ',' and depth == 0:
            args.append(cur); cur = ''
        else:
            cur += ch
    args.append(cur)
    return args

def transform_IF(expr):
    # convert IF(a,b,c) -> ((b) if (a) else (c)); innermost/rightmost first (lazy branches)
    while 'IF(' in expr:
        i = expr.rfind('IF(')
        j, depth = i + 3, 1
        while depth:
            if expr[j] == '(':
                depth += 1
            elif expr[j] == ')':
                depth -= 1
            j += 1
        a, b, c = _split_args(expr[i + 3:j - 1])
        expr = expr[:i] + f'(({b}) if ({a}) else ({c}))' + expr[j:]
    return expr

def translate(formula):
    s = formula[1:] if formula.startswith('=') else formula
    s = s.replace('""', 'None')
    s = REF.sub(lambda m: f'cv("{m.group()}")', s)
    s = s.replace('<>', '!=')
    s = re.sub(r'(?<![<>=!])=(?!=)', '==', s)
    return transform_IF(s)


# ---- collect + pre-translate laddered formulas per column (row-generic via row token) ----
fcache = {}
for name in NAMES:
    col = COL[name]
    for r in range(8, 868):
        v = outws_f[f'{col}{r}'].value
        if isinstance(v, str) and v.startswith('='):
            fcache[(name, r)] = translate(v)
        else:
            grid[f'{col}{r}'] = 0 if v is None else v   # init-row literals

ns = {'cv': cv, 'IF': IF, 'AND': AND, 'OR': OR, 'ISNUMBER': ISNUMBER, 'MIN': MIN, 'MAX': MAX, 'None': None}
for r in range(8, 868):
    for name in EVAL_ORDER:
        expr = fcache.get((name, r))
        if expr is None:
            continue
        grid[f'{COL[name]}{r}'] = eval(expr, ns)

# ---- compare to engine ----
dates, O, H, L, C = load('nvda_ohlc.csv'); p = Params()
r = run_ladder(dates, O, H, L, C, p, [p.k, 1.3*p.k, 1.7*p.k], [p.ou_buf_k], 'first',
               [0.80, 0.15, 0.05], None)
fund867 = grid[f'{COL["FUND"]}867']
buys = sum(grid.get(f'{COL["NB"]}{i}', 0) or 0 for i in range(9, 868))
print('EVALUATED WRITTEN FORMULAS vs ENGINE (NVDA, Option C):')
print(f'  Bayes terminal fund: formulas={fund867:.2f}  engine={r["bayes_fund"][-1]:.2f}')
print(f'  Bayes buys:          formulas={buys}  engine={r["bayes_trades"]}')
ok = abs(fund867 - r['bayes_fund'][-1]) < 1e-4 and buys == r['bayes_trades']
print('  RESULT:', 'FORMULAS VERIFIED' if ok else 'MISMATCH — transcription bug')

# ---- diagnostic: first divergence vs mirror (engine logic) ----
from mirror_ladder import mirror as _mirror
mir = _mirror(dates, O, H, L, C, p)
# map row r (8..) to engine index i = r-8
print('\n--- divergence scan (evaluator grid vs engine mirror) ---')
found = 0
for i in range(len(C)):
    r = i + 8
    ef = grid.get(f'{COL["FUND"]}{r}')
    es = grid.get(f'{COL["SH"]}{r}')
    if ef is None: continue
    if abs(ef - mir['FUND'][i]) > 0.01 or abs((es or 0) - mir['SH'][i]) > 1e-6:
        # also compare signal K/W/G sheet vs engine
        Kv = cache.get(f'K{r}'); Wv = cache.get(f'W{r-1}'); Gv = cache.get(f'G{r-1}')
        eK = mir_fair = None
        print(f'row {r} (i={i}) date signal:')
        print(f'  sheet K{r}={Kv}  engineFair(Lvl+Slp[i-1])... W{r-1}={Wv} G{r-1}={Gv}')
        print(f'  FUND eval={ef:.2f} mir={mir["FUND"][i]:.2f}   SH eval={es} mir={mir["SH"][i]}')
        found += 1
        if found >= 3: break
if not found:
    print('no divergence >0.01')

print('\n--- tail scan (rows 585-867): buys or fund changes ---')
prevf=None
for r in range(585, 868):
    nb = grid.get(f'{COL["NB"]}{r}', 0) or 0
    fu = grid.get(f'{COL["FUND"]}{r}')
    sh = grid.get(f'{COL["SH"]}{r}')
    Bv = cache.get(f'B{r}')
    if nb>0 or (prevf is not None and fu is not None and abs(fu-prevf)>0.01):
        print(f'  row {r}: B={Bv!r} NB={nb} SH={sh} FUND={fu:.2f} (prev {prevf if prevf else 0:.2f})')
    prevf = fu
