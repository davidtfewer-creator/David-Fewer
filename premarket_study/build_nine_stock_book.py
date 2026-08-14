"""
Add MRVL to the live eight-name workbook -> the nine-name book.

Follows build_eight_stock_book.py: EXTENDS the uploaded live file in place. The
trade log (26 rows), refreshed Query feed (through 2026-08-13), fees, notes and
every user-entered cell are preserved. MRVL enters per the August 2026 admission
round: reference vector frozen (lam normalised to 1 via the exact degeneracy),
planning figure 78%.

  Query           MRVL OHLC appended in AH:AK, matched by date (both run to 13 Aug).
  Feed/Model MRVL copies of the TSM pair, repointed; reference params written.
  Allocation      row 19; weights, floor and caps to 1/9; machine block rebuilt for
                  18 sleeves (rows 25-42); summary block shifted below (44-46).
  Dashboard       row 12.
  Active Trading  funds row 15, order rows 35-36, lookups widened;
                  IBKR_Orders -> $A$18:$J$36. Trade log untouched (starts row 40).
"""
import copy as pycopy
import datetime
import json
import math

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from fresh_opt_cands import daily_from_5min

SRC = ('/root/.claude/uploads/9e026445-9e62-588b-af81-7e0c231b0f24/'
       'a8444819-TradingExcel_8stock_live.xlsx')
OUT = '/home/user/David-Fewer/TradingExcel_9stock_live.xlsx'

OLD = ['TSM', 'VRT', 'VST', 'RKLB', 'MU', 'GM', 'VLO', 'CF']
NEW = 'MRVL'
ALL = OLD + [NEW]
W9 = 1.0 / 9.0
ANNRET = 0.78                       # MRVL planning figure (train-only variant floors)


def mrvl_params():
    vec = json.load(open('fresh_opt_cands.json'))['MRVL']['reference']['vec']
    return dict(lam=1.0, phi_L=math.exp(vec[0]), psi=vec[1], k=vec[2],
                premium=vec[3], peak_cap=vec[4], ou_buf_k=vec[5], ou_prem=vec[6],
                ou_cap=vec[7], ou_W=int(round(vec[8])))


def to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
    return None


def selfrow(f, src, dst):
    import re
    return re.sub(r"(?<![!$A-Za-z0-9_])(\$?)([A-Z]{1,2})" + str(src) + r"(?![0-9])",
                  lambda m: f"{m.group(1)}{m.group(2)}{dst}", f)


def snapshot(ws, row, cols):
    return {c: (ws.cell(row, c).value, ws.cell(row, c)._style) for c in cols}


def write_row(tmpl, src_row, ws, dst_row, subs):
    for c, (v, style) in tmpl.items():
        if isinstance(v, str) and v.startswith('='):
            v = Translator(v, origin=f'{get_column_letter(c)}{src_row}') \
                .translate_formula(f'{get_column_letter(c)}{dst_row}')
            for a, b in subs:
                v = v.replace(a, b)
        cell = ws.cell(dst_row, c)
        cell.value = v
        cell._style = pycopy.copy(style)


def main():
    wb = openpyxl.load_workbook(SRC)
    p = mrvl_params()

    # ------------------------------------------------------------------ Query
    q = wb['Query']
    dates = {}
    for r in range(2, q.max_row + 1):
        d = to_date(q.cell(r, 1).value)
        if d:
            dates[d] = r
    c0 = 34                                              # AH
    for j, sfx in enumerate(['_O', '_H', '_L', '_C']):
        q.cell(1, c0 + j).value = f'{NEW}{sfx}'
    dts, O, H, L, C = daily_from_5min(NEW)
    n = 0
    for k, d in enumerate(dts):
        r = dates.get(d)
        if r is None:
            continue
        q.cell(r, c0).value = O[k]
        q.cell(r, c0 + 1).value = H[k]
        q.cell(r, c0 + 2).value = L[k]
        q.cell(r, c0 + 3).value = C[k]
        n += 1
    print(f'Query: {NEW} cols AH:AK, {n} rows, last {dts[-1]}')

    # ------------------------------------------------------------------ Feed
    ws = wb.copy_worksheet(wb['Feed TSM'])
    ws.title = f'Feed {NEW}'
    letters = [get_column_letter(c0 + j) for j in range(4)]
    ws.cell(1, 1).value = '=Query!A1'
    for j in range(4):
        ws.cell(1, 2 + j).value = f'=Query!{letters[j]}1'
    ws.cell(1, 7).value = (f'Daily OHLC for {NEW}. Loaded from the Query sheet; '
                           f'the Model sheet reads columns A:E.')
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).value = '=INDEX(Query!$A:$A,ROW())'
        for j in range(4):
            Lc = letters[j]
            ws.cell(r, 2 + j).value = f'=INDEX(Query!${Lc}:${Lc},ROW())'
    print(f'Feed {NEW}: -> Query {letters[0]}:{letters[3]}')

    # ------------------------------------------------------------------ Model
    m = wb.copy_worksheet(wb['Model TSM'])
    m.title = f'Model {NEW}'
    for row in m.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and 'Feed TSM' in cell.value:
                cell.value = cell.value.replace('Feed TSM', f'Feed {NEW}')
    m['A1'] = (f'{NEW} Daily Bayesian – 2 Tranches, 50-Day Stop-Loss  '
               f'(live Feed | auto-extends)')
    m['B2'] = p['lam'];      m['D2'] = p['phi_L'];   m['F2'] = p['psi']
    m['H2'] = p['k'];        m['J2'] = p['premium']; m['L2'] = p['peak_cap']
    m['V2'] = 0.5
    m['B3'] = p['ou_W'];     m['D3'] = p['ou_buf_k']
    m['H3'] = p['ou_prem'];  m['J3'] = p['ou_cap']
    print(f'Model {NEW}: params written (lam=1 normalised vector)')

    # -------------------------------------------------------------- Allocation
    al = wb['Allocation']
    # capture the summary block and notes BEFORE anything moves
    summ = [(al.cell(r, 1).value, al.cell(r, 1)._style, al.cell(r, 2)._style)
            for r in (42, 43, 44)]
    note1 = al['A46'].value
    note2 = al['A47'].value
    al['A46'] = None
    al['A47'] = None

    al_t = snapshot(al, 11, range(1, 19))
    write_row(al_t, 11, al, 19, [("'Feed TSM'", f"'Feed {NEW}'")])
    al.cell(19, 1).value = NEW
    al.cell(19, 2).value = 1
    al.cell(19, 3).value = ANNRET
    al.cell(19, 17).value = 0.5                          # Q: per-stock Bayes share
    for r in range(11, 20):                              # widen 11:18 -> 11:19
        for c in range(1, 19):
            v = al.cell(r, c).value
            if isinstance(v, str) and '$18' in v:
                for Lc in 'FHJ':
                    v = v.replace(f'${Lc}$11:${Lc}$18', f'${Lc}$11:${Lc}$19')
                al.cell(r, c).value = v
        al.cell(r, 5).value = W9                         # caps 1/9
    al['B8'] = W9                                        # floor 1/9
    al['B21'] = '=SUM(B11:B19)'
    for c, Lc in ((11, 'K'), (12, 'L'), (13, 'M'), (14, 'N')):
        al.cell(21, c).value = f'=SUM({Lc}11:{Lc}19)'

    mac_style = {c: al.cell(25, c)._style for c in range(1, 7)}
    for j in range(18):                                  # machine block, 18 sleeves
        r = 25 + j
        i = j // 2
        s = ALL[i]
        tranche = 'Bayes' if j % 2 == 0 else 'OU'
        share = f'$R${11+i}' if tranche == 'Bayes' else f'(1-$R${11+i})'
        al.cell(r, 1).value = s
        al.cell(r, 2).value = tranche
        al.cell(r, 3).value = f'=IF(SUM($F$25:$F$42)=0,0,$B$5*F{r}/SUM($F$25:$F$42))'
        al.cell(r, 4).value = f"=IF('Active Trading'!$C${19+j}=\"HOLDING\",1,0)"
        al.cell(r, 5).value = f'=$K${11+i}*{share}'
        al.cell(r, 6).value = f'=E{r}*(1-D{r})'
        for c in range(1, 7):
            al.cell(r, c)._style = pycopy.copy(mac_style[c])

    formulas = ['=COUNT(D25:D42)-SUM(D25:D42)', '=IF(B44=0,0,$B$5/B44)',
                '=IF(ABS(SUM(C25:C42)-$B$5)<0.01,"OK","MISMATCH")']
    for k, (label, sa, sb) in enumerate(summ):
        r = 44 + k
        al.cell(r, 1).value = label
        al.cell(r, 1)._style = pycopy.copy(sa)
        al.cell(r, 2).value = formulas[k]
        al.cell(r, 2)._style = pycopy.copy(sb)
    al['A48'] = ('Held? is read from the blotter Status in Active Trading C19:C36; '
                 'a sleeve marked HOLDING gets nothing and its weight is spread '
                 'over the free sleeves.')
    al['A49'] = note2
    al['A1'] = ('PORTFOLIO ALLOCATION  —  nine daily names, equal weighted. MRVL '
                'admitted August 2026 (planning figure 78%, AI class).')
    al['A2'] = ('Edit only the blue cells. Nine daily names, equal weighted: the '
                'floor (B8) and the cap (E11:E19) are both 1/9, which pins each '
                'name to 11.1%.')
    print('Allocation: row 19 added, weights 1/9, machine block 18 sleeves (25-42)')

    # -------------------------------------------------------------- Dashboard
    db = wb['Dashboard']
    db_t = snapshot(db, 4, range(1, 18))
    for c, (v, style) in db_t.items():
        if isinstance(v, str) and v.startswith('='):
            v = v.replace("'Feed TSM'", f"'Feed {NEW}'") \
                 .replace("'Model TSM'", f"'Model {NEW}'")
            v = selfrow(v, 4, 12)
        cell = db.cell(12, c)
        cell.value = v
        cell._style = pycopy.copy(style)
    db.cell(12, 1).value = NEW
    db['A1'] = 'HYBRID DASHBOARD — daily pre-open levels (9 daily names)'
    print('Dashboard: row 12 added')

    # ---------------------------------------------------------- Active Trading
    at = wb['Active Trading']
    at_t = snapshot(at, 7, range(1, 7))
    for c, (v, style) in at_t.items():
        if isinstance(v, str) and v.startswith('='):
            v = selfrow(v, 7, 15)
        cell = at.cell(15, c)
        cell.value = v
        cell._style = pycopy.copy(style)
    at.cell(15, 1).value = NEW
    for r in range(7, 16):
        for c in range(1, 7):
            v = at.cell(r, c).value
            if isinstance(v, str):
                v = v.replace('$R$11:$R$18', '$R$11:$R$19') \
                     .replace('$A$11:$A$18', '$A$11:$A$19')
                at.cell(r, c).value = v

    ord_t = {t: snapshot(at, 19 + t, range(1, 13)) for t in (0, 1)}
    for t in (0, 1):
        src = 19 + t
        R = 35 + t
        for c, (v, style) in ord_t[t].items():
            if isinstance(v, str) and v.startswith('='):
                for Lc in 'CDEF':
                    v = v.replace(f'Dashboard!{Lc}4', f'<<{Lc}>>')
                v = v.replace("'Model TSM'", f"'Model {NEW}'")
                v = selfrow(v, src, R)
                for Lc in 'CDEF':
                    v = v.replace(f'<<{Lc}>>', f'Dashboard!{Lc}12')
            cell = at.cell(R, c)
            cell.value = v
            cell._style = pycopy.copy(style)
        at.cell(R, 1).value = NEW
        at.cell(R, 2).value = 'Bayes' if t == 0 else 'OU'
    for r in range(19, 37):
        for c in range(1, 13):
            v = at.cell(r, c).value
            if isinstance(v, str):
                v = v.replace('Allocation!$C$25:$C$40', 'Allocation!$C$25:$C$42') \
                     .replace('Dashboard!$B$4:$B$11', 'Dashboard!$B$4:$B$12') \
                     .replace('Dashboard!$A$4:$A$11', 'Dashboard!$A$4:$A$12')
                at.cell(r, c).value = v
    at['A1'] = ('ACTIVE TRADING BLOTTER  —  nine daily names  '
                '(TSM · VRT · VST · RKLB · MU · GM · VLO · CF · MRVL)')
    print('Active Trading: funds row 15, order rows 35-36, ranges widened')

    dn = wb.defined_names.get('IBKR_Orders')
    dn.value = "'Active Trading'!$A$18:$J$36"
    print('IBKR_Orders ->', dn.value)

    nt = wb['Notes']
    nt['B3'] = ('Nine daily names: TSM, VRT, VST, RKLB, MU, GM, VLO, CF + MRVL '
                '(admitted August 2026: planning 78% on train-only fits, AI class, '
                'beta 0.95). AMD, SMCI and CEG assessed in the same round and '
                'declined; AVGO on the watch list to early 2027.')
    nt['B4'] = ('Parameters frozen at admission. MRVL carries its reference vector '
                'normalised to lam=1 (the exact degeneracy); residual OU sigma; '
                'Bayes share 0.50 on all nine (split mode B9=2 reads Q11:Q19).')
    nt['B5'] = ('Allocation is equal weight: floor and cap both 1/9 (11.1%). Enter '
                'the CASH available this morning in Allocation!B5.')
    wb.save(OUT)
    print(f'\nwritten {OUT}')


if __name__ == '__main__':
    main()
