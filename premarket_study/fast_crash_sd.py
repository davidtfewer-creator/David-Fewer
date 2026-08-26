"""
Intraday fast-crash stand-down study (user request, 26 Aug 2026).

The one protection gap on the books: the 200dma gate is grinding-bear armour but
is too slow for a V-crash (April 2025 deepened 29->32% under the gate because the
pool concentrated into the last ungated names), and the daily circuit breaker
failed 2022 because once tripped against an all-time-high reference it stayed
tripped for months, half-sizing the diversifier rotation. Only automation can
watch the book INTRADAY -- this is that watcher, tested.

RULE (veto only, strictly ex ante):
  reference  = max of the last R daily close equities (ROLLING peak: decays in a
               grind, so the rule re-arms instead of staying dead -- speed, not
               depth, is what trips it)
  trigger    = first 5-minute bar where morning cash + held book (marked at bar
               opens) <= reference*(1-trip). From that bar, unfilled bids are
               cancelled (a bid whose first touch was at/before the trigger bar
               still fills -- pessimistic for the rule).
  stand-down = no new bids until a close recovers above reference*(1-reset).
  Resting exits, targets and stops run unchanged throughout.

Approximation (documented in book_sim): the intraday path ignores same-day fills
and exits; crash-day fills would only make the real trigger EARLIER, so the
measured benefit is conservative.

Evaluated on BOTH regimes:
  2024-26  live config (09:00/4% PM rule both sides): return cost on both
           halves + the April 2025 episode (DD over Feb-Jul 2025), with and
           without the 200dma gate.
  2022     the bear replay: must NOT reproduce the daily breaker's anti-synergy
           with the gate. Recorded matrix reproduced first as invariance check.
"""
import datetime
import json
import os
import pickle

import numpy as np

import bear_replay
import book_sim
from book_sim import NAMES as N8, load_all, simulate
from delayed_entry import full_bars
from engine import Params
from fresh_opt_cands import aw_params
from live5_load import load as load_book

APR_LO, APR_HI = datetime.date(2025, 2, 1), datetime.date(2025, 7, 31)
BEAR_BARS_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               'data_bear', 'bars_cache.pkl')
OUT = 'fast_crash_sd.json'


def window_dd(eq, cal, lo, hi):
    seg = [e for e, d in zip(eq, cal) if lo <= d <= hi]
    peak, dd = -1e30, 0.0
    for e in seg:
        peak = max(peak, e)
        if peak > 0:
            dd = max(dd, (peak - e) / peak)
    return dd


def attach_bars_2426(data):
    for s in data:
        fb = full_bars(s)
        data[s]['bars'] = {d: (v[0], v[1], v[3]) for d, v in fb.items()}  # tm, open, low


def bear_bars(stock):
    cache = {}
    if os.path.exists(BEAR_BARS_CACHE):
        with open(BEAR_BARS_CACHE, 'rb') as f:
            cache = pickle.load(f)
    if stock in cache:
        return cache[stock]
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(bear_replay.DIR, f'{stock}_5min.xlsx'),
                                read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = {}
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        dt, o, l = row[0], row[1], row[3]
        if isinstance(dt, str):
            try:
                dt = datetime.datetime.fromisoformat(dt)
            except ValueError:
                continue
        if not isinstance(dt, datetime.datetime) or o is None:
            continue
        t = dt.time()
        if t < bear_replay.RTH_START or t >= bear_replay.RTH_END:
            continue
        m = (t.hour - 9) * 60 + t.minute - 30
        by.setdefault(dt.date(), []).append((m, float(o), float(l)))
    wb.close()
    idx = {}
    for d, bars in by.items():
        bars.sort()
        idx[d] = tuple(np.array(x) for x in zip(*bars))
    cache[stock] = idx
    with open(BEAR_BARS_CACHE, 'wb') as f:
        pickle.dump(cache, f)
    return idx


def rep(label, r, cal):
    apr = window_dd(r['equity'], cal, APR_LO, APR_HI)
    print(f"  {label:30s} full {r['full']*100:6.1f}%  train {r['train']*100:6.1f}%  "
          f"test {r['test']*100:6.1f}%  DD {r['maxdd']*100:5.1f}%  "
          f"Apr25 {apr*100:5.1f}%  trips {r['sd_trips']:2d}  halt d {r['sd_days']:3d}",
          flush=True)
    return dict(full=r['full'], train=r['train'], test=r['test'], maxdd=r['maxdd'],
                apr=apr, sd_trips=r['sd_trips'], sd_days=r['sd_days'],
                fills=r['fills'])


def main():
    # ================= 2024-26, live configuration =================
    t0 = Params(capital=1_000_000, comm=0.005, interest=0.0314, stop_days=50,
                bayes_pct=0.5, years=2.2, ou_W=80)
    mrvl = aw_params(json.load(open('fresh_opt_cands.json'))['MRVL']['reference']['vec'], t0)
    names = N8 + ['MRVL']
    data, sleeves, cal = load_all(names=names, params_override={'MRVL': mrvl})
    pm = pickle.load(open('data_pm/pm_last_cuts.pkl', 'rb'))['09:00']
    def pm_rule(name, i, bid):
        pml = pm[name].get(data[name]['dts'][i])
        return pml is not None and bid < pml * (1 - 0.04)
    attach_bars_2426(data)
    gate = {s: bear_replay.dma_gate_dates(data[s]['dts'], data[s]['C']) for s in names}

    res = {'2024-26': {}}
    print('===== 2024-26, pooled nine-name book, 09:00/4% PM rule both sides =====')
    r = simulate(data, sleeves, cal, mode='pooled', excl_fn=pm_rule)
    res['2024-26']['baseline (PM rule)'] = rep('baseline (PM rule)', r, cal)
    r = simulate(data, sleeves, cal, mode='pooled', excl_fn=pm_rule, no_buy=gate)
    res['2024-26']['+ 200dma gate'] = rep('+ 200dma gate', r, cal)

    print('  --- stand-down alone (R=5, reset=trip/2) ---')
    for trip in (0.06, 0.08, 0.10, 0.12):
        r = simulate(data, sleeves, cal, mode='pooled', excl_fn=pm_rule,
                     intraday_sd=(trip, trip / 2, 5))
        res['2024-26'][f'SD {trip:.0%}/R5'] = rep(f'+ SD {trip:.0%} (R=5)', r, cal)
    r = simulate(data, sleeves, cal, mode='pooled', excl_fn=pm_rule,
                 intraday_sd=(0.10, 0.05, 10))
    res['2024-26']['SD 10%/R10'] = rep('+ SD 10% (R=10)', r, cal)

    print('  --- stacked on the 200dma gate (the April-2025 fix test) ---')
    for trip in (0.08, 0.10):
        r = simulate(data, sleeves, cal, mode='pooled', excl_fn=pm_rule, no_buy=gate,
                     intraday_sd=(trip, trip / 2, 5))
        res['2024-26'][f'gate + SD {trip:.0%}'] = rep(f'gate + SD {trip:.0%}', r, cal)

    # ================= 2022 bear replay =================
    print('\n===== 2022 bear replay (no PM data pre-2024; PM rule off, as recorded) =====')
    _, book_params, _ = load_book()
    bdata = {}
    for s in bear_replay.NAMES:
        (dts, O, H, L, C), idx = bear_replay._load(s)
        p = bear_replay.params_for(s, book_params)
        from engine import run_model
        chk = bear_replay.bear_checker(idx, dts, O)
        rm = run_model(dts, O, H, L, C, p, ou_sigma='resid', same_day_exit=chk,
                       collect=True)
        bdata[s] = dict(dts=dts, O=O, H=H, L=L, C=C, p=p,
                        idx={d: i for i, d in enumerate(dts)}, chk=chk,
                        X=rm.frames['X'], AM=rm.frames['AM'],
                        bars=bear_bars(s))
        print(f'  {s} loaded', flush=True)
    bsleeves = []
    for s in bear_replay.NAMES:
        p = bdata[s]['p']
        bsleeves.append(dict(name=s, kind='B', bids='X', prem=p.premium))
        bsleeves.append(dict(name=s, kind='O', bids='AM', prem=p.ou_prem))
    common = None
    for s in bear_replay.NAMES:
        common = set(bdata[s]['dts']) if common is None else common & set(bdata[s]['dts'])
    bcal = sorted(common)
    bgate = {s: bear_replay.dma_gate_dates(bdata[s]['dts'], bdata[s]['C'])
             for s in bear_replay.NAMES}
    brk = (0.15, 0.075, 0.5)
    Y22a, Y22b = bear_replay.Y22a, bear_replay.Y22b

    res['2022'] = {}
    configs = [('nothing', {}),
               ('200dma gate', dict(no_buy=bgate)),
               ('daily breaker 15%/half', dict(breaker=brk)),
               ('gate + daily breaker', dict(no_buy=bgate, breaker=brk)),
               ('SD 8% (R=5)', dict(intraday_sd=(0.08, 0.04, 5))),
               ('SD 10% (R=5)', dict(intraday_sd=(0.10, 0.05, 5))),
               ('gate + SD 8%', dict(no_buy=bgate, intraday_sd=(0.08, 0.04, 5))),
               ('gate + SD 10%', dict(no_buy=bgate, intraday_sd=(0.10, 0.05, 5)))]
    print(f"  {'config':30s}{'2022 total':>11s}{'2022 DD':>9s}{'trips':>6s}"
          f"{'halt d':>7s}{'fills':>7s}")
    for label, kw in configs:
        r = simulate(bdata, bsleeves, bcal, capital=9_000_000, mode='pooled',
                     date_lo=Y22a, date_hi=Y22b, **kw)
        eq = r['equity']
        tot = eq[-1] / eq[0] - 1
        res['2022'][label] = dict(total=tot, maxdd=r['maxdd'],
                                  sd_trips=r['sd_trips'], sd_days=r['sd_days'],
                                  fills=r['fills'])
        print(f"  {label:30s}{tot*100:>10.1f}%{r['maxdd']*100:>8.1f}%"
              f"{r['sd_trips']:>6d}{r['sd_days']:>7d}{r['fills']:>7d}", flush=True)

    with open(OUT, 'w') as f:
        json.dump(res, f, indent=1)
    print(f'\nsaved {OUT}')


if __name__ == '__main__':
    main()
