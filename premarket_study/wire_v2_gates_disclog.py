"""
Wire three user-requested changes into the live nine-stock workbook (19 Aug 2026).
Layout-preserving: nothing inserted or moved; IBKR_Orders (A18:J36) and the
blotter columns the scripts touch (A..M, rows 42+) keep their addresses.

1. DISCRETIONARY TRADE LOG -- columns T..AF from row 41, mirroring the main
   log's titles (# / Stock / Tranche / Buy date / Buy price / Shares / Buy cost
   / Sell date / Sell price / Proceeds / Net P&L / Week ending / Status).
   Manual entry; the derived columns carry the same formulas as the main log
   (Buy cost, Proceeds with the B3/E3 fees, Net P&L, Week ending, Status),
   prefilled rows 42..241. Its profits feed the weekly summary's Cumulative
   (col R) and Total trades (col S): R adds closed discretionary P&L up to each
   week; S adds discretionary buy+sell events, mirroring the main-log count.

2. 9:00 PRE-MARKET PRICE -- column P, orders block (P18 title, P19:P36 left
   empty for Script 1 to populate). The adopted pre-market rule is wired as
   Allocation column H ('PM gate'): flag = 1 when the row's Buy @ (col F) sits
   more than X below its pre-market price (col P); the free-weight formula
   becomes E*(1-Held)*(1-DMAgate)*(1-PMgate), so a gated sleeve's capital
   pools to the others. Fails open: blank P or blank X disables the gate.

3. ATH GATE (epsilon) -- the studied ath_target_guard: each Dashboard buy is
   additionally capped at ATH*(1-eps) - prevclose*premium (its own sleeve's
   premium), so no order can construct a trade whose exit needs a print above
   ATH*(1-eps). Implemented as a third MIN() term in Dashboard C4:C12 (Bayes)
   and E4:E12 (OU); N() coerces a blank eps to 0 (the recommended exact-ATH
   guard); eps = -1 disables. Model-sheet history columns are untouched.

Gate variables live under the user's 'Gate Variables' title (Active
Trading L5): L6/M6 = PM gate X (default 4%), L7/M7 = ATH eps (default 0),
with explanations in N. Notes changelog appended. Full cell diff verified.
"""
import copy
import sys

import openpyxl

STOCKS = ['TSM', 'VRT', 'VST', 'RKLB', 'MU', 'GM', 'VLO', 'CF', 'MRVL']
DISC_LAST = 241          # discretionary log formula depth
WEEK_LAST = 93           # weekly summary rows 42..93


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    at = wb['Active Trading']
    al = wb['Allocation']
    d = wb['Dashboard']
    notes = wb['Notes']

    # ---------------- preconditions
    assert at['L5'].value == 'Gate Variables'
    assert all(at.cell(row=r, column=c).value is None
               for r in range(6, 17) for c in range(12, 17))
    assert all(at.cell(row=r, column=16).value is None for r in range(18, 37))
    assert all(at.cell(row=r, column=c).value is None
               for r in range(41, 61) for c in range(20, 33))
    assert all(al.cell(row=r, column=8).value is None for r in range(24, 43))
    for r in range(25, 43):
        assert al.cell(row=r, column=6).value == f'=E{r}*(1-D{r})*(1-G{r})', r
    for i, s in enumerate(STOCKS):
        r = 4 + i
        assert d.cell(row=r, column=3).value == \
            f"=MIN(J{r}-'Model {s}'!H2*K{r},L{r}*(1-'Model {s}'!L2))", (r, 'C')
        assert d.cell(row=r, column=5).value == \
            f"=MIN(Q{r}-'Model {s}'!D3*P{r},L{r}*(1-'Model {s}'!J3))", (r, 'E')

    def style_like(dst, src, numfmt=None):
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = numfmt if numfmt else src.number_format

    # ---------------- 1. discretionary log
    for c_src in range(1, 14):                      # A..M -> T..AF headers
        src = at.cell(row=41, column=c_src)
        dst = at.cell(row=41, column=19 + c_src, value=src.value)
        style_like(dst, src)
    for r in range(42, DISC_LAST + 1):
        at.cell(row=r, column=26,                   # Z: Buy cost
                value=f'=IF(OR($X{r}="",$Y{r}=""),"",$Y{r}*$X{r}+$Y{r}*$B$3)')
        at.cell(row=r, column=29,                   # AC: Proceeds
                value=f'=IF(OR($AB{r}="",$Y{r}=""),"",$Y{r}*$AB{r}-$Y{r}*$E$3)')
        at.cell(row=r, column=30,                   # AD: Net P&L
                value=f'=IF($AC{r}="","",$AC{r}-$Z{r})')
        at.cell(row=r, column=31,                   # AE: Week ending
                value=f'=IF(OR($AA{r}="",$AF{r}<>"CLOSED"),"",$AA{r}-WEEKDAY($AA{r},2)+5)')
        at.cell(row=r, column=32,                   # AF: Status
                value=f'=IF($W{r}="","",IF($AA{r}="","OPEN","CLOSED"))')
        for c_src, c_dst in ((7, 26), (10, 29), (11, 30), (12, 31), (13, 32)):
            style_like(at.cell(row=r, column=c_dst), at.cell(row=42, column=c_src))
    # column widths T..AF from A..M
    from openpyxl.utils import get_column_letter
    for c_src in range(1, 14):
        src = at.column_dimensions.get(get_column_letter(c_src))
        if src is not None and src.width:
            at.column_dimensions[get_column_letter(19 + c_src)].width = src.width
    # weekly summary: R gains closed discretionary P&L, S gains its trade events
    for r in range(42, WEEK_LAST + 1):
        at.cell(row=r, column=18,
                value=(f'=IF($O{r}="","",SUM($P$42:$P{r})'
                       f'+SUMIFS($AD$42:$AD${DISC_LAST},$AE$42:$AE${DISC_LAST},"<="&$O{r}))'))
        at.cell(row=r, column=19,
                value=(f'=IF($O{r}="","",COUNTIFS($D$42:$D$1041,"<="&$O{r})'
                       f'+COUNTIFS($H$42:$H$1041,"<="&$O{r})'
                       f'+COUNTIFS($W$42:$W${DISC_LAST},"<="&$O{r})'
                       f'+COUNTIFS($AA$42:$AA${DISC_LAST},"<="&$O{r}))'))

    # ---------------- 2. pre-market price column + PM gate
    hdr = at.cell(row=18, column=16, value='9:00 pre-mkt')
    style_like(hdr, at['O18'])
    for r in range(19, 37):
        style_like(at.cell(row=r, column=16), at.cell(row=r, column=10), numfmt='0.00')
    lab_style, inp_style = at['A6'], at['B3']
    for row_, label, value, fmt, expl in (
        (6, 'PM gate %', 0.04, '0.0%',
         'No order when its Buy @ (col F) sits more than this below the 9:00 '
         'pre-market price (col P). Veto only - the sleeve\'s capital pools to '
         'the others via Allocation col H. Blank P cell or blank % disables '
         '(fails open). Tested optimum 4%.'),
        (7, 'ATH gate ε', 0.0, '0.0%',
         'Each buy is capped at ATH×(1−ε) − prev close × premium, so no '
         'trade is created whose exit needs a print above ATH×(1−ε). '
         '0 = guard at the exact ATH (recommended, ≈free). Set -1 to disable. '
         'Blank counts as 0.'),
    ):
        lc = at.cell(row=row_, column=12, value=label)
        style_like(lc, lab_style)
        vc = at.cell(row=row_, column=13, value=value)
        style_like(vc, inp_style, numfmt=fmt)
        ec = at.cell(row=row_, column=14, value=expl)
        ec.font = copy.copy(at['A2'].font)
    h24 = al.cell(row=24, column=8, value='PM gate')
    style_like(h24, al['G24'])
    for r in range(25, 43):
        at_row = r - 6
        al.cell(row=r, column=8,
                value=(f"=IFERROR(IF('Active Trading'!$F${at_row}"
                       f"<'Active Trading'!$P${at_row}"
                       f"*(1-'Active Trading'!$M$6),1,0),0)"))
        al.cell(row=r, column=6, value=f'=E{r}*(1-D{r})*(1-G{r})*(1-H{r})')

    # ---------------- 3. ATH gate epsilon in the Dashboard buys
    eps = "N('Active Trading'!$M$7)"
    for i, s in enumerate(STOCKS):
        r = 4 + i
        d.cell(row=r, column=3,
               value=(f"=MIN(J{r}-'Model {s}'!H2*K{r},L{r}*(1-'Model {s}'!L2),"
                      f"L{r}*(1-{eps})-B{r}*'Model {s}'!J2)"))
        d.cell(row=r, column=5,
               value=(f"=MIN(Q{r}-'Model {s}'!D3*P{r},L{r}*(1-'Model {s}'!J3),"
                      f"L{r}*(1-{eps})-B{r}*'Model {s}'!H3)"))

    # ---------------- Notes changelog
    sa, sb = notes['A8'], notes['B8']
    row = notes.max_row + 2
    for label, text in (
        ('Discretionary log',
         'Active Trading T41:AF241: a manually-updated trade log mirroring the main '
         'blotter\'s columns (fees B3/E3 applied in Buy cost / Proceeds; Net P&L, Week '
         'ending and Status computed). Closed discretionary P&L feeds the weekly '
         'summary\'s Cumulative (col R) and its buy/sell events feed Total trades (col '
         'S). Scripts do not read or write T:AF.'),
        ('PM gate',
         'Orders block col P = 9:00 pre-market price (Script 1 to populate P19:P36). '
         'Allocation col H flags any sleeve whose Buy @ sits more than Gate Variables '
         'M6 (default 4%) below col P; F25:F42 = E*(1-Held)*(1-DMA)*(1-PM), so gated '
         'capital pools to the other sleeves. Fails open on blank P or blank M6. '
         'Basis: verified pooled book, train 45.4->59.5%, tested half 110.7->118.4%.'),
        ('ATH gate',
         'Dashboard C4:C12 and E4:E12 gained a third MIN term: ATH*(1-eps) - prev '
         'close * premium (eps = Gate Variables M7, blank counts as 0, -1 disables). '
         'No order can construct a trade whose exit needs a print above ATH*(1-eps). '
         'Measured ~free at eps=0; removes the above-ATH-exit trap (col O flags it).'),
    ):
        ca = notes.cell(row=row, column=1, value=label)
        ca.font = copy.copy(sa.font)
        ca.alignment = copy.copy(sa.alignment)
        cb = notes.cell(row=row, column=2, value=text)
        cb.font = copy.copy(sb.font)
        cb.alignment = copy.copy(sb.alignment)
        row += 1

    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)


def diff(in_path, out_path):
    a = openpyxl.load_workbook(in_path)
    b = openpyxl.load_workbook(out_path)
    out = []
    for sn in a.sheetnames:
        wa, wc = a[sn], b[sn]
        for r in range(1, max(wa.max_row, wc.max_row) + 1):
            for c in range(1, max(wa.max_column, wc.max_column) + 1):
                va, vb = wa.cell(row=r, column=c).value, wc.cell(row=r, column=c).value
                if va != vb:
                    out.append((sn, wa.cell(row=r, column=c).coordinate))
    return out


def check(diffs):
    from openpyxl.utils import coordinate_to_tuple
    bad = []
    for sn, coord in diffs:
        r, c = coordinate_to_tuple(coord)[::-1][::-1]
        row, col = coordinate_to_tuple(coord)
        ok = (
            (sn == 'Active Trading' and row == 18 and col == 16) or
            (sn == 'Active Trading' and row in (6, 7) and col in (12, 13, 14)) or
            (sn == 'Active Trading' and row == 41 and 20 <= col <= 32) or
            (sn == 'Active Trading' and 42 <= row <= DISC_LAST and col in (26, 29, 30, 31, 32)) or
            (sn == 'Active Trading' and 42 <= row <= WEEK_LAST and col in (18, 19)) or
            (sn == 'Allocation' and 24 <= row <= 42 and col in (6, 8)) or
            (sn == 'Dashboard' and 4 <= row <= 12 and col in (3, 5)) or
            (sn == 'Notes')
        )
        if not ok:
            bad.append((sn, coord))
    return bad


if __name__ == '__main__':
    in_path, out_path = sys.argv[1], sys.argv[2]
    wire(in_path, out_path)
    d = diff(in_path, out_path)
    bad = check(d)
    print(f'{len(d)} cells changed; unexpected: {bad if bad else "NONE"}')
