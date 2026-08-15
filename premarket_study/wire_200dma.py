"""
Wire the 200dma gate into the live nine-stock workbook (bear-protection adoption,
15 Aug 2026), plus the two Active Trading display columns the user asked for.

Layout-preserving by construction -- nothing is inserted or moved, so the order
block (A18:J36), the IBKR_Orders range and the blotter keep their addresses and
the execution scripts read exactly what they read before. Four changes:

  1. THE GATE. Allocation G25:G42 (previously empty) gets a per-sleeve flag:
     1 if the stock's last close sits below its 200-day average, else 0
     (IFERROR-wrapped to fail OPEN -- a broken lookup means no gate, not a dead
     sleeve). F25:F42 becomes E*(1-D)*(1-G): a gated sleeve drops out of the
     free-weight renormalisation exactly like a HOLDING one, so its cash pools
     into the ungated names' bids -- the mechanism that made the gate free in
     the pooled 2024-26 measurement. Its Active Trading row then shows Tranche
     fund 0 and Qty 0. Exits, targets and stops are untouched.

  2. 200 DMA column: Active Trading N19:N36 (right of the Override column),
     the trailing 200-close average of the row's stock, straight off the Query
     close column, growing with the data (COUNT-anchored, expanding window
     until 200 closes exist).

  3. SALE %-BELOW-ATH column: Active Trading O19:O36, (ATH - sale)/ATH where
     sale is the row's live sale price (col G: the calculated target on CASH
     rows, the resting sell on HOLDING rows) and ATH is Dashboard column L.
     Negative means the sale price sits ABOVE the all-time high -- the ATH trap
     visible live.

  4. Notes sheet: a changelog section describing all of the above.

Verification: full-workbook cell diff against the input (only the intended cells
may change), plus a Python mirror computing every 200dma, gate flag, allocation
share and ATH from the Query data.
"""
import copy
import datetime
import sys

import openpyxl

STOCKS = ['TSM', 'VRT', 'VST', 'RKLB', 'MU', 'GM', 'VLO', 'CF', 'MRVL']
QCOL = {s: 5 + 4 * i for i, s in enumerate(STOCKS)}       # Query close column
QLET = {s: openpyxl.utils.get_column_letter(QCOL[s]) for s in STOCKS}


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    at = wb['Active Trading']
    al = wb['Allocation']
    q = wb['Query']
    notes = wb['Notes']

    # ---- preconditions
    for i, s in enumerate(STOCKS):
        assert q.cell(row=1, column=QCOL[s]).value == f'{s}_C', (s, q.cell(row=1, column=QCOL[s]).value)
        assert at.cell(row=19 + 2 * i, column=1).value == s
        assert at.cell(row=20 + 2 * i, column=1).value == s
    assert all(at.cell(row=r, column=c).value is None for r in range(18, 37) for c in (14, 15))
    assert all(al.cell(row=r, column=7).value is None for r in range(24, 43))
    assert at['M18'].value == 'Override'
    assert al['F25'].value == '=E25*(1-D25)'

    # ---- 2. Active Trading N: current 200 DMA per stock
    hdr_style = at['M18']
    for col, title in ((14, '200 DMA'), (15, 'Sale % below ATH')):
        c = at.cell(row=18, column=col, value=title)
        c.font = copy.copy(hdr_style.font)
        c.fill = copy.copy(hdr_style.fill)
        c.border = copy.copy(hdr_style.border)
        c.alignment = copy.copy(hdr_style.alignment)
    for i, s in enumerate(STOCKS):
        L = QLET[s]
        dma = (f'=AVERAGE(INDEX(Query!${L}:${L},MAX(2,COUNT(Query!${L}:${L})-198)):'
               f'INDEX(Query!${L}:${L},COUNT(Query!${L}:${L})+1))')
        for r in (19 + 2 * i, 20 + 2 * i):
            c = at.cell(row=r, column=14, value=dma)
            c.number_format = '0.00'
            # ---- 3. sale price as % below the all-time high
            ath = f'INDEX(Dashboard!$L$4:$L$12,MATCH($A{r},Dashboard!$A$4:$A$12,0))'
            o = at.cell(row=r, column=15, value=f'=IFERROR(({ath}-$G{r})/{ath},"")')
            o.number_format = '0.0%'

    # ---- 1. the gate: Allocation G flags + F formula
    g24 = al.cell(row=24, column=7, value='DMA gate')
    ref = al['F24']
    g24.font = copy.copy(ref.font)
    g24.fill = copy.copy(ref.fill)
    g24.border = copy.copy(ref.border)
    g24.alignment = copy.copy(ref.alignment)
    for r in range(25, 43):
        at_row = r - 6                                     # alloc 25 <-> AT 19
        al.cell(row=r, column=7,
                value=(f'=IFERROR(IF(INDEX(Dashboard!$B$4:$B$12,'
                       f'MATCH($A{r},Dashboard!$A$4:$A$12,0))'
                       f"<'Active Trading'!$N${at_row},1,0),0)"))
        al.cell(row=r, column=6, value=f'=E{r}*(1-D{r})*(1-G{r})')

    # ---- 4. Notes changelog
    style_a = notes['A8']
    style_b = notes['B8']
    row = notes.max_row + 2
    entries = [
        ('200dma gate',
         'A stock whose last close sits below the average of its last 200 closes places no new '
         'bids; exits, targets and stops run unchanged. Wiring: Active Trading N19:N36 computes '
         'each stock\'s 200-day average off the Query closes; Allocation G25:G42 flags a gated '
         'sleeve (fails open on any lookup error) and F25:F42 is now E*(1-Held)*(1-Gate), so a '
         'gated sleeve drops out of the cash renormalisation and its capital pools into the '
         'ungated names the same morning.'),
        (None,
         'A gated sleeve shows Tranche fund 0 and Qty 0 in the order block (layout unchanged; '
         'the order script must skip zero-quantity BUY rows). Basis: 2022 replay on verified '
         'fills -- unprotected book -21.9%, gated -2.6%, max drawdown 28.1% -> 17.4%; cost in '
         'the pooled 2024-26 sample ~0 (75.8% -> 75.9%/yr). Caveat: in a fast crash the gate '
         'concentrates capital in whichever names remain ungated (April 2025 episode -29.2% -> '
         '-32.1%); it is armour for the grinding bear, not the gap crash.'),
        ('New AT columns',
         '200 DMA (col N): the current 200-day average for the row\'s stock. Sale % below ATH '
         '(col O): (ATH - sale price)/ATH using Dashboard col L and the row\'s live sale price '
         '(calculated target on CASH rows, resting sell on HOLDING rows). A NEGATIVE value '
         'means the sale price sits above the all-time high -- the ATH-trap flag.'),
    ]
    for label, text in entries:
        if label:
            ca = notes.cell(row=row, column=1, value=label)
            ca.font = copy.copy(style_a.font)
            ca.alignment = copy.copy(style_a.alignment)
        cb = notes.cell(row=row, column=2, value=text)
        cb.font = copy.copy(style_b.font)
        cb.alignment = copy.copy(style_b.alignment)
        row += 1

    wb.save(out_path)
    return out_path


def diff(in_path, out_path):
    """Every cell that differs between input and output -- must be exactly the set
    of intended edits."""
    a = openpyxl.load_workbook(in_path)
    b = openpyxl.load_workbook(out_path)
    out = []
    for sn in a.sheetnames:
        wa, wc = a[sn], b[sn]
        for r in range(1, max(wa.max_row, wc.max_row) + 1):
            for c in range(1, max(wa.max_column, wc.max_column) + 1):
                va = wa.cell(row=r, column=c).value
                vb = wc.cell(row=r, column=c).value
                if va != vb:
                    out.append((sn, wa.cell(row=r, column=c).coordinate, str(va)[:40], str(vb)[:60]))
    return out


def mirror(in_path):
    """Compute today's 200dma / gate / allocation state from the Query data."""
    wb = openpyxl.load_workbook(in_path, data_only=False)
    q = wb['Query']
    last = max(r for r in range(2, q.max_row + 1) if q.cell(row=r, column=1).value is not None)
    print(f'Query data to row {last}')
    print(f'{"name":6s}{"last close":>11s}{"200dma":>10s}{"gate":>6s}{"ATH":>10s}')
    gates = {}
    for s in STOCKS:
        col = QCOL[s]
        closes = [q.cell(row=r, column=col).value for r in range(2, last + 1)]
        closes = [v for v in closes if v is not None]
        highs = [q.cell(row=r, column=col - 2).value for r in range(2, last + 1)]
        highs = [v for v in highs if v is not None]
        dma = sum(closes[-200:]) / len(closes[-200:])
        gates[s] = closes[-1] < dma
        print(f'{s:6s}{closes[-1]:>11.2f}{dma:>10.2f}{"GATED" if gates[s] else "-":>6s}'
              f'{max(highs):>10.2f}')
    n_free_effect = sum(1 for s in STOCKS if not gates[s])
    print(f'gated today: {sum(gates.values())}/9; ungated names split the gated cash '
          f'({n_free_effect} names carry the book)')
    return gates


if __name__ == '__main__':
    in_path, out_path = sys.argv[1], sys.argv[2]
    wire(in_path, out_path)
    d = diff(in_path, out_path)
    print(f'{len(d)} cells changed:')
    for row in d:
        print('  ', row)
    print()
    mirror(in_path)
