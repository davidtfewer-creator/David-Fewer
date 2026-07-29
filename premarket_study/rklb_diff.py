import json, datetime, re, openpyxl
from openpyxl.utils import get_column_letter
from engine import Params
from mirror_ladder import mirror
from multi_stock import params_for
import verify_all_ladder as V

# extract RKLB data from this workbook
q = V.srcdo['Query']; i = V.STOCKS.index('RKLB')
dts,O,H,L,C=[],[],[],[],[]
for row in q.iter_rows(min_row=2, values_only=True):
    d=V.to_date(row[0]); o=row[1+4*i]
    if d is None or not isinstance(o,(int,float)) or o<=0: continue
    dts.append(d);O.append(o);H.append(row[2+4*i]);L.append(row[3+4*i]);C.append(row[4+4*i])
p=params_for('RKLB', json.load(open('params_all.json')))
mir=mirror(dts,O,H,L,C,p)

# evaluate RKLB sheet keeping grid
do=V.srcdo['Model RKLB']; fo=V.srcf['Model RKLB']
cache={}
for row in do.iter_rows():
    for c in row:
        if c.value is not None and c.value!='' : cache[c.coordinate]=V.to_serial(c.value)
for r in range(2,7): cache[f'{V.CFG_VAL_COL}{r}']=fo[f'{V.CFG_VAL_COL}{r}'].value
grid={}
def cv(ref):
    ref=ref.replace('$',''); col=re.match(r'[A-Z]+',ref).group()
    return grid.get(ref) if col in V.LAD_COLS else cache.get(ref)
ns={'cv':cv,'IF':V.IF,'AND':V.AND,'OR':V.OR,'ISNUMBER':V.ISNUMBER,'MIN':min,'MAX':max,'None':None}
fc={}
for n in V.NAMES:
    for r in range(8,868):
        v=fo[f'{V.COL[n]}{r}'].value
        if isinstance(v,str) and v.startswith('='): fc[(n,r)]=V.translate(v)
        else: grid[f'{V.COL[n]}{r}']=0 if v is None else v
for r in range(8,868):
    for n in V.EVAL_ORDER:
        e=fc.get((n,r))
        if e is not None: grid[f'{V.COL[n]}{r}']=eval(e,ns)

# compare per row (i -> row i+8)
for idx in range(len(C)):
    r=idx+8
    ef=grid.get(f'{V.COL["FUND"]}{r}'); es=grid.get(f'{V.COL["SH"]}{r}')
    if ef is None: continue
    if abs(ef-mir['FUND'][idx])>0.01 or abs((es or 0)-mir['SH'][idx])>1e-6:
        print(f'first divergence row {r} (i={idx}) date {dts[idx]}')
        print(f'  eval FUND={ef:.2f} mir={mir["FUND"][idx]:.2f}  eval SH={es} mir SH={mir["SH"][idx]}')
        pr1 = grid.get(f'{V.COL["PR1"]}{r}'); pr2 = grid.get(f'{V.COL["PR2"]}{r}'); pr3 = grid.get(f'{V.COL["PR3"]}{r}')
        f1 = grid.get(f'{V.COL["FIL1"]}{r}'); f2 = grid.get(f'{V.COL["FIL2"]}{r}'); f3 = grid.get(f'{V.COL["FIL3"]}{r}')
        print(f'  pr1={pr1} pr2={pr2} pr3={pr3}  low={cache.get(f"D{r}")} open={cache.get(f"B{r}")}')
        print(f'  FIL1={f1} FIL2={f2} FIL3={f3}  fresh={grid.get(f"{V.COL[chr(70)+chr(82)+chr(69)+chr(83)+chr(72)]}{r}")}')
        break
