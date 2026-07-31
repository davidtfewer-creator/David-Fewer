"""
Build the feed host workbook by MODIFYING a known-good Excel file (the uploaded old_stocks
feed, which opens cleanly), rather than generating from scratch. Preserves the valid package
skeleton Excel accepts. Repurposes it into: Config / Tickers (named ranges) / Data (target) /
ReadMe. No API key written.
"""
import openpyxl
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName

SRC = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/0a153e7b-Hybrid10_Feed_old_stocks.xlsx'
OUT = '/home/user/David-Fewer/Hybrid10_Feed_PowerQuery.xlsx'
TICKERS = ['NVDA','TSM','TSLA','VRT','VST','AVGO','PLTR','RKLB','SOFI','SPOT']
HDR = Font(name='Arial', size=10, bold=True)
MONO = Font(name='Consolas', size=9)

wb = openpyxl.load_workbook(SRC)          # known-good skeleton
print('src sheets:', wb.sheetnames)

# ---- Data: keep the header row (Date, NVDA_O ... SPOT_C), clear the old data ----
data = wb['Data']
if data.max_row > 1:
    data.delete_rows(2, data.max_row - 1)
data['A3'] = ('Load the Hybrid10Feed query here (Close & Load To > Table > =Data!$A$1); '
              'it overwrites this header with the live feed.')

# ---- repurpose Sheet1 -> Config ----
cfg = wb['Sheet1']; cfg.title = 'Config'
cfg['A1'] = 'Setting'; cfg['B1'] = 'Value'
cfg['A2'] = 'ApiKey'                       # B2 left empty for the key
cfg['A3'] = 'StartDate'; cfg['B3'] = '2024-04-01'
cfg['A1'].font = HDR; cfg['B1'].font = HDR
cfg['D2'] = 'paste your Massive API key in B2 (kept out of the query code)'
cfg['D3'] = 'feed start date yyyy-mm-dd (matches the script)'
cfg.column_dimensions['A'].width = 12; cfg.column_dimensions['B'].width = 26
cfg.column_dimensions['D'].width = 58

# ---- Tickers ----
tk = wb.create_sheet('Tickers')
tk['A1'] = 'Ticker'; tk['A1'].font = HDR
for i, t in enumerate(TICKERS, start=2):
    tk.cell(i, 1, t)
tk.column_dimensions['A'].width = 12

# ---- ReadMe ----
rm = wb.create_sheet('ReadMe')
rm.column_dimensions['A'].width = 120
steps = [
    'Hybrid10 Feed - Power Query replica of the Google Apps Script',
    '',
    'SECURITY: the key is NOT in the query - paste it into Config!B2. Rotate the key you shared.',
    '',
    'SETUP:',
    '1. Config!B2: paste your Massive API key.  Config!B3: start date (default 2024-04-01).',
    '2. Data tab: Get Data > Launch Power Query Editor > New Query > Blank Query.',
    '3. Advanced Editor: delete all, paste the M code below, Done. Rename it Hybrid10Feed.',
    '4. Close & Load To... > Table > Existing worksheet > =Data!$A$1.',
    '5. First refresh: set api.massive.com credential to Anonymous (key rides in the URL).',
    '',
    'The ranges FeedConfig (Config!A1:B3) and FeedTickers (Tickers!A1:A11) are already named.',
    'Refresh later with Data > Refresh All. Date is emitted as text dd/MM/yyyy.',
    '',
    '==================  M CODE (paste into Advanced Editor)  ==================',
]
r = 1
for s in steps:
    if s:
        rm.cell(r, 1, s)
    r += 1
with open('/home/user/David-Fewer/premarket_study/HybridFeed.m') as f:
    for line in f.read().splitlines():
        if line != '':
            rm.cell(r, 1, line).font = MONO
        r += 1

# ---- named ranges (NOT matching any sheet name) ----
for nm in ('FeedConfig', 'FeedTickers'):
    if nm in wb.defined_names:
        del wb.defined_names[nm]
wb.defined_names.add(DefinedName('FeedConfig',  attr_text="Config!$A$1:$B$3"))
wb.defined_names.add(DefinedName('FeedTickers', attr_text=f"Tickers!$A$1:$A${1+len(TICKERS)}"))

wb.save(OUT)
print('saved', OUT, '| sheets:', wb.sheetnames, '| names:', list(wb.defined_names))
