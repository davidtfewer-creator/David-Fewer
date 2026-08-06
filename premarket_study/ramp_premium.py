"""
Time-decayed ("ramped") take-profit premium on NVDA.

Question
--------
The book buys on a limit and rests a GTC sell at bid + prev_close*premium. Capital sits
trapped whenever the stock approaches the target and turns away. Since a large part of the
return is compounding through repeated round trips, does it pay to ACCEPT LESS EARLY --
sell at a reduced premium on the fill day, a slightly larger one the next session, the full
premium from then on -- in exchange for recycling the capital sooner?

  premium 2%, ramp (0.5, 0.75, 1.0)  ->  1.0% today, 1.5% tomorrow, 2.0% thereafter

Implementation is an upward-only amendment of the resting sell, decided the evening before,
so nothing here needs information the model does not already have.

The control that matters
------------------------
A ramp lowers the average premium. So does simply cutting the premium. Any test of the ramp
that does not also run a FLAT premium at a comparable level cannot tell the two apart -- it
would credit the ramp for a benefit that is really just "sell cheaper". Every ramp below is
reported against flat-premium controls at the same mean realised premium.

Fill realism
------------
The ramp works mainly by converting holds into same-day round trips, and a same-day round
trip is exactly what daily bars cannot adjudicate: the sheet books one whenever the low
reached the bid and the high reached the target, in either order. Everything is therefore
scored three ways:
    sheet     - the workbook's own optimistic rule (upper bound, not a forecast)
    verified  - 5-minute bars prove the low preceded the high (the honest number)
    none      - no same-day exits at all (hard lower bound)
The ramp is only interesting if it survives 'verified'.

Data: 'TradingExcel_s1.xlsx' (587 sessions, 2024-04-01 -> 2026-08-03) and NVDA 5-minute bars
over the same span. Parameters are the workbook's own and are held FROZEN throughout -- per
the standing finding that structural changes survive out of sample and parameter searches do
not, nothing here is re-fitted.
"""
import collections
import datetime
import os
import sys

import numpy as np
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from engine import Params, run_model

UPLOADS = '/root/.claude/uploads/822d405e-f99b-5b59-9c6b-87e725054402'
BOOK = f'{UPLOADS}/69b83a92-TradingExcel_s1.xlsx'
FIVE_MIN = {'NVDA': f'{UPLOADS}/88d53e8a-NVDA_5min_Apr2024Aug2026.xlsx'}

RTH_START = datetime.time(9, 30)
RTH_END = datetime.time(16, 0)

# half-sample blade used throughout the study
SPLIT = datetime.date(2025, 5, 23)


# ---------------------------------------------------------------- data

def load_feed(stock, path=BOOK):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[f'Feed {stock}']
    d, O, H, L, C = [], [], [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dt_, o, h, l, c = row[0], row[1], row[2], row[3], row[4]
        if dt_ is None or not all(isinstance(v, (int, float)) for v in (o, h, l, c)):
            continue
        if o <= 0:
            continue
        d.append(dt_.date() if hasattr(dt_, 'date') else dt_)
        O.append(float(o)); H.append(float(h)); L.append(float(l)); C.append(float(c))
    wb.close()
    return d, O, H, L, C


def load_params(stock, path=BOOK, years=None):
    """Read the Model sheet's own parameter block, so the run is the workbook's run."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[f'Model {stock}']
    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    kv = {}
    for r in (rows[1], rows[2]):
        for i in range(len(r) - 1):
            if isinstance(r[i], str) and isinstance(r[i + 1], (int, float)):
                kv[r[i]] = r[i + 1]
    cached = dict(profit=rows[3][24], buys=rows[3][26], ann=rows[4][24])
    wb.close()
    cap = kv.get('Total capital', kv.get('Fund per tranche'))
    p = Params(lam=kv['λ'], phi_L=kv['φ_L'], psi=kv['ψ'], k=kv['k'],
               premium=kv['Premium'], peak_cap=kv['Peak cap'],
               comm=kv['Commission ($/sh)'], capital=cap,
               interest=kv['IBKR interest (pa)'],
               stop_days=int(kv['Stop-loss (calendar days)']), bayes_pct=kv['Bayes %'],
               ou_W=int(kv['OU lookback W']), ou_buf_k=kv['OU buffer k'],
               ou_prem=kv['OU premium'], ou_cap=kv['OU cap'],
               years=years if years is not None else 2.2)
    return p, cached


def build_index(stock):
    """date -> (lows, suffix-max-highs) over regular-hours 5-minute bars."""
    wb = openpyxl.load_workbook(FIVE_MIN[stock], read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        dt_, o, h, l = row[0], row[1], row[2], row[3]
        if dt_ is None or o is None or not isinstance(dt_, datetime.datetime):
            continue
        t = dt_.time()
        if t < RTH_START or t >= RTH_END:
            continue
        by[dt_.date()].append((dt_, h, l))
    wb.close()
    idx = {}
    for day, rs in by.items():
        rs.sort(key=lambda x: x[0])
        highs = np.array([r[1] for r in rs], dtype=float)
        lows = np.array([r[2] for r in rs], dtype=float)
        idx[day] = (lows, np.maximum.accumulate(highs[::-1])[::-1])
    return idx


def make_checker(idx, dates, O):
    """f(i, bid, target) -> was a same-day round trip genuinely achievable?"""
    def check(i, bid, target):
        ent = idx.get(dates[i])
        if ent is None:                      # no bars: fall back to the provable at-open case
            return bid >= O[i] - 1e-9
        lows, suffix = ent
        hit = lows <= bid + 1e-9
        if not hit.any():
            return False
        j = int(np.argmax(hit))
        return bool(suffix[j] >= target - 1e-9)
    return check


# ---------------------------------------------------------------- diagnostics

def trades_of(tr):
    """Reconstruct round trips from a tranche's flag arrays. One position at a time."""
    Z, AD, AA, AC = tr['Z'], tr['AD'], tr['AA'], tr['AC']
    out, entry = [], None
    for i in range(len(Z)):
        if Z[i] == 1:
            entry = i
        if AD[i] == 1 and entry is not None:
            out.append((entry, i, i - entry + 1, AA[i - 1] if AD[i] and entry != i else AA[i], AC[i]))
            entry = None
    return out


def stats(res, p, n_sessions):
    """Round trips, holding period and how much of the time capital was idle."""
    t1, t2 = res.frames['t1'], res.frames['t2']
    tr = trades_of(t1) + trades_of(t2)
    held = [t[2] for t in tr]
    idle = sum(1 for i in range(n_sessions) if t1['AE'][i] == 0) + \
           sum(1 for i in range(n_sessions) if t2['AE'][i] == 0)
    return dict(
        trips=len(tr),
        same_day=sum(1 for h in held if h == 1),
        med_hold=float(np.median(held)) if held else float('nan'),
        mean_hold=float(np.mean(held)) if held else float('nan'),
        cash_pct=100.0 * idle / (2 * n_sessions),
    )


def realised_premium(res):
    """Mean ramp multiplier actually achieved at exit -- the ramp's effective premium cut."""
    mults = []
    for tr in (res.frames['t1'], res.frames['t2']):
        for entry, ex, sess, _sh, _px in trades_of(tr):
            mults.append(sess)
    return mults


# ---------------------------------------------------------------- runner

def run(stock='NVDA', ramp=None, mode='verified', p=None, data=None, idx=None,
        prem_scale=1.0):
    d, O, H, L, C = data
    if prem_scale != 1.0:
        p = Params(**{**p.__dict__, 'premium': p.premium * prem_scale,
                      'ou_prem': p.ou_prem * prem_scale})
    if mode == 'sheet':
        sde = True
    elif mode == 'none':
        sde = False
    elif mode == 'at_open':
        sde = 'at_open'
    elif mode == 'verified':
        sde = make_checker(idx, d, O)
    else:
        raise ValueError(mode)
    return run_model(d, O, H, L, C, p, ou_sigma='level', same_day_exit=sde,
                     ramp=ramp, collect=True)
