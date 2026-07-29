"""
Impact of the stop-loss period (T2) per stock, read straight from the original 10-name
Hybrid9 workbook. First validate the engine reproduces each Model sheet's cached
Y4/AA4/AC4, then sweep stop_days per stock and report profit / ann / Sharpe / maxDD /
trades / stops. Finally a split-sample robustness check: is the in-sample-best stop
trustworthy out of sample, or does frozen 50 hold up?
"""
import openpyxl, datetime, copy
from engine import Params, run_model

F = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/602b5e6f-Hybrid9_Bayesian_OULIVE_PLTR_nopq.xlsx'
STOCKS = ['NVDA', 'TSM', 'TSLA', 'VRT', 'VST', 'AVGO', 'PLTR', 'RKLB', 'SOFI', 'SPOT']
GRID = [10, 15, 20, 25, 30, 40, 50, 60, 75, 90, 120, 150, 10_000]  # 10000 = effectively no stop


def load_book():
    wbf = openpyxl.load_workbook(F, data_only=False)
    wbv = openpyxl.load_workbook(F, data_only=True)
    q = wbv['Query']
    # price data per stock
    data = {}
    for i, s in enumerate(STOCKS):
        co = 2 + 4 * i
        dts, O, H, L, C = [], [], [], [], []
        for r in range(2, q.max_row + 1):
            d = q.cell(r, 1).value
            o = q.cell(r, co).value
            if not isinstance(o, (int, float)) or o <= 0:
                continue
            if isinstance(d, datetime.datetime):
                d = d.date()
            elif isinstance(d, (int, float)):
                d = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(d))
            dts.append(d); O.append(o)
            H.append(q.cell(r, co + 1).value); L.append(q.cell(r, co + 2).value)
            C.append(q.cell(r, co + 3).value)
        data[s] = (dts, O, H, L, C)
    # params + cached results per Model sheet
    params, cached = {}, {}
    for s in STOCKS:
        ws = wbv[f'Model {s}']
        g = lambda a: ws[a].value
        params[s] = Params(
            lam=g('B2'), phi_L=g('D2'), psi=g('F2'), k=g('H2'), premium=g('J2'),
            peak_cap=g('L2'), comm=g('N2'), capital=g('P2'), interest=g('R2'),
            stop_days=int(g('T2')), bayes_pct=g('V2'), ou_W=int(g('B3')),
            ou_buf_k=g('D3'), ou_prem=g('H3'), ou_cap=g('J3'), years=2.2)
        cached[s] = dict(profit=g('Y4'), ann=g('Y5'), buys=g('AA4'), stops=g('AC4'))
    return data, params, cached


def run(s, data, p):
    dts, O, H, L, C = data[s]
    return run_model(dts, O, H, L, C, p, collect=False)


def validate(data, params, cached):
    print('=== ENGINE VALIDATION (this workbook, current T2) ===')
    print(f'{"stock":6s}{"profit eng/cache":>30s}{"buys":>11s}{"stops":>9s}   ok')
    allok = True
    for s in STOCKS:
        r = run(s, data, params[s]); c = cached[s]
        ok = abs(r.profit - c['profit']) < 1.0 and r.total_buys == int(c['buys']) \
            and r.stop_loss_exits == int(c['stops'])
        allok &= ok
        print(f'{s:6s}  {r.profit:12,.0f} / {c["profit"]:12,.0f}  '
              f'{r.total_buys:4d}/{int(c["buys"]):<4d} {r.stop_loss_exits:3d}/{int(c["stops"]):<3d}   '
              f'{"OK" if ok else "MISMATCH"}')
    print('VALIDATION:', 'ALL PASS\n' if allok else 'FAILURES\n')
    return allok


def sweep(data, params):
    print('=== STOP-LOSS SWEEP (per stock; T2 varied, all else frozen) ===')
    print('    ann% / Sharpe / maxDD% / trades / stops   (current T2 marked *, best-Sharpe #)\n')
    best = {}
    for s in STOCKS:
        p0 = params[s]; cur = p0.stop_days
        rows = []
        for sd in GRID:
            p = copy.copy(p0); p.stop_days = sd
            r = run(s, data, p)
            rows.append((sd, r.annual_return, r.sharpe, r.max_drawdown,
                         r.total_buys, r.stop_loss_exits))
        bysharpe = max(rows, key=lambda x: x[2])
        byret = max(rows, key=lambda x: x[1])
        best[s] = dict(cur=cur, bysharpe=bysharpe, byret=byret, rows=rows)
        print(f'--- {s}  (current T2={cur}) ---')
        for sd, a, sh, dd, tr, st in rows:
            tag = ''
            if sd == cur: tag += ' *'
            if (sd, a, sh, dd, tr, st) == bysharpe: tag += ' #Sharpe'
            if (sd, a, sh, dd, tr, st) == byret: tag += ' $ret'
            lbl = 'none' if sd == 10_000 else str(sd)
            print(f'   T2={lbl:>5s}: {a*100:7.1f}%  Sh {sh:4.2f}  dd {dd*100:5.1f}%  '
                  f'buys {tr:3d}  stops {st:2d}{tag}')
        print()
    return best


def robustness(data, params):
    """Split-sample: pick best-Sharpe stop on the FIRST half, apply to the SECOND half,
    compare its 2nd-half return to frozen-50's 2nd-half return. Trust check on tuning T2."""
    print('=== SPLIT-SAMPLE ROBUSTNESS OF TUNING T2 ===')
    print('pick best-Sharpe T2 on train (1st half); score both halves; compare to frozen 50\n')
    print(f'{"stock":6s}{"train-best T2":>14s}{"test ret(tuned)":>17s}{"test ret(T2=50)":>17s}   winner')
    print('-' * 74)
    tuned_wins = 0
    for s in STOCKS:
        dts, O, H, L, C = data[s]; N = len(C); mid = N // 2
        p0 = params[s]

        def seg_ret_sharpe(sd, lo, hi):
            p = copy.copy(p0); p.stop_days = sd
            r = run_model(dts, O, H, L, C, p, collect=True)
            eq = r.frames['equity']
            ret = eq[hi] / eq[lo] - 1 if eq[lo] > 0 else -1
            seg = [eq[i] / eq[i - 1] - 1 for i in range(lo + 1, hi + 1) if eq[i - 1] > 0]
            import statistics, math
            sh = (statistics.mean(seg) / statistics.pstdev(seg) * math.sqrt(252)
                  if len(seg) > 2 and statistics.pstdev(seg) > 0 else 0)
            return ret, sh

        train_best = max(GRID, key=lambda sd: seg_ret_sharpe(sd, 0, mid)[1])
        test_tuned = seg_ret_sharpe(train_best, mid, N - 1)[0]
        test_froz = seg_ret_sharpe(50, mid, N - 1)[0]
        win = 'tuned' if test_tuned > test_froz else 'frozen50'
        tuned_wins += (test_tuned > test_froz)
        lbl = 'none' if train_best == 10_000 else str(train_best)
        print(f'{s:6s}{lbl:>14s}{test_tuned*100:>16.1f}%{test_froz*100:>16.1f}%   {win}')
    print('-' * 74)
    print(f'Tuning T2 on the past beats frozen-50 out of sample in {tuned_wins}/10 names.\n')


if __name__ == '__main__':
    data, params, cached = load_book()
    if validate(data, params, cached):
        best = sweep(data, params)
        robustness(data, params)
