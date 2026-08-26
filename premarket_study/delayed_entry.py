"""
Delayed-entry study — is there merit in trading N hours after the open, using the
opening price as the model signal? (user, 26 Aug 2026)

Decomposition — the proposal bundles two ideas, tested separately and together:

  TIMING   the deployed bid (which already caps at the day's open, so the open is
           already in the order) is placed at open+delta instead of resting from
           the open. An entry counts only if the bid is touched AFTER the cutoff;
           if the tape at the cutoff already sits below the bid the limit is
           marketable and fills at the tape (better price), with the take-profit
           target set off the actual fill. Resting SELLS are unchanged — a held
           position's exit order is live from the open as deployed.

  SIGNAL   the open replaces the previous close as the model's freshest price:
           OU forecast anchored on today's open (OUf = mu + ar*(O_t - mu)) and the
           Bayes fair nudged toward the open (fair' = fair + g*(O_t - fair),
           g in {0.25, 0.5, 1.0} reported separately — no train-picking here,
           every cell is shown on both halves). Implementable live because the
           orders are placed after the open in any case.

  DIAGNOSTIC  when do the deployed model's fills actually happen? First-touch time
           of every filled bid, bucketed, from the 5-minute bars.

Everything on verified fills (post-cutoff same-day checker for delayed variants),
deployed parameters, both halves of the standard split (2025-05-23). Engine hooks:
entry_low / entry_tape (behaviour-preserving, before/after snapshot EXACT MATCH
across all nine names).
"""
import datetime
import json
import os
import pickle

import numpy as np

from book_sim import load_all
from engine import Params, run_model
from fresh_opt import SPLIT, annualise
from fresh_opt_cands import aw_params

DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data_5min')
CACHE = os.path.join(DIR, 'full_bars_cache.pkl')
RTH_START = datetime.time(9, 30)
RTH_END = datetime.time(16, 0)
NAMES = ['TSM', 'VRT', 'VST', 'RKLB', 'MU', 'GM', 'VLO', 'CF', 'MRVL']
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'delayed_entry.json')


def full_bars(stock):
    """{date: (times_min_after_930, opens, highs, lows)} regular hours only."""
    cache = {}
    if os.path.exists(CACHE):
        with open(CACHE, 'rb') as f:
            cache = pickle.load(f)
    if stock in cache:
        return cache[stock]
    import openpyxl
    path = None
    for fn in sorted(os.listdir(DIR)):
        if fn.upper().startswith(stock.upper() + '_') and fn.endswith('.xlsx'):
            path = os.path.join(DIR, fn)
            break
    if path is None:
        raise FileNotFoundError(stock)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = {}
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        dt, o, h, l = row[0], row[1], row[2], row[3]
        if not isinstance(dt, datetime.datetime) or o is None:
            continue
        t = dt.time()
        if t < RTH_START or t >= RTH_END:
            continue
        m = (t.hour - 9) * 60 + t.minute - 30
        by.setdefault(dt.date(), []).append((m, float(o), float(h), float(l)))
    wb.close()
    idx = {}
    for d, bars in by.items():
        bars.sort()
        idx[d] = tuple(np.array(x) for x in zip(*bars))
    cache[stock] = idx
    with open(CACHE, 'wb') as f:
        pickle.dump(cache, f)
    return idx


def postT(stock, dts, delay_min):
    """entry_low[], entry_tape[], and a post-cutoff verified same-day checker.
    Days without post-cutoff bar coverage forbid entry (inf) — counted."""
    idx = full_bars(stock)
    N = len(dts)
    entry_low = [None] * N
    tape = [None] * N
    perday = {}
    missing = 0
    for i, d in enumerate(dts):
        ent = idx.get(d)
        if ent is None:
            entry_low[i] = float('inf')
            missing += 1
            continue
        tm, o, h, l = ent
        k = int(np.searchsorted(tm, delay_min))
        if k >= len(tm):
            entry_low[i] = float('inf')     # session over before the cutoff (half day)
            missing += 1
            continue
        lows = l[k:]
        highs = h[k:]
        entry_low[i] = float(lows.min())
        tape[i] = float(o[k])
        perday[i] = (lows, np.maximum.accumulate(highs[::-1])[::-1])

    def check(i, bid, target):
        ent = perday.get(i)
        if ent is None:
            return False
        lows, suffix = ent
        hit = lows <= bid + 1e-9
        if not hit.any():
            return False
        j = int(np.argmax(hit))
        return bool(suffix[j] >= target - 1e-9)

    return entry_low, tape, check, missing


def fill_time_diag(stock, dts, frames):
    """First-touch time of every filled bid, bucketed."""
    idx = full_bars(stock)
    buckets = {'<10:30': 0, '10:30-11:30': 0, '11:30-13:00': 0, '>=13:00': 0}
    for tkey, bkey in (('t1', 'X'), ('t2', 'AM')):
        Z = frames[tkey]['Z']
        bids = frames[bkey]
        for i, z in enumerate(Z):
            if z != 1:
                continue
            ent = idx.get(dts[i])
            if ent is None:
                continue
            tm, o, h, l = ent
            hit = l <= bids[i] + 1e-9
            if not hit.any():
                continue
            m = tm[int(np.argmax(hit))]
            if m < 60:
                buckets['<10:30'] += 1
            elif m < 120:
                buckets['10:30-11:30'] += 1
            elif m < 210:
                buckets['11:30-13:00'] += 1
            else:
                buckets['>=13:00'] += 1
    return buckets


def score(res, dts, cut):
    eq = res.frames['equity']
    N = len(dts)
    full = annualise(eq[N - 1] / eq[0] - 1, dts, 0, N - 1)
    train = annualise(eq[cut] / eq[0] - 1, dts, 0, cut)
    test = annualise(eq[N - 1] / eq[cut] - 1, dts, cut, N - 1)
    return full, train, test


def main():
    t0 = Params(capital=1_000_000, comm=0.005, interest=0.0314, stop_days=50,
                bayes_pct=0.5, years=2.2, ou_W=80)
    mrvl = aw_params(json.load(open(
        os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     'fresh_opt_cands.json')))['MRVL']['reference']['vec'], t0)
    data, _, _ = load_all(names=NAMES, params_override={'MRVL': mrvl})

    results = {}
    agg = {}
    for s in NAMES:
        d = data[s]
        dts, O, H, L, C, p, chk = d['dts'], d['O'], d['H'], d['L'], d['C'], d['p'], d['chk']
        N = len(dts)
        cut = next(i for i, x in enumerate(dts) if x >= SPLIT)
        rows = {}

        base = run_model(dts, O, H, L, C, p, ou_sigma='resid', same_day_exit=chk,
                         collect=True)
        rows['baseline'] = score(base, dts, cut) + (base.total_buys, base.stop_loss_exits)
        diag = fill_time_diag(s, dts, base.frames)

        # -------- TIMING: delay-only, 1h / 2h / 3h after the open
        for lab, dm in [('delay 1h', 60), ('delay 2h', 120), ('delay 3h', 180)]:
            el, tp, pchk, miss = postT(s, dts, dm)
            r = run_model(dts, O, H, L, C, p, ou_sigma='resid', same_day_exit=pchk,
                          collect=True, entry_low=el, entry_tape=tp)
            rows[lab] = score(r, dts, cut) + (r.total_buys, r.stop_loss_exits)
            if lab == 'delay 2h':
                rows['_miss2h'] = miss

        # -------- SIGNAL: open as the model's freshest price (no delay)
        ou_anchor = [O[i] for i in range(N)]
        for g in (0.25, 0.5, 1.0):
            r = run_model(dts, O, H, L, C, p, ou_sigma='resid', same_day_exit=chk,
                          collect=True, ou_anchor=ou_anchor,
                          bayes_signal=O, bayes_gain=g)
            rows[f'open-signal g={g}'] = score(r, dts, cut) + (r.total_buys,
                                                               r.stop_loss_exits)

        # -------- BOTH: 2h delay + open signal (g=0.5 representative)
        el, tp, pchk, _ = postT(s, dts, 120)
        r = run_model(dts, O, H, L, C, p, ou_sigma='resid', same_day_exit=pchk,
                      collect=True, entry_low=el, entry_tape=tp,
                      ou_anchor=ou_anchor, bayes_signal=O, bayes_gain=0.5)
        rows['delay 2h + signal'] = score(r, dts, cut) + (r.total_buys,
                                                          r.stop_loss_exits)

        results[s] = dict(rows=rows, fill_times=diag)
        print(f"\n===== {s} ({N} sessions, split at row {cut}) =====")
        tot = sum(diag.values())
        print('  fill first-touch times: ' + ', '.join(
            f'{k} {v} ({v/tot*100:.0f}%)' for k, v in diag.items()) +
            f'  [n={tot}]')
        print(f"  {'variant':22s}{'full':>8s}{'train':>8s}{'test':>8s}"
              f"{'buys':>6s}{'stops':>6s}")
        for lab, v in rows.items():
            if lab.startswith('_'):
                continue
            f_, tr, te, nb, st = v
            print(f"  {lab:22s}{f_*100:>7.1f}%{tr*100:>7.1f}%{te*100:>7.1f}%"
                  f"{nb:>6d}{st:>6d}")
        for lab, v in rows.items():
            if lab.startswith('_'):
                continue
            agg.setdefault(lab, []).append(v)

    print('\n===== BOOK AVERAGE (equal-weight mean of per-name annualised) =====')
    print(f"  {'variant':22s}{'full':>8s}{'train':>8s}{'test':>8s}"
          f"{'names test-better':>19s}")
    b = agg['baseline']
    for lab, vs in agg.items():
        f_ = np.mean([v[0] for v in vs]); tr = np.mean([v[1] for v in vs])
        te = np.mean([v[2] for v in vs])
        nb = sum(1 for v, vb in zip(vs, b) if v[2] > vb[2])
        print(f"  {lab:22s}{f_*100:>7.1f}%{tr*100:>7.1f}%{te*100:>7.1f}%"
              f"{nb:>12d}/9")

    with open(OUT, 'w') as f:
        json.dump({s: dict(rows={k: list(v) if isinstance(v, tuple) else v
                                 for k, v in r['rows'].items()},
                           fill_times=r['fill_times'])
                   for s, r in results.items()}, f, indent=1)
    print(f'\nsaved {OUT}')


if __name__ == '__main__':
    main()
