"""
Can pre-market data predict which names will FILL today?

If so, idle cash can be steered toward likely fillers -- raising capital utilisation with no
leverage at all, since total capital is unchanged.

A fill occurs exactly when the session low reaches the UNCAPPED bid
    b_t = min( fair_t - k*sigma_t , G_{t-1}(1-c) )
(if the stock gaps below b_t the open cap binds and it fills at the open; otherwise the low must
reach b_t). b_t is known before the bell, so conditioning on pre-market data is legitimate.

Signals tested, all observable at 09:29:
    d_vwap = (PM_VWAP - b) / sigma      how far the pre-market VWAP sits above the bid
    d_low  = (PM_Low  - b) / sigma      how far the pre-market LOW sits above the bid
    touched = PM_Low <= b               the pre-market already traded at the bid
"""
import openpyxl, datetime, statistics
from stop_sweep import load_book
from engine import run_model

PM = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/39840daf-PreMarket_04000900ET_20240401_to_20260728.xlsx'
data, params, cached = load_book()
STOCKS = list(data)


def load_pm():
    wb = openpyxl.load_workbook(PM, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    col = {h: i for i, h in enumerate(hdr) if h}
    out = {}
    for row in it:
        d = row[0]
        if not isinstance(d, datetime.datetime):
            continue
        rec = {}
        for s in STOCKS:
            hi, lo, vw = col.get(f'{s}_PM_High'), col.get(f'{s}_PM_Low'), col.get(f'{s}_PM_VWAP')
            if hi is None: continue
            H, L, V = row[hi], row[lo], row[vw]
            if isinstance(L, (int, float)) and isinstance(V, (int, float)):
                rec[s] = (H, L, V)
        out[d.date()] = rec
    wb.close()
    return out


def uncapped_bids(s):
    """Return per-day (uncapped bayes bid, uncapped ou bid, sigma, actual fill flags)."""
    dts, O, H, L, C = data[s]
    p = params[s]
    fr = run_model(dts, O, H, L, C, p, collect=True).frames
    Lvl, Slp, W, G = fr['Lvl'], fr['Slp'], fr['W'], fr['G']
    OUf, OUsig = fr['OUf'], fr['OUsig']
    rows = []
    for i in range(1, len(C)):
        bb = min(Lvl[i-1] + Slp[i-1] - p.k*W[i-1], G[i-1]*(1-p.peak_cap))
        ob = None
        if OUf[i] is not None and OUsig[i] is not None:
            ob = min(OUf[i] - p.ou_buf_k*OUsig[i], G[i-1]*(1-p.ou_cap))
        rows.append((dts[i], bb, ob, W[i-1], L[i] <= bb,
                     (ob is not None and L[i] <= ob)))
    return rows


if __name__ == '__main__':
    pm = load_pm()
    print('=== DOES PRE-MARKET PREDICT A FILL? (Bayes sleeve, all ten names pooled) ===\n')
    buckets = {}          # bucket -> [fills, total]
    touched = [0, 0]; nottouched = [0, 0]
    base = [0, 0]
    for s in STOCKS:
        for (d, bb, ob, sig, fillb, fillo) in uncapped_bids(s):
            rec = pm.get(d, {}).get(s)
            if rec is None or sig <= 0: continue
            _, plow, pvwap = rec
            base[0] += fillb; base[1] += 1
            z = (pvwap - bb)/sig                      # sigmas the PM VWAP sits above the bid
            b = min(int(z//1), 5) if z >= 0 else -1   # bucket
            f, t = buckets.get(b, (0, 0))
            buckets[b] = (f + fillb, t + 1)
            if plow <= bb: touched[0] += fillb; touched[1] += 1
            else:          nottouched[0] += fillb; nottouched[1] += 1
    print(f'base rate: {base[0]}/{base[1]} = {base[0]/base[1]*100:.0f}% of name-days fill\n')
    print(f'{"PM VWAP above bid":>20s}{"fills":>9s}{"days":>8s}{"fill rate":>11s}')
    print('-'*48)
    for b in sorted(buckets):
        f, t = buckets[b]
        lbl = 'below bid' if b < 0 else f'{b}-{b+1} sigma'
        print(f'{lbl:>20s}{f:>9d}{t:>8d}{f/t*100:>10.0f}%')
    print()
    print(f'PM low ALREADY at/below bid : {touched[0]}/{touched[1]} = '
          f'{touched[0]/touched[1]*100:.0f}% fill')
    print(f'PM low above bid            : {nottouched[0]}/{nottouched[1]} = '
          f'{nottouched[0]/nottouched[1]*100:.0f}% fill')
