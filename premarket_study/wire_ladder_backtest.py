"""
Wire the laddered (Option C) Bayes backtest into a Model sheet as LIVE, self-recalculating
formulas. Adds a block of laddered-state columns (transcribed cell-for-cell from the
mirror-verified recursion) and repoints the summary cells (profit / return / buys / stops /
interest) and the Bayes equity column to the laddered columns. OU + Kalman signal untouched.

Column block starts at index 60 (BH). Config cells hold the ladder depths/weights (editable).
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

SRC = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/2323ed8f-TradingExcel_s1_laddering.xlsx'
OUT = '/home/user/David-Fewer/TradingExcel_s1_laddering_OptionC_backtest.xlsx'
R0, R1 = 8, 867                     # data block rows (init row 8, extends to 867)

# laddered-state columns (name -> absolute column index), contiguous from 60
NAMES = ['FRESH','F0','PR1','PR2','PR3','B1','B2','B3','FIL1','FIL2','FIL3','DSH',
         'SHMID','FUNDMID','ANCM','TGT','STOP','EXIT','SALE','SH','FUND','F1','F2','F3',
         'ANC','BD','NB','INT','EQ','SH1','SH2','SH3']
COL = {n: get_column_letter(60 + i) for i, n in enumerate(NAMES)}
# readable headers for the visible ladder columns (others get their code name)
HEADERS = {'PR1':'Rung1 bid','PR2':'Rung2 bid','PR3':'Rung3 bid','SH1':'Rung1 shares',
           'SH2':'Rung2 shares','SH3':'Rung3 shares','SH':'Bayes shares','FUND':'Bayes cash',
           'EQ':'Bayes equity','TGT':'Target','NB':'Rungs filled','SALE':'Sale px',
           'B1':'Rung1 $','B2':'Rung2 $','B3':'Rung3 $'}
CFG = get_column_letter(60 + len(NAMES) + 1)          # config label column
CFV = get_column_letter(60 + len(NAMES) + 2)          # config value column
c2, c3 = f'${CFV}$2', f'${CFV}$3'                      # m2, m3
w1, w2, w3 = f'${CFV}$4', f'${CFV}$5', f'${CFV}$6'     # weights


def wire(ws):
    L = COL
    # ---- config block ----
    ws[f'{CFG}1'] = 'LADDER C'
    for row, (lab, val) in {2: ('m2 (xk)', 1.3), 3: ('m3 (xk)', 1.7),
                            4: ('w1', 0.80), 5: ('w2', 0.15), 6: ('w3', 0.05)}.items():
        ws[f'{CFG}{row}'] = lab
        cell = ws[f'{CFV}{row}']; cell.value = val
        cell.fill = PatternFill('solid', fgColor='FFD9E1F2')
    # ---- headers ----
    for n in NAMES:
        ws[f'{L[n]}7'] = HEADERS.get(n, n)
        ws[f'{L[n]}7'].font = Font(name='Arial', size=8, bold=True)
    ws[f'{L["PR1"]}6'] = 'LADDER — Bayes 3 rungs (bid, $ budget, shares per rung)'
    ws[f'{L["PR1"]}6'].font = Font(name='Arial', size=9, bold=True, color='FF0B1F3A')
    # ---- init row 8 ----
    z = {n: 0 for n in NAMES}
    ws[f'{L["FUND"]}8'] = '=$P$2*$V$2'
    ws[f'{L["EQ"]}8'] = f'={L["FUND"]}8'
    for n in ['SH','F1','F2','F3','B1','B2','B3','ANC','NB','INT','STOP','EXIT','FRESH',
              'SH1','SH2','SH3']:
        ws[f'{L[n]}8'] = 0
    ws[f'{L["BD"]}8'] = ''
    ws[f'{L["SALE"]}8'] = ''
    ws[f'{L["TGT"]}8'] = 0

    def f(name, r, expr):
        ws[f'{L[name]}{r}'] = expr

    for r in range(9, R1 + 1):
        p = r - 1
        B, D, Cc, E, K, W, G, A, AN = (f'B{r}', f'D{r}', f'C{r}', f'E{r}', f'K{r}',
                                       f'W{p}', f'G{p}', f'A{r}', f'AN{r}')
        blank = f'B{r}=""'
        f('FRESH', r, f'=IF({blank},"",IF({L["SH"]}{p}<=0,1,0))')
        # interest exactly as the engine: fund + fund*rate*days/365 (NOT fund*(1+...), which
        # differs by a float ULP and flips razor-edge fills over the full history)
        f('F0',    r, f'=IF({blank},"",{L["FUND"]}{p}+{L["FUND"]}{p}*$R$2*{AN}/365)')
        f('PR1',   r, f'=IF({blank},"",MIN({K}-$H$2*{W},{B},{G}*(1-$L$2)))')
        f('PR2',   r, f'=IF({blank},"",MIN({K}-{c2}*$H$2*{W},{B},{G}*(1-$L$2)))')
        f('PR3',   r, f'=IF({blank},"",MIN({K}-{c3}*$H$2*{W},{B},{G}*(1-$L$2)))')
        f('B1', r, f'=IF({blank},{L["B1"]}{p},IF({L["FRESH"]}{r}=1,{L["F0"]}{r}*{w1},{L["B1"]}{p}))')
        f('B2', r, f'=IF({blank},{L["B2"]}{p},IF({L["FRESH"]}{r}=1,{L["F0"]}{r}*{w2},{L["B2"]}{p}))')
        f('B3', r, f'=IF({blank},{L["B3"]}{p},IF({L["FRESH"]}{r}=1,{L["F0"]}{r}*{w3},{L["B3"]}{p}))')
        # fund-availability guard: a rung fills only if enough cash remains (sequential depletion)
        av1 = f'{L["F0"]}{r}'
        av2 = f'({L["F0"]}{r}-{L["FIL1"]}{r}*{L["B1"]}{r})'
        av3 = f'({L["F0"]}{r}-{L["FIL1"]}{r}*{L["B1"]}{r}-{L["FIL2"]}{r}*{L["B2"]}{r})'
        f('FIL1', r, f'=IF({blank},0,IF(AND(IF({L["FRESH"]}{r}=1,0,{L["F1"]}{p})=0,{D}<={L["PR1"]}{r},{av1}>={L["B1"]}{r}-0.000000001),1,0))')
        f('FIL2', r, f'=IF({blank},0,IF(AND(IF({L["FRESH"]}{r}=1,0,{L["F2"]}{p})=0,{D}<={L["PR2"]}{r},{av2}>={L["B2"]}{r}-0.000000001),1,0))')
        f('FIL3', r, f'=IF({blank},0,IF(AND(IF({L["FRESH"]}{r}=1,0,{L["F3"]}{p})=0,{D}<={L["PR3"]}{r},{av3}>={L["B3"]}{r}-0.000000001),1,0))')
        f('DSH', r, (f'=IF({blank},0,{L["FIL1"]}{r}*{L["B1"]}{r}/({L["PR1"]}{r}+$N$2)'
                     f'+{L["FIL2"]}{r}*{L["B2"]}{r}/({L["PR2"]}{r}+$N$2)'
                     f'+{L["FIL3"]}{r}*{L["B3"]}{r}/({L["PR3"]}{r}+$N$2))'))
        f('SHMID', r, (f'=IF({blank},{L["SH"]}{p},IF({L["FRESH"]}{r}=1,0,{L["SH"]}{p})'
                       f'+{L["FIL1"]}{r}*{L["B1"]}{r}/({L["PR1"]}{r}+$N$2)'
                       f'+{L["FIL2"]}{r}*{L["B2"]}{r}/({L["PR2"]}{r}+$N$2)'
                       f'+{L["FIL3"]}{r}*{L["B3"]}{r}/({L["PR3"]}{r}+$N$2))'))
        # sequential subtraction (fund -= bj per rung) to match the engine's float path exactly
        f('FUNDMID', r, (f'=IF({blank},{L["FUND"]}{p},{L["F0"]}{r}-{L["FIL1"]}{r}*{L["B1"]}{r}'
                         f'-{L["FIL2"]}{r}*{L["B2"]}{r}-{L["FIL3"]}{r}*{L["B3"]}{r})'))
        f('ANCM', r, (f'=IF({blank},{L["ANC"]}{p},MAX(IF({L["FRESH"]}{r}=1,0,{L["ANC"]}{p}),'
                      f'{L["FIL1"]}{r}*{L["PR1"]}{r},{L["FIL2"]}{r}*{L["PR2"]}{r},{L["FIL3"]}{r}*{L["PR3"]}{r}))'))
        f('TGT', r, f'=IF({blank},"",{L["ANCM"]}{r}+E{p}*$J$2)')
        f('BD', r, (f'=IF({blank},{L["BD"]}{p},IF({L["FRESH"]}{r}=1,IF({L["DSH"]}{r}>0,{A},""),'
                    f'IF({L["SH"]}{p}>0,{L["BD"]}{p},"")))'))
        f('STOP', r, (f'=IF({blank},0,IF(AND({L["SHMID"]}{r}>0,'
                      f'IF(ISNUMBER({L["BD"]}{r}),{A}-{L["BD"]}{r},-1)>=$T$2),1,0))'))
        f('EXIT', r, f'=IF({blank},0,IF(AND({L["SHMID"]}{r}>0,OR({Cc}>={L["TGT"]}{r},{L["STOP"]}{r}=1)),1,0))')
        f('SALE', r, (f'=IF({blank},"",IF({L["EXIT"]}{r}=1,'
                      f'IF(AND({L["STOP"]}{r}=1,{Cc}<{L["TGT"]}{r}),{B},{L["TGT"]}{r}),""))'))
        f('SH', r, f'=IF({blank},{L["SH"]}{p},IF({L["EXIT"]}{r}=1,0,{L["SHMID"]}{r}))')
        f('FUND', r, (f'=IF({blank},{L["FUND"]}{p},IF({L["EXIT"]}{r}=1,'
                      f'{L["FUNDMID"]}{r}+{L["SHMID"]}{r}*({L["SALE"]}{r}-$N$2),{L["FUNDMID"]}{r}))'))
        f('F1', r, f'=IF({blank},{L["F1"]}{p},IF({L["EXIT"]}{r}=1,0,MAX(IF({L["FRESH"]}{r}=1,0,{L["F1"]}{p}),{L["FIL1"]}{r})))')
        f('F2', r, f'=IF({blank},{L["F2"]}{p},IF({L["EXIT"]}{r}=1,0,MAX(IF({L["FRESH"]}{r}=1,0,{L["F2"]}{p}),{L["FIL2"]}{r})))')
        f('F3', r, f'=IF({blank},{L["F3"]}{p},IF({L["EXIT"]}{r}=1,0,MAX(IF({L["FRESH"]}{r}=1,0,{L["F3"]}{p}),{L["FIL3"]}{r})))')
        f('ANC', r, f'=IF({blank},{L["ANC"]}{p},IF({L["EXIT"]}{r}=1,0,{L["ANCM"]}{r}))')
        f('NB', r, f'=IF({blank},0,{L["FIL1"]}{r}+{L["FIL2"]}{r}+{L["FIL3"]}{r})')
        f('INT', r, f'=IF({blank},0,{L["FUND"]}{p}*$R$2*{AN}/365)')
        f('EQ', r, f'=IF({blank},{L["FUND"]}{r},{L["FUND"]}{r}+{L["SH"]}{r}*{E})')
        # per-rung shares (visible ladder detail): shares bought at each rung today
        f('SH1', r, f'=IF({blank},0,{L["FIL1"]}{r}*{L["B1"]}{r}/({L["PR1"]}{r}+$N$2))')
        f('SH2', r, f'=IF({blank},0,{L["FIL2"]}{r}*{L["B2"]}{r}/({L["PR2"]}{r}+$N$2))')
        f('SH3', r, f'=IF({blank},0,{L["FIL3"]}{r}*{L["B3"]}{r}/({L["PR3"]}{r}+$N$2))')

    # ---- repoint summary cells to laddered columns ----
    Ff, Nb, In, Eq, Ex, Sa, Tg = (L['FUND'], L['NB'], L['INT'], L['EQ'], L['EXIT'], L['SALE'], L['TGT'])
    # terminal profit/return use marked-to-market EQUITY at the LAST DATA ROW (not fund on the
    # blank terminal row): LOOKUP grabs the last row with a real Open. Bayes equity = EQ column,
    # OU equity = BD column (both are cash + shares*close on data rows).
    bayes_eq = f'LOOKUP(2,1/($B$8:$B$867<>""),{Eq}$8:{Eq}$867)'
    ou_eq = 'LOOKUP(2,1/($B$8:$B$867<>""),$BD$8:$BD$867)'
    ws['Y4'] = f'={bayes_eq}+{ou_eq}-$P$2'
    ws['Y5'] = f'=IFERROR((({bayes_eq}+{ou_eq})/$P$2)^(1/2.2)-1,"")'
    ws['AA4'] = f'=SUM({Nb}8:{Nb}867)+SUM(AG8:AG867)'
    ws['AA5'] = f'=SUM({In}8:{In}867)+SUM(AS8:AS867)'
    ws['AC4'] = (f'=SUMPRODUCT(({Ex}8:{Ex}867=1)*({Sa}8:{Sa}867<{Tg}8:{Tg}867))'
                 '+SUMPRODUCT((AK8:AK867=1)*(AJ8:AJ867<AI8:AI867))')
    # ---- repoint the FAMILIAR Bayes columns to the laddered values (so the main sheet area
    # shows the ladder, not the orphaned single-bid formulas) ----
    for r in range(R0, R1 + 1):
        ws[f'BC{r}'] = f'={Eq}{r}'                      # T1 equity chart -> laddered equity
        ws[f'X{r}'] = f'={L["PR1"]}{r}'                 # Buy price -> rung1 bid
        ws[f'Y{r}'] = f'={L["FUND"]}{r}'                # Fund -> laddered cash
        ws[f'Z{r}'] = f'={L["NB"]}{r}'                  # Buy fl -> rungs filled today
        ws[f'AA{r}'] = f'={L["SH"]}{r}'                 # Shares -> laddered total shares
        ws[f'AB{r}'] = f'={L["TGT"]}{r}'                # Target -> laddered target
        ws[f'AC{r}'] = f'={L["SALE"]}{r}'               # Act.sale -> laddered sale
        ws[f'AD{r}'] = f'={L["EXIT"]}{r}'               # Sell fl -> laddered exit
        ws[f'AE{r}'] = f'=IF({L["SH"]}{r}>0,1,0)'       # Hold fl -> holding laddered position


if __name__ == '__main__':
    sheets = sys.argv[1:] or ['Model NVDA']
    wb = openpyxl.load_workbook(SRC, data_only=False)
    for s in sheets:
        wire(wb[s]); print('wired', s)
    wb.calculation.calcMode = 'auto'; wb.calculation.fullCalcOnLoad = True
    wb.save(OUT); print('saved', OUT)
