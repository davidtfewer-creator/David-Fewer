"""
The 2022 bear replay: the deployed nine-name book pushed through Aug 2021 - Jun 2023.

The one question the 2024-26 sample cannot answer: what does this book do when the
trend flattens and reverses? The user exported 5-minute bars for all nine names
(Aug 2021 - Jun 2023, the earliest available), so the replay runs on the same
verified-fill standard as everything else -- no floor/bracket compromises.

Everything is out of sample in the harshest direction: every parameter was fitted
on 2024-26 data, then driven backwards through a year in which these names fell
40-80%.

Outputs:
  1. Per name: buy-and-hold vs the model over calendar 2022 (return, max
     drawdown, time stops) -- the unprotected damage.
  2. Book level (pooled, equal weights): the protection matrix over 2022 and over
     the full span --
       nothing | 200dma gate | breaker 15%/7.5% half-size | gate+breaker
       | gate+breaker+price stop -25%
     The gate uses each name's own trailing-close average (200 days, expanding
     from a 100-day minimum so it is live by Jan 2022), strictly ex ante.

Conventions: deployed/reference parameters untouched; residual OU sigma; verified
same-day fills from the bear-period 5-minute index; interest on idle cash kept at
the model's 3.14%/yr (2022 rates rose 0->4.5%, so pooled cash income is, if
anything, understated late in the year).
"""
import collections
import datetime
import json
import os
import pickle
import sys

import numpy as np

from engine import Params, run_model
from fresh_opt import annualise
from fresh_opt_cands import ref_params, aw_params
from live5_load import load as load_book, STOCKS as BOOK
import book_sim

DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data_bear')
CACHE = os.path.join(DIR, 'bear_cache.pkl')
RTH_START = datetime.time(9, 30)
RTH_END = datetime.time(16, 0)
NAMES = BOOK + ['GM', 'VLO', 'CF', 'MRVL']
Y22a, Y22b = datetime.date(2022, 1, 3), datetime.date(2022, 12, 30)


def _load(stock):
    """(daily bars, minute index) for a bear-period file; cached together."""
    cache = {}
    if os.path.exists(CACHE):
        with open(CACHE, 'rb') as f:
            cache = pickle.load(f)
    if stock in cache:
        return cache[stock]
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(DIR, f'{stock}_5min.xlsx'),
                                read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        dt, o, h, l, c = row[0], row[1], row[2], row[3], row[4]
        if isinstance(dt, str):
            try:
                dt = datetime.datetime.fromisoformat(dt)
            except ValueError:
                continue
        if not isinstance(dt, datetime.datetime) or o is None:
            continue
        t = dt.time()
        if t < RTH_START or t >= RTH_END:
            continue
        by[dt.date()].append((dt, float(o), float(h), float(l), float(c)))
    wb.close()
    dts, O, H, L, C = [], [], [], [], []
    idx = {}
    for d in sorted(by):
        bars = sorted(by[d])
        dts.append(d)
        O.append(bars[0][1])
        H.append(max(b[2] for b in bars))
        L.append(min(b[3] for b in bars))
        C.append(bars[-1][4])
        highs = np.array([b[2] for b in bars])
        lows = np.array([b[3] for b in bars])
        idx[d] = (lows, np.maximum.accumulate(highs[::-1])[::-1])
    cache[stock] = ((dts, O, H, L, C), idx)
    with open(CACHE, 'wb') as f:
        pickle.dump(cache, f)
    return cache[stock]


def bear_checker(idx, dates, O):
    def check(i, bid, target):
        ent = idx.get(dates[i])
        if ent is None:
            return bid >= O[i] - 1e-9
        lows, suffix = ent
        hit = lows <= bid + 1e-9
        if not hit.any():
            return False
        j = int(np.argmax(hit))
        return bool(suffix[j] >= target - 1e-9)
    return check


def params_for(s, book_params):
    if s in BOOK:
        return book_params[s]
    if s == 'MRVL':
        t0 = Params(capital=1_000_000, comm=0.005, interest=0.0314, stop_days=50,
                    bayes_pct=0.5, years=2.2, ou_W=80)
        return aw_params(json.load(open('fresh_opt_cands.json'))['MRVL']['reference']['vec'], t0)
    return ref_params(s)


def dma_gate_dates(dts, C, n=200, min_n=100):
    """Dates where the previous close sits below the trailing mean of the previous
    min(n, >=min_n) closes -- strictly ex ante, live once min_n days exist."""
    out = set()
    for i in range(min_n, len(dts)):
        w = C[max(0, i - n):i]
        if C[i - 1] < sum(w) / len(w):
            out.add(dts[i])
    return out


def wstat(eq, dts, lo_d, hi_d):
    """(annualised return, max drawdown) of an equity curve over a date window."""
    i0 = next(i for i, d in enumerate(dts) if d >= lo_d)
    i1 = max(i for i, d in enumerate(dts) if d <= hi_d)
    seg = eq[i0:i1 + 1]
    peak, dd = -1e30, 0.0
    for e in seg:
        peak = max(peak, e)
        if peak > 0:
            dd = max(dd, (peak - e) / peak)
    return annualise(seg[-1] / seg[0] - 1, dts, i0, i1), dd


def main():
    _, book_params, _ = load_book()
    data = {}
    print('===== per-name replay, calendar 2022 (deployed params, verified fills) =====')
    print(f'{"name":6s}{"B&H 2022":>10s}{"model 2022":>11s}{"model DD":>9s}{"stops":>6s}'
          f'{"buys":>6s}{"span":>25s}')
    for s in NAMES:
        (dts, O, H, L, C), idx = _load(s)
        p = params_for(s, book_params)
        chk = bear_checker(idx, dts, O)
        r = run_model(dts, O, H, L, C, p, ou_sigma='resid', same_day_exit=chk, collect=True)
        eq = r.frames['equity']
        m_ann, m_dd = wstat(eq, dts, Y22a, Y22b)
        i0 = next(i for i, d in enumerate(dts) if d >= Y22a)
        i1 = max(i for i, d in enumerate(dts) if d <= Y22b)
        bh = C[i1] / C[i0] - 1
        buys = (sum(r.frames['t1']['Z'][i0:i1 + 1]) + sum(r.frames['t2']['Z'][i0:i1 + 1]))
        stps = sum(1 for t in ('t1', 't2') for i in range(i0, i1 + 1)
                   if r.frames[t]['AD'][i] == 1 and r.frames[t]['AC'][i] is not None
                   and r.frames[t]['AC'][i] < r.frames[t]['AB'][i] - 1e-9)
        print(f'{s:6s}{bh*100:>9.1f}%{m_ann*100:>10.1f}%{m_dd*100:>8.1f}%{stps:>6d}'
              f'{buys:>6d}{str(dts[0]) + ".." + str(dts[-1]):>25s}')
        data[s] = dict(dts=dts, O=O, H=H, L=L, C=C, p=p,
                       idx={d: i for i, d in enumerate(dts)},
                       chk=chk,
                       X=r.frames['X'], AM=r.frames['AM'])

    # ---------------- book level
    sleeves = []
    for s in NAMES:
        p = data[s]['p']
        sleeves.append(dict(name=s, kind='B', bids='X', prem=p.premium))
        sleeves.append(dict(name=s, kind='O', bids='AM', prem=p.ou_prem))
    common = None
    for s in NAMES:
        common = set(data[s]['dts']) if common is None else common & set(data[s]['dts'])
    cal = sorted(common)
    gate = {s: dma_gate_dates(data[s]['dts'], data[s]['C']) for s in NAMES}
    brk = (0.15, 0.075, 0.5)
    configs = [('nothing', {}),
               ('200dma gate', dict(no_buy=gate)),
               ('breaker 15%/half', dict(breaker=brk)),
               ('gate + breaker', dict(no_buy=gate, breaker=brk)),
               ('gate + breaker + stop -25%', dict(no_buy=gate, breaker=brk, price_stop=0.25))]
    for label, lo, hi in [('CALENDAR 2022', Y22a, Y22b),
                          ('FULL SPAN (Jan 2022 - Jun 2023)', Y22a, datetime.date(2023, 6, 30))]:
        print(f'\n===== book, pooled equal weights, {label} =====')
        print(f'{"config":30s}{"ann":>8s}{"total":>8s}{"maxDD":>7s}{"fills":>7s}{"stops":>6s}')
        for name_, kw in configs:
            r = book_sim.simulate(data, sleeves, cal, capital=9_000_000,
                                  collect_trades=True, date_lo=lo, date_hi=hi, **kw)
            eq = r['equity']
            print(f'{name_:30s}{r["full"]*100:>7.1f}%{(eq[-1]/eq[0]-1)*100:>7.1f}%'
                  f'{r["maxdd"]*100:>6.1f}%{r["fills"]:>7d}{r["stops"]:>6d}')


if __name__ == '__main__':
    main()
