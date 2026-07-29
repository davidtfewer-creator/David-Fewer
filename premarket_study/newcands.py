"""
Load the four candidate LIVE workbooks (CCL, LLY, CVNA, MU), read each stock's
own fitted parameters + OHLC, and run the validated engine. First job: confirm the
engine reproduces each workbook's cached summary (Y4 profit / Y5 ann.ret / AA4 buys
/ AC4 stops) before it is trusted for optimisation.
"""
import openpyxl, datetime
from engine import Params, run_model

BASE = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff'
FILES = {
    'CCL':  f'{BASE}/b676f5a4-Daily_trading_Bayesian__OU_tranche_LIVE_CCL.xlsx',
    'LLY':  f'{BASE}/279271c2-Daily_trading_Bayesian__OU_tranche_LIVE_LLY.xlsx',
    'CVNA': f'{BASE}/bf72c12c-Daily_trading_Bayesian__OU_tranche_LIVE_CVNA.xlsx',
    'MU':   f'{BASE}/bd831b15-Daily_trading_Bayesian__OU_tranche_LIVE_MU.xlsx',
}


def load(stock):
    """Return (dates,O,H,L,C, Params, cached-dict) for a stock."""
    f = FILES[stock]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb['2 Tranche StopLoss 50d']
    g = lambda a: ws[a].value
    p = Params(
        lam=g('B2'), phi_L=g('D2'), psi=g('F2'), k=g('H2'),
        premium=g('J2'), peak_cap=g('L2'), comm=g('N2'),
        capital=g('P2') * 2, interest=g('R2'), stop_days=int(g('T2')),
        bayes_pct=0.5, ou_W=int(g('B3')), ou_buf_k=g('D3'),
        ou_prem=g('H3'), ou_cap=g('J3'), years=2.2,
    )
    cached = dict(profit=g('Y4'), annret=g('Y5'), buys=g('AA4'), stops=g('AC4'),
                  fund=g('P2'))
    q = wb['Query']
    dts, O, H, L, C = [], [], [], [], []
    for r in range(2, q.max_row + 1):
        d = q.cell(r, 1).value
        o = q.cell(r, 2).value
        if not isinstance(o, (int, float)) or o <= 0:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        dts.append(d); O.append(o)
        H.append(q.cell(r, 3).value); L.append(q.cell(r, 4).value); C.append(q.cell(r, 5).value)
    return dts, O, H, L, C, p, cached


if __name__ == '__main__':
    print(f'{"stock":5s}{"N":>5s} | {"engine profit":>15s}{"cached profit":>15s} | '
          f'{"eng ann":>9s}{"cache ann":>10s} | {"eng buys":>9s}{"cache":>7s} | '
          f'{"eng stops":>10s}{"cache":>7s}')
    for s in FILES:
        dts, O, H, L, C, p, c = load(s)
        r = run_model(dts, O, H, L, C, p, collect=False)
        print(f'{s:5s}{len(O):>5d} | {r.profit:>15,.0f}{c["profit"]:>15,.0f} | '
              f'{r.annual_return:>9.2%}{c["annret"]:>10.2%} | '
              f'{r.total_buys:>9d}{int(c["buys"]):>7d} | '
              f'{r.stop_loss_exits:>10d}{int(c["stops"]):>7d}')
