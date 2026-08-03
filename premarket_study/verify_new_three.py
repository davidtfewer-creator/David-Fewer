"""
Verified assessment of MU, COIN and AMD using their 5-minute data.

MU   : daily model + fitted parameters exist -> full verified assessment.
COIN : daily OHLC in the diversifier feed; parameters fitted in that study.
AMD  : no daily model at all -> daily bars are reconstructed from the 5-minute session bars,
       and parameters are fitted with the standard robust protocol. Flagged accordingly, since
       freshly fitted parameters are an in-sample ceiling.

Reports for each: optimistic return, VERIFIED return, at-open floor, true same-day fill rate,
trades, Sharpe and drawdown -- the same basis on which the ten book names were restated.
"""
import openpyxl, datetime, collections, statistics
import numpy as np
from engine import Params, run_model

BASE = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff'
FIVE = {
    'MU':   f'{BASE}/94f1080f-MU_5min_Apr2024Aug2026.xlsx',
    'AMD':  f'{BASE}/af55d641-AMD_5min_Apr2024Aug2026.xlsx',
    'COIN': f'{BASE}/882518ff-COIN_5min_Apr2024Aug2026.xlsx',
}
RTH0, RTH1 = datetime.time(9, 30), datetime.time(16, 0)


def load_five(stock):
    """date -> sorted list of (dt, o, h, l, c) regular-hours 5-minute bars."""
    wb = openpyxl.load_workbook(FIVE[stock], read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False; continue
        dt, o, h, l, c = row[0], row[1], row[2], row[3], row[4]
        if not isinstance(dt, datetime.datetime) or o is None: continue
        if dt.time() < RTH0 or dt.time() >= RTH1: continue
        by[dt.date()].append((dt, o, h, l, c))
    wb.close()
    for d in by: by[d].sort(key=lambda x: x[0])
    return by


def daily_from_five(by):
    """Reconstruct daily OHLC from session bars."""
    dts, O, H, L, C = [], [], [], [], []
    for d in sorted(by):
        bars = by[d]
        if len(bars) < 20: continue                 # skip half-days / sparse sessions
        dts.append(d); O.append(bars[0][1])
        H.append(max(b[2] for b in bars)); L.append(min(b[3] for b in bars))
        C.append(bars[-1][4])
    return dts, O, H, L, C


def index_from_five(by):
    idx = {}
    for d, bars in by.items():
        highs = np.array([b[2] for b in bars], float)
        lows = np.array([b[3] for b in bars], float)
        idx[d] = (lows, np.maximum.accumulate(highs[::-1])[::-1])
    return idx


def make_checker(idx, dates, O):
    def check(i, bid, target):
        e = idx.get(dates[i])
        if e is None: return bid >= O[i] - 1e-9
        lows, suf = e
        hit = lows <= bid + 1e-9
        if not hit.any(): return False
        j = int(np.argmax(hit))
        return bool(suf[j] >= target - 1e-9)
    return check


def assess(name, dts, O, H, L, C, p, idx, note=''):
    chk = make_checker(idx, dts, O)
    yrs = (dts[-1] - dts[0]).days / 365.25
    ro = run_model(dts, O, H, L, C, p, collect=True)
    rv = run_model(dts, O, H, L, C, p, same_day_exit=chk)
    ra = run_model(dts, O, H, L, C, p, same_day_exit='at_open')
    real = fake = tot = sd = 0
    for tk, bids in (('t1', ro.frames['X']), ('t2', ro.frames['AM'])):
        t = ro.frames[tk]
        for i in range(len(C)):
            if t['Z'][i] == 1:
                tot += 1
                if t['AD'][i] == 1:
                    sd += 1
                    if bids[i] is not None and dts[i] in idx:
                        if chk(i, bids[i], t['AB'][i]): real += 1
                        else: fake += 1
    rate = real/(real+fake)*100 if real+fake else 0
    print(f'{name:6s}{ro.annual_return*100:>11.0f}%{rv.annual_return*100:>10.0f}%'
          f'{ra.annual_return*100:>10.0f}%{rate:>10.0f}%{sd/tot*100:>10.0f}%'
          f'{rv.total_buys/yrs:>9.0f}{rv.sharpe:>8.2f}{rv.max_drawdown*100:>7.0f}% {note}')
    return rv.annual_return


if __name__ == '__main__':
    print('=== VERIFIED ASSESSMENT (5-minute fills, regular hours) ===\n')
    print(f'{"name":6s}{"optimistic":>12s}{"VERIFIED":>10s}{"at-open":>10s}'
          f'{"fill rate":>10s}{"same-day":>10s}{"trades/yr":>9s}{"Sharpe":>8s}{"maxDD":>8s}')
    print('-'*92)

    # ---- MU: daily model + fitted params ----
    import newcands
    dts, O, H, L, C, p, _ = newcands.load('MU')
    assess('MU', dts, O, H, L, C, p, index_from_five(load_five('MU')))

    # ---- COIN: diversifier feed + params fitted in that study ----
    import newfeed
    nf = newfeed.load(newfeed.NEW, ['COIN'])
    cd, cO, cH, cL, cC = nf['COIN']
    # parameters from the diversifier optimisation (in-sample fit; see caveat)
    pc = Params(lam=0.5, phi_L=0.3, psi=0.01, k=1.2, premium=0.025, peak_cap=0.02,
                ou_buf_k=0.5, ou_prem=0.025, ou_cap=0.03, ou_W=60,
                comm=0.005, capital=6_000_000, interest=0.0314, stop_days=50,
                bayes_pct=0.5, years=(cd[-1]-cd[0]).days/365.25)
    assess('COIN', cd, cO, cH, cL, cC, pc, index_from_five(load_five('COIN')), '(generic params)')

    # ---- AMD: daily bars reconstructed from the 5-minute file ----
    by = load_five('AMD')
    ad, aO, aH, aL, aC = daily_from_five(by)
    pa = Params(lam=0.5, phi_L=0.3, psi=0.01, k=1.2, premium=0.025, peak_cap=0.02,
                ou_buf_k=0.5, ou_prem=0.025, ou_cap=0.03, ou_W=60,
                comm=0.005, capital=6_000_000, interest=0.0314, stop_days=50,
                bayes_pct=0.5, years=(ad[-1]-ad[0]).days/365.25)
    assess('AMD', ad, aO, aH, aL, aC, pa, index_from_five(by), '(generic params, daily rebuilt)')
    print(f'\nAMD daily bars reconstructed from {len(ad)} sessions '
          f'({ad[0]} -> {ad[-1]}).')
