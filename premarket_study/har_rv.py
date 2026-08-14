"""
HAR-RV: realised-volatility forecasts from the 5-minute bars (Corsi 2009).

The deployed model scales every noise in the Bayes filter -- q_L, q_b, r -- by the
daily range H-L, and the OU buffer by a window residual std. Both are volatility
proxies built from daily data. We own two years of 5-minute bars; realised variance
(the sum of squared 5-minute log returns inside the session) is a far less noisy
measurement of the same quantity, and the HAR regression

    sigma_t = b0 + bd*sigma_{t-1} + bw*mean(sigma_{t-5..t-1}) + bm*mean(sigma_{t-22..t-1})

is the standard forecaster on top of it: three horizons of volatility memory,
fitted by OLS. Everything here is strictly ex ante -- the forecast for day t uses
bars through day t-1 only, and the coefficients are fitted on the TRAIN half only
(boundary 2025-05-23, same split as everything else).

Two per-row series are produced for the engine hooks:
  F_har[i]    dollar range-equivalent = RANGE_EQ * sigma_rel[i] * C[i-1], feeding
              the Kalman noise scalings in place of H-L. RANGE_EQ = 2*sqrt(2/pi)
              (the Brownian E[range]/sigma ratio) keeps phi_L / psi / k on the
              same scale as the deployed model, so the existing search bounds
              remain appropriate.
  ou_sig[i]   dollar sigma = sigma_rel[i] * C[i-1], feeding the OU buffer in
              place of the AR(1) residual std (already daily-sigma scale).
Rows without a forecast (22-day warmup, dates before 5-minute coverage) are None:
the engine falls back to the deployed proxy on those rows.
"""
import collections
import datetime
import math
import os
import pickle

import numpy as np

from minute_index import DIR, RTH_START, RTH_END

RANGE_EQ = 2 * math.sqrt(2 / math.pi)   # ~1.596: E[H-L] / sigma for Brownian motion
HAR_LAGS = (1, 5, 22)


def rv_daily(stock):
    """date -> realised variance of RTH 5-min log returns (relative units).
    Includes the open-to-first-close return; cached beside the minute index."""
    cache = os.path.join(DIR, f'{stock}_rv.pkl')
    if os.path.exists(cache):
        with open(cache, 'rb') as fh:
            return pickle.load(fh)
    import openpyxl
    path = os.path.join(DIR, f'{stock}_5min.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        dt, o, c = row[0], row[1], row[4]
        if not isinstance(dt, datetime.datetime) or o is None:
            continue
        t = dt.time()
        if t < RTH_START or t >= RTH_END:
            continue
        by[dt.date()].append((dt, float(o), float(c)))
    wb.close()
    rv = {}
    for d, bars in by.items():
        bars.sort()
        px = [bars[0][1]] + [b[2] for b in bars]
        rv[d] = sum(math.log(px[j + 1] / px[j]) ** 2
                    for j in range(len(px) - 1) if px[j] > 0 and px[j + 1] > 0)
    with open(cache, 'wb') as fh:
        pickle.dump(rv, fh)
    return rv


def har_fit(rv, train_end):
    """OLS of sigma_t on (1, sigma lag-1, lag-5 mean, lag-22 mean), train rows only.
    Returns (beta, diagnostics dict)."""
    ds = sorted(rv)
    sig = [math.sqrt(rv[d]) for d in ds]
    rows, ys, dates = [], [], []
    m = max(HAR_LAGS)
    for i in range(m, len(ds)):
        rows.append([1.0,
                     sig[i - 1],
                     sum(sig[i - 5:i]) / 5.0,
                     sum(sig[i - 22:i]) / 22.0])
        ys.append(sig[i])
        dates.append(ds[i])
    X = np.array(rows); y = np.array(ys)
    tr = np.array([d < train_end for d in dates])
    beta, *_ = np.linalg.lstsq(X[tr], y[tr], rcond=None)

    def r2(mask, pred):
        e = y[mask] - pred[mask]
        return 1 - (e @ e) / (((y[mask] - y[mask].mean()) ** 2).sum())
    pred = X @ beta
    naive = X[:, 1]                       # random walk: yesterday's sigma
    diag = dict(beta=[float(b) for b in beta],
                n_train=int(tr.sum()), n_test=int((~tr).sum()),
                r2_train=float(r2(tr, pred)), r2_test=float(r2(~tr, pred)),
                r2_naive_test=float(r2(~tr, naive)))
    return beta, diag


def forecast_series(dts, C, rv, beta):
    """Per-row sigma_rel forecast for the model calendar dts, strictly ex ante:
    day i uses RV observations on dates < dts[i]. None where history < 22 days."""
    ds = sorted(rv)
    sig = [math.sqrt(rv[d]) for d in ds]
    out = [None] * len(dts)
    j = 0
    m = max(HAR_LAGS)
    for i, d in enumerate(dts):
        while j < len(ds) and ds[j] < d:
            j += 1
        if j < m:
            continue
        s1 = sig[j - 1]
        s5 = sum(sig[j - 5:j]) / 5.0
        s22 = sum(sig[j - 22:j]) / 22.0
        pred = beta[0] + beta[1] * s1 + beta[2] * s5 + beta[3] * s22
        out[i] = max(pred, 0.1 * s22)     # OLS floor: never a degenerate noise scale
    return out


def engine_series(dts, C, sig_rel):
    """(F_har, ou_sig) dollar series for the engine hooks; None rows fall back."""
    F_har, ou_sig = [None] * len(dts), [None] * len(dts)
    for i in range(1, len(dts)):
        if sig_rel[i] is None:
            continue
        F_har[i] = RANGE_EQ * sig_rel[i] * C[i - 1]
        ou_sig[i] = sig_rel[i] * C[i - 1]
    return F_har, ou_sig
