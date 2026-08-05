"""
Allocate capital only to sleeves the blotter says are in cash.

The live holding state is 'Active Trading'!C19:C28, the Status column the IBKR automation drives
off the blotter. Allocation!C25:C34 ignored it entirely -- they read M11:N15, a fixed tenth of
capital each, whatever the book was holding -- so a sleeve that bought yesterday still showed a
full dollar allocation.

The rule is a weight renormalisation over the sleeves in cash:

    held_i    from 'Active Trading'!C19:C28 = "HOLDING"
    w_i       = K_stock x Bayes share       (or x (1 - Bayes share) for the OU sleeve)
    f_i       = w_i x (1 - held_i)
    C_i       = B5 x f_i / sum(f)

Written this way rather than as "cash / number of free sleeves" so it stays correct if the
equal-weighting is relaxed; under the deployed floor = cap = 0.20 and a 50/50 split every w_i is
0.10 and it reduces exactly to B5 divided by the number of free sleeves.

The blotter's share count is brought onto the same definition. 'Active Trading'!H19:H28 already
carried a copy of this logic --

    MIN( Allocation!$B$5 * $D19 / SUMIF($C$19:$C$28,"CASH",$D$19:$D$28) , $D19 )

-- but D19:D28 sum to exactly B5, so B5*D19/SUMIF(cash) is never less than D19 and the MIN always
picks D19. The redistribution could not fire. Rather than repair the duplicate, H now reads the
allocation block directly, so there is one definition of how much money a sleeve gets and the
IBKR script and the blotter cannot disagree. D19:D28, whose only consumer was that formula, shows
the same figure.

Not changed: Allocation L11:L15 and M/N. 'Active Trading'!B7:B11 read L11:L15 as each stock's
Day-1 starting fund and derive P&L and return from it, so a baseline that moved with the holding
state would make the blotter's return column drift every morning. The standing capital plan and
today's order sizes are different quantities.
"""
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = ('/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/'
       '3e1130ba-TradingExcel_5stock_live.xlsx')
OUT = '/home/user/David-Fewer/TradingExcel_5stock_live_freesleeves.xlsx'
NAMES = ['TSM', 'VRT', 'VST', 'RKLB', 'MU']

BLUE = PatternFill('solid', fgColor='DCE6F1')
HDR = Font(bold=True, color='0B1F3A')
GREY = Font(italic=True, color='5A6270', size=9)

# Allocation row 25..34 lines up one for one with 'Active Trading' row 19..28
AT_ROW = lambda r: r - 6


def held_f(r):
    return f'=IF(\'Active Trading\'!$C${AT_ROW(r)}="HOLDING",1,0)'


def base_f(r, srow, tranche):
    share = f'$R${srow}' if tranche == 'Bayes' else f'(1-$R${srow})'
    return f'=$K${srow}*{share}'


def free_f(r):
    return f'=E{r}*(1-D{r})'


def alloc_f(r):
    return f'=IF(SUM($F$25:$F$34)=0,0,$B$5*F{r}/SUM($F$25:$F$34))'


def shares_f(r):
    """r is the Active Trading row. One definition of the money, read from Allocation."""
    return (f'=IF($C{r}="CASH",IFERROR(FLOOR(INDEX(Allocation!$C$25:$C$34,ROW()-18)'
            f'/($F{r}+$B$3),1),0),SUMIFS($F$42:$F$1041,$B$42:$B$1041,$A{r},'
            f'$C$42:$C$1041,$B{r},$M$42:$M$1041,"OPEN"))')


def fund_f(r):
    return '=INDEX(Allocation!$C$25:$C$34,ROW()-18)'


def build_allocation(ws):
    ws['A23'] = ('MACHINE-READABLE OUTPUT  (stock x tranche dollar allocation - read by the IBKR '
                 'script). A sleeve the blotter shows as HOLDING is allocated nothing; the cash '
                 'in B5 is divided across the sleeves still in cash.')
    for col, txt in ((4, 'Held?'), (5, 'Base wt'), (6, 'Free wt')):
        c = ws.cell(24, col, txt)
        c.font = HDR
        c.alignment = Alignment(horizontal='right')

    for i, s in enumerate(NAMES):
        srow = 11 + i
        for t, tranche in enumerate(('Bayes', 'OU')):
            r = 25 + 2*i + t
            ws.cell(r, 4, held_f(r))
            ws.cell(r, 5, base_f(r, srow, tranche))
            ws.cell(r, 6, free_f(r))
            ws.cell(r, 3, alloc_f(r))
            ws.cell(r, 3).number_format = '#,##0.00'
            ws.cell(r, 5).number_format = '0.0000'
            ws.cell(r, 6).number_format = '0.0000'

    ws['A36'] = 'Sleeves in cash today'
    ws['B36'] = '=COUNT(D25:D34)-SUM(D25:D34)'
    ws['A37'] = '$ per free sleeve (equal weights)'
    ws['B37'] = '=IF(B36=0,0,$B$5/B36)'
    ws['A38'] = 'Check: allocated = cash available'
    ws['B38'] = '=IF(ABS(SUM(C25:C34)-$B$5)<0.01,"OK","MISMATCH")'
    for c in ('A36', 'A37', 'A38'):
        ws[c].font = HDR
    ws['B37'].number_format = '#,##0.00'

    ws['A40'] = ('Held? is read from the blotter Status in Active Trading C19:C28, which the IBKR '
                 'automation drives. Nothing here needs touching by hand.')
    ws['A41'] = ('B5 is the CASH available this morning, not the total value of the book. The '
                 'sleeves in cash take the whole of B5 between them, so entering total portfolio '
                 'value would commit capital that is already invested in the held sleeves.')
    for c in ('A40', 'A41'):
        ws[c].font = GREY
    ws['A5'] = 'Cash available today ($)'
    ws['B5'].fill = BLUE


def build_active(ws):
    for r in range(19, 29):
        ws.cell(r, 4, fund_f(r))
        ws.cell(r, 8, shares_f(r))
        ws.cell(r, 4).number_format = '#,##0.00'


if __name__ == '__main__':
    shutil.copy(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    build_allocation(wb['Allocation'])
    build_active(wb['Active Trading'])
    n = wb['Notes']
    n['A19'] = 'Free-sleeve allocation'
    n['B19'] = ('Allocation C25:C34 skip any sleeve the blotter reports as HOLDING and divide the '
                'cash in B5 across the sleeves still in cash. Previously every sleeve showed a '
                'full tenth of capital whatever the book held. Column D carries the held flag '
                'read from Active Trading C19:C28, E and F the base and free weights. Active '
                'Trading D19:D28 and H19:H28 now read the same block, so the order size and the '
                'share count cannot disagree; the MIN cap that used to sit in H could never fire, '
                'because D19:D28 sum to B5 and the cap therefore always bound. L11:L15 are '
                'unchanged and still feed the blotter as the Day-1 baseline.')
    wb.save(OUT)
    print('wrote', OUT, flush=True)
    print('DONE', flush=True)
