"""
Traffic-light colour coding for the gate monitor columns (19 Aug 2026).

Orders block, rows 19-36:
  N (200 DMA)          green: close (col J) comfortably above the average;
                       orange: within 5% above it; red: at/below (DMA gate fires).
  O (Sale % below ATH) green: sale price comfortably below the ATH;
                       orange: within 2pp of the ATH gate epsilon (M7);
                       red: at/below epsilon (the ATH guard is binding).
Pure conditional formatting -- no cell values change, thresholds reference the
live Gate Variables cell so the colours track M7 automatically.
"""
import sys

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

# NB: conditional-formatting (dxf) fills need BOTH colours set or Excel renders
# only the font part -- cell backgrounds, no font colouring (user preference).
def _fill(rgb):
    return PatternFill(fill_type='solid', start_color=rgb, end_color=rgb, bgColor=rgb)

GREEN = dict(fill=_fill('C6EFCE'))
ORANGE = dict(fill=_fill('FFD9B3'))
RED = dict(fill=_fill('FFC7CE'))


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    at = wb['Active Trading']
    assert at['N18'].value == '200 DMA' and at['O18'].value == 'Sale % below ATH'
    assert at['M7'].value is not None            # ATH epsilon lives here

    def rule(formula, style):
        return FormulaRule(formula=[formula], stopIfTrue=True, fill=style['fill'])

    at.conditional_formatting.add(
        'N19:N36', rule('AND(ISNUMBER($N19),$J19<=$N19)', RED))
    at.conditional_formatting.add(
        'N19:N36', rule('AND(ISNUMBER($N19),$J19<=$N19*1.05)', ORANGE))
    at.conditional_formatting.add(
        'N19:N36', rule('ISNUMBER($N19)', GREEN))

    at.conditional_formatting.add(
        'O19:O36', rule('AND(ISNUMBER($O19),$O19<=N($M$7))', RED))
    at.conditional_formatting.add(
        'O19:O36', rule('AND(ISNUMBER($O19),$O19<=N($M$7)+0.02)', ORANGE))
    at.conditional_formatting.add(
        'O19:O36', rule('ISNUMBER($O19)', GREEN))

    notes = wb['Notes']
    import copy
    row = notes.max_row + 1
    ca = notes.cell(row=row, column=1, value='Gate colours')
    ca.font = copy.copy(notes['A8'].font)
    cb = notes.cell(
        row=row, column=2,
        value=('Cols N/O are traffic-lit: N green = close well above the 200dma, orange = '
               'within 5% above, red = at/below (DMA gate firing). O green = sale price well '
               'below the ATH, orange = within 2pp of the ATH epsilon (M7), red = at/below '
               'epsilon (ATH guard binding). Thresholds follow M7 live.'))
    cb.font = copy.copy(notes['B8'].font)

    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)


def verify(path):
    wb = openpyxl.load_workbook(path)
    at = wb['Active Trading']
    for rng in ('N19:N36', 'O19:O36'):
        rules = None
        for cf in at.conditional_formatting:
            if str(cf.sqref) == rng:
                rules = cf.rules
        print(rng, '->', len(rules), 'rules:',
              [r.formula[0][:44] for r in rules])


if __name__ == '__main__':
    wire(sys.argv[1], sys.argv[2])
    verify(sys.argv[2])
