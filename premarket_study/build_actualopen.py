"""
Non-destructive edit: add an actual-open input to the Dashboard so the live BUY
levels include the model's open cap (currently omitted pre-open).

For each stock the validated model's bid is  MIN(fair-buffer, OPEN, peak*(1-cap)).
The Dashboard currently computes MIN(fair-buffer, peak*(1-cap)) only. We add a
'Open @9:30' input column (R) and fold it into the Bayes BUY (C) and OU BUY (E)
formulas, binding only when an open has been entered.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName

SRC = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/d0c3ec6f-Hybrid9_Bayesian_OULIVE_PLTR_nopq_test_copy.xlsx'
OUT = '/home/user/David-Fewer/Hybrid9_Bayesian_OULIVE_PLTR_actualopen.xlsx'
ROWS = range(4, 14)  # NVDA..SPOT
BLUE = 'FFD9E1F2'     # existing "type the blue cells" input fill

wb = openpyxl.load_workbook(SRC, data_only=False)
ws = wb['Dashboard']

# --- open-term appended before the final ')' of each MIN() ---
def add_open(formula, r):
    assert formula.rstrip().endswith(')')
    return formula[:-1] + f",IF($R{r}>0,$R{r},9.9E+307))"

for r in ROWS:
    c = ws.cell(row=r, column=3)   # Bayes BUY
    e = ws.cell(row=r, column=5)   # OU BUY
    c.value = add_open(c.value, r)
    e.value = add_open(e.value, r)

# --- input column R: header + blue input cells ---
hdr = ws.cell(row=3, column=18, value='Open @9:30')       # R3
hdr.font = Font(name='Arial', size=10, bold=True)
hdr.alignment = Alignment(horizontal='center')
for r in ROWS:
    cell = ws.cell(row=r, column=18)                       # R4:R13
    cell.value = None
    cell.font = Font(name='Arial', size=10)
    cell.number_format = '"$"#,##0.00'
    cell.fill = PatternFill('solid', fgColor=BLUE)
    cell.alignment = Alignment(horizontal='right')
ws.column_dimensions['R'].width = 12

# short legend under the block
note = ws.cell(row=15, column=18,
               value='^ type each stock’s actual open at 9:30:05 (blue). '
                     'Blank = pre-open estimate (no open cap).')
note.font = Font(name='Arial', size=9, italic=True)

# --- refreshed instruction line ---
ws['A2'] = ('At 9:30:05 enter each stock’s actual open in the blue "Open @9:30" column, '
            'then place a limit BUY at the buy level for any tranche in cash and a GTC SELL at its target.')

# --- named range for the automation to write opens into ---
try:
    del wb.defined_names['IBKR_OpenIn']
except KeyError:
    pass
wb.defined_names.add(DefinedName('IBKR_OpenIn', attr_text='Dashboard!$R$4:$R$13'))

wb.save(OUT)
print('saved', OUT)
print('edited Bayes BUY (C4:C13) and OU BUY (E4:E13); added open input R4:R13 + IBKR_OpenIn')
