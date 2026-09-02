"""
RKLB -> AVGO composition swap (user decision, 2 Sep 2026).

Rationale (user's): RKLB's post-IPO behaviour is partly SpaceX-linked and it
carries binary event risk (Neutron); its admission was likely volatility- and
excitement-driven ahead of the SpaceX IPO. AVGO takes its slot. The user has
already renamed the Feed/Model sheets and rewired Dashboard row 7, Allocation
row 14 and order-block rows 25/26; this script finishes the job:

  1. Query N:Q      — headers RKLB_* -> AVGO_*; data replaced with AVGO daily
                      RTH OHLC (verified research series, exact date match) for
                      2024-04-01..2026-08-03; the 20 sessions 2026-08-04..31
                      are CLEARED (no AVGO source here) — backfill via
                      ops/backfill_query.py (IBKR) before AVGO trades.
                      Feed AVGO pulls from Query by INDEX, so the whole
                      Feed -> Model -> Dashboard chain updates itself.
  2. Model AVGO     — RKLB's fitted parameters replaced with AVGO's reference
                      vector (fresh_opt_cands, lam=1 family, verified fills:
                      captive 51-53%/yr, train 49 / test 54).
  3. Labels         — Dashboard A7, Allocation A31/A32, Active Trading A1
                      banner and A10, Feed AVGO note, Notes roster text.
  4. Allocation C14 — annual-return assumption 1.15 (RKLB) -> 0.51 (AVGO
                      verified reference), feeding the efficiency split.
  5. Performance    — hold-table references: AVGO 6.5 d (captive back-test,
                      144 trades), BOOK 5.9 -> 6.5 (pooled book re-run with
                      AVGO in the slot).
  6. Notes          — changelog block for the swap.

NOT touched, by design: historical RKLB rows in the blotter and the open
discretionary RKLB position (history and manual book respectively); every
sheet address, row and range the Python scripts read (layout identical — the
only script-side change is the ticker string RKLB -> AVGO wherever Scripts 1/2
carry their own ticker list).

Usage: python wire_avgo_swap.py <in.xlsx> <out.xlsx>
"""
import copy
import datetime as dt
import json
import sys

import openpyxl

sys.path.insert(0, '.')

EPOCH = dt.date(1899, 12, 30)
QCOLS = (14, 15, 16, 17)                     # N O P Q
AVGO_HOLD, BOOK_HOLD_OLD, BOOK_HOLD_NEW = 6.5, 5.9, 6.5

# RKLB's fitted values still sitting in 'Model AVGO' (preconditions)
OLD = {'B2': 0.352847, 'D2': 0.237406, 'F2': 0.113048, 'H2': 0.712293,
       'J2': 0.026842, 'L2': 0.008015, 'B3': 85, 'D3': 0.25,
       'H3': 0.022147, 'J3': 0.033647}


def avgo_params():
    from engine import Params
    from fresh_opt_cands import aw_params
    t0 = Params(capital=1_000_000, comm=0.005, interest=0.0314, stop_days=50,
                bayes_pct=0.5, years=2.2, ou_W=80)
    vec = json.load(open('fresh_opt_cands.json'))['AVGO']['reference']['vec']
    return aw_params(vec, t0)


def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, (int, float)):
        return EPOCH + dt.timedelta(days=int(v))
    return v


def wire(in_path, out_path):
    from fresh_opt_cands import daily_from_5min
    dts, O, H, L, C = daily_from_5min('AVGO')
    bars = {(d.date() if hasattr(d, 'date') else d): (o, h, l, c)
            for d, o, h, l, c in zip(dts, O, H, L, C)}

    wb = openpyxl.load_workbook(in_path)
    q, at = wb['Query'], wb['Active Trading']
    m, dash = wb['Model AVGO'], wb['Dashboard']
    al, pf, notes = wb['Allocation'], wb['Performance'], wb['Notes']

    # ---- 1. Query block N:Q
    assert [q.cell(row=1, column=c).value for c in QCOLS] == \
        ['RKLB_O', 'RKLB_H', 'RKLB_L', 'RKLB_C']
    for c, h in zip(QCOLS, ['AVGO_O', 'AVGO_H', 'AVGO_L', 'AVGO_C']):
        q.cell(row=1, column=c, value=h)
    filled = cleared = 0
    for r in range(2, q.max_row + 1):
        d = as_date(q.cell(row=r, column=1).value)
        if d in bars:
            for c, v in zip(QCOLS, bars[d]):
                q.cell(row=r, column=c, value=round(v, 4))
            filled += 1
        else:
            for c in QCOLS:
                # NB: ws.cell(..., value=None) silently skips assignment —
                # must set .value directly to actually blank the cell
                q.cell(row=r, column=c).value = None
            cleared += 1
    assert filled == 587 and cleared == 20, (filled, cleared)

    # ---- 2. Model AVGO parameters
    p = avgo_params()
    for cell, old in OLD.items():
        assert abs(m[cell].value - old) < 1e-6, (cell, m[cell].value)
    m['B2'] = p.lam
    m['D2'] = p.phi_L
    m['F2'] = p.psi
    m['H2'] = p.k
    m['J2'] = p.premium
    m['L2'] = p.peak_cap
    m['B3'] = p.ou_W
    m['D3'] = p.ou_buf_k
    m['H3'] = p.ou_prem
    m['J3'] = p.ou_cap
    assert m['A1'].value.startswith('RKLB')
    m['A1'] = ('AVGO Daily Bayesian – 2 Tranches, 50-Day Stop-Loss  '
               '(live Feed | auto-extends)')

    # ---- 3. labels
    assert dash['A7'].value == 'RKLB'
    dash['A7'] = 'AVGO'
    for cell in ('A31', 'A32'):
        assert al[cell].value == 'RKLB'
        al[cell] = 'AVGO'
    assert at['A10'].value == 'RKLB'
    at['A10'] = 'AVGO'
    assert '· RKLB ·' in at['A1'].value
    at['A1'] = at['A1'].value.replace('· RKLB ·', '· AVGO ·')
    f = wb['Feed AVGO']
    assert 'RKLB' in f['G1'].value
    f['G1'] = ('Daily OHLC for AVGO. Loaded from the Query sheet; the Model '
               'sheet reads columns A:E.')

    # ---- 4. allocation assumption
    assert abs(al['C14'].value - 1.15) < 1e-9
    al['C14'] = 0.51

    # ---- 5. performance hold references
    assert pf['I12'].value == 'AVGO' and abs(pf['J12'].value - 3.1) < 1e-9
    pf['J12'] = AVGO_HOLD
    assert pf['I18'].value == 'BOOK' and abs(pf['J18'].value - BOOK_HOLD_OLD) < 1e-9
    pf['J18'] = BOOK_HOLD_NEW

    # ---- 6. notes: roster + buf-note + changelog
    assert 'RKLB' in notes['B3'].value
    notes['B3'] = ('Nine daily names: TSM, VRT, VST, AVGO, MU, GM, VLO, CF + MRVL '
                   '(MRVL admitted August 2026: planning 78% on train-only fits, AI '
                   'class, beta 0.95; AMD, SMCI and CEG assessed in the same round '
                   'and declined). AVGO replaced RKLB on 2 September 2026 — a risk '
                   'decision, not a model one: SpaceX-linked performance since IPO '
                   'and binary launch risk (Neutron). See the changelog below.')
    assert 'RKLB 0.2206 -> 0.25, ' in notes['B9'].value
    notes['B9'] = notes['B9'].value.replace('RKLB 0.2206 -> 0.25, ', '') + \
        ' (AVGO, GM, VLO, CF, MRVL were fitted directly under the new convention.)'

    style_hdr, style_a, style_b = notes['A32'], notes['A33'], notes['B33']
    row = notes.max_row + 2
    h = notes.cell(row=row, column=1, value='UPDATE — 2 SEPTEMBER 2026')
    h.font = copy.copy(style_hdr.font)
    h.alignment = copy.copy(style_hdr.alignment)
    row += 1
    entries = [
        ('RKLB -> AVGO swap',
         'AVGO takes RKLB\'s slot everywhere: Query cols N:Q (AVGO daily OHLC, '
         'verified research series to 3 Aug 2026), Feed/Model AVGO (sheets renamed '
         'earlier; fitted parameters now AVGO\'s reference vector — verified '
         'captive ~51%/yr), Dashboard row 7, Allocation row 14 (annual-return '
         'assumption 0.51) and rows 31/32, order-block rows 25/26, stock-funds '
         'row 10. Historical RKLB rows in the blotter are untouched (their P&L '
         'stays in the weekly log; the per-stock funds row now shows AVGO). The '
         'open discretionary RKLB position remains in the disc log until closed.'),
        ('BEFORE AVGO TRADES',
         '1) Backfill Query N:Q for 4-31 Aug 2026 (ops/backfill_query.py via '
         'IBKR, or Script 1\'s historical pull) — until then AVGO\'s last close, '
         '200dma and ATH read as of 3 Aug and its buy levels are stale. '
         '2) Update the ticker list in Scripts 1 and 2 (RKLB -> AVGO) — sheet '
         'layout and every cell address are unchanged, so that one string is the '
         'only script edit. 3) AVGO reports ~3-4 Sep 2026: the earnings map '
         'measured week-of model entries adverse for AVGO (the MU profile) — '
         'consider keeping AVGO bids off until the print clears.'),
        ('Hold references',
         'Performance hold table: AVGO historical 6.5 d (captive back-test, 144 '
         'trades, 39% same-day); BOOK reference 5.9 -> 6.5 (pooled book re-run '
         'with AVGO in the slot). Swapped-book pooled baseline: 72.4%/yr full, '
         '43.9 train / 104.9 test (was 87.3 / 59.5 / 118.4 with RKLB) — the '
         'documentation re-cast will carry these as the new reference numbers.'),
    ]
    for label, text in entries:
        ca = notes.cell(row=row, column=1, value=label)
        ca.font = copy.copy(style_a.font)
        ca.alignment = copy.copy(style_a.alignment)
        cb = notes.cell(row=row, column=2, value=text)
        cb.font = copy.copy(style_b.font)
        cb.alignment = copy.copy(style_b.alignment)
        row += 1

    wb.save(out_path)


def diff(in_path, out_path):
    a = openpyxl.load_workbook(in_path)
    b = openpyxl.load_workbook(out_path)
    changed = []
    for ws_name in b.sheetnames:
        wa, wb_ = a[ws_name], b[ws_name]
        for r in range(1, max(wa.max_row, wb_.max_row) + 1):
            for c in range(1, max(wa.max_column, wb_.max_column) + 1):
                va = wa.cell(row=r, column=c).value
                vb = wb_.cell(row=r, column=c).value
                if va != vb:
                    # openpyxl's save re-serialises floats to shortest repr;
                    # ignore that noise so real deviations stand out
                    if (isinstance(va, float) and isinstance(vb, float)
                            and abs(va - vb) <= 1e-8 * max(abs(va), 1.0)):
                        continue
                    changed.append((ws_name, wb_.cell(row=r, column=c).coordinate))
    return changed


ALLOWED = {('Dashboard', 'A7'), ('Allocation', 'A31'), ('Allocation', 'A32'),
           ('Allocation', 'C14'), ('Active Trading', 'A1'), ('Active Trading', 'A10'),
           ('Feed AVGO', 'G1'), ('Performance', 'J12'), ('Performance', 'J18')}

if __name__ == '__main__':
    src, dst = sys.argv[1], sys.argv[2]
    wire(src, dst)
    ch = diff(src, dst)
    outside = [x for x in ch
               if not (x[0] == 'Query' and x[1][0] in 'NOPQ')
               and not (x[0] == 'Model AVGO' and x[1] in
                        ('A1', 'B2', 'D2', 'F2', 'H2', 'J2', 'L2', 'B3', 'D3', 'H3', 'J3'))
               and x[0] != 'Notes' and x not in ALLOWED]
    print(f'changed cells: {len(ch)}; outside intended set: {outside or "NONE"}')
