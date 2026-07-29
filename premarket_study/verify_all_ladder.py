"""Verify the written laddered formulas for ALL ten Model sheets against the engine
(run on each sheet's own data from this workbook). Confirms transcription across the book."""
import re, datetime, openpyxl
from openpyxl.utils import get_column_letter
from engine import Params
from ladder_engine import run_ladder

OUT = '/home/user/David-Fewer/TradingExcel_s1_laddering_OptionC_backtest.xlsx'
STOCKS = ['NVDA','TSM','TSLA','VRT','VST','AVGO','PLTR','RKLB','SOFI','SPOT']
LAD_START = 60
NAMES = ['FRESH','F0','PR1','PR2','PR3','B1','B2','B3','FIL1','FIL2','FIL3','DSH','SHMID',
         'FUNDMID','ANCM','TGT','STOP','EXIT','SALE','SH','FUND','F1','F2','F3','ANC','BD',
         'NB','INT','EQ','SH1','SH2','SH3']
COL = {n: get_column_letter(LAD_START + i) for i, n in enumerate(NAMES)}
EVAL_ORDER = ['FRESH','F0','PR1','PR2','PR3','B1','B2','B3','FIL1','FIL2','FIL3','DSH','SHMID',
              'FUNDMID','ANCM','TGT','BD','STOP','EXIT','SALE','SH','FUND','F1','F2','F3','ANC',
              'NB','INT','EQ','SH1','SH2','SH3']
LAD_COLS = set(COL.values()); EPOCH = datetime.date(1899, 12, 30)
CFG_VAL_COL = get_column_letter(LAD_START + len(NAMES) + 2)


def to_serial(v):
    if isinstance(v, datetime.datetime): v = v.date()
    if isinstance(v, datetime.date): return (v - EPOCH).days
    return v

def to_date(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)): return v.date() if isinstance(v, datetime.datetime) else v
    if isinstance(v, (int, float)): return EPOCH + datetime.timedelta(days=int(v))
    try: return datetime.date.fromisoformat(str(v)[:10])
    except Exception: return None

def IF(a, b, c): return b if a else c
def AND(*a): return all(a)
def OR(*a): return any(a)
def ISNUMBER(x): return isinstance(x, (int, float)) and not isinstance(x, bool)
MIN, MAX = min, max
REF = re.compile(r'\$?[A-Z]{1,3}\$?[0-9]+')

def _split(content):
    args, depth, cur = [], 0, ''
    for ch in content:
        if ch == '(': depth += 1; cur += ch
        elif ch == ')': depth -= 1; cur += ch
        elif ch == ',' and depth == 0: args.append(cur); cur = ''
        else: cur += ch
    args.append(cur); return args

def transform_IF(e):
    while 'IF(' in e:
        i = e.rfind('IF('); j, d = i + 3, 1
        while d:
            if e[j] == '(': d += 1
            elif e[j] == ')': d -= 1
            j += 1
        a, b, c = _split(e[i+3:j-1]); e = e[:i] + f'(({b}) if ({a}) else ({c}))' + e[j:]
    return e

def translate(f):
    s = f[1:] if f.startswith('=') else f
    s = s.replace('""', 'None'); s = REF.sub(lambda m: f'cv("{m.group()}")', s)
    s = s.replace('<>', '!='); s = re.sub(r'(?<![<>=!])=(?!=)', '==', s)
    return transform_IF(s)


srcdo = openpyxl.load_workbook(OUT, data_only=True)     # cached signal values (OU/signal recalc'd? no)
srcf = openpyxl.load_workbook(OUT, data_only=False)
# NOTE: OUT was written by openpyxl so formula cells have NO cache; signal values must come
# from the ORIGINAL source workbook instead.
SRC = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/2323ed8f-TradingExcel_s1_laddering.xlsx'
srcdo = openpyxl.load_workbook(SRC, data_only=True)


def eval_sheet(sheet):
    do = srcdo[sheet]; fo = srcf[sheet]
    cache = {}
    for row in do.iter_rows():
        for c in row:
            if c.value is not None and c.value != '':
                cache[c.coordinate] = to_serial(c.value)
    for r in range(2, 7):
        cache[f'{CFG_VAL_COL}{r}'] = fo[f'{CFG_VAL_COL}{r}'].value
    grid = {}
    def cv(ref):
        ref = ref.replace('$', ''); col = re.match(r'[A-Z]+', ref).group()
        return grid.get(ref) if col in LAD_COLS else cache.get(ref)
    ns = {'cv': cv, 'IF': IF, 'AND': AND, 'OR': OR, 'ISNUMBER': ISNUMBER, 'MIN': MIN, 'MAX': MAX, 'None': None}
    fc = {}
    for n in NAMES:
        for r in range(8, 868):
            v = fo[f'{COL[n]}{r}'].value
            if isinstance(v, str) and v.startswith('='): fc[(n, r)] = translate(v)
            else: grid[f'{COL[n]}{r}'] = 0 if v is None else v
    for r in range(8, 868):
        for n in EVAL_ORDER:
            e = fc.get((n, r))
            if e is not None: grid[f'{COL[n]}{r}'] = eval(e, ns)
    buys = sum(grid.get(f'{COL["NB"]}{i}', 0) or 0 for i in range(9, 868))   # Bayes rung-fills
    # Bayes terminal marked-to-market equity at the last data row (what Y4/Y5 LOOKUP grabs)
    last = max(r for r in range(8, 868) if cache.get(f'B{r}') is not None)
    bayes_eq = grid.get(f'{COL["EQ"]}{last}', 0) or 0
    return buys, bayes_eq


import json
from multi_stock import params_for
_PJ = json.load(open('params_all.json'))

def engine_full(sheet):
    q = srcdo['Query']; s = sheet.split()[1]; i = STOCKS.index(s)
    dts, O, H, L, C = [], [], [], [], []
    for row in q.iter_rows(min_row=2, values_only=True):
        d = to_date(row[0]); o = row[1 + 4*i]
        if d is None or not isinstance(o, (int, float)) or o <= 0: continue
        dts.append(d); O.append(o); H.append(row[2+4*i]); L.append(row[3+4*i]); C.append(row[4+4*i])
    p = params_for(s, _PJ)                                   # each stock's own parameters
    return run_ladder(dts, O, H, L, C, p, [p.k, 1.3*p.k, 1.7*p.k], [p.ou_buf_k], 'first', [0.80, 0.15, 0.05], None)


if __name__ == '__main__':
    print(f'{"stock":6s}{"Bayes buys f/e":>16s}{"Bayes equity f":>16s}{"Bayes equity e":>16s}   ok')
    allok = True
    for s in STOCKS:
        sheet = f'Model {s}'
        fb, feq = eval_sheet(sheet)
        r = engine_full(sheet)
        eeq = r['eqB'][-1]
        ok = fb == r['bayes_trades'] and abs(feq - eeq) < 1e-2
        allok = allok and ok
        print(f'{s:6s}{str(fb)+"/"+str(r["bayes_trades"]):>16s}{feq:>16,.2f}{eeq:>16,.2f}   '
              f'{"OK" if ok else "MISMATCH"}')
    print('\nRESULT:', 'ALL TEN VERIFIED — laddered Bayes buys + terminal equity match the engine'
          if allok else 'MISMATCHES PRESENT')
