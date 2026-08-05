"""
GM, VLO, CF and ALNY as diversifiers, against the current five-name daily book.

The protocol is the strict one that decided the book, not the looser screen the first diversifier
shortlist got:

    fit on the FIRST half only  ->  freeze  ->  score the tested half

with the boundary at 2025-05-23. Nothing that scores the tested half has seen it. A full-sample
fit is also reported, purely to show the gap between what fitting produces and what survives --
the first shortlist averaged 27% fitted and 4% tested, and PLTR earned 365% on the untested half
against 3.1% on the tested one. That gap is the whole reason this test exists.

Three things are better here than in the earlier screen, all of which make the comparison fairer
rather than kinder:

  VERIFIED FILLS. These four come with 5-minute bars, so a same-day round trip is kept only where
  the low provably preceded the high. MRNA, OXY, FSLR and DVN had no intraday coverage and were
  scored on the at-open floor, a hard lower bound. Like-for-like against the book at last.

  THE DEPLOYED CONFIGURATION. Residual OU sigma and a 50/50 sleeve split, which is what the book
  runs now. The earlier screen used the mis-specified level sigma at a 75% Bayes tilt. ou_buf_k is
  inside the fitted vector, so it lands on the residual scale by construction.

  THE MARGINAL TEST. Clearing 50% is not the bar. The bar is whether the book is better with the
  name than without it: at equal weight a sixth name takes a sixth of the capital, so it has to
  beat what it dilutes. Return, drawdown and Sharpe are reported for the five-name book and for
  each candidate six-name book over the tested half, using the candidate's frozen first-half
  parameters -- never its full-sample fit.
"""
import collections
import copy
import datetime
import math
import os
import pickle
import statistics
import sys

import openpyxl
from scipy.optimize import differential_evolution

from engine import Params, run_model
from optimise_candidates import BOUNDS, POLICY, PERTURB, NAMES, mp, bvec
import five_min

BASE = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff'
CAND = {
    'GM':   f'{BASE}/845d1a52-GM_5min_Apr2024Aug2026.xlsx',
    'VLO':  f'{BASE}/7ebcbcf6-VLO_5min_Apr2024Aug2026.xlsx',
    'CF':   f'{BASE}/e0771e4b-CF_5min_Apr2024Aug2026.xlsx',
    'ALNY': f'{BASE}/f9813c50-ALNY_5min_Apr2024Aug2026.xlsx',
}
for k, v in CAND.items():
    five_min.FILES.setdefault(k, v)

BOOK = ['RKLB', 'TSM', 'VST', 'VRT', 'MU']
BUF_RESID = {'RKLB': 0.25, 'TSM': 0.20, 'VST': 0.65, 'VRT': 0.40, 'MU': 0.75}
CUT = datetime.date(2025, 5, 23)
SEED = Params(capital=6_000_000, comm=0.005, interest=0.0314, stop_days=50,
              bayes_pct=0.5, years=2.2)
DAILY_CACHE = '/home/user/David-Fewer/premarket_study/newdiv_daily.pkl'
RTH0, RTH1 = datetime.time(9, 30), datetime.time(16, 0)


# ---------------------------------------------------------------- data
def daily_from_5min(stock):
    """Aggregate regular-hours 5-minute bars into daily OHLC.

    The supplied files carry pre- and post-market prints -- GM's first row is 04:00 -- which have
    to go: the model trades the session, its open is the 09:30 print, and a 04:00 bar would
    corrupt both the open and the range.
    """
    cache = pickle.load(open(DAILY_CACHE, 'rb')) if os.path.exists(DAILY_CACHE) else {}
    if stock in cache:
        return cache[stock]
    wb = openpyxl.load_workbook(CAND[stock], read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        dt, o, h, l, c = row[0], row[1], row[2], row[3], row[4]
        if not isinstance(dt, datetime.datetime) or o is None:
            continue
        if not (RTH0 <= dt.time() < RTH1):
            continue
        by[dt.date()].append((dt, o, h, l, c))
    wb.close()
    dts, O, H, L, C = [], [], [], [], []
    for d in sorted(by):
        bars = sorted(by[d], key=lambda x: x[0])
        dts.append(d)
        O.append(bars[0][1])
        H.append(max(b[2] for b in bars))
        L.append(min(b[3] for b in bars))
        C.append(bars[-1][4])
    out = (dts, O, H, L, C)
    cache[stock] = out
    pickle.dump(cache, open(DAILY_CACHE, 'wb'))
    return out


# ---------------------------------------------------------------- engine wrappers
def run(data, vec, t, sde):
    return run_model(data['dts'], data['O'], data['H'], data['L'], data['C'], mp(vec, t),
                     collect=True, same_day_exit=sde, ou_sigma='resid')


def seg(data, vec, t, lo, hi, sde):
    fr = run(data, vec, t, sde).frames
    buys = sum(fr['t1']['Z'][lo:hi+1]) + sum(fr['t2']['Z'][lo:hi+1])
    eq = fr['equity']
    return (eq[hi]/eq[lo] - 1.0 if eq[lo] > 0 else -1.0), buys


def robust(data, vec, t, lo, hi, floor, sde):
    def one(v):
        rr, b = seg(data, v, t, lo, hi, sde)
        return -5.0 + b*1e-3 if b < floor else rr
    base = one(vec)
    s = []
    for i in POLICY:
        for f in PERTURB:
            v = list(vec)
            v[i] = min(max(v[i]*f, BOUNDS[i][0]), BOUNDS[i][1])
            s.append(one(v))
    return 0.5*base + 0.5*sum(s)/len(s)


def optimise(data, t, lo, hi, floor, sde, maxiter=12, popsize=10, seed=42):
    res = differential_evolution(lambda v: -robust(data, v, t, lo, hi, floor, sde), BOUNDS,
                                 x0=bvec(t), init='sobol', seed=seed, maxiter=maxiter,
                                 popsize=popsize, mutation=(0.5, 1.0), recombination=0.7,
                                 tol=1e-3, polish=False, disp=False, updating='immediate',
                                 workers=1)
    return list(res.x)


def ann(r, dts, lo, hi):
    y = max((dts[hi] - dts[lo]).days/365.25, 1e-6)
    return (1+r)**(1/y) - 1 if r > -1 else -1.0


def stats(eq, lo, hi):
    """Annualised return, max drawdown and Sharpe over a slice of an equity curve."""
    seg_ = eq[lo:hi+1]
    peak, dd = seg_[0], 0.0
    for v in seg_:
        peak = max(peak, v)
        if peak > 0:
            dd = max(dd, 1 - v/peak)
    rs = [seg_[i]/seg_[i-1] - 1 for i in range(1, len(seg_)) if seg_[i-1] > 0]
    sd = statistics.pstdev(rs) if len(rs) > 2 else 0.0
    sh = (statistics.mean(rs)/sd*math.sqrt(252)) if sd > 0 else float('nan')
    return dd*100, sh


def corr(a, b):
    n = len(a)
    if n < 3:
        return 0.0
    ma, mb = statistics.mean(a), statistics.mean(b)
    cov = sum((a[i]-ma)*(b[i]-mb) for i in range(n))
    va = sum((x-ma)**2 for x in a)
    vb = sum((x-mb)**2 for x in b)
    return cov/math.sqrt(va*vb) if va > 0 and vb > 0 else 0.0


# ---------------------------------------------------------------- the book as it runs now
def book_curves():
    """Each daily name's equity curve on the deployed configuration: resid sigma, 50/50."""
    from daily_window_split import data as bdata, params as bparams
    from five_min import make_checker as fm
    from mu_rerun import from_workbook
    bdata = dict(bdata)
    bdata['MU'] = from_workbook()
    out = {}
    for s in BOOK:
        dts, O, H, L, C = bdata[s]
        chk = fm(s, dts, O)[0]
        p = copy.copy(bparams[s])
        p.bayes_pct = 0.5
        p.ou_buf_k = BUF_RESID[s]
        eq = run_model(dts, O, H, L, C, p, collect=True, same_day_exit=chk,
                       ou_sigma='resid').frames['equity']
        out[s] = (dts, eq)
    return out


def daily_returns(dts, eq):
    return {dts[i]: eq[i]/eq[i-1] - 1 for i in range(1, len(eq)) if eq[i-1] > 0}


def equal_weight(curves, lo_date, hi_date):
    """Rebalanced-daily equal-weight curve over the union of dates in [lo_date, hi_date]."""
    rets = [daily_returns(d, e) for d, e in curves]
    dates = sorted(set().union(*[set(r) for r in rets]))
    dates = [d for d in dates if lo_date <= d <= hi_date]
    eq, v = [1.0], 1.0
    for d in dates:
        vals = [r[d] for r in rets if d in r]
        v *= (1 + (statistics.mean(vals) if vals else 0.0))
        eq.append(v)
    return dates, eq


if __name__ == '__main__':
    names = [a for a in sys.argv[1:] if a in CAND] or list(CAND)

    print('=== data ===', flush=True)
    feed = {}
    for t in names:
        dts, O, H, L, C = daily_from_5min(t)
        feed[t] = (dts, O, H, L, C)
        rng = statistics.mean((H[i]-L[i])/C[i] for i in range(len(C)))*100
        print(f'  {t:5s} {len(C):>4d} sessions  {dts[0]} to {dts[-1]}  '
              f'avg daily range {rng:.2f}%  close {C[0]:.2f} -> {C[-1]:.2f} '
              f'({(C[-1]/C[0]-1)*100:+.0f}%)', flush=True)

    bc = book_curves()
    bk_dates, bk_eq = equal_weight([bc[s] for s in BOOK], datetime.date(2000, 1, 1),
                                   datetime.date(2100, 1, 1))
    bkret = daily_returns(bk_dates, bk_eq[1:])

    print(f'\n=== half-sample split, boundary {CUT}: fit the first half, score the second ==='
          , flush=True)
    print(f'{"name":6s}{"fit 1st half":>14s}{"TESTED 2nd":>13s}{"full fit":>11s}'
          f'{"buys/yr":>9s}{"tested DD":>11s}{"corr book":>11s}', flush=True)
    print('-'*76, flush=True)
    rows = []
    for t in names:
        dts, O, H, L, C = feed[t]
        N = len(C)
        k = next(i for i, d in enumerate(dts) if d >= CUT)
        tmpl = copy.copy(SEED)
        tmpl.years = (dts[-1]-dts[0]).days/365.25
        d = dict(dts=dts, O=O, H=H, L=L, C=C)
        chk = five_min.make_checker(t, dts, O)[0]

        th = optimise(d, tmpl, 0, k-1, max(8, int(0.03*k)), chk)
        r1, _ = seg(d, th, tmpl, 0, k-1, chk)
        r2, b2 = seg(d, th, tmpl, k, N-1, chk)
        thf = optimise(d, tmpl, 0, N-1, max(8, int(0.03*N)), chk)
        rf, _ = seg(d, thf, tmpl, 0, N-1, chk)

        eq = run(d, th, tmpl, chk).frames['equity']
        dd, _ = stats(eq, k, N-1)
        pa, pb = [], []
        for i in range(1, N):
            if dts[i] in bkret and eq[i-1] > 0:
                pa.append(eq[i]/eq[i-1]-1)
                pb.append(bkret[dts[i]])
        c = corr(pa, pb)
        a1, a2 = ann(r1, dts, 0, k-1)*100, ann(r2, dts, k, N-1)*100
        af = ann(rf, dts, 0, N-1)*100
        yrs = max((dts[N-1]-dts[k]).days/365.25, 1e-6)
        rows.append(dict(t=t, a1=a1, a2=a2, af=af, c=c, th=th, dd=dd,
                         curve=(dts, eq), buys=b2/yrs))
        print(f'{t:6s}{a1:>13.1f}%{a2:>12.1f}%{af:>10.1f}%{b2/yrs:>9.0f}'
              f'{dd:>10.1f}%{c:>11.2f}', flush=True)
    print('-'*76, flush=True)

    print('\n  for reference, the book on the same tested half:', flush=True)
    lo = next(i for i, dd_ in enumerate(bk_dates) if dd_ >= CUT)
    bdd, bsh = stats(bk_eq[1:], lo, len(bk_dates)-1)
    br = ann(bk_eq[1:][len(bk_dates)-1]/bk_eq[1:][lo] - 1, bk_dates, lo, len(bk_dates)-1)*100
    print(f'    five-name book  {br:.1f}%  max DD {bdd:.1f}%  Sharpe {bsh:.2f}', flush=True)
    for s in BOOK:
        dts, eq = bc[s]
        j = next(i for i, dd_ in enumerate(dts) if dd_ >= CUT)
        print(f'      {s:5s} tested half {ann(eq[-1]/eq[j]-1, dts, j, len(dts)-1)*100:>7.1f}%',
              flush=True)

    print('\n=== the marginal test: is the book better with the name than without it? ===',
          flush=True)
    print('Equal weight, tested half only, candidate on its FROZEN first-half parameters.\n',
          flush=True)
    print(f'{"book":22s}{"return":>10s}{"max DD":>10s}{"Sharpe":>9s}{"vs 5-name":>12s}',
          flush=True)
    print('-'*63, flush=True)
    print(f'{"five names":22s}{br:>9.1f}%{bdd:>9.1f}%{bsh:>9.2f}{"--":>12s}', flush=True)
    for r in rows:
        cur = [bc[s] for s in BOOK] + [r['curve']]
        ds, eq6 = equal_weight(cur, CUT, datetime.date(2100, 1, 1))
        a = ann(eq6[-1]/eq6[0]-1, ds, 0, len(ds)-1)*100
        dd6, sh6 = stats(eq6, 0, len(eq6)-1)
        print(f'{"+ " + r["t"]:22s}{a:>9.1f}%{dd6:>9.1f}%{sh6:>9.2f}{a-br:>+11.1f}pp',
              flush=True)

    print('\nparameters fitted on the first half (the frozen set scored above):', flush=True)
    for r in rows:
        print(f'  {r["t"]:5s}' + '  '.join(f'{n}={v:.4g}' for n, v in zip(NAMES, r['th'])),
              flush=True)
    print('DONE', flush=True)
