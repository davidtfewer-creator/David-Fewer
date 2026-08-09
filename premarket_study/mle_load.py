"""
Data loaders for the rolling-MLE drift study. Two sources:

  load_workbook_ohlc(path)  the 5-stock live workbook: Query sheet, Date in column A,
                            then 4 columns per stock headed NAME_O / _H / _L / _C.
                            Headers are read, not assumed, so column order is free.
  load_csv_dir(path)        NAME.csv per stock with Date,Open,High,Low,Close.

Both return {name: (dates, O, H, L, C)} with rows filtered to valid positive opens,
matching the convention in newcands.load / stop_sweep.load_book.
"""
import csv
import datetime
import os


def _to_date(d):
    if isinstance(d, datetime.datetime):
        return d.date()
    if isinstance(d, datetime.date):
        return d
    if isinstance(d, (int, float)):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(d))
    return datetime.date.fromisoformat(str(d)[:10])


def load_workbook_ohlc(path, sheet='Query'):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    q = wb[sheet]
    rows = list(q.iter_rows(values_only=True))
    wb.close()
    head = rows[0]
    stocks = {}
    for c, h in enumerate(head):
        if isinstance(h, str) and h.endswith('_O'):
            stocks[h[:-2]] = c
    data = {}
    for s, c0 in stocks.items():
        dts, O, H, L, C = [], [], [], [], []
        for row in rows[1:]:
            if row[0] is None:
                continue
            o = row[c0]
            if not isinstance(o, (int, float)) or o <= 0:
                continue
            dts.append(_to_date(row[0]))
            O.append(float(o))
            H.append(float(row[c0 + 1]))
            L.append(float(row[c0 + 2]))
            C.append(float(row[c0 + 3]))
        data[s] = (dts, O, H, L, C)
    return data


def load_csv_dir(path):
    data = {}
    for fn in sorted(os.listdir(path)):
        if not fn.endswith('.csv'):
            continue
        name = fn[:-4].upper()
        dts, O, H, L, C = [], [], [], [], []
        with open(os.path.join(path, fn)) as fh:
            for row in csv.DictReader(fh):
                keys = {k.lower().strip(): k for k in row}
                try:
                    o = float(row[keys['open']])
                except (KeyError, ValueError):
                    continue
                if o <= 0:
                    continue
                dts.append(_to_date(row[keys['date']]))
                O.append(o)
                H.append(float(row[keys['high']]))
                L.append(float(row[keys['low']]))
                C.append(float(row[keys['close']]))
        if dts:
            data[name] = (dts, O, H, L, C)
    return data
