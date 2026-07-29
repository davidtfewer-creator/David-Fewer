"""
Add the 3-rung live-order plumbing to the laddered backtest workbook so limit prices can be
placed live: Dashboard rung2/rung3 bid columns + weight config, and an expanded Active Trading
TODAY'S ORDERS block showing, per stock's Bayes tranche, three rung buy prices + weighted share
sizes (0.80/0.15/0.05) sharing one first-rung sell target. OU rows unchanged.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

SRC = '/home/user/David-Fewer/TradingExcel_s1_laddering_OptionC_backtest.xlsx'
OUT = '/home/user/David-Fewer/TradingExcel_s1_laddering_OptionC_full.xlsx'
BLUE = 'FFD9E1F2'; HDR = Font(name='Arial', size=10, bold=True)
STOCKS = ['NVDA', 'TSM', 'TSLA', 'VRT', 'VST', 'AVGO', 'PLTR', 'RKLB', 'SOFI', 'SPOT']

wb = openpyxl.load_workbook(SRC, data_only=False)

# ---------- Dashboard: rung2/rung3 bid columns + weight/depth config ----------
d = wb['Dashboard']
d['U1'] = 'LADDER C';                     d['U1'].font = HDR
for r, (lab, val) in {2: ('m2 (xk)', 1.3), 3: ('m3 (xk)', 1.7),
                      4: ('w1', 0.80), 5: ('w2', 0.15), 6: ('w3', 0.05)}.items():
    d[f'U{r}'] = lab
    c = d[f'V{r}']; c.value = val; c.fill = PatternFill('solid', fgColor=BLUE)
    c.number_format = '0.00' if r < 4 else '0%'
d['R3'] = 'Rung2 bid'; d['S3'] = 'Rung3 bid'
d['R3'].font = HDR; d['S3'].font = HDR
for row in range(4, 14):                                   # rung1 = existing C; add R=rung2, S=rung3
    c1 = d.cell(row=row, column=3).value                   # C{row} formula (rung1)
    d.cell(row=row, column=18, value=c1.replace(f'J{row}-', f'J{row}-$V$2*', 1))   # R = rung2
    d.cell(row=row, column=19, value=c1.replace(f'J{row}-', f'J{row}-$V$3*', 1))   # S = rung3
    for col in (18, 19):
        d.cell(row=row, column=col).number_format = '"$"#,##0.00'

# ---------- Active Trading: 3-rung Bayes orders ----------
at = wb['Active Trading']
for col, lab in {'K': 'Rung2 @', 'L': 'Rung2 sh', 'M': 'Rung3 @', 'N': 'Rung3 sh'}.items():
    at[f'{col}18'] = lab; at[f'{col}18'].font = HDR
at['A17'] = ('TODAY’S ORDERS  —  Bayes ladders into 3 rungs (F/H = rung1 @0.80, K/L = rung2 @0.15, '
             'M/N = rung3 @0.05); place all three as GTC limit BUYs and one GTC SELL at the shared '
             'target (col G). OU unchanged.')
for i in range(10):
    r = 19 + 2*i                                           # Bayes rows: 19,21,...,37
    dr = 4 + i                                             # Dashboard row for this stock
    deploy = (f'MIN(Allocation!$B$5*$D{r}/SUMIF($C$19:$C$38,"CASH",$D$19:$D$38),$D{r})')
    hold = (f'SUMIFS($F$42:$F$1041,$B$42:$B$1041,$A{r},$C$42:$C$1041,$B{r},$M$42:$M$1041,"OPEN")')
    # rung1 shares now weighted (0.80); price stays F = Dashboard!C{dr}
    at[f'H{r}'] = (f'=IF($C{r}="CASH",IFERROR(FLOOR(Dashboard!$V$4*{deploy}/($F{r}+$B$3),1),0),{hold})')
    at[f'K{r}'] = f'=IF($C{r}="CASH",Dashboard!R{dr},"")'
    at[f'L{r}'] = f'=IF($C{r}="CASH",IFERROR(FLOOR(Dashboard!$V$5*{deploy}/($K{r}+$B$3),1),0),"")'
    at[f'M{r}'] = f'=IF($C{r}="CASH",Dashboard!S{dr},"")'
    at[f'N{r}'] = f'=IF($C{r}="CASH",IFERROR(FLOOR(Dashboard!$V$6*{deploy}/($M{r}+$B$3),1),0),"")'

# expand the orders named range to include the rung columns
try:
    del wb.defined_names['IBKR_Orders']
except KeyError:
    pass
wb.defined_names.add(DefinedName('IBKR_Orders', attr_text="'Active Trading'!$A$18:$N$38"))

wb.calculation.calcMode = 'auto'; wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print('saved', OUT)
