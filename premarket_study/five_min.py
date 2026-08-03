"""
General 5-minute fill verifier for the book.

Loads each name's 5-minute bars (regular hours only, 09:30-16:00 ET -- the supplied files also
carry pre/post-market bars which must be excluded, since the model trades the session), builds a
per-session index of (lows, suffix-max-highs), and exposes a checker f(i, bid, target) that
decides whether a same-day exit was genuinely achievable:
    first bar whose LOW <= bid  ->  was the target reached at/after that bar?
Earlier calibration on NVDA showed 5-minute bars reproduce the 1-minute answer exactly.
"""
import openpyxl, datetime, collections, pickle, os
import numpy as np

BASE = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff'
FILES = {
    'RKLB': f'{BASE}/8c607cec-RKLB_5min_Apr2024Jun2026.xlsx',
    'AVGO': f'{BASE}/f20b9007-AVGO_5min_Apr2024Jun2026.xlsx',
    'SOFI': f'{BASE}/9c2cff23-SOFI_5min_Apr2024Jun2026.xlsx',
    'SPOT': f'{BASE}/35247e40-SPOT_5min_Apr2024Jun2026.xlsx',
    'VRT':  f'{BASE}/976bd5af-VRT_5min_Apr2024Jun2026.xlsx',
    'TSM':  f'{BASE}/bbf336cd-TSM_5min_Apr2024Jun2026.xlsx',
    'VST':  f'{BASE}/ec3a3432-VST_5min_Apr2024Jun2026.xlsx',
}
CACHE = '/home/user/David-Fewer/premarket_study/five_min_index.pkl'
RTH_START = datetime.time(9, 30)
RTH_END = datetime.time(16, 0)


def build_index(stock):
    """date -> (lows, suffix_max_high) for regular-hours 5-minute bars."""
    cache = {}
    if os.path.exists(CACHE):
        with open(CACHE, 'rb') as f:
            cache = pickle.load(f)
    if stock in cache:
        return cache[stock]
    wb = openpyxl.load_workbook(FILES[stock], read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False; continue
        dt, o, h, l = row[0], row[1], row[2], row[3]
        if dt is None or o is None or not isinstance(dt, datetime.datetime):
            continue
        t = dt.time()
        if t < RTH_START or t >= RTH_END:          # regular hours only
            continue
        by[dt.date()].append((dt, h, l))
    wb.close()
    idx = {}
    for d, rows in by.items():
        rows.sort(key=lambda x: x[0])
        highs = np.array([r[1] for r in rows], dtype=float)
        lows = np.array([r[2] for r in rows], dtype=float)
        idx[d] = (lows, np.maximum.accumulate(highs[::-1])[::-1])
    cache[stock] = idx
    with open(CACHE, 'wb') as f:
        pickle.dump(cache, f)
    return idx


def make_checker(stock, dates, O):
    idx = build_index(stock)
    def check(i, bid, target):
        ent = idx.get(dates[i])
        if ent is None:                             # no bars: allow only the provable at-open case
            return bid >= O[i] - 1e-9
        lows, suffix = ent
        hit = lows <= bid + 1e-9
        if not hit.any():
            return False
        j = int(np.argmax(hit))
        return bool(suffix[j] >= target - 1e-9)
    return check, idx


if __name__ == '__main__':
    from stop_sweep import load_book
    from engine import run_model
    data, params, cached = load_book()
    print('=== TRUE FILL RATES AND VERIFIED RETURNS (5-minute bars, regular hours) ===\n')
    print(f'{"stock":6s}{"sessions":>9s}{"cover":>7s}{"same-day real":>15s}'
          f'{"optimistic":>12s}{"VERIFIED":>10s}{"at-open":>9s}{"buys o/v":>11s}')
    print('-' * 79)
    for s in FILES:
        dts, O, H, L, C = data[s]
        chk, idx = make_checker(s, dts, O)
        cover = sum(1 for d in dts if d in idx)
        # same-day trades: how many verify as real
        r = run_model(dts, O, H, L, C, params[s], collect=True)
        real = fake = 0
        for tkey, bids in (('t1', r.frames['X']), ('t2', r.frames['AM'])):
            t = r.frames[tkey]
            for i in range(len(C)):
                if t['Z'][i] == 1 and t['AD'][i] == 1 and bids[i] is not None and dts[i] in idx:
                    if chk(i, bids[i], t['AB'][i]): real += 1
                    else: fake += 1
        rate = real/(real+fake)*100 if real+fake else 0
        ro = run_model(dts, O, H, L, C, params[s], same_day_exit=True)
        rv = run_model(dts, O, H, L, C, params[s], same_day_exit=chk)
        ra = run_model(dts, O, H, L, C, params[s], same_day_exit='at_open')
        print(f'{s:6s}{len(idx):>9d}{cover/len(dts)*100:>6.0f}%{f"{real}/{real+fake} = {rate:.0f}%":>15s}'
              f'{ro.annual_return*100:>11.0f}%{rv.annual_return*100:>9.0f}%{ra.annual_return*100:>8.0f}%'
              f'{f"{ro.total_buys}/{rv.total_buys}":>11s}')
