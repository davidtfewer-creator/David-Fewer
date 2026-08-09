"""
5-minute bar index for verified same-day fills, fed from local files in data_5min/
(extracted from the Box "minute data" folder — xlsx or csv, one file per name,
columns Datetime, Open, High, Low[, Close...]).

Same semantics as five_min.py, which was calibrated against 1-minute NVDA data:
regular hours only (09:30-16:00 ET); per session, the checker finds the first bar
whose low reaches the bid and asks whether any bar from then on reaches the target.
Sessions without bar coverage allow only the provable at-open case (bid >= open).
"""
import collections
import datetime
import os
import pickle

import numpy as np

DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data_5min')
CACHE = os.path.join(DIR, 'index_cache.pkl')
RTH_START = datetime.time(9, 30)
RTH_END = datetime.time(16, 0)


def _rows_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        yield row[0], row[2], row[3]      # dt, high, low
    wb.close()


def _rows_csv(path):
    import csv
    with open(path) as fh:
        rd = csv.DictReader(fh)
        keys = {k.lower().strip(): k for k in rd.fieldnames}
        kd = next(keys[k] for k in ('datetime', 'date', 'timestamp', 'time') if k in keys)
        kh, kl = keys['high'], keys['low']
        for row in rd:
            try:
                dt = datetime.datetime.fromisoformat(row[kd].replace('Z', ''))
                yield dt, float(row[kh]), float(row[kl])
            except (ValueError, KeyError):
                continue


def build_index(stock):
    cache = {}
    if os.path.exists(CACHE):
        with open(CACHE, 'rb') as f:
            cache = pickle.load(f)
    if stock in cache:
        return cache[stock]
    path = None
    for fn in sorted(os.listdir(DIR)):
        if fn.upper().startswith(stock.upper()) and fn[-5:].lower() in ('.xlsx',) or \
           fn.upper().startswith(stock.upper()) and fn[-4:].lower() == '.csv':
            path = os.path.join(DIR, fn)
            break
    if path is None:
        raise FileNotFoundError(f'no 5-minute file for {stock} in {DIR}')
    rows = _rows_xlsx(path) if path.endswith('.xlsx') else _rows_csv(path)
    by = collections.defaultdict(list)
    for dt, h, l in rows:
        if not isinstance(dt, datetime.datetime) or h is None or l is None:
            continue
        t = dt.time()
        if t < RTH_START or t >= RTH_END:
            continue
        by[dt.date()].append((dt, float(h), float(l)))
    idx = {}
    for d, bars in by.items():
        bars.sort(key=lambda x: x[0])
        highs = np.array([b[1] for b in bars])
        lows = np.array([b[2] for b in bars])
        idx[d] = (lows, np.maximum.accumulate(highs[::-1])[::-1])
    cache[stock] = idx
    with open(CACHE, 'wb') as f:
        pickle.dump(cache, f)
    return idx


def make_checker(stock, dates, O):
    idx = build_index(stock)
    def check(i, bid, target):
        ent = idx.get(dates[i])
        if ent is None:
            return bid >= O[i] - 1e-9
        lows, suffix = ent
        hit = lows <= bid + 1e-9
        if not hit.any():
            return False
        j = int(np.argmax(hit))
        return bool(suffix[j] >= target - 1e-9)
    return check


def coverage(stock, dates):
    idx = build_index(stock)
    return sum(1 for d in dates if d in idx), len(dates)
