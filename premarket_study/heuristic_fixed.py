"""
The old fixed-parameter mean-reversion heuristic (maths doc §2), modelled and optimised on
the nine-name daily book — with the phantom same-day trades it used to book now controlled.

When the heuristic was live there was no minute data, so the backtest recorded a round trip
whenever the day's low touched the bid AND the day's high touched the target, which daily
bars cannot order. Here every figure is quoted on three fill bases:
    sd=Y      same-day exits allowed   — the old, phantom-inclusive basis
    at-open   same-day exit only where the bid filled AT the open (the high provably came
              after the buy) — verifiable without minute data, a hard lower bound
    sd=N      no same-day exits at all — the absolute floor

Protocol (house conventions):
  * fit on the first half only (boundary 2025-05-23), freeze, score the second half
  * robust objective = 0.5·base + 0.5·mean(±3% on all five parameters), min-trade floor
  * differential evolution, sobol init, workers=1
  * fitting is done ON THE AT-OPEN BASIS so the optimiser cannot tune itself to phantoms
  * incumbent comparison: deployed Bayes+OU params from the 9-stock workbook, same data,
    same bases, same test window (deployed params are full-sample fits, so the incumbent
    carries a mild lookahead advantage on the tested half — stated, not corrected)
  * a full-sample fit is also reported as the in-sample ceiling, with its phantom share

Data/params come from the uploaded TradingExcel_9stock_live.xlsx via data_9stock.csv /
params_9stock.json (extraction kept in extract_9stock(); derived files are gitignored).
"""
import csv, datetime, json, math, os, sys

from scipy.optimize import differential_evolution

from engine import Params, run_model
from heuristic_engine import HeurParams, heuristic_bid, run_bid, run_heuristic

HERE = os.path.dirname(os.path.abspath(__file__))
CSV = os.path.join(HERE, 'data_9stock.csv')
PARAMS = os.path.join(HERE, 'params_9stock.json')
CUT = datetime.date(2025, 5, 23)          # the house half-sample boundary
NAMES = ['TSM', 'VRT', 'VST', 'RKLB', 'MU', 'GM', 'VLO', 'CF', 'MRVL']

HNAMES = ['w', 'a', 'rho', 'prem', 'cap']
BOUNDS = [(0.50, 1.10), (0.50, 1.10), (-30.0, 10.0), (0.005, 0.05), (0.002, 0.07)]
PERTURB = [0.97, 1.03]


def load(stock):
    with open(CSV) as f:
        rows = list(csv.DictReader(f))
    dts, O, H, L, C = [], [], [], [], []
    for r in rows:
        vals = [r[f'{stock}_{c}'] for c in 'OHLC']
        if any(v in ('', None) for v in vals):
            continue
        dts.append(datetime.date.fromisoformat(r['Date']))
        o, h, l, c = map(float, vals)
        O.append(o); H.append(h); L.append(l); C.append(c)
    with open(PARAMS) as f:
        d = json.load(f)[stock]
    p = Params(lam=d['lam'], phi_L=d['phi_L'], psi=d['psi'], k=d['k'],
               premium=d['premium'], peak_cap=d['peak_cap'], comm=d['comm'],
               capital=d['capital'], interest=d['interest'], stop_days=int(d['stop_days']),
               bayes_pct=d['bayes_pct'], ou_W=int(d['ou_W']), ou_buf_k=d['ou_buf_k'],
               ou_prem=d['ou_prem'], ou_cap=d['ou_cap'])
    return dts, O, H, L, C, p


def validate(stock='TSM'):
    """run_bid must reproduce engine.run_model's Bayes tranche to the penny."""
    dts, O, H, L, C, p = load(stock)
    for sd in (True, False, 'at_open'):
        import copy
        p1 = copy.copy(p); p1.bayes_pct = 1.0
        ref = run_model(dts, O, H, L, C, p1, collect=True, same_day_exit=sd)
        mine = run_bid(dts, O, H, L, C, ref.frames['X'], p.premium, p, same_day_exit=sd)
        assert abs(mine.fundY_final - ref.fundY_final) < 1e-6, (sd, mine.fundY_final, ref.fundY_final)
        assert mine.bayes_buys == ref.bayes_buys
    print(f'validated: run_bid == engine Bayes tranche on {stock} (all three fill bases)')


def hp(vec, base: Params):
    return HeurParams(w=vec[0], a=vec[1], rho=vec[2], prem=vec[3], cap=vec[4],
                      comm=base.comm, capital=base.capital, interest=base.interest,
                      stop_days=base.stop_days)


def seg_ann(eq, dts, lo, hi):
    if eq[lo] <= 0:
        return float('nan')
    yrs = max((dts[hi] - dts[lo]).days, 1) / 365.25
    return (eq[hi] / eq[lo]) ** (1 / yrs) - 1


def hseg(data, vec, lo, hi, sd):
    dts, O, H, L, C, p = data
    r = run_heuristic(dts, O, H, L, C, hp(vec, p), same_day_exit=sd)
    eq = r.frames['equity']
    buys = sum(r.frames['Z'][lo:hi + 1])
    return eq[hi] / eq[lo] - 1.0 if eq[lo] > 0 else -1.0, buys, r


def robust(data, vec, lo, hi, floor, sd='at_open'):
    def one(v):
        rr, b, _ = hseg(data, v, lo, hi, sd)
        return -5.0 + b * 1e-3 if b < floor else rr
    base = one(vec)
    s = []
    for i in range(5):
        for f in PERTURB:
            v = list(vec)
            v[i] = min(max(v[i] * f, BOUNDS[i][0]), BOUNDS[i][1])
            s.append(one(v))
    return 0.5 * base + 0.5 * sum(s) / len(s)


def optimise(data, lo, hi, floor, maxiter=35, popsize=10, seed=42):
    neg = lambda v: -robust(data, v, lo, hi, floor)
    res = differential_evolution(neg, BOUNDS, init='sobol', seed=seed, maxiter=maxiter,
                                 popsize=popsize, mutation=(0.5, 1.0), recombination=0.7,
                                 tol=1e-4, polish=False, disp=False,
                                 updating='immediate', workers=1)
    return list(res.x), -res.fun


def incumbent(data, lo, hi):
    dts, O, H, L, C, p = data
    out = {}
    for sd, tag in ((True, 'sdY'), ('at_open', 'open'), (False, 'sdN')):
        r = run_model(dts, O, H, L, C, p, collect=True, same_day_exit=sd)
        eq = r.frames['equity']
        buys = sum(r.frames['t1']['Z'][lo:hi + 1]) + sum(r.frames['t2']['Z'][lo:hi + 1])
        out[tag] = (seg_ann(eq, dts, lo, hi), buys, eq)
    return out


def corr(x, y):
    n = min(len(x), len(y))
    x, y = x[:n], y[:n]
    mx = sum(x) / n; my = sum(y) / n
    sx = math.sqrt(sum((v - mx) ** 2 for v in x))
    sy = math.sqrt(sum((v - my) ** 2 for v in y))
    if sx == 0 or sy == 0:
        return float('nan')
    return sum((x[i] - mx) * (y[i] - my) for i in range(n)) / (sx * sy)


def daily_rets(eq, lo, hi):
    return [eq[i] / eq[i - 1] - 1 for i in range(lo + 1, hi + 1) if eq[i - 1] > 0]


def run_stock(stock):
    dts, O, H, L, C, p = load(stock)
    data = (dts, O, H, L, C, p)
    N = len(dts)
    cut = next(i for i, d in enumerate(dts) if d >= CUT)
    print(f'\n===== {stock}  (N={N}, fit [0:{cut}], test [{cut}:{N - 1}]) =====', flush=True)

    inc_te = incumbent(data, cut, N - 1)
    inc_buys_tr = incumbent(data, 0, cut)['open'][1]
    floor = max(8, int(0.20 * inc_buys_tr))

    vec, obj = optimise(data, 0, cut, floor)
    print('  fitted (first half, at-open basis):',
          {n: round(v, 4) for n, v in zip(HNAMES, vec)}, f'robust obj {obj * 100:.1f}%', flush=True)

    row = dict(stock=stock, vec=vec)
    for tag, sd in (('sdY', True), ('open', 'at_open'), ('sdN', False)):
        rtr, btr, _ = hseg(data, vec, 0, cut, sd)
        rte, bte, res = hseg(data, vec, cut, N - 1, sd)
        eq = res.frames['equity']
        a_tr = seg_ann(eq, dts, 0, cut)
        a_te = seg_ann(eq, dts, cut, N - 1)
        row[tag] = (a_tr, a_te, btr, bte)
        if tag == 'open':
            row['te_eq'] = eq
            row['stops'] = res.frames['stops']
    inc = {t: inc_te[t][0] for t in inc_te}
    row['inc'] = inc
    row['inc_buys'] = inc_te['open'][1]

    print(f'  {"basis":8s}{"fit-half ann":>14s}{"TESTED ann":>12s}{"buys te":>9s}'
          f'{"incumbent te":>14s}', flush=True)
    for tag, lbl in (('sdY', 'sd=Y'), ('open', 'at-open'), ('sdN', 'sd=N')):
        a_tr, a_te, btr, bte = row[tag]
        print(f'  {lbl:8s}{a_tr * 100:>13.1f}%{a_te * 100:>11.1f}%{bte:>9d}'
              f'{inc[tag] * 100:>13.1f}%', flush=True)

    hr = daily_rets(row['te_eq'], cut, N - 1)
    ir = daily_rets(inc_te['open'][2], cut, N - 1)
    row['corr'] = corr(hr, ir)
    print(f'  tested-half corr(heuristic, Bayes+OU) = {row["corr"]:.2f}   '
          f'stops {row["stops"]}   floor {floor}', flush=True)

    vec_f, _ = optimise(data, 0, N - 1, max(8, int(0.20 * row['inc_buys'] + 0.20 * inc_buys_tr)))
    _, _, r_open = hseg(data, vec_f, 0, N - 1, 'at_open')
    _, _, r_sdy = hseg(data, vec_f, 0, N - 1, True)
    row['full'] = (r_open.annual_return, r_sdy.annual_return)
    print(f'  full-sample fit (in-sample ceiling): at-open {r_open.annual_return * 100:.1f}%  '
          f'sd=Y {r_sdy.annual_return * 100:.1f}%  '
          f'params {[round(v, 3) for v in vec_f]}', flush=True)
    return row


if __name__ == '__main__':
    validate()
    only = sys.argv[1:] or NAMES
    rows = [run_stock(s) for s in only]
    print('\n===== SUMMARY (tested half, annualised) =====', flush=True)
    print(f'{"stock":6s}{"heur sd=Y":>10s}{"heur open":>10s}{"heur sd=N":>10s}'
          f'{"phantom pp":>11s}{"inc open":>10s}{"gap open":>10s}{"corr":>6s}', flush=True)
    for r in rows:
        ph = (r['sdY'][1] - r['open'][1]) * 100
        gap = (r['open'][1] - r['inc']['open']) * 100
        print(f'{r["stock"]:6s}{r["sdY"][1] * 100:>9.1f}%{r["open"][1] * 100:>9.1f}%'
              f'{r["sdN"][1] * 100:>9.1f}%{ph:>10.1f}{r["inc"]["open"] * 100:>9.1f}%'
              f'{gap:>+9.1f}{r["corr"]:>6.2f}', flush=True)


def extract_9stock(xlsx_path):
    """Regenerate data_9stock.csv / params_9stock.json from the uploaded workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    params = {}
    for s in NAMES:
        ws = wb[f'Model {s}']
        r2, r3 = list(ws.iter_rows(min_row=2, max_row=3, max_col=22, values_only=True))
        params[s] = dict(lam=r2[1], phi_L=r2[3], psi=r2[5], k=r2[7], premium=r2[9],
                         peak_cap=r2[11], comm=r2[13], capital=r2[15], interest=r2[17],
                         stop_days=r2[19], bayes_pct=r2[21], ou_W=r3[1], ou_buf_k=r3[3],
                         ou_prem=r3[7], ou_cap=r3[9])
    ws = wb['Query']
    rows = list(ws.iter_rows(values_only=True))
    epoch = datetime.date(1899, 12, 30)
    with open(CSV, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['Date'] + list(rows[0][1:37]))
        for r in rows[1:]:
            if r[0] is None:
                continue
            d = r[0].date() if isinstance(r[0], datetime.datetime) else \
                epoch + datetime.timedelta(days=int(r[0]))
            w.writerow([d.isoformat()] + list(r[1:37]))
    with open(PARAMS, 'w') as f:
        json.dump(params, f, indent=1)
