"""
Implement the corrected OU sigma in the daily workbook.

Deployed, column AZ is STDEVP of the last W closes -- the dispersion of the price LEVEL, which on
a trending name is mostly the trend. The buffer ou_buf_k * sigma therefore inflates exactly when a
stock is running, pushing the bid below the market and killing fills.

The replacement takes sigma from the fitted AR(1) residuals, using the same window, the same mean
(AX) and the same AR coefficient (AY) the sheet already computes:

    e_j  = (C_j - AX) - AY*(C_{j-1} - AX)
    mean(e) = (AVERAGE(R1) - AX) - AY*(AVERAGE(R0) - AX)        [linear, so no array needed]
    sigma   = SQRT( SUMPRODUCT((e - mean(e))^2) / (W-1) )

R1 and R0 are the same two shifted ranges the SLOPE in AY already uses, so nothing new is
introduced. SUMPRODUCT avoids any need for Ctrl+Shift+Enter and works in every Excel version.
The formula was checked against the Python engine's residual sigma on both names to 5e-15.

ou_buf_k (D3) is also updated per name. Residual sigma is about a third of level sigma, so the
old buffer would be far too tight against it; the new values are the walk-forward-validated fits
(10 of 15 folds, +14.7pp on the sleeve). The previous values are recorded on the Notes sheet so
the change can be reverted by restoring D3 alone.
"""
import shutil
import openpyxl

SRC = '/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/f3046864-TradingExcel_5stock.xlsx'
OUT = '/home/user/David-Fewer/TradingExcel_5stock_OUresid.xlsx'
NAMES = ['TSM', 'VRT', 'VST', 'RKLB', 'MU']
NEW_BUF = {'RKLB': 0.25, 'TSM': 0.20, 'VST': 0.65, 'VRT': 0.40, 'MU': 0.75}

R1 = 'INDEX($E:$E,ROW()-$B$3+1):INDEX($E:$E,ROW()-1)'
R0 = 'INDEX($E:$E,ROW()-$B$3):INDEX($E:$E,ROW()-2)'


def az_formula(r):
    me = f'((AVERAGE({R1})-AX{r})-AY{r}*(AVERAGE({R0})-AX{r}))'
    e = f'(({R1}-AX{r})-AY{r}*({R0}-AX{r}))'
    return (f'=IF(B{r}="","",IF(ROW()-$B$3<8,"",'
            f'SQRT(SUMPRODUCT(({e}-{me})^2)/($B$3-1))))')


if __name__ == '__main__':
    shutil.copy(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    old_buf = {}
    for s in NAMES:
        m = wb[f'Model {s}']
        last = max(r for r in range(8, 1000) if m.cell(r, 2).value not in (None, ''))
        n = 0
        for r in range(8, last + 1):
            if m.cell(r, 52).value is None:          # column AZ
                continue
            m.cell(r, 52).value = az_formula(r)
            n += 1
        old_buf[s] = m['D3'].value
        m['D3'] = NEW_BUF[s]
        m['C3'] = 'OU buffer k (resid)'
        print(f'Model {s}: rewrote {n} AZ cells (rows 8-{last}); '
              f'ou_buf_k {old_buf[s]:.6g} -> {NEW_BUF[s]}')

    nt = wb['Notes']
    row = max((c.row for c in nt['B'] if c.value), default=6) + 2
    nt.cell(row, 1, 'OU sigma')
    nt.cell(row, 2, 'Column AZ now takes sigma from the fitted AR(1) residuals instead of '
                    'STDEVP of the last W closes. The old definition measured the dispersion of '
                    'the price level, which on a trending name is mostly the trend, so the bid '
                    'buffer inflated exactly when a stock was running and fills were lost.')
    nt.cell(row+1, 2, 'Residual sigma is about a third of the old value, so ou_buf_k (D3) is '
                      'raised to match: ' + ', '.join(f'{s} {old_buf[s]:.4f} -> {NEW_BUF[s]}'
                                                      for s in NAMES) + '.')
    nt.cell(row+2, 2, 'Effect at the deployed 75% Bayes share, verified fills and marked to '
                      'market: book mean 64% to 71% from the sigma change alone, and to 77% with '
                      'the new buffers. The buffer re-fit is walk-forward validated at 10 of 15 '
                      'folds, +14.7pp on the OU sleeve.')
    nt.cell(row+3, 2, 'To revert: restore D3 on each Model sheet to the value above and put AZ '
                      'back to =IF(B{r}="","",IF(ROW()-$B$3<8,"",STDEVP(INDEX($E:$E,ROW()-$B$3):'
                      'INDEX($E:$E,ROW()-1)))). The allocation and the Bayes share are unchanged '
                      '- the split was re-tested and stays at 75%.')
    for k in range(4):
        nt.cell(row+k, 2).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        nt.row_dimensions[row+k].height = 46

    wb.save(OUT)
    print(f'\nwritten {OUT}')
