"""
Re-cast the Performance-tab plan for the RKLB -> AVGO swap (2 Sep 2026).

Final construction, settled the same day: the plan is cautious per-name
anchors plus the machine's two MEASURED mechanical effects, each taken
conservatively --

  anchors (captive, sheet convention, OOS haircuts)          53%/yr
  x gap-exit execution uplift (per name; AVGO x1.17)         66%/yr
  x pooling uplift at its WEAKER measured half (x1.12;
    captive book vs pooled book, identical frozen vectors:
    x1.23 full, x1.49 train, x1.12 test, exec-accurate)     ~74%/yr  <- C5

The pooled live-loop full-sample measurement (88.5%/yr exec) is the upper
reference, not the target -- it carries fit lookahead. Derivation:
verified_book_ranking.pdf, composition-change section.

The chart title drops the hard-coded '80%' so future re-casts are a one-cell
edit. A Notes changelog line records the change. Accepts a workbook whose C5
still holds 0.80 (pre-re-cast) or 0.66 (the interim anchors-x-execution plan
delivered earlier on 2 Sep, before the pooling tier was settled).

Usage: python wire_plan_update.py <in.xlsx> <out.xlsx>
"""
import copy
import sys

import openpyxl

NEW_PLAN = 0.74
OLD_PLANS = (0.80, 0.66)


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    pf, notes = wb['Performance'], wb['Notes']

    assert any(abs(pf['C5'].value - o) < 1e-9 for o in OLD_PLANS), pf['C5'].value
    pf['C5'] = NEW_PLAN

    ch = pf._charts[0]
    ch.title = 'Cumulative P&L — actual vs plan'

    style_a, style_b = notes['A33'], notes['B33']
    row = notes.max_row + 1
    ca = notes.cell(row=row, column=1, value='Plan re-cast (Performance C5)')
    ca.font = copy.copy(style_a.font)
    ca.alignment = copy.copy(style_a.alignment)
    cb = notes.cell(
        row=row, column=2,
        value=('Target return C5 set to 74%/yr on 2 Sep 2026 (RKLB -> AVGO swap; '
               'plan construction settled). Three tiers: per-name captive anchors '
               '53%/yr (sheet convention, OOS haircuts; AVGO anchor 51% replacing '
               'RKLB\'s 156%), x gap-exit execution uplift = 66%, x pooling uplift '
               'at its weaker measured half (x1.12; full-sample x1.23) = ~74%. '
               'Pooled live-loop measurement 72.4%/yr sheet / 88.5%/yr execution-'
               'accurate is the upper reference, not the target. Performance before '
               '2 Sep is judged against the old 79-80% plan, after it against 74%. '
               'Derivation: verified_book_ranking.pdf, composition-change section.'))
    cb.font = copy.copy(style_b.font)
    cb.alignment = copy.copy(style_b.alignment)
    wb.save(out_path)


def diff(in_path, out_path):
    a = openpyxl.load_workbook(in_path)
    b = openpyxl.load_workbook(out_path)
    changed = []
    for ws_name in b.sheetnames:
        wa, wb_ = a[ws_name], b[ws_name]
        for r in range(1, max(wa.max_row, wb_.max_row) + 1):
            for c in range(1, max(wa.max_column, wb_.max_column) + 1):
                va = wa.cell(row=r, column=c).value
                vb = wb_.cell(row=r, column=c).value
                if va != vb:
                    if (isinstance(va, float) and isinstance(vb, float)
                            and abs(va - vb) <= 1e-8 * max(abs(va), 1.0)):
                        continue
                    changed.append((ws_name, wb_.cell(row=r, column=c).coordinate))
    return changed


if __name__ == '__main__':
    src, dst = sys.argv[1], sys.argv[2]
    wire(src, dst)
    ch = diff(src, dst)
    outside = [x for x in ch if x != ('Performance', 'C5') and x[0] != 'Notes']
    print(f'changed cells: {len(ch)}; outside intended set: {outside or "NONE"}')
