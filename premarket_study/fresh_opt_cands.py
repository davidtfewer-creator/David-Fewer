"""
The fresh_opt protocol extended to the five candidates: GM, VLO, CF, NVDA, AVGO.

Identical blade to fresh_opt.py: verified same-day fills from the 5-minute bars,
residual OU sigma, fit on the first half only (boundary 2025-05-23), freeze, score
the tested half. Daily OHLC is derived from the 5-minute regular-hours bars (the
same basis the diversifier study used); the incumbents run off the workbook Query
sheet as before.

Reference vectors (the "deployed" role — full-sample fits, so their tested numbers
carry the same lookahead flavour as the incumbents' deployed rows):
  GM / VLO / CF   the diversifier-study vectors (HANDOVER section 6)
  NVDA            nvda_verified_bayesfloor.json
  AVGO            no daily vector exists — fitted here on the full sample with the
                  same optimiser (9-dim: the A space plus ou_W), flagged as such.

Variants per name, both trained on the first half only:
  A  8-dim (lam=1 exact degeneracy removed, ou_W frozen at reference);
     AVGO: 9-dim with ou_W searched, since its reference ou_W is not pre-existing.
  B  6-dim policy-only, filter pinned at the TRAIN-half MLE, ou_W frozen at
     reference; AVGO: 7-dim with ou_W searched.
"""
import collections
import datetime
import json
import math
import sys

import numpy as np
from scipy.optimize import differential_evolution

from engine import Params, run_model
from fresh_opt import (SPLIT, A_BOUNDS, A_POLICY, B_BOUNDS, B_POLICY,
                       a_params, a_x0, b_params, make_objective, seg, annualise)
from minute_index import build_index, make_checker
from rolling_mle import fit_window

CANDS = ['GM', 'VLO', 'CF', 'NVDA', 'AVGO']

REF = {
    'GM':   dict(lam=0.704472, phi_L=0.750723, psi=0.0805582, k=0.461711,
                 premium=0.00969721, peak_cap=0.0423717, ou_buf_k=1.06024,
                 ou_prem=0.0451112, ou_cap=0.0417662, ou_W=45),
    'VLO':  dict(lam=0.723393, phi_L=0.625222, psi=0.0200999, k=0.543763,
                 premium=0.0303035, peak_cap=0.0308569, ou_buf_k=1.17841,
                 ou_prem=0.0419514, ou_cap=0.0370821, ou_W=125),
    'CF':   dict(lam=0.736218, phi_L=0.106120, psi=0.0694359, k=2.03552,
                 premium=0.0465592, peak_cap=0.0286416, ou_buf_k=0.841237,
                 ou_prem=0.0590932, ou_cap=0.0831024, ou_W=66),
    'NVDA': dict(lam=0.7610628506573973, phi_L=0.2960566567669778,
                 psi=0.01749641179473716, k=1.5002333743247538,
                 premium=0.018810857338147094, peak_cap=0.005801033104238518,
                 ou_buf_k=0.7432317834252365, ou_prem=0.03737698844908682,
                 ou_cap=0.023213731637427307, ou_W=48),
    'AVGO': None,   # fitted on the full sample below
    'AMD':  None,   # Aug 2026 AI-candidate round: references fitted here, flagged
    'MRVL': None,
    'SMCI': None,
    'CEG':  None,
    'FCX':  None,   # Aug 2026 diversifier round (G1: AI-related, 50% gate)
    'NEM':  None,   # (G1: AI-related, 50% gate)
    'UAL':  None,   # (G1: AI-related, 50% gate)
    'LEN':  None,   # (G1: uncorrelated, 30% gate)
}

AW_BOUNDS = A_BOUNDS + [(30, 150)]              # + ou_W
BW_BOUNDS = B_BOUNDS + [(30, 150)]


def ref_params(s):
    r = REF[s]
    return Params(capital=1_000_000, comm=0.005, interest=0.0314, stop_days=50,
                  bayes_pct=0.5, years=2.2, **r)


def daily_from_5min(stock):
    """Daily RTH OHLC derived from the 5-minute index source file."""
    import openpyxl, os
    from minute_index import DIR, RTH_START, RTH_END
    path = os.path.join(DIR, f'{stock}_5min.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        dt, o, h, l, c = row[0], row[1], row[2], row[3], row[4]
        if not isinstance(dt, datetime.datetime) or o is None:
            continue
        t = dt.time()
        if t < RTH_START or t >= RTH_END:
            continue
        by[dt.date()].append((dt, float(o), float(h), float(l), float(c)))
    wb.close()
    dts, O, H, L, C = [], [], [], [], []
    for d in sorted(by):
        bars = sorted(by[d])
        dts.append(d)
        O.append(bars[0][1])
        H.append(max(b[2] for b in bars))
        L.append(min(b[3] for b in bars))
        C.append(bars[-1][4])
    return dts, O, H, L, C


def aw_params(vec, t):
    p = a_params(vec[:8], t)
    p.ou_W = int(round(vec[8]))
    return p


def bw_params(vec, t, mle):
    p = b_params(vec[:6], t, mle)
    p.ou_W = int(round(vec[6]))
    return p


def de(obj, bounds, x0=None):
    kw = dict(init='sobol', seed=42, maxiter=8, popsize=8, mutation=(0.5, 1.0),
              recombination=0.7, tol=1e-3, polish=False, updating='immediate',
              workers=1)
    if x0 is not None:
        kw['x0'] = x0
    return differential_evolution(obj, bounds, **kw)


def main(only=None):
    try:
        results = json.load(open('fresh_opt_cands.json'))
    except FileNotFoundError:
        results = {}
    for s in (only or CANDS):
        dts, O, H, L, C = daily_from_5min(s)
        chk = make_checker(s, dts, O)
        N = len(C)
        cut = next(i for i, d in enumerate(dts) if d >= SPLIT)
        trlo, trhi, telo, tehi = 0, cut - 1, cut, N - 1
        print(f'\n===== {s}  (N={N}, {dts[0]} to {dts[-1]}) =====', flush=True)

        # ---- reference vector (full-sample provenance)
        if REF[s] is None:
            # AVGO: fit the reference on the FULL sample, 9-dim, same optimiser.
            t0 = Params(capital=1_000_000, comm=0.005, interest=0.0314,
                        stop_days=50, bayes_pct=0.5, years=2.2, ou_W=80)
            _, allb = seg(dts, O, H, L, C, t0, chk, 0, N - 1)
            obj = make_objective(dts, O, H, L, C, t0, chk, 0, N - 1,
                                 max(8, int(0.5 * allb)),
                                 lambda v: aw_params(v, t0), A_POLICY)
            rr = de(obj, AW_BOUNDS)
            t = aw_params(rr.x, t0)
            print(f'  {s} reference fitted full-sample (flagged): '
                  f'phi_L {t.phi_L:.3f} psi {t.psi:.4f} k {t.k:.3f} ou_W {t.ou_W}',
                  flush=True)
        else:
            t = ref_params(s)

        ret_tr, buys_tr = seg(dts, O, H, L, C, t, chk, trlo, trhi)
        ret_te, buys_te = seg(dts, O, H, L, C, t, chk, telo, tehi)
        ret_f, buys_f = seg(dts, O, H, L, C, t, chk, 0, N - 1)
        print(f'  reference verified: full {annualise(ret_f, dts, 0, N-1)*100:6.1f}%/yr '
              f'({buys_f} buys) | train {annualise(ret_tr, dts, trlo, trhi)*100:6.1f}%/yr '
              f'| TEST {annualise(ret_te, dts, telo, tehi)*100:6.1f}%/yr ({buys_te} buys)',
              flush=True)
        res = dict(reference=dict(
            full=annualise(ret_f, dts, 0, N - 1), train=annualise(ret_tr, dts, trlo, trhi),
            test=annualise(ret_te, dts, telo, tehi), buys_test=buys_te, buys_full=buys_f))

        floor = max(4, int(0.4 * buys_tr))
        fixed_W = REF[s] is not None

        # ---- variant A
        if fixed_W:
            objA = make_objective(dts, O, H, L, C, t, chk, trlo, trhi, floor,
                                  lambda v: a_params(v, t), A_POLICY)
            rA = de(objA, A_BOUNDS, x0=a_x0(t))
            pA = a_params(rA.x, t)
        else:
            objA = make_objective(dts, O, H, L, C, t, chk, trlo, trhi, floor,
                                  lambda v: aw_params(v, t), A_POLICY)
            rA = de(objA, AW_BOUNDS)
            pA = aw_params(rA.x, t)
        a_tr, _ = seg(dts, O, H, L, C, pA, chk, trlo, trhi)
        a_te, a_b = seg(dts, O, H, L, C, pA, chk, telo, tehi)
        print(f'  A         : train {annualise(a_tr, dts, trlo, trhi)*100:6.1f}%/yr '
              f'| TEST {annualise(a_te, dts, telo, tehi)*100:6.1f}%/yr ({a_b} buys)', flush=True)
        res['A'] = dict(train=annualise(a_tr, dts, trlo, trhi),
                        test=annualise(a_te, dts, telo, tehi), buys_test=a_b,
                        vec=list(rA.x))

        # ---- variant B
        F = [H[i] - L[i] for i in range(N)]
        mle, _, _ = fit_window(np.array(C[:cut]), np.array(F[:cut]),
                               [np.array([0.1, 0.7, 0.005]),
                                np.array([t.lam, t.phi_L, max(t.psi, 1e-3)])])
        print(f'  B filter pinned at train-MLE lam/phi_L/psi = '
              f'{mle[0]:.3f}/{mle[1]:.3f}/{mle[2]:.4f}', flush=True)
        if fixed_W:
            objB = make_objective(dts, O, H, L, C, t, chk, trlo, trhi, floor,
                                  lambda v: b_params(v, t, mle), B_POLICY)
            rB = de(objB, B_BOUNDS,
                    x0=[t.k, t.premium, t.peak_cap, t.ou_buf_k, t.ou_prem, t.ou_cap])
            pB = b_params(rB.x, t, mle)
        else:
            objB = make_objective(dts, O, H, L, C, t, chk, trlo, trhi, floor,
                                  lambda v: bw_params(v, t, mle), B_POLICY)
            rB = de(objB, BW_BOUNDS)
            pB = bw_params(rB.x, t, mle)
        b_tr, _ = seg(dts, O, H, L, C, pB, chk, trlo, trhi)
        b_te, b_b = seg(dts, O, H, L, C, pB, chk, telo, tehi)
        print(f'  B         : train {annualise(b_tr, dts, trlo, trhi)*100:6.1f}%/yr '
              f'| TEST {annualise(b_te, dts, telo, tehi)*100:6.1f}%/yr ({b_b} buys)', flush=True)
        res['B'] = dict(train=annualise(b_tr, dts, trlo, trhi),
                        test=annualise(b_te, dts, telo, tehi), buys_test=b_b,
                        vec=list(rB.x), mle=[float(x) for x in mle])
        if REF[s] is None:
            res['reference']['fitted_here'] = True
            res['reference']['vec'] = list(rr.x)
        results[s] = res
        with open('fresh_opt_cands.json', 'w') as fh:
            json.dump(results, fh, indent=1, default=str)
        print(f'  wrote fresh_opt_cands.json ({len(results)} names)', flush=True)


if __name__ == '__main__':
    main(sys.argv[1:] or None)
