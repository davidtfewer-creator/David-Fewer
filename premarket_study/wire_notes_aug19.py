"""
Consolidated Notes-sheet explanation of the 19 August 2026 workbook updates
(user request): a titled section with practical how-to detail, complementing
the terse changelog lines the wiring scripts appended.
"""
import copy
import sys

import openpyxl

ENTRIES = [
    ('UPDATE — 19 AUGUST 2026', None),
    ('Gate Variables',
     'All gate settings live in Active Trading L5:N7. M6 = PM gate % (default 4%): how far a '
     'Buy @ may sit below the 9:00 pre-market before the order is withheld. M7 = ATH gate '
     'epsilon (default 0): how far below the all-time high every sale target must be '
     'achievable. Both are blue inputs; explanations sit beside them in column N.'),
    ('Discretionary trade log (cols T:AF)',
     'A second, manual-only blotter beside the main one, same thirteen columns. You type: # '
     '(T), Stock (U), Tranche/label (V), Buy date (W), Buy price (X), Shares (Y), and on exit '
     'Sell date (AA) and Sell price (AB). The sheet computes: Buy cost (Z, price*shares + buy '
     'fee B3), Proceeds (AC, less sell fee E3), Net P&L (AD), Week ending (AE) and Status (AF, '
     'OPEN/CLOSED from the dates). Closed discretionary P&L is added to the weekly summary '
     'Cumulative column (R) and its buy/sell events to Total trades (S); the weekly Realised '
     'P&L column (P) stays main-book-only so the model\'s own week remains visible. Rows '
     '42-241 are pre-wired; Scripts 1/2 never touch T:AF. Pairs with the Allocation '
     'discretionary carve-out (E5/E6): switch the flag on, trade here, sell, switch off.'),
    ('9:00 pre-market price + PM gate (col P + Allocation col H)',
     'Script 1 writes each stock\'s last pre-market print before 9:00 into P19:P36. Any sleeve '
     'whose Buy @ (col F) sits more than M6 below its col-P price is gated for the day: '
     'Allocation col H flags it and the free-weight formula (F25:F42) zeroes its allocation, '
     'so its cash pools into the other sleeves\' orders — same mechanism as the DMA gate. '
     'Fails open: a blank P cell or blank M6 means no gate. The rule is a veto only — never '
     'move a buy price toward the market instead (tested: re-pricing loses on every metric). '
     'Basis: verified pooled book, tested half 110.7% -> 118.4%/yr at the 4% threshold.'),
    ('ATH gate (Dashboard buys, epsilon = M7)',
     'Every Bayes and OU buy on the Dashboard (C4:C12, E4:E12) is now also capped at '
     'ATH*(1-eps) - prev close * that sleeve\'s premium, so no order can create a trade whose '
     'exit would need a print above ATH*(1-eps). eps = 0 guards at the exact ATH '
     '(recommended; measured ~free on verified fills). Blank counts as 0; set -1 to disable. '
     'Model-sheet history columns are untouched, so the sheet\'s historical stats are '
     'unchanged — the gate acts on live orders only.'),
    ('Traffic lights (cols N and O)',
     'Col N (200 DMA): green = close comfortably above the average, amber = within 5% above, '
     'red = at/below (the DMA gate is firing; fund and Qty go to 0). Col O (Sale % below '
     'ATH): green = target comfortably below the ATH, amber = within 2pp of epsilon, red = '
     'at/below epsilon (the ATH gate is binding on that order\'s buy price). Colours follow '
     'M7 live. Cell fills only; blank cells stay uncoloured.'),
]


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    n = wb['Notes']
    hdr_style = n['A2']
    lab_style, txt_style = n['A8'], n['B8']
    row = n.max_row + 2
    for label, text in ENTRIES:
        if text is None:
            c = n.cell(row=row, column=1, value=label)
            c.font = copy.copy(hdr_style.font)
            c.alignment = copy.copy(hdr_style.alignment)
        else:
            ca = n.cell(row=row, column=1, value=label)
            ca.font = copy.copy(lab_style.font)
            ca.alignment = copy.copy(lab_style.alignment)
            cb = n.cell(row=row, column=2, value=text)
            cb.font = copy.copy(txt_style.font)
            cb.alignment = copy.copy(txt_style.alignment)
        row += 1
    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)


if __name__ == '__main__':
    wire(sys.argv[1], sys.argv[2])
    wb = openpyxl.load_workbook(sys.argv[2])
    nn = wb['Notes']
    for r in range(31, nn.max_row + 1):
        a, b = nn.cell(row=r, column=1).value, nn.cell(row=r, column=2).value
        if a or b:
            print(r, '|', str(a)[:36], '|', str(b)[:60] if b else '')
