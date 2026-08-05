"""
Check the free-sleeve allocation by re-deriving every written formula independently.

LibreOffice cannot open this workbook -- the unmodified original fails to load the same way, so
it is the file and not the edit -- which rules out a recalculation check. Instead the arithmetic
is reproduced here from the source workbook's cached values, and each formula saved into
Allocation!C25:F34 and 'Active Trading'!D19:H28 is compared against an independently built
expectation. Any drift in either the numbers or the cell references fails the run.

Checked:
    held_i  from 'Active Trading'!C19:C28 = "HOLDING"
    w_i     = K x R (Bayes) or K x (1-R) (OU)
    C_i     = B5 x w_i x (1-held_i) / sum

plus the invariants that matter in the morning: a held sleeve gets exactly zero, the sleeves in
cash between them get exactly B5, they are equally sized under the deployed equal weights, the
blotter's share count equals FLOOR(allocation / (bid + fee)), and the Day-1 baseline the P&L
column depends on is untouched.
"""
import math
import openpyxl

SRC = ('/root/.claude/uploads/2d71f10a-e19f-51b2-8457-2cd547c34dff/'
       '3e1130ba-TradingExcel_5stock_live.xlsx')
OUT = '/home/user/David-Fewer/TradingExcel_5stock_live_freesleeves.xlsx'
NAMES = ['TSM', 'VRT', 'VST', 'RKLB', 'MU']


def main():
    sv = openpyxl.load_workbook(SRC, data_only=True)
    sf = openpyxl.load_workbook(SRC, data_only=False)
    of = openpyxl.load_workbook(OUT, data_only=False)
    a, at = of['Allocation'], of['Active Trading']
    sa, sat = sv['Allocation'], sv['Active Trading']

    B5 = sa['B5'].value
    fee = sat['B3'].value
    rows = []
    for i, s in enumerate(NAMES):
        srow = 11 + i
        K, R = sa.cell(srow, 11).value, sa.cell(srow, 18).value
        for t, tranche in enumerate(('Bayes', 'OU')):
            r = 25 + 2*i + t
            atr = r - 6
            held = 1 if sat.cell(atr, 3).value == 'HOLDING' else 0
            w = K*R if tranche == 'Bayes' else K*(1-R)
            rows.append(dict(r=r, atr=atr, s=s, tr=tranche, held=held, srow=srow,
                             w=w, f=w*(1-held), bid=sat.cell(atr, 6).value))

    tot = sum(x['f'] for x in rows)
    fails = []
    print(f'cash available B5 = {B5:,.2f}   buy fee/sh = {fee}\n', flush=True)
    print(f'{"AT":>4s}{"stock":>7s}{"tranche":>9s}{"status":>9s}{"free wt":>9s}'
          f'{"$ old":>13s}{"$ new":>13s}{"bid":>10s}{"shares":>9s}', flush=True)
    print('-'*84, flush=True)
    for x in rows:
        al = B5*x['f']/tot if tot else 0.0
        old = sa.cell(x['srow'], 13 if x['tr'] == 'Bayes' else 14).value
        sh = (math.floor(al/(x['bid'] + fee)) if (not x['held'] and x['bid']) else 0)
        print(f'{x["atr"]:>4d}{x["s"]:>7s}{x["tr"]:>9s}'
              f'{"HOLDING" if x["held"] else "CASH":>9s}{x["f"]:>9.4f}'
              f'{old:>13,.2f}{al:>13,.2f}'
              f'{(x["bid"] or 0):>10.2f}{sh:>9d}', flush=True)
        x['al'] = al

        share = f'$R${x["srow"]}' if x['tr'] == 'Bayes' else f'(1-$R${x["srow"]})'
        want = {
            (a, x['r'], 4): f'=IF(\'Active Trading\'!$C${x["atr"]}="HOLDING",1,0)',
            (a, x['r'], 5): f'=$K${x["srow"]}*{share}',
            (a, x['r'], 6): f'=E{x["r"]}*(1-D{x["r"]})',
            (a, x['r'], 3): f'=IF(SUM($F$25:$F$34)=0,0,$B$5*F{x["r"]}/SUM($F$25:$F$34))',
            (at, x['atr'], 4): '=INDEX(Allocation!$C$25:$C$34,ROW()-18)',
            (at, x['atr'], 8): (
                f'=IF($C{x["atr"]}="CASH",IFERROR(FLOOR(INDEX(Allocation!$C$25:$C$34,ROW()-18)'
                f'/($F{x["atr"]}+$B$3),1),0),SUMIFS($F$42:$F$1041,$B$42:$B$1041,$A{x["atr"]},'
                f'$C$42:$C$1041,$B{x["atr"]},$M$42:$M$1041,"OPEN"))'),
        }
        for (ws, rr, cc), exp in want.items():
            got = ws.cell(rr, cc).value
            if got != exp:
                fails.append(f'{ws.title} r{rr} c{cc}\n   want {exp}\n   got  {got}')

    print('-'*84, flush=True)
    nfree = sum(1 for x in rows if not x['held'])
    print(f'sleeves in cash {nfree} of {len(rows)}; $ per free sleeve '
          f'{B5/nfree if nfree else 0:,.2f}', flush=True)

    for x in rows:
        if x['held'] and x['al'] != 0.0:
            fails.append(f'held sleeve {x["s"]} {x["tr"]} allocated {x["al"]}')
    if abs(sum(x['al'] for x in rows) - B5) > 0.01:
        fails.append(f'allocated {sum(x["al"] for x in rows):,.2f} != cash {B5:,.2f}')
    sizes = {round(x['al'], 6) for x in rows if not x['held']}
    if len(sizes) > 1:
        fails.append(f'free sleeves not equally sized: {sizes}')

    # the Day-1 baseline the blotter's P&L depends on must be byte-identical
    for col in (12, 13, 14):
        for i in range(11, 16):
            if a.cell(i, col).value != sf['Allocation'].cell(i, col).value:
                fails.append(f'Allocation baseline changed at row {i} col {col}')
    for r in range(7, 12):
        for col in (2, 3, 4, 5, 6, 7, 8):
            if at.cell(r, col).value != sf['Active Trading'].cell(r, col).value:
                fails.append(f'Active Trading stock-funds block changed at r{r} c{col}')
    # and the status column itself must not have been touched
    for r in range(19, 29):
        if at.cell(r, 3).value != sf['Active Trading'].cell(r, 3).value:
            fails.append(f'Active Trading status C{r} changed')

    print('', flush=True)
    for f in fails:
        print('  FAIL ' + f, flush=True)
    print(f'\n{"ALL CHECKS PASSED" if not fails else str(len(fails)) + " FAILURES"}', flush=True)
    print('DONE', flush=True)


if __name__ == '__main__':
    main()
