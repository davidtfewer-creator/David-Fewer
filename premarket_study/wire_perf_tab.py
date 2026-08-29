"""
Two workbook modifications (user, 29 Aug 2026):

1. DISC P&L REWIRE. The discretionary log's Net P&L (AT col AD) currently feeds
   the Cumulative column (R) directly. Corrected: each closed discretionary
   trade's P&L joins the WEEKLY realised P&L (col P) of its week-ending (the
   log's own col AE), and Cumulative becomes a pure running sum of P.
     P42:P93: =IF($O{r}="","",SUMIFS(blotter K by week) + SUMIFS($AD$42:$AD$241,
               $AE$42:$AE$241,$O{r}))
     R42:R93: =IF($O{r}="","",SUM($P$42:$P{r}))
   (# trades col Q counts blotter trades only, as before; Total trades col S
   already counts the disc log separately — both untouched.)

2. PERFORMANCE TAB. New sheet "Performance" between Notes and Allocation:
   projected weekly P&L (yearly compounded target return, default 80%, starting
   capital $6,150,000 — both editable assumption cells) against actual weekly
   P&L (AT col P), with cumulative lines, a variance column and a line chart of
   cumulative projected vs cumulative actual. Rows mirror the 52 weekly rows
   (AT O42:O93) and fill automatically as the weekly log fills.

Layout-preserving: nothing inserted or moved inside existing sheets; the new
sheet only changes the tab order (scripts address sheets by name). Verified by
full-workbook cell diff (intended cells only) below.
"""
import copy
import sys

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

WROWS = range(42, 94)          # the 52 weekly rows in Active Trading


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    at = wb['Active Trading']
    notes = wb['Notes']
    assert wb.sheetnames[0] == 'Notes' and wb.sheetnames[1] == 'Allocation'
    assert 'Performance' not in wb.sheetnames

    # ---- 1. the rewire, with preconditions on every row touched
    for r in WROWS:
        p_old = at.cell(row=r, column=16).value
        r_old = at.cell(row=r, column=18).value
        assert p_old == f'=IF($O{r}="","",SUMIFS($K$42:$K$1041,$L$42:$L$1041,$O{r}))', (r, p_old)
        assert r_old == (f'=IF($O{r}="","",SUM($P$42:$P{r})'
                         f'+SUMIFS($AD$42:$AD$241,$AE$42:$AE$241,"<="&$O{r}))'), (r, r_old)
        at.cell(row=r, column=16,
                value=(f'=IF($O{r}="","",SUMIFS($K$42:$K$1041,$L$42:$L$1041,$O{r})'
                       f'+SUMIFS($AD$42:$AD$241,$AE$42:$AE$241,$O{r}))'))
        at.cell(row=r, column=18, value=f'=IF($O{r}="","",SUM($P$42:$P{r}))')

    # ---- 2. the Performance sheet, between Notes and Allocation
    pf = wb.create_sheet('Performance', 1)
    navy, gold = 'FF182644', 'FFC6A04A'
    hdr_fill = PatternFill(fill_type='solid', start_color=navy, end_color=navy)
    hdr_font = Font(bold=True, color='FFFFFFFF')
    title = pf.cell(row=1, column=1, value='PERFORMANCE — weekly realised P&L vs target')
    title.font = Font(bold=True, size=13, color=navy)
    pf.cell(row=2, column=1,
            value=('Actual = Active Trading col P (blotter + discretionary, by week ending). '
                   'Projected = starting capital compounding at the target rate, '
                   'week 1 aligned to the first logged week.')).font = Font(italic=True,
                                                                            color='FF5F5F5F')

    pf.cell(row=4, column=2, value='Starting capital').font = Font(bold=True)
    c = pf.cell(row=4, column=3, value=6150000)
    c.number_format = '$#,##0'
    pf.cell(row=5, column=2, value='Target return (p.a., compounded)').font = Font(bold=True)
    c = pf.cell(row=5, column=3, value=0.80)
    c.number_format = '0%'
    pf.cell(row=6, column=2, value='Weekly growth rate').font = Font(bold=True)
    c = pf.cell(row=6, column=3, value='=(1+$C$5)^(1/52)-1')
    c.number_format = '0.000%'

    heads = ['Week #', 'Week ending', 'Projected P&L', 'Actual P&L',
             'Cumulative projected', 'Cumulative actual', 'Actual vs projected (cum.)']
    for j, h in enumerate(heads, start=1):
        cell = pf.cell(row=8, column=j, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    widths = [8, 13, 15, 15, 19, 17, 22]
    for j, w in enumerate(widths, start=1):
        pf.column_dimensions[get_column_letter(j)].width = w

    for i, r_at in enumerate(WROWS):
        r = 9 + i
        pf.cell(row=r, column=1, value=f'=IF($B{r}="","",ROW()-8)')
        b = pf.cell(row=r, column=2,
                    value=f"=IF('Active Trading'!$O{r_at}=\"\",\"\",'Active Trading'!$O{r_at})")
        b.number_format = 'dd/mm/yyyy'
        pf.cell(row=r, column=3,
                value=f'=IF($B{r}="","",$C$4*(1+$C$6)^(ROW()-9)*$C$6)')
        pf.cell(row=r, column=4,
                value=f"=IF($B{r}=\"\",\"\",'Active Trading'!$P{r_at})")
        pf.cell(row=r, column=5, value=f'=IF($B{r}="","",SUM($C$9:$C{r}))')
        pf.cell(row=r, column=6, value=f'=IF($B{r}="","",SUM($D$9:$D{r}))')
        pf.cell(row=r, column=7, value=f'=IF($B{r}="","",$F{r}-$E{r})')
        for j in (3, 4, 5, 6, 7):
            pf.cell(row=r, column=j).number_format = '$#,##0'

    last = 9 + len(WROWS) - 1
    ch = LineChart()
    ch.title = 'Cumulative P&L — actual vs 80% target'
    ch.style = 2
    ch.height, ch.width = 9, 17
    data = Reference(pf, min_col=5, max_col=6, min_row=8, max_row=last)
    cats = Reference(pf, min_col=2, min_row=9, max_row=last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.y_axis.numFmt = '$#,##0'
    ch.series[0].graphicalProperties.line.solidFill = 'C6A04A'   # projected: gold
    ch.series[0].graphicalProperties.line.width = 20000
    ch.series[1].graphicalProperties.line.solidFill = '176E78'   # actual: teal
    ch.series[1].graphicalProperties.line.width = 26000
    pf.add_chart(ch, 'I4')

    # ---- Notes changelog
    style_hdr = notes['A32']
    style_a = notes['A33']
    style_b = notes['B33']
    row = notes.max_row + 2
    h = notes.cell(row=row, column=1, value='UPDATE — 29 AUGUST 2026')
    h.font = copy.copy(style_hdr.font)
    h.alignment = copy.copy(style_hdr.alignment)
    row += 1
    entries = [
        ('Discretionary P&L rewire',
         'A closed discretionary trade\'s Net P&L (col AD) now joins the WEEKLY realised P&L '
         '(col P) of its week ending (the log\'s col AE), and Cumulative (col R) is a pure '
         'running sum of col P. Previously AD fed col R directly, so weekly P&L and its '
         'cumulative could disagree. Note: a discretionary trade closed before the first '
         'blotter week in the log would predate row 42 and not be captured — close-outs '
         'should postdate the log\'s first week.'),
        ('Performance tab',
         'New sheet between Notes and Allocation: projected weekly P&L from a compounding '
         'target (assumptions in C4 starting capital $6,150,000 and C5 target 80%/yr, both '
         'editable) against actual weekly P&L (Active Trading col P), cumulative columns, '
         'variance, and a chart of cumulative actual vs target. Rows fill automatically as '
         'the weekly log fills. No script-read cells moved; scripts address sheets by name, '
         'so the new tab position is safe.'),
    ]
    for label, text in entries:
        ca = notes.cell(row=row, column=1, value=label)
        ca.font = copy.copy(style_a.font)
        ca.alignment = copy.copy(style_a.alignment)
        cb = notes.cell(row=row, column=2, value=text)
        cb.font = copy.copy(style_b.font)
        cb.alignment = copy.copy(style_b.alignment)
        row += 1

    wb.save(out_path)
    return out_path


def diff(in_path, out_path):
    a = openpyxl.load_workbook(in_path)
    b = openpyxl.load_workbook(out_path)
    changed, new_sheets = [], []
    for ws_name in b.sheetnames:
        if ws_name not in a.sheetnames:
            new_sheets.append(ws_name)
            continue
        wa, wb_ = a[ws_name], b[ws_name]
        for r in range(1, max(wa.max_row, wb_.max_row) + 1):
            for c in range(1, max(wa.max_column, wb_.max_column) + 1):
                va = wa.cell(row=r, column=c).value
                vb = wb_.cell(row=r, column=c).value
                if va != vb:
                    changed.append((ws_name, wb_.cell(row=r, column=c).coordinate, va, vb))
    return changed, new_sheets


if __name__ == '__main__':
    src, dst = sys.argv[1], sys.argv[2]
    wire(src, dst)
    ch, ns = diff(src, dst)
    print(f'new sheets: {ns}')
    print(f'changed cells in existing sheets: {len(ch)}')
    outside = [c for c in ch if not (c[0] == 'Active Trading'
                                     and c[1][0] in 'PR' and 42 <= int(c[1][1:]) <= 93)
               and c[0] != 'Notes']
    print('changes outside the intended set:', outside or 'NONE')
    for ws, coord, va, vb in ch[:4] + ch[-4:]:
        print(f'  {ws}!{coord}: {str(va)[:58]!r} -> {str(vb)[:70]!r}')
