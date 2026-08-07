"""
Book admission process, applied to new candidates end to end.

The gates, in the order they are cheap to fail:

  1. SCREEN -- no fitting required. Long-only cannot survive a net decliner, so buy-and-hold
     drift must be positive. A single session moving more than 40% disqualifies: violent gaps
     break limit fills, which is the mechanism that confirmed the MSTR rejection. Daily range is
     the harvest fuel and is reported, as is dollar volume, since the book must be able to get
     filled.

  2. THE BLADE -- fit on the FIRST half only, freeze, score the second on verified fills. This
     is the test that eliminated PLTR (365% fitted, 3.1% tested) and it is the number that
     decides admission. Nothing here is scored on data the parameters have seen.

  3. DEPLOYABLE VECTOR -- a full-sample robust fit, reported for completeness and as the vector
     that would actually be run. Its full-sample return is NOT evidence; the tested half is.

  4. PORTFOLIO FIT -- correlation of share returns to the nine-name book and beta to the AI
     factor. A high-return name that simply adds more of the same trade is a return name, not a
     diversifier, and the book is already concentrated.

Parameters are fitted with the repository's own BOUNDS and robust objective from
optimise_candidates, the same contract that produced GM, VLO and CF, so these candidates are
directly comparable to the ones already accepted. Residual OU sigma throughout, matching the
live book.

Three of these names -- FSLR, DVN and COIN -- were rejected in earlier work on the uncorrected
engine. Standing policy is that a material engine change triggers a re-test, so this is that
re-test rather than a fresh look.

Run:  python3 admit_candidates.py [screen|full]
"""
import collections
import datetime
import os
import sys

import numpy as np
import openpyxl

import ramp_premium as R
from engine import Params, run_model
from optimise_candidates import BOUNDS, NAMES, POLICY, PERTURB, mp

UP = '/root/.claude/uploads/822d405e-f99b-5b59-9c6b-87e725054402'
LIVE = f'{UP}/8d17afe4-TradingExcel_5stock_live.xlsx'
CACHE = '/home/user/David-Fewer/premarket_study/admit_cache.pkl'

CAND = {
    'FSLR': f'{UP}/c03a2e5f-FSLR_5min_Apr2024Aug2026.xlsx',
    'AMD':  f'{UP}/5125adc2-AMD_5min_Apr2024Aug2026.xlsx',
    'ARM':  f'{UP}/17358b4d-ARM_5min_Apr2024Aug2026.xlsx',
    'COIN': f'{UP}/c5667fad-COIN_5min_Apr2024Aug2026.xlsx',
    'DVN':  f'{UP}/0e0052e1-DVN_5min_Apr2024Aug2026.xlsx',
    'HOOD': f'{UP}/0c7bb0be-HOOD_5min_Apr2024Aug2026.xlsx',
    'OXY':  f'{UP}/69c0329d-OXY_5min_Apr2024Aug2026.xlsx',
    'SPOT': f'{UP}/70dfce7a-SPOT_5min_Apr2024Aug2026.xlsx',
    'RBLX': f'{UP}/cf6756ae-RBLX_5min_Apr2024Aug2026.xlsx',
}
BOOK = ('RKLB', 'TSM', 'VST', 'VRT', 'MU')       # live five, for the correlation gate
AI = ('TSM', 'VRT', 'VST', 'MU')                 # AI factor, per ai_concentration.py
RTH0, RTH1 = datetime.time(9, 30), datetime.time(16, 0)
CAPITAL = 1_000_000


def five_min(path):
    """Regular-hours 5-minute bars -> daily OHLC + dollar volume + same-day fill index."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by = collections.defaultdict(list)
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        dt, o, h, l, c = row[0], row[1], row[2], row[3], row[4]
        v = row[5] if len(row) > 5 else 0
        if dt is None or o is None or not isinstance(dt, datetime.datetime):
            continue
        if not (RTH0 <= dt.time() < RTH1):
            continue
        by[dt.date()].append((dt, o, h, l, c, v or 0))
    wb.close()
    d, O, H, L, C, DV = [], [], [], [], [], []
    idx = {}
    for day in sorted(by):
        bars = sorted(by[day], key=lambda x: x[0])
        d.append(day)
        O.append(bars[0][1]); H.append(max(b[2] for b in bars))
        L.append(min(b[3] for b in bars)); C.append(bars[-1][4])
        DV.append(sum(b[5] * b[4] for b in bars))
        hi = np.array([b[2] for b in bars], dtype=float)
        lo = np.array([b[3] for b in bars], dtype=float)
        idx[day] = (lo, np.maximum.accumulate(hi[::-1])[::-1])
    return (d, O, H, L, C), DV, idx


def screen(name, bars, dv):
    d, O, H, L, C = bars
    C = np.array(C, dtype=float)
    drift = C[-1] / C[0] - 1.0
    yrs = (d[-1] - d[0]).days / 365.25
    ann = (C[-1] / C[0]) ** (1 / yrs) - 1
    rng = float(np.mean((np.array(H) - np.array(L)) / C))
    step = np.abs(C[1:] / C[:-1] - 1.0)
    worst = float(step.max())
    med_dv = float(np.median(dv))
    fails = []
    if drift <= 0:
        fails.append('net decliner')
    if worst > 0.40:
        fails.append(f'{worst:.0%} single-day move')
    if med_dv < 20e6:
        fails.append('thin volume')
    return dict(name=name, drift=drift, ann=ann, rng=rng, worst=worst, dv=med_dv,
                pass_=not fails, fails=fails, sessions=len(d))


def seg(bars, chk, vec, t, lo, hi):
    """One evaluation on the SAME engine the candidate is scored on."""
    d, O, H, L, C = bars
    r = run_model(d, O, H, L, C, mp(vec, t), ou_sigma='resid',
                  same_day_exit=chk, collect=True)
    fr = r.frames
    eq = fr['equity']
    buys = sum(fr['t1']['Z'][lo:hi + 1]) + sum(fr['t2']['Z'][lo:hi + 1])
    return (eq[hi] / eq[lo] - 1.0 if eq[lo] > 0 else -1.0), buys


def robust(bars, chk, vec, t, lo, hi, floor):
    """0.5*base + 0.5*mean(+/-3% on the policy params), with a minimum-trade floor.

    Fitting MUST use the same engine configuration as scoring. Inheriting
    optimise_candidates.robust silently fits on level sigma with the optimistic same-day rule
    while the candidate is scored on residual sigma with verified fills -- the parameters then
    optimise a book of same-day round trips that the 5-minute bars refuse, and the fitted half
    comes back NEGATIVE. That mis-specification produced the first run of this script.
    """
    def one(v):
        rr, b = seg(bars, chk, v, t, lo, hi)
        return -5.0 + b * 1e-3 if b < floor else rr
    base = one(vec)
    s = []
    for i in POLICY:
        for f in PERTURB:
            v = list(vec)
            v[i] = min(max(v[i] * f, BOUNDS[i][0]), BOUNDS[i][1])
            s.append(one(v))
    return 0.5 * base + 0.5 * sum(s) / len(s)


def fit(bars, chk, t, lo, hi, floor, seed=42, maxiter=6, popsize=6):
    from scipy.optimize import differential_evolution
    res = differential_evolution(
        lambda v: -robust(bars, chk, v, t, lo, hi, floor), BOUNDS,
        maxiter=maxiter, popsize=popsize, seed=seed, tol=0.01,
        mutation=(0.5, 1.0), recombination=0.7, polish=False,
        init='sobol', workers=1)
    return list(res.x)


def score(bars, chk, vec, t, lo, hi):
    d, O, H, L, C = bars
    p = mp(vec, t)
    r = run_model(d, O, H, L, C, p, ou_sigma='resid', same_day_exit=chk, collect=True)
    eq = r.frames['equity']
    yrs = (d[hi] - d[lo]).days / 365.25
    ret = (eq[hi] / eq[lo]) ** (1 / yrs) - 1 if eq[lo] > 0 and yrs > 0 else float('nan')
    buys = sum(r.frames['t1']['Z'][lo:hi + 1]) + sum(r.frames['t2']['Z'][lo:hi + 1])
    peak, dd = -1e30, 0.0
    for e in eq[lo:hi + 1]:
        peak = max(peak, e)
        dd = max(dd, (peak - e) / peak) if peak > 0 else dd
    return ret, buys / max(yrs, 1e-9), dd, eq


def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else 'full'
    loaded = {}
    print('loading 5-minute data...', flush=True)
    for n, p in CAND.items():
        bars, dv, idx = five_min(p)
        loaded[n] = (bars, dv, idx)
        print(f'  {n}: {len(bars[0])} sessions {bars[0][0]} .. {bars[0][-1]}', flush=True)

    print('\n=== GATE 1: screen (no fitting) ===\n')
    print(f"{'name':6s} {'B&H total':>10s} {'B&H p.a.':>9s} {'avg range':>10s} "
          f"{'worst day':>10s} {'median $vol':>13s}  verdict")
    scr = {}
    for n in CAND:
        s = screen(n, loaded[n][0], loaded[n][1])
        scr[n] = s
        v = 'PASS' if s['pass_'] else 'FAIL: ' + ', '.join(s['fails'])
        print(f"{n:6s} {100*s['drift']:9.1f}% {100*s['ann']:8.1f}% {100*s['rng']:9.2f}% "
              f"{100*s['worst']:9.1f}% {s['dv']/1e6:12.0f}m  {v}")
    if mode == 'screen':
        return

    survivors = [n for n in CAND if scr[n]['pass_']]
    print(f'\n{len(survivors)} of {len(CAND)} pass the screen: {", ".join(survivors)}')
    print('(names failing the screen are still fitted below, so the rejection is evidenced '
          'rather than asserted)\n')

    print('=== GATE 2/3: fit on the first half, freeze, score the tested half ===')
    print(f'    split {R.SPLIT}; verified fills; residual OU sigma\n', flush=True)
    t0 = Params(capital=CAPITAL, years=1.0)
    out = {}
    for n in CAND:
        bars, dv, idx = loaded[n]
        d = bars[0]
        iS = next(k for k, x in enumerate(d) if x >= R.SPLIT)
        chk = R.make_checker(idx, d, bars[1])
        half = fit(bars, chk, t0, 0, iS, floor=8)
        full = fit(bars, chk, t0, 0, len(d) - 1, floor=15)
        f_fit, f_buys, f_dd, _ = score(bars, chk, half, t0, 0, iS)
        f_tst, t_buys, t_dd, _ = score(bars, chk, half, t0, iS, len(d) - 1)
        a_full, a_buys, a_dd, eqf = score(bars, chk, full, t0, 0, len(d) - 1)
        out[n] = dict(half=half, full=full, fitted=f_fit, tested=f_tst,
                      full_ret=a_full, buys=a_buys, dd=a_dd, eq=eqf, bars=bars)
        print(f'  {n:6s} fitted {100*f_fit:8.1f}%   tested {100*f_tst:8.1f}%   '
              f'full-sample fit {100*a_full:8.1f}%   {a_buys:5.0f} buys/yr   '
              f'DD {100*a_dd:5.1f}%', flush=True)

    print('\n=== GATE 4: portfolio fit ===\n')
    px = {}
    for n in BOOK:
        dd, _o, _h, _l, cc = R.load_feed(n, path=LIVE)
        px[n] = (dd, cc)
    for n in CAND:
        b = loaded[n][0]
        px[n] = (b[0], b[4])
    common = sorted(set.intersection(*[set(v[0]) for v in px.values()]))
    M = {}
    for n, (dd, cc) in px.items():
        ix = {t: k for k, t in enumerate(dd)}
        s = np.array([cc[ix[t]] for t in common], dtype=float)
        M[n] = s[1:] / s[:-1] - 1
    fac = np.mean([M[n] for n in AI], axis=0)
    print(f"{'name':6s} {'corr to book':>13s} {'beta to AI':>11s} {'max pair':>10s}")
    for n in CAND:
        cs = [np.corrcoef(M[n], M[b])[0, 1] for b in BOOK]
        beta = np.cov(M[n], fac)[0, 1] / np.var(fac)
        out[n]['corr'] = float(np.mean(cs)); out[n]['beta'] = float(beta)
        out[n]['maxpair'] = float(np.max(cs))
        print(f'{n:6s} {np.mean(cs):13.3f} {beta:11.3f} {np.max(cs):10.3f}')

    print('\n=== VERDICT ===\n')
    print(f"{'name':6s} {'screen':6s} {'tested':>9s} {'corr':>7s} {'beta':>7s} "
          f"{'buys/yr':>8s}  decision")
    for n in CAND:
        o = out[n]
        s = scr[n]
        why = []
        if not s['pass_']:
            why.append('screen: ' + '; '.join(s['fails']))
        if o['tested'] < 0.15:
            why.append(f"tested {100*o['tested']:.0f}% below the 15% bar")
        if o['corr'] > 0.45:
            why.append(f"corr {o['corr']:.2f} - return name, not a diversifier")
        if o['buys'] < 12:
            why.append(f"only {o['buys']:.0f} buys/yr - thin evidence")
        dec = 'ADMIT' if not why else 'REJECT'
        print(f"{n:6s} {'PASS' if s['pass_'] else 'FAIL':6s} {100*o['tested']:8.1f}% "
              f"{o['corr']:7.3f} {o['beta']:7.3f} {o['buys']:8.0f}  {dec}"
              + (f"  ({'; '.join(why)})" if why else ''))

    print('\nfor reference, names already accepted on the same basis:')
    print('  GM  tested 56.5%  corr 0.21  |  CF  tested 54.7%  corr -0.03  |  '
          'VLO tested 137.9% corr 0.11 (dropped: 11.4/137.9 half split)')
    print('  previously rejected on the UNCORRECTED engine: '
          'MRNA -1.0%, OXY 4.1%, FSLR 2.9%, DVN 15.1%, COIN 0.1%')


if __name__ == '__main__':
    main()
