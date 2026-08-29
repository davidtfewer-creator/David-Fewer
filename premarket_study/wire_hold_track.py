"""
Add the average-hold tracking block to the Performance tab (user, 29 Aug 2026;
re-placed same day at the user's request).

Table at Performance!I7:K19, BESIDE the weekly return table (headers on row 8,
aligned with the return table's); the cumulative-P&L chart moves down to I22.
Per name: historical average hold (the pooled back-test reference, verified
fills Apr 2024 - Aug 2026, HANDOVER 3.27 diagnostic run; calendar days,
same-day round trips = 0) vs the LIVE average hold computed from the blotter's
CLOSED trades (sell date minus buy date, via SUMIFS/COUNTIFS on Active Trading
B/D/H/M; open positions excluded until they close; the discretionary log is
deliberately excluded — this tracks the model book). A BOOK aggregate row
closes the table.

Rationale (ops): hold length is a low-noise drift indicator — CF creeping from
~17 toward 30+ days, or the AI names' same-day recycling fading, flags a
harvest change well before weekly returns can say anything honest.

Append-only in cells (columns I:K rows 7-19 + a Notes line); the only other
change is the chart anchor. Cell-diff verified.
"""
import copy
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HIST = [('TSM', 4.2), ('VRT', 4.4), ('VST', 4.4), ('RKLB', 3.1), ('MU', 6.5),
        ('GM', 6.2), ('VLO', 14.1), ('CF', 16.7), ('MRVL', 7.3)]
BOOK_HIST = 5.9
AT = "'Active Trading'"
RNG = lambda col: f"{AT}!${col}$42:${col}$1041"


def live_formula(name_cell):
    cnt = f'COUNTIFS({RNG("B")},{name_cell},{RNG("M")},"CLOSED")'
    return (f'=IF({cnt}=0,"—",'
            f'(SUMIFS({RNG("H")},{RNG("B")},{name_cell},{RNG("M")},"CLOSED")'
            f'-SUMIFS({RNG("D")},{RNG("B")},{name_cell},{RNG("M")},"CLOSED"))'
            f'/{cnt})')


def wire(in_path, out_path):
    wb = openpyxl.load_workbook(in_path)
    pf = wb['Performance']
    notes = wb['Notes']
    assert pf.max_row == 60 and pf.max_column == 7, (pf.max_row, pf.max_column)
    assert len(pf._charts) == 1
    pf._charts[0].anchor = 'I22'

    navy = 'FF182644'
    hdr_fill = PatternFill(fill_type='solid', start_color=navy, end_color=navy)
    hdr_font = Font(bold=True, color='FFFFFFFF')

    t = pf.cell(row=7, column=9, value='AVERAGE HOLD (calendar days)')
    t.font = Font(bold=True, size=11, color=navy)
    for j, h in enumerate(['Stock', 'Historical avg', 'Live avg'], start=9):
        c = pf.cell(row=8, column=j, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for i, (nm, hv) in enumerate(HIST):
        r = 9 + i
        pf.cell(row=r, column=9, value=nm).font = Font(bold=True)
        c = pf.cell(row=r, column=10, value=hv)
        c.number_format = '0.0'
        c = pf.cell(row=r, column=11, value=live_formula(f'$I{r}'))
        c.number_format = '0.0'
    r = 18
    pf.cell(row=r, column=9, value='BOOK').font = Font(bold=True, color=navy)
    c = pf.cell(row=r, column=10, value=BOOK_HIST)
    c.number_format = '0.0'
    cnt = f'COUNTIFS({RNG("M")},"CLOSED")'
    c = pf.cell(row=r, column=11,
                value=(f'=IF({cnt}=0,"—",(SUMIFS({RNG("H")},{RNG("M")},"CLOSED")'
                       f'-SUMIFS({RNG("D")},{RNG("M")},"CLOSED"))/{cnt})'))
    c.number_format = '0.0'
    pf.cell(row=19, column=9,
            value=('Same-day = 0; open positions and the discretionary log excluded. '
                   'Historical: pooled back-test, verified fills. A drifting hold flags a '
                   'harvest change before returns can — read monthly.')).font = \
        Font(italic=True, size=9, color='FF5F5F5F')
    for col, w in (('I', 8), ('J', 13), ('K', 12)):
        pf.column_dimensions[col].width = w

    style_a = notes['A33']
    style_b = notes['B33']
    row = notes.max_row + 1
    ca = notes.cell(row=row, column=1, value='Hold tracking (Performance tab)')
    ca.font = copy.copy(style_a.font)
    ca.alignment = copy.copy(style_a.alignment)
    cb = notes.cell(row=row, column=2,
                    value=('Performance!I7:K19, beside the weekly return table: per-stock '
                           'average hold, historical (back-test reference) vs live (closed '
                           'blotter trades, sell minus buy dates). Same-day = 0; open '
                           'positions and discretionary log excluded. The cumulative chart '
                           'sits below it. A drifting hold length flags a harvest change '
                           'before returns can — review monthly.'))
    cb.font = copy.copy(style_b.font)
    cb.alignment = copy.copy(style_b.alignment)

    wb.save(out_path)
    return out_path


def diff(in_path, out_path):
    a = openpyxl.load_workbook(in_path)
    b = openpyxl.load_workbook(out_path)
    changed = []
    for ws_name in b.sheetnames:
        wa, wb_ = a[ws_name], b[ws_name]
        for r in range(1, max(wa.max_row, wb_.max_row) + 1):
            for c in range(1, max(wa.max_column, wb_.max_column) + 1):
                va = wa.cell(row=r, column=c).value
                vb = wb_.cell(row=r, column=c).value
                if va != vb:
                    changed.append((ws_name, wb_.cell(row=r, column=c).coordinate, va, vb))
    return changed


if __name__ == '__main__':
    src, dst = sys.argv[1], sys.argv[2]
    wire(src, dst)
    ch = diff(src, dst)
    outside = [c for c in ch if not (c[0] == 'Performance' and c[1][0] in 'IJK')
               and c[0] != 'Notes']
    print(f'changed cells: {len(ch)}; outside intended set: {outside or "NONE"}')
